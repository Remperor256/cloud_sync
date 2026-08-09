"""
Tenant Monitoring & Management — Web Edition (single-file build)
==================================================================
A phone-friendly web app that shares data with the original Tkinter desktop
app, synced entirely through the cloud service (see CLOUD_MODE below) --
there is no LAN/local-network pairing path in this build.

Run:
    pip install flask openpyxl reportlab qrcode[pil]
    python app.py

Deploy this file with CLOUD_MODE=1 and a DATABASE_URL, then pair phones to
it from the desktop app's Settings -> Connect Phone (which shows the QR
code for this deployed service directly).

Everything -- backend logic, the REST API, and the mobile frontend -- lives
in this one file on purpose, so it's a single thing to copy/share.
"""
import os
import re
import io
import copy
import json
import time
import shutil
import hashlib
import secrets
import calendar
import threading
import webbrowser
from urllib.parse import quote
from datetime import datetime, date, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, HRFlowable)

from flask import Flask, request, jsonify, session, send_file, Response, g
import base64

# Created here (rather than further down) because some decorators/routes
# below (e.g. @app.before_request for the CLOUD_MODE gate) reference `app`
# before the file reaches the "Flask app + REST API" section.
app = Flask(__name__)


@app.errorhandler(Exception)
def _api_error_handler(err):
    """Without this, any unhandled exception in a route (a bug, a bad
    Postgres value, whatever) falls through to Flask's default error
    page -- plain HTML, not JSON. The frontend's api() helper can't
    parse that as JSON, so it silently swallows the real cause and
    just shows the generic 'Request failed.' toast, which is exactly
    what makes bugs like this hard to track down from a bug report
    alone. Logging the real traceback server-side and returning it as
    JSON means the *next* time this fires, the toast (and the server
    log) shows what actually broke instead of a dead end."""
    import traceback
    traceback.print_exc()
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": f"{type(err).__name__}: {err}"}), 500
    raise err

# Same brand icon used by the Tkinter desktop app (updates.py's
# APP_ICON_256_B64 house glyph), lifted onto a solid white background at
# 512px so the web/PWA icon matches the desktop app icon exactly instead
# of using its own separately-derived version. Solid (non-transparent)
# background is kept deliberately -- a mostly-transparent icon is why
# installs used to show a plain letter-avatar fallback instead of the
# real icon on the home screen. Served at native size and downscaled by
# the browser for smaller manifest sizes (192px etc).
APP_ICON_512_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAIAAAB7GkOtAABdV0lEQVR42u39Z5Ndx5UveK+1Mvfe55zyDih4EKRoRO9EUd50S2qp"
    "+9658cTzPDER82I+Sn+a29N9Z9pLYlOOhChaGIKgd3BVQHlfx+yduda8yHOKoAxFoQ5AgPz/oiMENosF1EZV/jNzZ67FZkYAAPDF"
    "I3gEAAAIAAAAQAAAAAACAAAAEAAAAIAAAAAABAAAACAAAAAAAQAAAAgAAABAAAAAAAIAAAAQAAAAgAAAAAAEAAAAIAAAAAABAAAA"
    "CAAAAEAAAAAAAgAAABAAAACAAAAAAAQAAAAgAAAAEAAAAIAAAAAABAAAACAAAAAAAQAAAAgAAABAAAAAAAIAAAAQAAAAgAAAAAAE"
    "AAAAIAAAAAABAAAACAAAAEAAAAAAAgAAABAAAACAAAAAAAQAAAAgAAAAAAEAAAAIAAAAQAAAACAAAAAAAQAAAAgAAABAAAAAAAIA"
    "AAAQAAAAgAAAAAAEAAAAIAAAAAABAAAACAAAAEAAAAAAAgAAABAAAACAAAAAAAQAAAAgAAAAAAEAAAAIAAAAQAAAAAACAAAAEAAA"
    "AIAAAAAABAAAAAIAAAAQAAAAgAAAAAAEAAAAIAAAAAABAAAACAAAAEAAAAAAAgAAABAAAACAAAAAAAQAAAAgAAAAAAEAAAAIAAAA"
    "QAAAAAACAAAAEAAAAIAAAAAABAAAACAAAAAAAQAAAAgAAAAEAAAAIAAAAAABAAAACAAAAEAAAAAAAgAAABAAAACAAAAAAAQAAAAg"
    "AAAAAAEAAAAIAAAAQAAAAAACAAAAEAAAAIAAAAAABAAAACAAAAAAAQAAAAgAAABAAAAAAAIAAAAQAAAAgAAAAEAAAAAAAgAAABAA"
    "AACAAAAAAAQAAAAgAAAAAAEAAAAIAAAAQAAAAAACAAAAEAAAAIAAAAAABAAAACAAAAAAAQAAAAgAAABAAAAAAAIAAAAQAAAAgAAA"
    "AAAEAAAAIAAAAAABAACAAAAAAAQAAAAgAAAAAAEAAAAIAAAA+JzweAQA14mZdVSJqOZc+sdTS/P/a/bc8+vLBxpD/+eh2384fTB9"
    "ZDCrTDMWz4znBggAgFubmhFRLiLMRLRVlZeam29uri1UZSvqRijf3Vrfv1o/1BgcyLJMxLEzMyMiIoQA3Bjc+5YDgP7N/YkqjcLs"
    "WYhovdM5vbr48vL8qxsr56tyrdMs2B+qDTw0MvaV0ckHxyYODg4zUTSLqsLsBXuzgBUAwC039JsRMxE5Fsesaq0YTq8tPzU/c3x5"
    "bk4Di9MqtMQ2yu0PLq9f6bQ2Y3iS6GBj0IuIiBJhHQAIAIBbjxKVIWTOpd38pVbzleWFX6/Mv7K5PE/K4rTZslCRcybczrPTrc3t"
    "+VBp/ObkvqNDw8xMZpUpmWUsjFcCgAAAuDVGfzNm9iye2cxaMZ5ZW/7lwuxLm2vzWjKRbW75KkaOHM0F1ZquDcjJ9eW6+IxFmA8N"
    "DjlmRxxI03oCGQAIAICbmpkpUaUxE5c5IaKFVvPVtaVfLs29vLU2FzpERmXJaipmJN1tnlBpm83511ub5YJWpt/m/UcGh4nIs1Sq"
    "apYxC14JAAIA4KYd/Y2IyDyzYzaiZqhOrS49PX/ppY21+dAhM95qsqoKKTMRmVGlUaqYayxZloZkdX3BM+XizOzQ4JBjyURKjXi8"
    "gACAm2u8S7/A7gQRWTq9Q5QLMwsRvbe++ura8nMrcyc3lufaHRbhsiQzYzUjErfzH5KZRhVhraros9e2N8K8rXTaj41P3jU8NlLU"
    "cnFBNZo6xioAEAAAN9non/b9hYiJiejS9uazi1d+MT/z9vbmtkaKSs0WEUUyEkdmdNXB68isZqzKW83oZLVee3F5e7HdWinbyvx4"
    "tseJOJFKlZgM54IAAQCf7Xhn1t3wSOcdmch9Ieem6SkoUTCtiXNEavbm6vLzy/O/WZ57Y3ur7cU6QUKV3uQSEzHT7127YTZmNmOL"
    "LphWldaKD9rbtO7Y+xj07pGxiXo9F4lEVYxexGHVBQgA+AwHvqt//YXdBbJeHHJvXn5he+PpuUtPz8+cC50gbO3SWm0l053/RvUP"
    "nyYRafpfi9xqaVXR4OCHZTMsz2+3O50QvrnvQOYcEymZmjJ1rxYDIADghk54zewP5/vafQNqX5DDKulRRDNhzkSIKKi+t7H226Ur"
    "zy7Nv9PckoE6tzsSKksB4ZiY/8jof/XnFGFVSRtEVRXz/FxzM0bNnK8V/u6hsfFave68GQVTIxK8gAEEANzICa+qMvPOFtDOcK+q"
    "RMS9D/jcD0xp50fJyChtyJzf3PjPmfO/XrryYXubMx+bLdepLD0Xod/b9/+Tn5Y5MpOa65TWKbnRmAnt364vVBTaVfjm9MFMhJii"
    "GqXRH1cEAAEA15uqErMReeeIKMb4zszM+fn5zLmj09MHJyfrRZE+sgqBmZ18bu+vGlEayx2zZ0dEa2V7trV9fGHu+MrC21vrUiuk"
    "0+FQpfM+JvJn5/4fYTYRUqUYHbGVpdaKi80tMhVicu7uodF99UYqLFqpChEuCsMuub//+7/HU4BP2O6IKQB6mz+Xl5b+/aWX/un4"
    "8Vfee6/Z6YwNDk6PjaWP7ITQG8o+n+uAaGlDP73T5fVO+/Tq4r9ePv/LhZkL7VZksqqkqiQzJTbeSY2/kIgRkaqVwZzbjLpQdhaa"
    "21F1b602mOVEFFRTxSFGBgBWAHC95v7pu0SEiKLqhfn5Z8+e/eWrr/7uzTeNebssRcTMDk1NTQwPN4qCiDS9Lfh8bQd1jz+ReZb0"
    "1nex1Ty9snh8df6Zpbl5DcYkMVoVKJqKGPf2/f+i8T+9E2YmIh9NSLVTxiKf12pzfSmQ1Zw8OTk9XR8onCOiYBrVnAgyALACgP7P"
    "/dOOv3OOiC4tLPzbCy/847PPvjkzU6mKc6ubm/Orqwtra6a6b3KynncnpzFGI2Lmz815lZDeb1C3PNtWVT6/cPmfZ8+9vLm2GCtV"
    "s7LDVWlqxmx9OK9vRmbpynCI5H3leLWq5lpbZRWma/XRokZEgTR2FwLEuCQAWAFAv0b/NP3PvCeiEOPF+fnjr7/+9OnTp86dY5HM"
    "jFVbnc7rFy+ubG62Oh1x7rEvfWlieLie55waYPXOjN7S64A0gxfmnQP4y+3WmdWlZ5bmX1hZbDfqplpUoQrBVM3YvCfdeVlwrYM/"
    "kTlnZqLqjbTTYco3czm1vhxjHK3VCu8n8lruHHFvyUWE18KAFQDseufHLMaoZsSc9v0vzM//24sv/q/f/vbNmZkqRo6RYpQUEiLb"
    "7fby5ub5K1dW1tfreT45PJxiI8YYYrT0rvLWHJhSXxc12tlmWWu3fjs/+2+zF17ZWlvXYFGt06EQTE27Lwe4n7+/kQmRKqkqC2X5"
    "elmudTrtqhzPiolaPf0hQy9scEUAsAKAXc3905x9Z+4/s7j4zNmzPz9x4uQHH7CINxOiYGZEWdoZZ55bX59bXl5cXQ1EVYz3Hjky"
    "OjjonfPOfXRt+FZbCuzM/VNXr0rjUqv12vryr5euHF+80qkVHNWFSqtKyZTYxJMp9avFXnpuIsbMMUpQbnck05b3J9eWWlVZuCxz"
    "froxUHOuYE5FKT4HSy7ACgA+y7l/GrDT3P/c3Ny/vvDCv/7ud2/MzlYxpstK1ns7fHVlCBJZb7VWtrevLC5G1eFGY2xoKA1GIcYq"
    "RmbmW+d1pZoFMyZyLEzUDuHd9dVfzM/+bG7mbGtr3SKpWqdDqqraO/F5Hf4caYdfldjYiFWNScVtxLBQtVfa7YxpT62eieO0DmA2"
    "M6wDACsAuMbpf5bO+6vOLi4eP3v2P15++eT774uIVyXVGCOlkyrpLpKZYxZV9r5UfWd29vyVK82qiqqqenjv3sy5zHuJkUVurdIR"
    "TN17Xu0Q3ttc/+3S3H8tzr62vsa1nKvgQgwxKrPubPtcj/baV68DNEqM3Ikuy1s+e721PddsVqaFc/eOjA9keS6iZpFQNg6wAoC/"
    "dO6v2j26I0JEH1y+/B8vvvjvL7109sKFSESqziy1PemOd93jJ8Rm6WxMjJFFgura9vbS1tba5mZZVSMDAwO1WvqcVQhpm+JmPiGq"
    "ZKUaUXffv11V726s/XJ+9lfLc++1mtE7CxVXFZmaUfdp3ICvRYTImFiMu297i3w7xk2N681mzcmeWj13jplj966y4ZoYYAUAn3bu"
    "b2bprq+qzi4tPXv27L+8+OKp998X57yqxhhUjdl+b3JKFFMRiBiFSKqqYl7a2PjdG2/MzM/PLi2p2ZP33DMyMMDMO68Ebva5P1PG"
    "QkSdGN/bXDu+NPeblStvbG+xc7Sx5UjVSFONz6sexXXOJSViFTZVicqmohqcP9feXt7aECeDWXH/6ETd+5xZiaLhmxoQAPAp5v5p"
    "F1t6JRzem539r5Mnf3H69FuXLon3FKNLk8qrR//fyw8iYmYzEXGqqqoiH87PB9VOCHMrK4996Ut3Hz6cXiyralBNrxluqvpx0Sya"
    "eeE8jf6hentj7VcLl3+3uvhBu8XeUbsjpByJhOyzmV5zumfA0ZyoUkWcbzh5cX2FjDY6nSempofyPJWKi2ZE5LAOAAQA/OmZpVJv"
    "s5uILs3PP3v27D8dP/7GxYtKJDGqamlmf3r038mAyBxjZCIXo8ZIzl1aXl7d2ppZWFjd3CSiOw8eLPJcRMgsvUu4idZAROkNanoQ"
    "ZQxvr68dX77y9OLsB51WJOOtbVKKUY35M9tjN6XuCxijGCgSb21H7y9xa/XKxY6GepY9ODYxmOVMxMTRFHVD4RPgHcAXe+7f2/lJ"
    "0/93Ll166sSJn58+ffK999Q5irFwLu37f6odBeY02Hhm51w6HlSpLq2tbZXlerNZVlUty4YHBoTZpTeWMVJ6kfCZDlLBNJJ5Zs/M"
    "xJe2Nk6uLD67svDC+uLbW+vqHLU7PkZStRuz4/9nnzMRqTJxmuBb5jvMW6HaClUnVDnJYJZlIo45FS5lxAAgAOBjOx6pjDNzuqg1"
    "s7j41IkT//Dss6fPnQsxUgjpjKP+5fXMNO0/qLIqm1Vmi5ub783MrG5uZiLDAwNjg4M7f4adF8Kf1QhlRJFMjTwzE6102s8tXn5q"
    "bubZ1cXznXaIFW83OQQ1UiO7efasWCyVjCDjEE11y9GHm+uL7VbGPOTzsaKeTt0Gu4Xv4gG2gOA6bPswm1nalDezd2dmfnPmzNNn"
    "zpy9cIG8Z9XCucosXtNL23STQIiciPO+E2MZ48LW1ovvvqtE6+3243feece+fePDw0WWEVEVY+onc4PHKSOKpkbke2WLLm9vnlpd"
    "fmZl4cTW6jKZmXKIzlSVIqejODcPJmFVFTWxlKbWruWvba/ny0UgYqajg8O5cxkxkQUzYRIcEAWsAL7Idqq8Ua+d7+zi4k9ffvkf"
    "nnnmtQsXPjb3v/qgyzXvMqmaqsVIZu0QLq+uXlxc3Gg2G3l+cHIynTsqq8p6G0E3NADMgpkZeREmWmxtH5+//PT87Mnt9eVYWRXi"
    "5iarxm5T35ty6EzvZpjMjEK0ThWLfEWrlVarIB7x+Whec8LGVJkyMwIAsAL4os/9qVflTVXfv3z5+Nmzvzh9+sz585JlziwTKWNU"
    "IhLZzQFH62WAY3Yixlypbnc6b164UHY6McYqxnsOHTowNZVayphZuo7AV7Ubu55zfxOiQlz6/8w1t0+uLT6zunCquboYI8XoylLI"
    "TCk4SUPsTfqXyqxEQuSCOmchxlaWvd3aHFhdJGFjOjY4nDlXE5e+akbJIMAK4As79ydmU00VnmcWF//jxRf/r+PHX5+ZCTFSjBZj"
    "Goj7eMzRdl44p0JyzBut1pXl5dnl5XanMz40NDUykj4yqKa6y9c1AFJ7y2C2U7B6obX9zMLlXyzOvbK5uhyDVUHabY3RiDVttvR9"
    "9+Y6/OWmyqAUVcsyerdCcbnVzJhGszyVjSOiMrUUxgsBwArgizb6p20fL0IiUfXclSvHX3/9l2fOnDl3jr13qplIGULs92Ut6xUn"
    "yJidSDQLqnMbG6tvvLHdbivzxvb27fv3jw4O5umdRGo0f326zBtRVBXmmggRtUJY7LReWl741eKV17bXlmNlUbndlhgjkTpHZKTW"
    "x7+GdIInnejv+zqAzHyMhVJVVduZP7ux2vCemLdjODowNFbUauJSGDPeBwBWAF+c0T/E+LHuLvPz//7SS//429++NTtbVhWl6f/O"
    "ic/rNkFMo8/OfH+12fzgypW55WUzGx8cHBkYSH/aKsZ0JL+/U1U1U6LKVHpz/7Mri/8xe+G/FmffaG6utNsUIndKC8FSX0bma+np"
    "+AmjtJkIiQinT8v9TjhmIlYzMrUqmHfrpjOt7dVO22KcKup175m5UlXGHTHACuCLMfr/fneXhYXjZ8/+18mTpz74QJzzZo65IorX"
    "+Zx7qiEqRBlz6iW53el8sLCwuL5eqQbVx++88/CePY2iKHrHk7R/MaBm0cwx151Pc/9Lzc3ji1d+sXD5vc62eE8hShUohsikqT+7"
    "9XXub8YiaspMysJmHGPfH7gyE5kPmqtFpk3vN2NYrzqdGAeK4lGZHMrywrlIlp4JrghgBQCf39GfKKTuLkRp7n9hYeFfn3/+fz33"
    "3FtXrlQhsCrFSL0qkjfsrEuKpbTbVIaw0mzOLixst9u1LNszOup7FUnLEGjXp4Os99a3Uk133ojo1aX5/5y98OuVuQ/bLcoz7ZRU"
    "VmTRJN2i7d9zYO6VzCMWEZ9p5oXYMUkq4MzS18fe7Q5pxmZmIXCWNVU3NLSrSswm8qKRZUIcVCMR3gljBQCf27m/fby7y+zS0m/O"
    "nPnZyZMnP/yQmb2qmFWqsf/drD5pOI5mbCZEmZljLlUX1tcXlpc32+0yBCN64OjRRq3mnXPOpSrVu2x1YmZeJJNuibeL25vHl+ae"
    "np+9GNrELM2WK0uKGpnMuT7P/Xt36SILi5gIs5BQjEailqr89zN6jYhUhIgkxkxj3N6ixsBsp/XC+lKIsXD+ySyvO5eJqBkzKxne"
    "B2AFAJ8rf7S7y788//y/Pv/8m7OzlSqriurHKjzfWOn6L6UdKjNybmNra3lra25lZaPZHCiKyZGRtEHRqapUteIabopFs0gUzTwz"
    "EanqyaX5p+YuPbO88GGnyVlm7Q7HSFEtlXjr15NIKwlVYRNxLvPmM67lVlWD7DLvQ+akVhdViUox9rms9E6LApH05tlEtsk2QihV"
    "ndmIzwaynJnVLKimrxt7QVgBwOdn+k+97i5qNrO4ePz11//j5ZdPfvCBiGSq1uvu8hkVtiRNdaSZWdWbxRjbWfb27OzFpaXZpaUq"
    "hHpR7J+czJzLnOM/V43uT06GiTynIppUxfj2+uovF2Z/OXd5JnaMSFot6VRkpkKW7j30racjkRozG7OKqDg2K1rlwUZjf22w8vxB"
    "q7VqGpxTqlIQUh+TuNdJJopwCFlZWahsYHA2tp9fXShDpaZf834wy13q1MZsZmgrjxUAfC7m/qpqJr0D9e/Pzv7Hiy/+x8svv/an"
    "urt8ptLNL3bOVMm5SnVlc3N1c3NteztU1djQUKNWS2f2o2pQ/TRzVUt1LFJre2YmWmo1n1u88rO5iy+urZ7bXqfcW1mKqmk0ZhOm"
    "3tmcXc/9hcxYzTF7n1HmZHiQfTYp/ptTe3+8/+iTU/uODY3UnW+GsFJ1ZGCAmL0pG3Vzrl9/Kdwt052WW2ZEIltlua5hpSy3Op1C"
    "ZE+9kR5mGSPLZ1mUCbACgP7M/c0szf3NbGZx8dmzZ//5+edPnzsnzN7sj3R3+WwTi1nNWNUxW1WRyHaML7z99qWlpbnlZTV7+I47"
    "JkdGUkuZkFYtf26uml78+t6+ynK7eXJ18V9mLjy/Ot8Wx8zcakkIlrbL0wRc+9fP3dKIyypCTkitaJUPTu752wO3fXPP/oEs2yjL"
    "vXnNm7bK9mwMnOdaBWYzMu7jPQwjMjXmyMyqUpZUlZYXVzrt5c7c5e3NlsbRojjQGORuBzQ2dJTECgBu6bl/6u6SSqu9Nzv705df"
    "/vmJE6+eO5eONubM11bj8zqvAjgdRxEiTscTvd9oNtfb7eX19c1mU0SGGo08y9KapgzhT7WWVLKQ2tSnQqdEc83tZxdmn7py6eT2"
    "+pZzxExlR1KRIpbuIZy+jLrCZOSMxFic0yLjeo3Nbm8MPTk29dd7Dz4xNT1aFERUODfsswGXMXMgWq1KLXITFjMxTRlO0r+3AiJk"
    "JNpbYjin9WK51W7FKsaY1gGOWZjLVCMW6wCsAODWCwBVM3O93l4zi4vPnDnzT8ePvzk7G2NM3V06n6K7y2ezbCEKZkQk6XC6mYlc"
    "Wlq6vLDwweXLS5ubVQgP3X57kec72xQ7x5z+YP79UYub1U77xMriv1+++MLygtUKa7VJlTVEJWXXHfr7t+/f3W8RtjzjzEuMR+oD"
    "3xgZ//6eA/eNTgznBRGl0j1jtfqjk3saLht2WQzV2+urVBQUAqlji91I4r59ZxBT9EJqLkRqNlWj5dmZzbXNdtmsqpG8ODQw9NGm"
    "EdYBWAHArTX3/73uLu/OzPzsxImnTp069f77kbnb3YVo93P/VOdZmKlXWq7PiwEiEcm859RwWGRpdXWrqlplGVVz70cHB52IE+k1"
    "LOgOW0oWzZg4427Xrgsb688tzv1iYfbExkqnlluMeYhs0VTThni/li8kwlEdsxibF6sXVOQN4QeGxr83Of3NiekHxyeH84KJStNg"
    "SsTCnIsbzfLhLBsQ78RthqrtHOWZxuiYSY16VVL7tg4gEiJh8yRGpkWx0m61owpzRjySZTXvhVl7R2+xFEAAwM1u567vx7q7nDz5"
    "P5955sy5c5Uqx0gx6s7Nq93uc7AT4TQBvw6jg6U+VqoWI6tajOb9ytbWxYWFta0tIZoaGRlqNIgo9gpLpL0gJYqqjrtnWi5tbf12"
    "YfbfZs+/uLKwTWSdDnU6Fioji8TkXN8m/qkVmppnNhHKPGW+Hvme+tCP9x78yf4jd46MNnxmvXjLRDitA5gzcZO1+pGBwVGfrbY7"
    "FzfXuF5nVWdiZmxEZLssy/qxxRGRpVfdUS0EM6PMr4Vysd1qV9V4UUzW6ikAUolpRicZBADc7Ns+RKm7S9off2929henTv389OmX"
    "33lHRTjGQuQv6Oz4iXN/JlIRFdHeLoEQsVnflwJp+ulFMu/NLDBvtlprW1tb7XZZVcJcy/OBdFlMJKpquvImjpm3qvLttdXjS/O/"
    "XV06tbbUyT2FWKilWqdGRL0q0P2ZWJuJGjkXs4wbdWOezotHRye+P7nv61PTx4ZHfO8CGqXXM73bcGrGwp55MMvHs8KiVmaBbbMs"
    "1bu0u8XdoThdLOjTOoCZyVgpHf+P3i21Wi2zVqxMbchng1ku6f15ygysAxAAcHPO/dNbX+p1d5lZXPzZK6/8z2eeee3ChdjX7i7d"
    "zR/nlJm9T9NS75yY7ewa93UniK13V4DMLEYianU6V1ZXz8/NrWxsDNZqEyMjtTwnoqpb50eEuVVVb62vPj0/89OFi++ETjNW1mxx"
    "Valp+oT9LMHGzGZOzTNT5nhwgJyb9v6Jkcm/nT70/f2Hp+uNVGjBeisn6f13OzWZ0/82smyqqE/VahrjSqe9qYGdE2bSKCm1XD/X"
    "Ad1OMmYUI5mSyJroBxtrq+3WWF7srdVz54xJtVuOCeUiEABw0839qTf3T7PgDy5f/uWrrz516tRL77wTmZ1qwaw73V12841CRMwq"
    "YiL1PJ8YHKx7X1VVSIc4d5YC/c6ANE45kdw5Jopmpdni2trK1lZUjWa5941areZ9JiLM26F6fX3lNwtXnltbfrezHbxwpyzS9YEU"
    "gf3a9zdiUyFyLOol5p6KWsPnR4v6V8cmvz2x98nJveNFXZijWSdGZpaPD6PpyH00VSMlcsxDeT6Z13Jiz9whW6tKzTMzpqjprXh3"
    "O6gvTzmtA4xEzbGwcMyyTuaXtreYKSMezYoBn6W9vu5fBK6JIQDgppr7p190u7ssLKTuLm9culRd3d1ll3d9mbvVhJgty4j5zgMH"
    "Hrvjjv3j42VZbrZa5j2JeGbfr63qP/bFptaS3RM7ItudzqXFpZmV5SrGnZYygez1laWnrlx6eunyjJYhVrq1zSHGGNVSMbZ+XbBi"
    "1nR1jL13muc8NGid8oGh0b+amP7R9IGHxqfGi/pHP2YirleA+mOfhoipuxRI/7bm/WStfrg+KEbzneZaVUqek0YxkmjGZCkArE+P"
    "lcjSM9WoUUm1TTbXaa+12iN5sa82kDtJnQbSAS2HAEAAwM0w+ndve3nvnAsxnpub+/WZMz995ZWX3nsvqHqzPM39d7fzw8xepHur"
    "VqQoir3Dw9998MHv3X//bXv2DDcaPs8r1U5ZVmYxxtjrNtX3cSK9b/DMmXOZSGm2rXF2fr7d6lhVFeKil3e21p9ZnP3t6uK5TisK"
    "y3aziEGVInF/Js68U2ZTRZhE1PnoHWV+xGd3Noa+N773O5PTD4xNple+pWoqQP0JlffTW9ari5XWnN9Tb+REad3SDqE0M3GUzu6a"
    "cR8PX3WXRMwaxKwwUaFW7pe2NpzzBdOQzxreO+ZUSYmxDkAAwGe+85O6ultv3//iwsK/vfDCPz333FuXL1dVRTGmAx677e7CLGZ5"
    "lhkRZRkz37537w8feeQHjzzylbvuum3//rsOHZoeG8tENpvNlc1NyTISyb33kt509ru+NHNaCqQDr+QkKq2ur126cnml057R6uXN"
    "1ePL87OxCjFau00xqJIxW9/qu6VjlOaInTjJPdVqnPkRpa+OTv1gav+39+w/NjScO78T1C5dbftzD4F7qbnzYmA4y48MDI/5rAxh"
    "XWM7E8qcxJipkZr2bTXDRHbVekKjGjF3yJY67aXmtieeKGqDWZ7+lEFjKh2KDPh8wEWwW2/un36xU+H50uLib86cefrUqf52d+Fu"
    "eTAKqoGo5v14o/G1e+754SOPPHj77RPDw0S0d2xsanR0sFbLvS+cu7K+vtlsdkJIWzbd00G9+Xt/1j3p9amZN+MqhqraEnnHYnP2"
    "4r7R2tpQfdEZsVCrxVHNegPl7jemmLtTbzUWUmH1wt47lmHVh0fGvzs2+dXJfYeHholIzaqdif+ne/47x+2jWTB1LANZfkeWD4iw"
    "mvf+9Nb6ciy1yMtqm4Ul9crpw7mg9EKYTBypxqhOScoqCi/E6rnlOe89e/eITu2vN3LncnHButcTkABYAcCNHf2JUmdH6s39L8zP"
    "/+sLL/w/zz//xsxMiLFv3V2YHVHq0KveE9GX9u//6wcf/PHjjz96553jQ0M7H9goivGhoUNTU4empobq9e12e3lri7OMRByRF9k5"
    "I9S/LQsm5zRE65QagwwPZof3d/ZNbk+ObDpmYitLTqUumPt7hcopsZE4z5mngYaZHS0afzW174d7Dz4+vmffwODOgZ/UcfcvPTnT"
    "TYur/tuhvBjPipE8L2O52u5sk3FRiFnWLXPXx8pxTExkxpK6ICuZRefXyGa2t6sYhnw2muVO0isBi6bCggzACgBu3Nw/zX9Tt6wQ"
    "4+zi4jOvvfbTEydOvPcei2Sq3KfuLqlTVVSNzJlzk6Oj37j77h8//vgjX/rSyOCgmXVCiKqZSOb95MjI5MjIbdPT+8fH61nWqNcv"
    "LCxsNZsVcwiBzKR34qU/Vc7UTIOFSKp+sJ7vmyxuOyD79oSBulQVtdtUBXWOxPXxjTSno6hMJC54xyJZGQ7XGk+OTv5o3+H7RyYG"
    "sizt4wfVTMRf60lTYWZiMytViS1jd2BwqJ5lIVQW9OTaygJVVZaXMbLj9GxVrR+vXLrlo9WMSSVaTlQZLZMubq4Gi+ne350jY7lz"
    "wlJZJJSLwAoAbow/7O7y4ZUr//7SS//6wgtvzsx81N2le8792if+qXhy7r32ukrddfDg9x944EePPfbonXeODg4SUUxF/FNBCOlO"
    "A+tFMTY4uG98/PCePQN53qmqxbU1znN2ToiKLEu723rNV8aYyQmJWAjW7pCaGx3Kj+wvbjvk901yrUYxWlmxqmk6K9mPoSnVeFBl"
    "M8fC3luj4Rp17nQeHhn77wdv+86e/XcPjQ7l+U5I71zG3sVU/KO/wfSJ6t6PZPlEXjS8i2RzrW0ZHSEiH5W1ty3Yxy/ZjIgtHXQ1"
    "48xvaFhut81sJOuuA1KR6SqlO3aDsAKA6z39p6u7uywsHH/99X976aWT77+/090lxEi73GxJV/9FUg8TcW56aOjrd9/9t48//shd"
    "dw03GmZWVhUz51lGvT9MiDHNBMeHh8eHh7985Mj+sbHBWo3M3pudVecic6lqMXan5Nc8MU8BVwUykuFGfmg6v/2Q7J2yPHchUKdS"
    "jSpi3u/qd/n4A+l9Hg7CknnnJK/K2wdHvzu+72/3HZmuN4ioNGVi6u3792M7hjyzEQfVwJwxTzcGx4r6eFFruGyrU35YlepdReqE"
    "yYT7U+Oj90Wna31qQjFvtap2uzU0/Pr2ujAXXojoSyNjhThiSnU4cC4IKwC4XmK6xptefqbuLpcv/8eLL/7nyy+fPX8+dXfxvQrP"
    "u5r7E3nm3DlljiIksm9s7Gt33fWTr3zlsbvu6s79zchMeod80iZAt1tv77VE7v3Y4ODkyMje0dGxwUElWtvcjM6l2jsZs9+JtE/5"
    "p01HFYW1ClZWTOLHhrPD+4tjh/y+KakVFJWqaNbrqtiXq17MaSvMkYnzWuQyOMBEB4rGE6OTf73nwDem9h0dGu7+HZGRkbvqcm9/"
    "9p16Dyp9Wi8ymuVDWT7gvWdZbbdL7yjPxcyzkFZMqa1Nn25jdN9Kd0szSZZtxrhadsoQGuJGfZ4555iVqDLlv/yFByAA4FPMelW7"
    "NT57FZ5/+eqr/3j8+EvvvRdVvSrFGEPYbZ0fZiJyzCJSmVGWTQ4PP3nnnT985JEnv/zlqdFRIuqEwER5r+LQzijRrQzKbL22vfWi"
    "ODA5eWzfvumxMTJrdjpb7XYVQvpy+C8d/dM4qEohshM3NpIfPZAfO+j3TnKtoBCtU5qZOdfnvvapFiaxMkuj4VgO1gaeGB7/4Z4D"
    "358+dHBwiIhCKtJD7EXkOkyDhdmlZpaqjtmLm6zV99caA8Srze257Y3oMzIjUzLlnTM9/XoAxEpEqk7Vyqrybq7V3ChLJzLg/HhR"
    "T191+rvEyVAEAPR16DdL/+ev6u7ys5dffurUqdMffqgiFGOeSiLv+pxluuljIhUROzfUaDx87NjffeUrX7vnngOTk8yc9nl2ekz+"
    "0cmi9UpSpz9toyjGBwdHBgb2jo2NDgyY2frWljqXdhhS5Qb5hD2rdOBHhIysrKwKkmfZ3qn8tgP50f1+cpyLPO1AdatSu75sgnev"
    "ejkix8zOxaKQeq0mcufg8LfG93xvat9Do5MTtTp1O8/oH9Z46P8GIFG0VOyOhXk4z4ecr7M479pq6xa4KDRGR8xRmVLFi760t+T0"
    "hYkZM6kZ5/l6KLfJtkMIIRTMI0XhmAmdZG5NeAdws8/9r+7u8pvXXvvH48ffmp0NqbuLWSeEvnR3SdO3yEzONYrigYMHv3fvvd+4"
    "995Dk5PM3KkqYc7dJ1xoJWGW9FZANWp3Njo0MPD4XXfde+TIfUeO/Hp0VMxeO3+es6ybGTGm0fbP7ICoWlRxzk+M5Xcczg7vc0OD"
    "xGQhUhWIiLz7pM/zFw62nF6EGAuLOs95TlXnjuHxb49MfnfPgftGx9MJn2jGRBnLDdj6EKJCnJpFI89ERPsHh/4qOzI5MFC/MtNe"
    "ai9VgbPMLBAbR2WXTiz1I3qMjDmocVSJpVUhDDTebm0uttpzW5tbnX1DeT5a1LgX/5L6ISMDsAKAa9a96HtVd5d3Ll166sSJp06f"
    "Pvneex91d+lHhWdH5ESMKIqQc4NF8cixYz9+7LFv3nffnQcPioiahRCcc+7T7a2ny7qpUlv68xd5PjUyMlCrDdbrhffEvL65GZhV"
    "RM0kzbV33g939/GZmEzV2iVVFRd5Nj2V33YgO7zfjY2QF6q6hU53Pni3U11mI0tzeUei3ulAnbJs1Gf3D418d2rfNyemvzwyVjjP"
    "zKFbfJRv5IS3WxuVSIk8c91no1nREDfgs06oVsrSGjVzwkRiRBqJqG+V49IbETJHQsKW+W22japTqtVEGiLDeZGJCEs0M8Y6ACsA"
    "uOafczNN5yx7P0Izi4u/On36H5599t35+WjGZWmqZYx96eubrvakt74DtdoDhw//6JFHfvKVr+yfnGTmdOKzyLJP//O8sxl+dVvB"
    "WlE8dMcd+ycnv3TgwDNnz1KMH87PW2rJEqMQaa+pQPdmExOpUlQKkXKXTY0Xdx7JDk7L8ACpWVVRFSg1delXZbQ09ydjFudZvacs"
    "k+b2w5P7vzu178nJvYcaQ7lzRBSJHLNj4ht7CH7nHYOmnsnMY0Xt63v23zY0OuJ9pzr/QdXhImdi2m6LcUxX4frX9EaZ1ZTLjlSV"
    "NmpLzp3YWtXLoVVV382y8aK+015T+KMlHGAFAH/Jtg+Rqn6su8vp0z8/deqld9+NRHJVd5ddbnk7ImE2kUhkzg3V6w8fO/ajhx/+"
    "1v3333XoUKovXYUgIu4vfMHZTS/mGGMVY9oZ8M6NDAxMDA0N1esDtVqtKNSs3emE1KVS1e28XmamGK0KpCr1mp+eLI4ezA7v86PD"
    "JGSdimIgNWIhJ7sf17orj6jMws5pkVuem9rhxuDjQ+Pf27Pvycnp24dGnUiq8aDduf9nML51A8AsVY7LRDKR8aKosTBzJG2Ttk3N"
    "iZGxmRilWv/9eTcuQsxi5M1IzZxrmq5XZWmaiauJDPosE3HM0YyYGOUiEADwF839U3cXvqq7y09feun/On787KVLoaooBOs1d9ll"
    "d5dU4ZmZTcScGyiKBw8f/sljj/3tE08cnZ72zqW5f3cH6pp+L+6dEXLdFlTdpcD02NidBw4cnJzMvN9qtxdWV8V7Jsq8Z8dkpFEt"
    "KpFJo5Yd2FvccTQ/lPb9zUKkqMSSDob2Z3ODiaNyVPFOsowbdXayx+Rv9hz4bweOfGVi777GoPSqYH5yac8bkwHcqzK08xczVhQH"
    "GwMD4ptlZz1WbSYWEVWhSETWxwY4ZkSmzKTKIRpxm3m5qlbLTghhoijG8lp3jUKW3o4gA7AFBJ9q7p9+kW57RdUPr1x59rXXfvHq"
    "q2fOnWPvnVkmUsaovNuz3mKWtjIiM4kMFMXDR4/+8NFHv3HvvUemp4koxFiGUGSZ292Z+p3TQTHG9GIjz7KBWm2gVhsbGiryfLBW"
    "Gx8cvLiwMLe2VsZoahaClRUJ++Ehv28yO7zf798jjQaZWVlRagnA/TjtnvaO1Ditpry3LCfvhsVPsTy2Z/w7E9OPT+xN2z7BLMTo"
    "3bXXeOj7OkCIgqoSOaLC+WNDo0JMZnWfnd5cXbCODtSt1aYqiEWLZrtfB1i3XIQROY1cMVtHiZeJT62vVtFYpFVVRwdHGllG1O0i"
    "4HA0CCsA+FRzf7OPdXd56aV/PH78jZmZKoSPdXfZ9ZCc78z9vW8UxUNHj/7dV77ykyeeOLJ3r3cuxijMzjnpU0PwNFWVnvQZiyyb"
    "HBm5be/ew1NTg7XaarO5vLFO4ihEUnWDjeLIgfzO27L9e7lepAdEqulgYp9qHhARcYxMbOIkzynPhry/d2jsJ9OHfjR96J6R8XqW"
    "7XywY7nZyp+lh7nz1zSY5fvrA6N53irL1aqzzSJ5xhqzdE+A+lrB08iEzYyjGlPJvKrV7NbmdtmZyIvpxgClY7LdqiEY/7ECgD89"
    "+ncrPfQqPF+Yn3/27NlfvPrq6XPnxDlP5EXKNIneXXeXVI/tY3P/22776wcf/OZ99x3du3dn7l/b9dz/j/zWRErUbc5FlIkM1+vD"
    "9frekZGxwcHMZ9ODg+/NzKwTtYcHZO9kfuygm56UotCqtLLq5l4f997NyEi8IxHJCyIa89n9I6PfmZz+zp4DRwd27nmZpasYN9kY"
    "xr27e2rWjtGJZCJ76o3HZU+zLL24VzfXFkK5leWllWzkyDTGfqwDiIjMiRFzNAkh1T/azvxb7a3OUjmY5Y08O9QYqnvvnLPukVnG"
    "8VCsAOAPdn5Slber5v6XFhf/5YUX/tdzz701O1tVFan2s7uL92ZGzplzA3nenfs//viR6WnvXBruXO/e2XWZtKalAPNOwPjMjwwO"
    "3jY9fd+hQzWXNYXbeyb4tgPZ3gnJM9KoZaRU8JKpb/XOiCgqmfkip8xbnul2666s8eOpA9+e3HdkeHTndSszp0vYfBPvZvfWAcTE"
    "de8n8tq+emPYZ2XUy80tHhpkouzqXvB9+stlM5NUP1qVTDK/qdVyp7VddkZ9vq8xmD6sTAuQG3tqFrACuAXm/vTxCs+XFhaeee21"
    "p06dOrnT3YVo991dqNebJZpFERYZrNUePHz4rx966Dv333/bvn1EVIUQVAvv+zv3/8NJa9p7j6pVjKbqnIwNDIwNDNw+PX25034r"
    "szmJNj5EmbNWiyv96HJA31I33Rx2LEI+oxh0fbO6vGj1li9G/NB4HI6UrlmYeWa5uV9kMrNnVrNSVc1qzk/VGxNFbcR5R7RZdc6F"
    "TjPLOiGwGpOJkaUS0rv5slIVuPQyRlWiSqcjZMH5d5ubnRAGfD6UFQcGBho+q4szo9TQGZ1ksAKAj839qXfm5/z8/L+98ML//dxz"
    "6a5vH7u7CFHuHIuoCDmX3vr+t69+9YePPnp0Z+7PnKbnN2COxsxqFsmcSO67++wXO80T22tvc7XqnXqxELkKpunMD/f3uROz5Bnn"
    "Gavq6lb7wmz5/oVqdi42m7W82Dc52ajVhFlVhW+ZMmfMLNz9+2PmkTwf9nkuUkVdbDVDrSARrySkpGRkJNKXynGcDn0qkVE04zzf"
    "LDsboVqvSjGbKGqF88zd6wuGsnFYAUDa89mZ+0fVmcXFZ8+e/emJEyc++KC/3V12djMCM6W5/5Ej33/wwe899NCXDhwgojIEVS2y"
    "jEVuzNdOzCKSp0pnZltV+db6yitryyer5mLNV5VZq0Mhpgkj9autI6WbxsTek3ckomWpaxtxdqE8NxsWli92SivLicnJB++5e2Jk"
    "hG6pI4zM7IiIKKhGMs+SO3/H0IgQFcQWwtnN1W2RIEwmJpGNrS/N5c2M2UQ4RonqgrKRenm7ubEZQztGI75/ZGyy0fDMxK6ytJ1n"
    "jAOiCIAv9Nz/qvHl3Nzcf7700k9PnHhzZoad4xjTsVDb9c+nmLl028s5Ehksioduu+3vvvKVb917b3rre/W1gxv05ZMFVcfimYlo"
    "vd0+s7b0y8XLp1tb87Ha7HRYTWMkU0tT1N0Ojd2i1WRGqpRlUuQkotvNMLcUL8/HK0thZZ1i5DyrVCsNV9fW72Od/RtDRKx3pDh3"
    "7rbB4UxkvKgdXB387fLcnHcsohubXjWGYETa28bpQwyQERnHaBatVsy0tp/bWFlqtZf3TH9v36FuBT2zYEpEjglLAQTAF3T6Tzsd"
    "d81mFhaePXv2X1988dQHH3S7u6QuK32q8kZEkShVenjwyJG/euCBv3744WO9fX81S6dCb+ggRdQd/Tud19dXjq8sPLu8OCtKqrTd"
    "5BDJehsUO+UFrv1xf/QsyLmUKNZqVXOL5YeXbHZeN5tGJPValmcjQ0NDtUa6ikF/UdnqmycAiFIFp1LVseTe3z48Ol0fmChqnar6"
    "9eLlVp6r9xorU+a0Lb/7kDOjlCVmHCsXWYSjuEtVZ3Z9LVoczovHJ/aM1Wp5KqWHkR8B8AXUPex/1VmI92dnf/7KK784ffrtmRlx"
    "jmJkZt396M/crfTArEQkMtRoPHT06E/Sba8090+17G/Ipj/t7E6YpQIGRLRRdl5dXfrFwuwr2+szZYvzzDodCyHVk+zDyMtEaqTW"
    "ve6befZOo4aV9TC/VF26Ul1Z5M0tMuLMd+ud7eLm8023KZSKeRIT0UCWPTA6UcYwmhcnt9bf7nRirU5SiiqVgYJFR31YB1A6ZcRk"
    "akGNlWpRa8WrzU03c34zhG/t2ben3iAiRxSpW1BPMC4gAL5Qc3/f29a4tLBw/OzZf37++TPnzxORN9MYK9U+zP3N0uvAQJTe+t53"
    "4MBfPfDAjx577MjevcxchsBEmbtxdQ26nWt771Q3y87ZteVnly//Znlu3iKZyfqmaiRmNfrorW+f5v7shJwzNd3YKi9cLi/MxsUV"
    "65RORPLchMlMQ9QYta/tFT+zdUCv10JUTZE2nBff2HNgb21g+Mql0C4/LJtllsUqEEdm495llP6sA0g0RmbjjS1ybqVW/9Xi5Yps"
    "MPNP8J6xWp2IHHFIkxBsBCEAPvfSNd/U3SWNue9cuvSLU6d+cebMmzMz5BxVlRcJRGE3oz8zmUmvJYs5R8xsdvu+fX/z2GPfeeCB"
    "o9PT6Q+jqu5Gzv3N1DSX7tbK7ObGa+vLv11ZOLm5cqXssPccgmjUqOSkO/2/5vEofVGqpEZEJMJ5xt5pFeLyajUzX16YDYsr1u6w"
    "c+lt8EfvCT6X6850zZA5d+6ekbEqxoEsf3b5yqmNlXKgYd5bqy2mZKaq5v1u99zSOsCUjVhDOvFljfqrG6v1K7LWaT8yPnVkYLjw"
    "PmOOZuGWOmqFAIBrCoA/6O7y6zNn/vG55966eDGocozWv+4uIuKcC6munNmRiYlv3H33Dx59tHvmp6pE5C+q8LzbeSGRku78dgut"
    "5ssr879avPzi6tI6m8XIzZaZVmbK1I+Jv330GzthJ8RkVRmX18sPL1UXr4SVdQqBM8+Z57TvYUbOff6moqmTcLpFmKYgTuS+8cnp"
    "gcFceLtTvdnaCM5xkXNZclQTsd0HoXVzV41MlS1Sq6XCa1n2m/krC63mZqi+MUXHhkYK54Q59KueNyAAbtq5f+rusjP3//Vrrz11"
    "+vSbFy9GZlItnKtijLse/aVXeCuqqvdGdGRq6q8efPCvH3nkjv37uwWNY8xv1Da3EUU1YsrZEVEZwsXtzdPrK79ZXjy9vbFCSmoS"
    "KjFV05jO++/mEaSXkKqkSiKUe8oyJtNmq1pcqS7NVTPzcW3Dqoq948x/dMTocz0CpRIgqTQbk3mR6Xrja5PTZrZnZemt5sZlVR1o"
    "WKcjVeVUzUy5H0dvmY2ZVCVGTy6otjP/+vZGtrHcNPtGDHcNjQzmRXcdYCq4IoAA+JxJhTB/v7vLq6/+w/Hj7165ElUpBDLrV3eX"
    "7vpCxJjJbGJg4Gt33/0/vvGNR+64wzunZkRUu5Fz/+4PNhE7Irq4tfnLKzO/XV94q9PcNKWq4nbbjCOzdpvx7nqwS2c9Vck5csJM"
    "2uyEucXyvYvV7FxsdlhY8qw731cl61/brJuYMLMZi+z06Dk2NLK33jg8MPRfVy69tLkyEytuNHhrS7RUY939AzEiMmI2x0YWYiBT"
    "7VSdwcarzY2F7e1OCKT65dGJepYxczA1IsHlAATA52rbh8jMUpU3M3t3ZuaZM2f+69VXXzt3jr0X1TzN/Xe9+yxEZJb6+poImR2c"
    "nPzGl7/8N48++uCxY/U8V9VOCJlzvnfG8frP/aOw1JwjoirGmebW71bmj68tnm1tNJ2zqpIqsJGaRna7v5LKqbmYEy4yE0+OrQq6"
    "sVXOLVWX5qrLC7q5TcSUFZQ5Iu4WJvjC7D6k5vXRtEqXMERG8uLRsckQY70oXtpcudRutooiknFQ1y0X0YebKMSkRqzqyLz4EGOH"
    "+IPOdmNjJdUH/PLo+ECWF9wtG0dEDusABMDnYe6vSh+f+//XyZP/dPz4+4uLzExlqaql2e5/zNLPNjFz7+D8RKPxnXvv/d+/+937"
    "jhwZrNXSv8yckxu186NmwSzr/W4XtjZ+NT/7zOr8G52tbSLd2nZlFUlZXPfW0m5Hf0pvv0mJawVn3spSNzarC5c75y6H5TUrS8oL"
    "ZmJh0qsG/i/SUMOUilp/9LRHa/Wv7d0/2WgMO398ee7s1jY36twuszIGNrVUg2R3MZmqhzIrcaWB2kZlSbX8vfb2ZrulRMJ03+hk"
    "4T0RRVOzG3o0GQEA12Xun37xUXeXy5efff31p0+ffu38ec4yp/pRd5fdbbZ25/4iwSwyE9H+sbGv33PPDx555Ct33ZV5b2btssy9"
    "vwHHftKpvmjmiGrOE1EzVHOt5nPL888szb3R3mwScVW5spQYlUUd92XuL0zGos5R5sk7raq4vFZeuBzOz8YrS1SWXORUZMxkmq4E"
    "2xdzsyHVVFWijkYmzkSGsvz+4XELKqaO6FzVXs+5I4HLUqKyqWoqRrLbBUhq+ywx5qpVhzp5fr7sPL++zGSl2t0jY2NFLRdnRkam"
    "ZoLSoQiAW3T0j6p0VRWBmcXFn73yyj8///z78/MsQlUVY0xdtm03k19mNnPMxCzOpev4U4OD33/wwf/ft75175Ejaetpp/vjDfhx"
    "sl7H2tSrRE3fWl/97eLc71YW3m5vNUO0spJQRTNlIbLdnTjs9o4XJmIWFi0ycl7b7XJuqbp0pZpdlPUtT8QD9UDdyv7d455f7IGF"
    "iTKW7pNgyr3/8tj4YJ6PFfXjy/PPrczr0JA1m9JqWsXd55beH+xuy8xYNMaSxUKwjS0u8vfbm+vt5mq7s1G2H5vYO1FvMFMkCqr5"
    "TdZ+BwEAn2oK/IfdXY6fPfv06dP97e6SZr5sRiJVr9XGgdHRJ++664ePPvrkl7+c+vqWIeTeO+eu9yH3VNwumnnpFvnZKMt3NlaP"
    "L889uzz/9vZGYOYqcKdkM2Iz57rnL68p+ciMTUnT3N9b7th7IR4hmeQiyxvLKue3Ws0qcpGTTyUK0l4TDh12+/MYUTDrhJg718jy"
    "u0fHcxEh6kR9N7RXmStx5JnNnKlqtNTmfXcLVnMumklUb5WZtUQuWrCNJXVs4h7jqbG88CIiLtXBFXQVRgDcMnN/sxjj7839//2F"
    "F/7zlVfenZsTEYoxhpDKM+y2s6OZ783900/I1NDQDx566H984xv3HT2a3vSKiHeu24XxOk//lSiqBlUnzMSdEF5YmP3l/OXT2+tX"
    "qioQW7tDVUWkUVK1ANvNl0/dVyeciQTvrNGwqPt9/sTo5FeGxoa222/URn+2HU5eOMdFbmQcQ+6dhqBMaqkLMH9eb359+nWAI5Kr"
    "boMfHhz6jshoXnt+ZfHZ1uxCUZCIdEoXQygrZVOW/hwPNYtmFKNub0utuKLVcxsrGzEuN5tfn5o+OjIqRIGsDDG/OfowIwDgz0+B"
    "0y925v6XFhaeOXv2v06fPvH+++ycV3VEpZma7b67SxrTKzMmEqI9w8Nfu+eeHz322Dfvu09E0ljsRXwvHq7v3J+IiHKRXISItqrq"
    "zfXlXy9c+c3y3KoXMpOypLI0syi9mjO7GES4++pElKzjHHtfRNvr86+O7/ne1L6vTe0b9NnBvFGquaH6+fn5ja2tVuR2CNTbWebe"
    "MPQF/6ZNhwKCammWiXhxhweHR7JigLnTaT2/srDGXDmJRsbGRKxK/birmDrLi1keVTtl8H5OQ3NtqVWWhfeDRTGaF16EXbeTKKOd"
    "JALglpj777i4sPBvzz//0xMn3p6dZRGOUdPiYJfjTm/fn5nFezYj5qnh4R899th/++pXH7jtNpHegaDU3eV6j/5p7m9GvXr07So8"
    "Nz/7q8XLL60trpgySWy3XBWNuqm1i3ceRCwcg0Ri53zmy8zJ4IBuN+8dGPrOxPTjE3vuGBwZ9BkR3XH08H+vFV86fOiV99898fY7"
    "pz/4QPOciTiEzDlTjaqR2dKlgS/24CLMadcuXRIYKYqHx/c4kYl644W1pbfKUhoDZuQixRhI6aMU3/XUIZKQRmp1NNqG92eb6/my"
    "D6RPjO25bWTUMStRqdGzOGQAAuBmnv7/fneX1177zxMnTrz3Xn+7u6RRlIkqMzYTosnBwa/dc8+PH3vsuw88wCJqlub+Tm7EO7T0"
    "hWfERhRNO1FfW136xfzM8ZW5ZTUhtWbTlaUpqXAf5v6qZGRCyqmQqSuqeKCof2t0z99MH7p9eDQ9fyOq1WpfPnrk9gP794+PjRQ1"
    "Zn5nZqYTY6XaiZG0e9VJ0q2IL/ZSQJiFSM2qtEIiGq/Xv5rtG84LMYohnGttVz4LFMkcaWAjUu3LOiCKcIyuXVkVeKCx5t3J9eUQ"
    "KlMbKPKxvMjEFSJmhHNBCICbdO6fLnztDCIfXr7801deeerkyTcvXUrdXSxVgd7l925v7i8i6ci/Ee0ZHv6bxx//8WOPPXLHHamf"
    "l/XO1fD1/8KjGTEJsTAx0fnNjVdWl15YWXhpbXUlRhJnnZKimlI3ta55nE3LmqhCRs5T5rleM7URkkdHJ58cGf/q+N6jg0PdYeWq"
    "Teoiy+4/enSwKI7t33/yvfdOf/jh6+fPt9NSLEZnlrv0gjMada8+fWExs9hHB31q3t89MibE+xuDv5qfPbG+WA0OaBW43eIqkJKm"
    "N/m7rxxHnHrGU6djmq2JvLq1XhGtx/DY2OTDE1OOhZnaqo7I9YqdYORBANwsc/+du76qOru8fPz113e6u3hVSv2/+lLh2YyZK1Vy"
    "jokmBwaevOuu//7Vr377gQe69X9UnXP+hlz4Sl9OxkJEQXV2e+vFlYX/nLv46vpq08ipaqdDISqZ7fR1ueY3370hxliU2eWZEI2w"
    "e3Bg+Efj09/cuz9VmY9EasrEjjn1XHPODdbrD9x++5cOHbpz//6JoaFM5O2ZmfXtbWUOMabMTGsyli/0y0Ym8qmCtFmqSd7w2UMT"
    "U4cGhoSsFau3mlsdJssyi8rWO8K7y9FfNdUnIjNXBglq3q2ZvbS6tNTubIZyIMvuHB5zzFkqcHtDJjcIAPjz/kh3l8uXnzpx4ulT"
    "p3a6u0gaiXY9+nP3utNHPa32jY391YMP/vjxxx88dqzb9qu3J3sD5v5qpmTSO55xbnP9l/Ozv1tbfH17s+Udh2hVRdq7cLXLNDJz"
    "psxiXmKWSVE4s9tqA4+PTHx1ZOLhsck0+qeRi68qL69m3XsSRPU8v/foURE5tGfPqfffP3Pu3FsXL24zV0QUghMRZte7UP0Fj4Fe"
    "bezuo5yo1b4xtS8X9+La0mtbaxeb2zbQoKp0ZUWqMVWQ3uU2WnoDwZGMLQRzEuu1DzpNv7oYVL810Xp0ck/NZ0TU1sjEXgTbQQiA"
    "z37uT0TpfayZzS4tPfvaa//373539tw5Y3YxWm8m1YfOjmYiErv7QDY1NPS1u+/+/3zjG9+8/35J9dbNukc/b9Tc33erTNrl5tbz"
    "S/M/vXLx7XbTnHDZoaqyUBmLstvtqcF014GJiM15yfOC+EjR+Pro5A/2HnxwdKJI711Sp52ryso751yqRhdj+nMWWfbgsWN3Hz58"
    "98GD+8fHc+bTH37YrCoWSb0Mcdiw++jSK4Gr3u3fNjQyVR/YWx8YXPRCfL7T1CK3quq+Zdr9GxRVIlJxZCaq0inVjIv87e2Nte3t"
    "7bIcyvMHxqcoNa/u7VNh+EcAfDY+6u7S22x559KlX54+/fSpU29evGjeU1VlzoUQQl/m/kTinHMuEinRwcnJ7917748ef/yhY8fS"
    "717FKMz++u9gxNRFknmnXNe7G6vPLlx+dmXh3daW1WrWbmdRo6oR29WNva5hSpjecquxsDqxWo0zP8Ryz9DIk6N7Hh+ZuG9kvJZO"
    "3JoFjZn88QOvqSjTTjnuWpY9cNttTmRiaOiOgwffunTpndnZzU5HRTohaCqc+YWfWu5UkI4xmrBnGcqyh8YmvHDN+eMr8x+2t214"
    "WNtt7nScmpIZifHuygalx87MpF6jVZXW67Nl9fzq0kCeN2N4YHSykWXEFM2CmUvXxLAOQADc6ADY6e5CRL3uLv/w7LPvzM72v7sL"
    "EaeeHsymOjE8/PW77/7/f+c7j991V5FlaQ8qdXa8AT8H+vEGfjNbm8eXrvzz7LkPqnZgps1NiqGKQY1NZFcnPnsvPMQRsXCem3O1"
    "oPcOj3xvbPp704cONQZS4EUzJsrlj7/2SGUw0oQxbWsIc+b9fUeP3rZv30O33/67N974xauvnjl3bq0sSSSo2u774n4upKNBzrl0"
    "acsxj+bFVyenC5FQBgv6YXubi4JilBDIRFMrYN7VzxURqYiqSlVRUA2Ri/xSaP18fmaj0ynEPTyxZ2d+kA5WYPhHANzQuX+aVO50"
    "d3l3dvZXp08/derUmxcuqPds1q/uLmTGzE6EmIP3UWSkXv/Wvff++LHHHrr99iLLiKhVlpn3N+C2V1RNf5i0yljttD/YXD+5tvTb"
    "5fl3tzep0aBmq1ArVc3YmImvqcQbM5GxaqraScSxVnCRW4iHivqX68NfG596fGLqtt6BnzJN7T/FgVfuXVROnZC9cyONxoPHjmXO"
    "DdbrB6amzl64cH5+vlOWG61WGQK+1XeemxEFUzXORDLme4fH2nvDYJGf2lj7oL21VItBc+6UTtUiR0sFTnaxFGAmEYsqpE6JQ4h5"
    "PlN2XlhbGiiKrRjuGBze1xh0Iumtj6W7kAiCa9vu+/u//3s8hU89IHdve6VrVsw8s7Dws1de+Z/PPPPapUshRqoqUtU+tRWX3uYP"
    "M6tzFONjd9zxf3zve99/6KGhRuOjDxO53t/9aQhgSXUAaLndOrO69PT8zM/mZ95rb1VG1mpTCKlcjPbG8WubdqYvXIi8YxVHAw3O"
    "/B5x353c+6PpQ1+b2ndwYEi4e3BQiHyqHfwpnkDaufLO7by3Z+axwcE7DhzYPz5e977dbjPzoamph++4Y3J4OOX9DXi8N/92kGNJ"
    "r5qEOXd+T71xsDEw4n07VJdbW1qvs5GoUnrnb7TLmy47e59GxlFNA7FsaZxpby20mwXLRF4MZPnOwo4QAFgB3Ihtn6u6u6jquzMz"
    "x8+effrMmdfOn5cs+6i7y+5vlvbm/tGsYibmwaL40p49P3zkkSfuuiuN/s1OJ/deROR6bv1bOljJ7Nml32al0z61svjM0txLG0vn"
    "yxY7JyE61WiqRsZ0jeWd02vYtJfsnJKVecbiakwHffHo2Ph3JqYfGZscq9XTtk9l6kXcX7gDkD5cVStVJcqdy7zPvH/0jjty5wZr"
    "tSsrK0f37m0URXd+hML0V60DUuPGQlzd+9uHRnImVa3M3uo057kVMm9ELqpZUBIzI+Fr3wYUMSKKgYgkqEiMzi2SvbS27F1WqX11"
    "Yu+RoeF0bjX0yq8jBhAA123u//HuLrNLSz8/ceKfn3/+/YUFEUndXSpm3XWNz/TzlrZfiSg4pzHee+jQ//frX//ugw/uGRtLH5M5"
    "xzs9IK/nF16pOpH01nel1Tq9uvTU/Mzx1cUNEVLT5japUiqzn/4s11rgk1XZSByL8ybM9bo2m3e4+g+mDjw5vveOoeHhvEiZJMw5"
    "yzUfL3UiTkSvGi8y7+85fHjP2Nhms1lk2djg4M6iAQGw8z2ZiXTPPRMR0YGBoW+KG8yKqeW5Z9vNBQtU5NIJSlHUoiqJv/atIFMi"
    "Jul2EdCyYxpJ803vX9hYmd/aqlTr3vdOAFtQYyK+/nuhCIAv3ty/d9vlY91dzp795ZkzZ86dY+93urtEot22N+nN/dWsY0YiA0Vx"
    "bGrqrx966PsPP3zb9DQRlSGk+8DXdXjaubpcOM+9uf/J1aXfLF45sbGyIsYWpFNmRuXVo/81Df3dZl7MxBxYOHNMPMr+0NDotyf2"
    "fHt875fHJkV4ZxKaEbtdrHvSQ3NEMW3WmXnn6kVxsDfx39nAw+j/exnAzKFXa8SzHGgM5iyZmYXwytryTOjEWmaBrd124rpr5muv"
    "IG1EbCIUI5s51TxaSXGjoFfXV0ZrRc27Jyb2HGwM5eLSz+nO5hEgAPomxHj1cDO7uPjTV175l+ef/2B+npn71t1l52fMLMuySjVm"
    "mYXwwOHD/+PrX//O/fcfmpq6+ufwejd4ST0dPXfveq20m6dWl59amHl+dXnJKis7VAVTrYx2Nfr35v5iZiycZ+IdOTdK8sTQ6Lcm"
    "9jw6NnV4YCi95TUyYc6ob3teqV6eqf7e1gHG/U9aP6XuQ72T+JO1+hNT05P1+r7GwE+vzJzTkouCzHzQUFbKZLK7FgJmxqwmrFaG"
    "UiMTGWX+1PbGxsX3m6H6wfSh6cZg+muLqkTmBesABEAfJz7MadNgp7vLL06fPn3unHPOm/Wru8vOkKRE7RjJuXqeHztw4K8efvgH"
    "Dz98dHqaiKoQevP+6zv3J2ZhLpiJqIxxqdM6s7r0u5XFl9eXFyiSRml3nFplFni3F305dU9j0sxxkYnavix/cGjs2xN7v7V3/976"
    "ABGFVOKILLvqqldf/maJiEVUNaimcHU3qnPyLbwOIFKiUtUxC/NoUXu0qGUsG50yLF1eUtoW13FMTtmUoxLTta8D0rk75yiqxsjE"
    "GZXB+43cn1hZbGRFw2ePjk3trTfqzmciSrgmhgDo75SnN+O8tLDwL7/73c9OnHh3fn6nuwvtvrvLVeNROvLPeW5l+dixYz95/PHv"
    "PvDAoT17rv6Y670xbcxRYyapujO9t77y4vL8C6tLbzU3F8qOmVKnQxqDsu36ZQebCpEyW5ZJvWZEU5l8f3LfD/YcPDY4vKfW2NkM"
    "EN59TYlPevK+VxYUA8enjIFUQXrncX1pePR/O3TbgcHBZ5cXT7bb7VrNvJeqsqpksvjxIn3XtA7ojutBTUOgdoeK2mvbG9uz5y9s"
    "b31zYu/dI2PDRU2IiagyzdBGBgHQr42CMoQry8u/PnPm56dOnfzgg6u7u8Rdd3fZ+V2MiETErOb9senpv3rwwb957LEje/emuf/V"
    "+z/XcfS3bs02NSs1zrVbL6wu/nLpytmN9bYTC0E6HTKNStaHatPWnYaLsJOa6p68eGx48nuT+7+2Z1/6121V6V5Kuo5f9k6iolvk"
    "pw8Ax5xqj6cHWPfZA+NT40VNolmMbzQ314U1zzQEp0YWOR0Su9bvSyIyFiM2VTGSTmUurDl3YnV5M5TBNJDdNzY56DPpxwVMBAB8"
    "9M13YX7+6VOn/p/nnnt7bi51d+kWGOhTV6n041SZGVGjVvvKnXd+5/77//qRR3b2/c3Mpeu+13P0j6Zm7ISJuBnKd9fXTm2s/mZt"
    "8Z2q3c7FWqV1OhbVyPpw0smMhYzZvHMiQ1l2R2Pkm1N7vzYxffvg8NWPRa4aoOEmjIGrfwAODgx9d/rgZGPgVwuzz68tLqnKQINa"
    "bReUVKPw7itId5sRWCQWK0vO/IVQ/Wp1cVuVmL88PDacFw47QAiAftluty8sLj73xhsnz5//qLtLCLT77i5XrwBESFWJhoriiS99"
    "6W+feOLYvn3Uqzl6A3Z+uj9aZJ7YsWxW1etry88uzp0tt5tCXEXXaitpJLJ+vIXt9iUXsSzzIpN57f6h0e9N7r9vfJKIgmokc8yO"
    "+EbuyCNoruUvsXdkQIky5qPDI1P1BhMF1VObq0uxDHmmqkJGpJxuSV77czZiNmY146g+dqwsy7GRD2OHVhb3F7U9ea2R5blg/wcB"
    "0CfL6+vL6+uLGxuS5xRCd6+AuS9dxbl3dn7np6KWZfsnJm6bnk7/KoTgr//c/6N5eW8XpBnCxa3NDzfXtsXYOQuBmEz78YV3v2Rl"
    "ZfKO89yb7CnqhxsDe2r1nSgyuo6b/tDfdQBf9WsiGsiyxyemhvN8evHKi6tL72xv6tiIdTqu1eYQI5nu+rvImEmVhMmcqXKtttKq"
    "Lja3FjutfQODCAAEQN9stlpVCEWes6rFaKl7VL+nUtbb6vEi6SDKTrP1GzsF7v4imDZDCBo4RGEXgyoLie2qzMvvz+RS4RfniAez"
    "vOF8GWM6SsvMjshh/L9FSO+FcCoi7Zn21gf21geGnHdGJvRB1a7yLJalxEhGfdg5ZbbumR+1Tmkh5kWtFcJ2VanhFQACoM8LXZZ0"
    "P+uGDElRNcTIvROfN3Br4qNoY2bnWJjZpHu/Ia1RrM+/3w41i6mLC+Nt7K26FEjlUpQ4HcW/e3gsmg7Xas+vL721tb5W7x4N4krV"
    "ohLvekHJZJRapLKxI0z8EQD9VmRZOiMYQqCrzxdfj4kGczSrYqxirHu/M7e68YVvHXPhpOa8xY4yGZOj9O6OyfozQpsymaVWvJ0Q"
    "OhqNug28jK6qOQC31F5QxmJklalnrmfZIxN7x4p6nVmqcKLV6nhvahSr7jGy3S0FmMiIjdmYRKzwrua84PsGAdDvBcCNO4fSnUPt"
    "9Hf8jLZBuPuGjz/6yb6eNPU1w8z/8/DzQmakZpVZJk6Yjw0Ohxgy4ppzb7S25pmtVmir7auoqaJzOllgdC0Ti7SGSJdomHGJDwHQ"
    "77Gp54ZNo5xI917SVa1Zb+AivrshE9RSq8U+9P7+cwnrOTV6xU/v54Ew5yREFFVTRdU7RsaH8mKsqA0uXn5m4fJG7jnzHI1jJDJi"
    "sV3cEUv/F6NFUyVMIxAAn4tlx2f+Z0g9l75QXzL0928zzZwyEce8vzEoxEJUsJxtbb5frVW1gqKTsmQzSnX5Pn1B2bTQMCYmUiVl"
    "iyGGGEKIMZpHrzAEwK3MboKTDHxjiyEbDm987mS9AhtpPJ6qN76598DBxtBvl678e6vzXtXmohAmKkO6BNN986O95eif+Y7pRYUa"
    "qRkbmaUcwZNHAADAZz2HIKJeEW8vzjEPZfkD45Nsttnp8NL8LFXN1LRZA6fT0NKt+mGfXNWte3bIiJiKjL3zvpbVanmeZ1mG6T8C"
    "AABuCiKSmRGRUvec5pdGRv9OZLJef3rpyunVZc0yZ2ZqotHYWTpDqvbnAyC9IfOOvHdZ5vPMZZ6ZcYIMAQAAN8s6gJmVqFJNtYNq"
    "Pvvy6HjhRMkKlje311fZUc5WEpORGu+8D/ikN7p2VaaYkWnsNuTG6I8AAICbKwYySR2PulVgjwwO/8RnRwdH/nPu0rMLlzc1Ullq"
    "s0VVRczGQvKn74gxUzotzWLtjnnXcdlGKy4GmSvjeK0+Nj6OZ44AAICbZh1AnLZ1SlU1qzl3oDE4lhVNjWVVnTl/bnFmtlxZs3ZH"
    "zNQJiZDan/x01jupVq+R8Cb7CwODp0YmV4ZHxmr1//Hf/w7PHAEAADdXDMjHO8k0suzJyekx8SPnLz/73qWZ9z8IzbYjM+9YxKL+"
    "iVM9TN3+AkxOiGiT5e0sW6g1BvIs9xkCAAEAADdvBlSqVahqeb43r01MTl/2tdeXN2YuzlmzqWTGJMIaP/F8cEoAIyLriHSybMV7"
    "IsYrYAQAANzsGWBG6fq3937I+UbmcieVEyMyUmY2NvuE4fyjf8Mfu/qFmwAIAAC4qTOgV/FbzaoYShFpDLjhQfaOyNiUmVk/ubJD"
    "ep+cDgIpG+dZVi/qucMS4FNB9VQA+Cx1i8EZ6e/XgPg0NQi5u+HTLZre7TGBPSCsAADgFlgHpOY/kdnMogaNkUIgMjKyndten0y7"
    "w70SReJQlSQOzxYrAAC4lZYCtFNydxeb+KZmpnieCAAAuGWWApIK+3f/+dp3cVi491YAEAAAAIAAAAAABAAAAAIAAAAQAAAAgAAA"
    "AAAEAAAAIAAAAAABAAAACAAAgF2zno/+kT6xGcDH/8PePyhKQXx6KAYHADfHbFREnIikmp7ELEJsIn82A9LHpw9i58SJoBgcAgAA"
    "biHOuTzLvM+8K02JhRwTq33aAFA1syLL8jzLswzPEwEAALeMGLXsVKGqYoxmSsbWHf8/VQCkj+yEQMwRO0AIAAC42V01uLda7fXN"
    "9c3tbXGeKbUHMGIm/jN1QbubPyJMFEKoypIMPSERAABwS6WBiMvyXLU7fItj/fQj+U5OpDYygAAAgJvZTiN3Zp4YG7vj6G1mFELs"
    "5QGr0qd+CWzUazUMCAAAuNlHf5HucC3Mhw4cePLxR48cOlCWgciITET0078EJjM1713m/c6nBQQAANykdkZqZt4/vTfL/Nb2l0JV"
    "ETGxMbPZX7YCcCJOnDgEAAIAAG6dABgdGRkeHk6nOemq3aFPGwBX/VfMjGeLAACAWykMhIgc7nDdwGeORwAAgBUAAMBnRs1ijKpG"
    "af/mLz3J+dF/ZURU5DkeKQIAAG4NTOSdM8ERfgQAAHzRAgDvb284vAMAAEAAAAAAAgAAABAAAACAAAAAAAQAAAAgAAAAAAEAAAAI"
    "AAAAQAAAAAACAAAAEAAAAIAAAAAABAAAACAAAAAAAQAAAAgAAABAAAAAAAIAAAAQAAAAgAAAAAAEAAAAIAAAAAABAACAAAAAAAQA"
    "AAAgAAAAAAEAAAAIAAAAQAAAAAACAAAAEAAAAIAAAAAABAAAACAAAAAAAQAAAAgAAABAAAAAAAIAAAAQAAAAgAAAAAAEAAAAIAAA"
    "AAABAAAACAAAAEAAAAAAAgAAAAGARwAAgAAAAAAEAAAAIAAAAAABAAAACAAAAEAAAAAAAgAAABAAAACAAAAAgJuQxyP4S1l/P5vZ"
    "df38u/yzGZERExMxG5sZEzMR7+rzMhORERsbcfcT/uFzAAAEwE2H+/vZmK/r59/ln42JOIWUGRsTWS+gbNeP0diEzMjsjz4HAEAA"
    "fPZExIlkImJGZt57IqIY+xYAZkTkRFSVzGreO/mstubs6hzKxGXC3jiYWTRHpP1boQiRJ4oaOUZP4okdM2IAAAFwE077+ffm7P0a"
    "qlikOwV2zmLkm2YFwMJMJGQsxkYmwpGE1LgfIcAkxMxGRGrKIv19qgCAAOiPMoRWWW51OlarkWqr2SQicq7P6RIj1WrEPL++vt1u"
    "fzQnv9Gb490huIpxo+wsl2XZaJBzVoVOu52Cinh3EcBMTLEKFoMMDfLgwMbm1oZWzRCq3roKrwQAEAA3hRCjqppz7D2pchr6+xgA"
    "O4Od9yzSbLerED6j8Y+5N7RH1bIKHYtcK9h785WxkRF5R71tq10FnndUOanVpFbXTlkxVTFGVXy/ASAAbiKNPJ8cHr5r//7NqrIY"
    "/egoOxf79A4gERHnXBkjM48XxfT4+M5myI3cFeGrpvZ17w8NDt0bJzbYC+cqRN6IyER2vQJI/+PYZcaejRsDI3f4YqreqHn/0VeN"
    "RQAAAuAzN1ivT4+PP3DkiBOxGBtZxiKhr3NVZnbMnRDIbM/IyMGpqZ33wHIDXwjzVcPzYJbfNTZhmdtQI2GKyhrNSDl9zG4TgE3F"
    "KIhYljVq7lBeO1gfaPis+1Wn9wMAgAD4bNXyfKTR2D82tt1sWoxDjQYxa//mp8xsZsLcKUsz2zc5OTE09Nl/1d7vqTU6pptlGczE"
    "mRAZUd++cDPH3FE15weybG9eGytqRR831gDgzw4+eNsGAPDFhFIQAAAIAAAAQAAAAAACAAAAEAAAAIAAAAAABAAAACAAAAAAAQAA"
    "AAgAAABAAAAAAAIAAAAQAAAAgAAAAAAEAAAAIAAAAAABAAAACAAAAEAAAAAAAgAAABAAAACAAAAAAAQAAAAgAAAAEAAAAIAAAAAA"
    "BAAAACAAAAAAAQAAAAgAAABAAAAAAAIAAAAQAAAAgAAAAAAEAAAAIAAAAAABAAAACAAAAEAAAAAAAgAAABAAAACAAAAAAAQAAAAg"
    "AAAAAAEAAAAIAAAAQAAAACAA8AgAABAAAACAAAAAAAQAAAAgAAAAAAEAAAAIAAAAQAAAAAACAAAAEAAAAIAAAAAABAAAACAAAAAA"
    "AQAAAAgAAABAAAAAAAIAAAAQAAAAgAAAAAAEAAAAIAAAAAABAAAACAAAAEAAAAAgAAAAAAEAAAAIAAAAQAAAAAACAAAAEAAAAIAA"
    "AAAABAAAACAAAAAAAQAAAAgAAABAAAAAAAIAAAAQAAAAgAAAAAAEAAAAIAAAAAABAAAACAAAAEAAAAAAAgAAABAAAACAAAAAQAAA"
    "AAACAAAAEAAAAIAAAAAABAAAACAAAAAAAQAAAAgAAABAAAAAAAIAAAAQAAAAgAAAAAAEAAAAIAAAAAABAAAACAAAAEAAAAAAAgAA"
    "ABAAAACAAAAAAAQAAAAgAAAAAAEAAAAIAAAABAAAACAAAAAAAQAAAAgAAABAAAAAAAIAAAAQAAAAgAAAAAAEAAAAIAAAAAABAAAA"
    "CAAAAEAAAAAAAgAAABAAAACAAAAAAAQAAAAgAAAAAAEAAAAIAAAAQAAAAAACAAAAEAAAAAgAAABAAAAAAAIAAAAQAAAAgAAAAAAE"
    "AAAA3NL+X92Q+hZSdN/IAAAAAElFTkSuQmCC"
)


# Splash-screen-only variant: same house glyph, chroma-keyed to a
# transparent background instead of the opaque white square above.
# Kept separate from APP_ICON_512_B64 -- that one stays deliberately
# opaque (see comment above it) for the PWA/home-screen icon, but the
# splash screen sits on the page's own themed background, so a
# transparent icon here lets it blend with light/dark mode instead of
# always showing a hard-coded white box.
APP_ICON_SPLASH_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAQAAAAEACAYAAABccqhmAAA4y0lEQVR42u3deZxkVXn4/8855y5VvczCIg5xZVOnwSDdxl9MkGo2GYIm5mcVISZqEHqM"
    "skQNQU2kqlGT4IIsgnYDbmgSqtwiEVzQ7gEliZk2UWdwQUeRsMgwzNBLLffec57vH7d6INEgUz0zzAzn/XrNyEuG6rm37nnuc7bngOd5nud5nud5nud5"
    "nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5"
    "nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nud5nuc9AYy/Bd4uJaKAgDe+UdFo"
    "iL8hnvekafyo//X/KH9TPG9fV60G3Tc/fPA9T+WfJv6AxsSLF/91WcSU63WfffougLevpfv1oSHzwNBmddezRx0T71pFf1+ZOLwQ9HP4vRN/fNxvvmjr"
    "F0ul7Kqhzfq1PIt169b5bsETyKdk3k57lsr1ul5dPlBdrEYz+fiVRxCqtRSKL0GkH0Fw7i5azU+Sdm5UZ104d8zERLh+bMwqpZy/fT4AeHvrQ6QUF33v"
    "e9HQ0JCtKGW57D3P56AVaymGJ5HZmGbTEQVQ7FOIvYdt8x9lc/KP6oILFo5Zvz48ZHiTa6iK9XfSBwBvr8v6RVUajbBeLqdKKWHy0iEGB95AFJwMzjC7"
    "kOFsADgKsWWgr0iS/ZT51id4sPNPXHDBQl3E3DIzoydHRlJ/R3evwN8CbwmNX9dAl8GqSkVz3RVHUtRvIA6OxzrD7KzDqIDAaEQ0SQceTlosX/FslhXP"
    "RCk38Kn33VhR6kGmppSIGMAppfy4gM8A9v434778IFdF9MEzM2bl8LCrKGW5/rITCKI3EoVH0kkD0syBGMSpRx4zJYhzBKFQLBice4hO58Ziy13bGjv/"
    "v1fX61ENbKXiuwO7i/a3YOc3fBExk5OTwYYNGyIR0fvgNeqH7rwzXDs8kp2ulOUjl/4+UeHNxIUXkKYhnY5gnclfMOrRrxuFUoZOW7Gw4NDqAPoH/rA1"
    "EL8+uuTiw79fqSQ1MNUN9ci/mXwGsFc2fkBrrS2Sv/wvqlb10NCQKpfL+0RqW52aCoYOPFCffuRRyUVTXw/Gf/6dkwmjN1GMn0eatJhrKYw2aK1wjzG4"
    "n2WOOLT0DQRAm1by+biz7ZrkdW/bdJGIftnMjBkeHs58d8AHgL2m8U9PT5tSqWSf9drXxvfdf/8zE2vn+OpX75P8PisApZTdW68P0A1gY60m48tmY1Ye"
    "Okqk/4pC+GySTkYrFYw2KNRiAPy/nzwFIoKTjOWDBkWLduuLsYomOmeMbSrXG7peLivA+iDguwB7fEo8MzMTbN68WZRS3PXznx+fiFyDMZdw2mlDSimn"
    "SiV1+HnnBSKildor466ZBnW6Una82B7kKc94NQPh24jCZ9LpZHQ6EHQbP4+zvSoFCsPsgqVt+4iLL+uIvYhrLn3hpysVOzI5qabB7IvdqD3mS/W3YOmN"
    "v9FoBN/4xjfsOeec4zjppJcRBG/FmGG0fhbOPYNDD23yu79750PveU92B4SFI47Q98/M7BWLX5RSTKxfH84cfLB6pVIZ1eoKnnXQGH0DryIwv0GnZekk"
    "YAINj+PN/78zUK0V1ilsKgRBRDF+Oko/05xSevCesfM2PasGl9zcCu/85CdlfHzcZwK+C7BnNf5aoxGMl8tpuVLRn9269RSr9YUEwW+RprNobdA6xrn/"
    "xLnrsPZmvv71XwCqWq2aWq3m9uhVcCLquOlpM10q5X/P91UP4ICVr6K/788Iw5XMz7VJOwEYgw5AerwUrSHLBLAMDEBoYlqtbwWd5NL0Zw//qxofd2Pr"
    "14cnbtrk/AyB7wLsSWm/Ga9UsmqtZhoPPniCDYJ3EATHkKazZJnJR8NtG2Oej1LvJAjOYXT00G5/2jUaDVWtVrUsbpzZwxp/dXrarNu8WbTSrm/iXavY"
    "f+Vr6CuOofUgD88mpEmAigxK9974AZwDYxRaGxYWFEnWISoek0XRW9XTVryEaj2aHBnONh54oKrnawW8ncQvBOqt8ZuZmRm9adMmJyKiTj75JKLonWh9"
    "OGnaJssMWmuUAmsFkRRjBtH6dSh1uDr55A/w1a9+SynlXnnDDVEDLPmvPSYt/PDMTHBvqWRRysl1l69uxvrPKRaPx9kiC02HzQzKaBSPu8v/uH60Ek1r"
    "3hEVLIXCaqK4RnTfddRrN4yPjicT69eHIrLXDqb6LsC+8OYHM5yPTjte+tKXA3+FMUeTph2yTDAm3w6bD/YJzqUoBcbEBEGKyL+SJHWU+mduueVhjjsu"
    "KJdKul6rZU90l0BE1OE3XxHduea8TCllufL9R3FA/xuJ4+NxWcxCM8E5jVKmO5K/k3NSBc4KjoxCrCgWY1J3F83Zj7Hl7k/xlg+0yhs2RPWhIaeUyvwT"
    "6QPAbm385AOnmarVlPnmN0+xxryNIHghnc4c1mq0/r+yKodzGWEYEAQRWfYzRK4ttNufTm677W77yDSiPFFvt6qInp6e1uvyN7/wsfc8n3hwjELxpaRp"
    "wNxsijYBahd3HZUCJyA2Iwwd/QNFsvS/aSUf5v7mZ7nwwrm6iCnnD7DDTxP6ALC7+vybNm1yV111lVqn9fFE0bvQejVp2kLEoJTmsfrzSgnOWbQWwjBC"
    "ZA5rPwtMkCQ/kOlpOw2mBA6Q3Tn/XZWqnqam1zUawtZbNOqwIZYNrKVQPJlOGxZagiJA6d34aDqQzBLElr6+EGu30GpfzUMPfZY3jW+rTk0FtVLJaaWd"
    "4GNATwmXvwWP/82/qTsKvS4MTyaO34nWzyXL8pQYDL9uME8kH+gS0SRJhnP9hGEZY96HMS9TlUo4qlS25rzzwunpabO7BgdFRH3qzleF67TOU+rgOaex"
    "Yr93UCgcT7sFzaYAevc1froDCwoIDJk1NFsWpVYy0HcO+x3wlvjqSw+5eHQ0K01P64vE+bEsnwHsusa/cePGYGhoKN/ums/zXwCMkGVt0tRhTLhD6a0I"
    "iFiUsoRhTL4y6HuINEiSOl//+j3let30bdwYfqxWS3dll6Bcr5unHj0QfPCIUztSLUc8/cWnsrz/bMJgiHYiNBcsKIPR6gl5yW6/X9ZijNDfH4Gep5N8"
    "gftbk+rNb970SqmbEzlEj+GXDvsAsAvS/uHh4axUKpl1xpxEGP41xoyQJLNYG2CMptc3tYgAKcZowrCAcw9i7Sex9lN8/es/BFx1asoA1Eqlnb4ktjo1"
    "FQCMj45azj034rcOeSmF/j8nCp5Ls5Wx0FSYMEDvgsG+XlgnoBJWLgvRqk07/Vzcbn8kee35PxGp6iolPV6bdoyP+wpDPgDsnMY/cskl7rgHHlDrouh4"
    "guBdKPW87VN9oNF6afdQKUFEEMkHCLVOSdNvAJc+M03X/2x6ulOq1cx0rbZTxwWkXjdrD9mqJ4bHMnXppQUOCNfQVxwjjp7D3Jyj08lX9+2MbohisYu+"
    "E543JziVMThgiIKU+YXPY7mGn2zeRAk9/KOD1fqxMZ8JPE6+7/RrGj+Nhl134omnEgQXodQRpGlKlimU0uyMRf2L04UiGmsdEBKGLwF+4y7nPqH+4A8+"
    "pf75n7eN3HdfmMcLlcoS38YiYtbcfHPw5ZFKZ/L6e5dxwAFn0Vf8Q3Cr8safKLRRO6XBOitocWBUPse31HapFYhhft7RF0f0959Kkq7i6Ss/zuhffG1G"
    "RA81GqGCxEcAnwH01PgrtVrQqNVStBZOPvn3UeotaH1Md57foVTErtjQo5TdPlVoTAGRnyHyGUQ+yZe//EMFvLJajcq1Wl57r4cv+8Pr14cbHnhAX3nq"
    "qR3eVz2AA/d/DYX41UTRfszOJ2SJEBgDSvWc9i/22511GO0IIg1W6CSgjEHrpXUplALrLODo7wsohNCx32Jh/kPcdPtt0mi42oYNIUND2bgvOOoDwA6/"
    "+UdGsuOqVbPutttOIor+BmOOodPJ+/xKBezq3XzOpWhtieM+RFpY+xngo6t/+7f/645aLT2uVjMA07Xa4x4XWNyufPXmzdKoVCxXXrk/y/hTCtEYge5j"
    "vpWSppowMCC9v6hVnqWDOJSGOM6XCjvrUCJ0UsG5ndF1ypcQO5sy0K+IC5qk8206yeV84bbbpV53telpc8fi9Xq/Op/yt+BRjR/MjcPDtlwu63X/+q+j"
    "RNHFKPV8Op15nAu3r37b1YwJgIhmM8HakDCsYMz777jttlPLtVo4XavZp+Tfnd6BIK6ngbzx/+3+DMpaCvHrMLrIfCsjs4Yw0PmI+1JuZB7CiEOhWFTY"
    "bI7WwgxG/ZBCnyUKFM7JkvYO5F9YHgSCMGC+Ca2WEMXHEBfO55TfHlW1mmF62j1w4IHKbyf2GcCvbfyTk5PmlpUrXaNSsZxwwmmE4TtQ6nlkWUqW8Rgr"
    "/HZlJuAwRjBGYYzGuZ9i7T/Qan1UfeMbm52ImpycDMYeY9Cru5bAdAcQLZPvO4K+vnOIglGcHSRtZ2SiQesljfYvrt6zYilElsG+As3OFprzk7jkqxQG"
    "n4Yxf0YhPI7MOeYXLFoHqKV0Nbq/5YMijjBSRKEg7qcsdP6R/i3XUxlPzr3ppviKNWtSf/6ADwC/ckCs+ybNts/za/1XGDNMmrbIMkHr8In5dlT+NrXW"
    "YowiCIoo9Quy7J9J039kevpbCuR55XL0iQsvlBeOjKTyK7Ka4QZOVZTlI39/FHHfORQHTkTEMDeb4qzBhHr7W7XXcTkrDhFHsaiIwgBrf8J889Ok7U+x"
    "9q0PA/Dhv/1tlu9/JrE5jszFtJpJvnzaLG1fgVbgRMisI45hoC8mzf6bhblrWJi9QZ0zPn/oTZfHd645z/r9Az4A/FKff3h4OFOlkiEMT0DrKkHwApJk"
    "HucMSpkeGq4g4gDVfcPthNF0Z9E6I45jnNM49zXgKkRu56tfXZiamgpKpRJ0S2hJd13/aHcvv3zw0iH2i88ljk8lcwnteUcmIUGwxME+6F6r0FfURJEj"
    "STfSbH6c2Z/8C+df2TnsppviV61Zk44r5bj2siMphmcShSehVD/NtpClCtOddVjK3wUgySyFgqO/ENFJHyJNLmfrls9y/visPLKV2Jcef7KPATx6bb9S"
    "SrE4z2/MkWTZAiL5lt7e39622zB20jelDRDSbmc4lxFFL8GYdyFyBi9+8WCpVLK1RkMvLiFeOzNppqenXbnSUHLtlSMMRm8ljE7GpgntBXAuWFLjfyRj"
    "EFCWYgECY2l31tPM/pabvvl5zrsymZKp4G/n57MhUPV63eiz37RhYL71t7Ta16NUm4F+CEOL6waRXsdY8tWVEAaGTqKZX8iIwpXExTcweMCfLf/U1Su1"
    "1rb70tO+7/skzgAetbw3336b9/nfjlJHYm1GmjqUCnt6GEUsWiuiKEDEddcNsL2/u/S/vANsd9GQwrl7cO4mkuRa1q37QbleNw8cuBDeevyZbfnwh0PC"
    "9ikMDpyJUUfSagakWYZzwU4ZhU9TizGWFSsjsB2arZvCNPvo2396/obxceWqU1PBUKkklTwYqqludaFxpVz/ZZcdtHBgUCYK/xhtnkY7Seh0QBEsaZow"
    "z0oE64RQC4WiRptttJtf5eFtV/LGd9w9IRLOTE4y4RcMPfkCwGINv3K5nGqtRU444eUEwVvQ+rdI0wXSFIwJd7hPmk9LpQRBhNYJzm3CuZUYswqlMpIk"
    "624GCpbc8EQWf5YiDGOsbeK4mWJ8PXF8K41GMjwxEc7ErdMo9p9NMTySdidhdi4fzzBmCfPwClwmoDKiUFEoxmhzDy65iYe2/BNjb//Ro44L+x/1DfIZ"
    "QgkmZ2bU2pGRlMury1h5wKn0xX9GEK6m1U5pthxKaQKjH7Os+GNnS/nAYJo64tAxsCzCZR2a7c+RzE/wurdtqk5NBQcPDqqx4eEn9eGk5knW+NXMzExw"
    "0kknWcBMB8FJaP03GDNMls31PM+fP0AWrTVhCPBtsuzDWHsnxgwSBCswJu4u+bXdz9dLDAQGESHL2oRhQBgcRZIezgMPPMQxx8w+cNTTjpNC3xsoRquZ"
    "m+vQaoIOwp6TX9Ut/SPiUMpRiDR9/Qqt7mNu4R+Wb+PazuvffA/1ejS9ebO6+WUvS2u12i9FmfHxcXfjxAR3DA2Fd7zunCafu3kjpx0/hza/QRCsRGmT"
    "r4gUUKJ6ekflwS2fOckyTZKkBIGhEK5GggKl0U23/tGrHrrxxK1qaKisGqB4kh5Trp5Ejd/MzMzoSzZtcqvLZRk/6aSTCIJ3ofURJEna3dK748t78zdy"
    "htaOIChg7b8D76PTmabTUSxb9gLgzzHmJCDG2gznVHeMQeGcWkIQyPcRKGUJQk2WBSA/5ejV/8Yxhz+HqPBMWm0hTVja0t5uO3TOAY6+oiOKIlrtH5Ok"
    "kzz80JfU+eOzL5maCqZLpV+7Z6E7Nalq09N6qFSSytq1mv/v8KMpDPw5/f2/Q9YO2Dbr8lWDO2EPsnMOpR39RYUio5XcQkuuZO15P6RaVWMvO9hMDD85"
    "uwPqSdL4Fyv5WKWUMyeffJrV+u0oNUSWZaSp9NRHz9P+pPvmj7D2dtL0cpz7CuvWPbL67LjjnkOhcAIiZxAERwOWTifB2rxLoHXvg3GBAeeE+ZawfACO"
    "PDxh9WELDBYVrTZkVrr9/d6/a60htQ7BMVhUhKGhk97O/PxHnzZ7zzf++y0faJVFzBtAje7ANFtdxBzCjB5hOEMp4SNXHsVAOJYPVmYRc3Mpog16KeXH"
    "tmcuFm3yZc7FuE3GembnruGsN31DRKg0GkGjXE6fbNWF1JOl8WutU+ecUieffBpKXUAQjJAkC1irdnjA79H9cGM0xiiUWk+n8wGmpr4sIq5Uq8U/+9nP"
    "+NnHPtZRSgnHHVcgDE8gin4fa49H64NxLiHLUkT09q7HjjzkRoNzinaq6CtYhg5Z4AWr2wz0K+YWhDRR+eq+JY43WOsIQ0exP8QmCeK+TJpcz6vf9O+L"
    "+wtW9liyW0TUeTffHB27Zk1WUcpy7XuPpFA4g0LhFLQ6kFanQ7uj0CpY8h4Ch6BcwsBAQBQZOu3/oNm5mteev05EpJa/JNyTaf+AejI0fiCrNBq68aEP"
    "nUAcX4zWzydJ5nAuROsd3/K6mHZr7TAmxtrvkCTvBm6R6Wl7+Jo10Z0335x05+PNeVdcEVxxXrfI5rHHrqJQeC1KVQiCp+FciLXuUV2Bx5fyag2Iop1A"
    "HAlHH9HiqOe0iCKh1XLkW5WXmj7nXYwggGIBtJ5jfmEdreb7WPvWn7OhHq0fKsv2AqlLeAg/LBJOTk4ys3ZtyuXvPpCVy17D8sHTydL9aXYcWZafIpQf"
    "MbqE5coCYhPioqK/P6Kd/Dtp60Pc+olvyMT6bBrM1Y3Gk2b/gN6XG//MzMxi2k9jYuIk4vjdwGrStIlI2B1I66XPbxGxRFEBrX+ItRNs3vx1SiVXazTC"
    "V73oRemj+pPu2FWrsunpaVWv1w233XYfSXIlWXYeWfZFtG52p/Qszll+3V5fpRYH5BSdTDHQ5zjmuS2Oek6bOIRmW6O16S6sWVLbz3fzGceKZQqtfkGz"
    "cx2d4rtZ+9afT8lUUB0quxvzDUluiT+JMchOGxuzTEyEnP/Xm5fb7Fpm5y+l3bmXvr6AuOBwmQUnS9uPoUCCkHbHsLCQEpoXEsZv5nf+5Hg1ORmMKpWt"
    "Xv3kOY5M7auNv9ZoBGzcmI2Pj+fHdRlzIVr/Znde3vW8vNe5DGOEMIwRuRe4jFbrekqlZnloKADsr3p7iIgu1Wp6ulZzWmuHCOGJJx6eKHU8Sr0KrYe7"
    "04UdRDTGBL80QKhUnvZnDtptzUH7Jxz9nA7P+o2EKBTaicJJ/mceGQ3f8UfCiSAupVg0REEI8n3t5OODC82bHv6zN22ri5iNTKsaO7dKkQJeMjUVlIDx"
    "0dGMN72pyPNWncKyZWcSBr9JkqQstB1KAsIgv9beuzaCyxxRQdNXFJT6AQ/Nf4bZn3xGvenybc+Ymiq8FrLx0dHMB4C9q/GbRqNhyuVyqpRSnHjiGox5"
    "K8aMkKaLff6gxw/PUAriOELkHrJsctC5j89/7WtbDj333Phvjz02+3X9YBHRa847L2x997v21nXrMoGQ0dE1hOGfYMxvI7ISkYQ0zatiLi5F1t2R+NRq"
    "0gwO3D9l5HlNDn9WirPQaitE5XU3pMdHQaOw1iLaMthvMIGQtL5Hu/XxZ/70uzf/fPzj7bPXrw+3Dg+7TytlZRd+hzOgX6h1KheJ5lmXjRKHryEIfwel"
    "NPPNDHE6HxxcYnfAuoxiEZb1hbSze2m1ruPBBz/Nm8a3jU1MhGeMjcnoPrx/YJ8KAPV63WzdulWPjY1ZVSppjBntlu4+kk6nuX15b299/nz+O4oilLqX"
    "JLkOYyb0l788+/9fdFFUz4t3yg58pFYjI4aZmQwQTjjhCLQ+E2NejtYHkWWGLLMo9Uj1IWsVSao4cEXGb/3mAoesSskcdNK8qlCvg2TbG5ETTGAJQ0Vc"
    "sHRa36bduZq+279Jue5WNxpheTGr2g1ZXGXjxqBxxx2WSsVy3XtXExcuZLD/RXRsxPyCQy3WFVhi8RLnLNo4lvcHpO5+Wu2Priws3LCt8taHZWoqkPys"
    "Bgf7Xu3xfSYALA74NRoNV6lUHCec8FLi+GLgOSSJxVrp7ufvrc+/WMFX5H6s/RBJ8lGZnn547eRkMDE2tsP9YBFRtVpNDQ0NqUqlgog4deyxKygUfpsg"
    "OButX4J1GlxeE8BKAA5WHZhw9Oo2Bx+YogSSbuPv+atUdD9HCELHioE8EMwnn6OzcG192bN/vHFjWe4YaqhGuSzsxhHyqogugR5tNEQqFac+cslhFPrO"
    "pFD8PawMsrCQ5bsZu3UMeg0A+elNjsA4isUAHfyCVutz2IWPqldf8MDZsj48sbHJlcvlfW4TkdpXGv/i8l6llJiXvvTlNj+uK+/zL5bu7m15b4bWQhRF"
    "iPyCNL0Kkev52te2TExMhPfee6+Mj49nS/m7V2q14AFw68bHM8plQ7N5NGl6Clq/EmOeR6eTkqTCYU9r84KhNk9ZbkmcIs0e2drS6xS5k/z8wiAQikVN"
    "oO+n0/5n7p+/gTe/bRMiamxyMlg1NmafiOkxEdGTM5h7b6zZi8fHXTxx2TPay8M/IorOQJsDmJ/v0EogNEHPj7NerC6kLMZAXzFA64doNr/Cw7PXcM7b"
    "fzy2fn04PDzM65VKxQeAPavxz8zMmJHh4awKavzEE08lDC9A6xeRpnNkmeppwC9fZZf356MoRKl7SdNJ5uY+pv7t3x465/LL42NXrcp2xnHVSiluuOEG"
    "U9u40dwBGePjbmW5vHzrXPM0Os3TUHqYZ//Gco55XounHmBptTRpppZ0UMcjdfuEKLbEcYBzPyezn6AV3cBZZ82t707NXfP61y+5EOlSja1fHw4Da0dG"
    "Ui5720Hsf/AZDPSdipPD6LSFdpIXVO11n4MCVHfBU2AcywcMmQid1mdoz32I171tU1VEs3FjUFvcROYDwB7S+G+80R4Het2ttx5PofAulFpNkjS7a/t7"
    "LeNlUcoRhhFwL2l6LbOzk/pb35o9+fLL45vPOy/VWrud2TC63QIz/rPpYPWznuK+P95IwtNOe27yjFXv5ujnHsN+A8JCS7Bilr6hVUAUhIEjCjMemh3k"
    "R3ffzi23V81312+4ev36EOD1L3zhE974F+9Ng4Z++833BT859fyOXH55zHL3UgYGzyEyhzG34EgyQDRqCZFRK7DOoZVjoF+Bg4X2jWTp+znzL++mXjdT"
    "5bIq5TUF9vogEOztjX/Tpk1OarXFI7rHu6W721hruoNnPQ4KaUsUxTh3P2k6gVIf4VvfmnvB2Fh45qpVu+QNoJSSqakpxqdxd4yPJ1x+7rJk1ZFHI+yH"
    "OGGhDUmmCcxSbx44URRjIVDC3feF/NePIu6+74UsK7y6cNJJ7xsbHt48VKuFe8r3rbpFVmQNWUPOM6dr3ZHLLvsyquWwwRjF4m+i0oR2c3ERVG+ZgEB3"
    "j4Zmoe2II0OxcApZGHPt5ZNUKv91db2uB8uHGBGRvX1MQO+tjb/WaATDN95oK5WKDU455WUEwTsw5nkkiSXLwBjT0/57a/OTccMwBu7DuatXFYvXq1tu"
    "eXhsYiI47YwzpLKLVolNrF8fTpeA8fGEd7/tIPY//PUUCm8gClfRaucDfmGgespoFv+TxRoly/ryRTXf/0mBf/tOP3ffH6DYn77icxaUipVSMgS4pWxW"
    "2vlRYPvJyR/+j/8IOf/8Dl9/8BaarUtod75EITQsX2ZQKl/roWCH71VeWCSfeclSTbNpcTYmCk+gv/BWrn7vaKNSsSMM28bGjeHevmBor8oAuvvJ9caN"
    "G4OLTz89GX/lK40ZHT3Vwl+hdV662zmD1mFv0V/yWvNRFKL1PXQ61y2P44/f9y//svXcyy+P71+5MpscHbW7IKCp0vS0WTk87NYqZfnge57Kiv4/oVg4"
    "A6X2Z3YuxTkIjOppU4xafLABo4UohIW25s67Yr77g5jNDysG+iEKmqTZg7RbGcDcQw/tkV3EilK2KiL1DRui8tBQRyn1TT7+gYdQOKL4WAqFAZJWRuby"
    "XZC9Fh41RmMzxUI7o18bCoXfZcVAwMT7LbXabaePX5wcKC7IEwFlfQDYxS6qVvXMzIy58cYbM3nlKw0PPniCjeMqSh3ZbfxBd56/tz6/1hlhWATuJ0k+"
    "sn9f38SWL3xh7jXVauGK887r7Ip0T0RUhYb+48FDVEWpjEvftYoVxVfRX/xTnCxjy5Y2QkBgdLf+Xm9prQAmEOIQ5hc03/lRkR/8NKbVhr44IzQKJ5pH"
    "6ubt0caVckqp5EDngqoI40p9n4lLLyZ2Y/RFf0j/4DIW2habLp5qrHr4csAECpGQZtNh7QKF4giov+TQIJWxs//jeKVSJxKIiN4bxwT2mgCw2Oe/5JJL"
    "XKPRcJx44u8Rxxd1j+judN/8uqe1/c5Z8jJbISJtsqyxPI6v3fKFL8yNTUyEJ65cme6Kxq+AtTOTweq5I2Tt6EjKxDsPpX/wLPoKa8jsMloti10Maj2l"
    "zHnK74AgEAqBsPXhgG//sMBPfl4gSSAOXHf5L7vlzIOd+0xQAlsCPSRi/kip+/o/fuUH55rNbfT1n81AcZCWSuh0FNoEqB7OPBBZDB2Kdtugg4zArAbz"
    "N/zWcz4qw9UvKqWa69evD/P1YntXENB7U+MfHh7OGo2G5aSTXkYYvhWljiJNbbeGX9BT9d18A05eYy9PFeuFKJqY++IXt5ar1Wh4eJhd0ecXEfXKDRui"
    "yeF77fjoaMaVlxzDfk89h4HBl5G5lTTnLVmal8bqqc+vIE2FzAqFSCiEwt2/CLn9O33ceVeBNM13EWqdN/69dChLdccFts7M6BdMTIRzrzl3y0An+xTz"
    "zUtJ0h/SV4wY6NM4m5L1uJFIAGUUShuaTUg7EMVD7L/89fQd9MdUz18xMjKSrmXGLJ647APAzmsoplarBcPDw1mlUtGcdNIawvBCtM7382eZdE/S6eXD"
    "81r2cRyiVIZznweuaN90012HnHtuXB4asmtHRtJdEdCmp6dN4447rKhx4br3/yb7LXsjheg0HH0szCVYp/MVbr39CLLMERpLX9HiMuGn90Ssv6OPH98d"
    "kVlFHOWvtsx1H3D2amtHRtILx8ZcuV6Pmmv/8kE+/tlPMN+6GsW3iWJHHAUoZbE9FhoUyacINYZOB9pJhyA8nGL0Op51aJm/e+vKSTWSDpU2S13EyM4o"
    "Bf9k7wJ03/x6aGjIlkolsy6Kjkfri4Hn0uksILJYJKKXm52X7Y4iA1jS9ObC8uXV9mc/exdjY+GdV1yR7qI+v65NT+s7Nm8WKhWnrnv/8+krns9A8Xdo"
    "tw3zTYeOgm4Vm94av3OC0cLgoJCkmu/fFfGfPyiydS6kGNPdevzI8sF9ZEF4BZyUy05NTQVMTzvu3vYvaLOFSL+B/sILCQJFu93dct3DeQ/dUoNoo8kS"
    "2LKtzYqBpzLYdzZqlfB3f/eZiqpsZWIirK9cmS/v3sOnCfUe3PjNYp+/UqnYb4ThSzHmnWh9OGmadstpqR2+hv+5n990j9z6XJSm7+x87nN3Ua3qqTPO"
    "kEcNne3UtH/tzIyhVMqPIPvkB4+lf+AvKcS/y3w76FbEzWsU9BTU8h09RKGjv98wvxDx3R/2871N/WydE0KdEWjX/fx9b2tL3h2Q+ubNMlWracBx8223"
    "kyy8l3b6JYpFxeCgwUp+DoHqOfrlNR1tpplbcCgOYKD/bFb1jTHxrlWsXZvecsghurYXFN0N9tDGrxuNhqlUKqkCMaOjL7dB8BaUOoosy4/rUmrHp/ry"
    "AT/XbfwRSjnS9IZI5APprbfeeczYWDh28MEcPzq609d71+t1s+bOO4Mvj4x05LjjAq677PcJwj8lCo6mlSraLYdCY/SO7XNf7NNKJmSSMdCvKUQRnezH"
    "bJu9hQ0/GmBb+3QGBw7GqA7tTopzeeBbykk8e3ImUKlYEdHVoaFgvFZLUerb4ccu7aQ6nMXweywbXEknaZO081JsJmCHS5ArDQEBaWpZaDqKhVUsGzid"
    "pNDPle/75OTIyI9ERN+xYUP06SOPTMQHgMff+Ddu3BhUKpV08YhuG4b5WX2dzlx3hV/vh3ZA3vihQ5bdHCXJe5Nbb73zlHPPjc+84ors9F2wz70uYq6a"
    "nlbrjjgioVqNeObyl1KIzqdYOISFdkKzCaE2oHaw8bO4qs+itDAQB4SBI83+C9f5FOf+7mc47x0DnHjKLFq/HOcOw5gikHY3OZnup+xzdSFUd5qwvJqo"
    "LGIrSm3k3e++nFV9TQYHXkGxsB9YS5JanOie1goIoI0hTR3Odli5ciWxqbAiG4ivu/SDWqlNF01NuYumpoJaaecWT9knuwCLo/2NoaGsXC7rdbfffjxh"
    "mNfw63TmcG4pa/sdzrnuaH+GyJdI02py6613Do+NhTdfcUVWpucaM495TVdNT6tSqeSGJyYCnr7iVPqKFxBHz6TdTGgvaEJjet7Tbm3eVYkLQrHPkaTf"
    "Ybb5bmY3fFYoORF5mLR1GWnyVqz9GlrPE4aLZ/EJWu/0rs4e9DxRHyqnkG8mUn/917/gF+2raLavIU0fIC4Y4j6Hc7KkaqPGaFCGLdssNosY6HtpJ47O"
    "lWvef9g7R0ezcWDxyDafAfyaxr9p0yY3PjLiOPnk0zDmHWh9RLfPn6etvc3zu+19fq01WfZpnPs7tW7dz26o181VGzfukoIPIhJMA7eNjmbrPlBdwVP2"
    "O4Ow/49APZ1O25GkanvJ7h3dpqwUZNYRhMJgv8FJxnzrZlrN6+Sst/xX90pMo9EQdeutbRH5JqOj9xCGJ6P1awiCoe7pOZ3uOgjDXjJyvYOZgIiIKw8P"
    "q7H168ORF75wK9f8/aewhXuIg7+gv/9wNCmttqC1Riu1490BBYJGrGJhwRHHfcTxKSi9Qn/iqmtkdPT20XrdrGcmEJE96vwBvac0/sbGjYtTffk8P1yI"
    "UkeSZWn3iO5e5/ml2/hDtBayrE6SvJ+vfW3TMWNj4S1bt+pbL74425mNX0TU1NRUUNnY0KNKZX0T7zuA/Q9YS6HwOkLzLJJWRrujALPDqacCxApZmhGF"
    "MNAXothCaj/Fw7NXctZf/qcSVG3jxhBw5XJZzv7wh0MRcWpq6icsX/4xRN6Jc/8IbCEM+7ubnjKsze9DX98+FwSUUtnw8LAc9sMfxuqsC+e466EvknQu"
    "JencRhhmDA4ocIvnQu7oF54/PmGQH/Sy0LI4V6BYGM1Cdb6efO/xVCp2mOFszc03R3vS/oEnPOJXRfTBk5NmbGwsK5VKZl1+RHeNIDi6m/abns/Ty/ex"
    "pt3Gn5EkX4y1fmfnK1/58Sn5lt5sZ6/hXjz1pgGqopQdmHjfAfOF8NX09Z9NYIrMz3XIkgAV7NhhF4/U6BfAUogUhaJGcTft5LOk9nr1mnO3nL1+fXjG"
    "3JwcPzqaPfqTp0SCWq0WrBsf7wCy3yte8bSH5uZeQxC8HK2fjUiMcynOfYEkeTvr1t1/yrnnxjdfcUWyL1XByQuMzJjtZwJ+5MpjGAjOJopOIM1CFpoO"
    "Jyovqd5DwUGt85eOE8tgUQiiiPnWBpL07xj85r9Kuc40qM2QH5r6ZM4ARES9DMwtK1c6VaupdWF4Asa8C2OGSNMmEGBMr1Mpi6foKpQSsuwm0rTa+cpX"
    "flytVoPult6d3vinwUxPT+uKUparr145XzRnM1A8C0XM/Gy6vfGzg5t6RPKDLbTOiGMo9GnSbBOzzQ8dlA1cw2vO3XLMxER4xvCcjI6O/tJA5ijYdePj"
    "nWq1qo6rVoNtn//8fx/01KdehnNvwtqbMGaOIIgQCdmTdgDugsHBtcPD2TTT+ripqUCdee63o7n5K5ltfh0TtunvB60somxPWaFzeYUXowzzLU2nY+mP"
    "n0d//HYWXnysqtXkahpyy8yM3hMyAf1ERuKZmZlgeLGM9je+cSphWEOp52Ct7b7peu+XOufQOkPrfpy7jTS9hHXrflYul83Q0JA+fRcs7107MxmUwI6O"
    "jlom/v4oVti3Uew/HUc/zaYjy/T2Nz878OZX5DX6sQkDA5plAwFpcjvtziUHbe584YFXv3qhLGJm7r1XSpR+9YObv8Vl8cDOd1x0kf7FJz+5IF/+8r/G"
    "Sr0Ta99Lmv4ApQYYHIwABvfbT3Q+SLivRQEZpWTfWCrJRSI6ed0F3yez7ydJr0Wrrey/IsJoi7O2p0RZJP/ilNK020I7AROspli4kEOXn9Go3RJfMzKS"
    "zoCRJ3jzlX6CGr+pNRrByMhIprV23bX9F6DUC7A2odOxKBX2+OH5eX350dl9WPtvZNkVTE1tHB4eDi+88EL9R6efvlPnZatS1ROyPlw1fG8+1TNx6QtY"
    "vvwcisU/AFay0EywVqNMD2k/gsMShpaB/hhNQpJ9iU77g7z2/K88cMEFC6+s16M6DRgf/7UDTEopWTc+ng0NDanDzj03bjQaOvnSl35y2H77fZQsew/w"
    "zywsLACUh4asyD56VJ5SUlHK/svMpKFe15z5F3fS2fIxOu4TZPJz+uOYOAZrM8Sxw4VXRUBpBUrTaQlpkhFHR9I/+HoOe/6fSrW6bESptLaxYapTU4F6"
    "gjZi7fafWq1W9ctqNTNcq9kS6HXr1p1AoTDeneqbx7mgp9H+xT6/iMUYRRhqnPs+Ihfxla/cQrVqpkolSqXSTi3ltLi0l1LJXayUk2vf9wL6+l/PQPFE"
    "OgnMzluMDnf4Xi8eXiE4ggAKBY1mjvmFr9PmStae/5O8dHYjqA+Ve1q2rJTCOWdG1q7VM5OTKaA47jhTLZXc7ij9vaeo1+umBub7lUoil1eXMbDsD+gr"
    "vIYofjZpIrQ6+Rei0TvcdVPdxRqSOsJiysBATLvzMJ3OlWzd0uD88dlyvW7q5TJPRE2B3R0A1NjERDAxNmYrlYpqPPjgCUTRu9H6cLIsy1NkpXuq5JN/"
    "KXkprzguYu13sfZvjnvxi2+drtXc9PS03tmNf/tJNtMlN04NDl12FPHAORQLo7QXYL4NSmtUD5mWUmAzSxxDHGuELXQ6/8jCtn9Qa//mvhukbhqUqedH"
    "cbslBDBVq9UUwPgddyj96U9bJ0+uU7IXB25L09P61uOPz+Syc2IGnn0sfYUL6e87jIVWRqutQBlML+s1FCjJBwZDk1dftm6W1sIED9xzPRe8f2GxK7C7"
    "g4DazXdaKa1FRDAnnPD7NoouROshOh2LtQ5jgp77/PnBHRAERZy7HWsvkd/5na+S93kDYKfOvy4O+I0qlTE2FjJ86BqWL/8jougYsiRmfsEiSmN6OaBT"
    "CWKFKIZiISBL7yFJr2Tr3Jc59+1byiLmQtAjtZplJ72pq9WqHhoaUuVyWdST6HTc/3EPpqaCUqnE8VpnctFFAU9fdix9fX9GGB6LQpibd4jTBIHuba2A"
    "COIsUShEcQT6fjrtG5nd/A+8YXwT5KtGd+fsgNrNwUYolyO2bXtp96y+Y0jTJmkKxiylAKVDazDGIvIdrL1U3XLLF52IajQaYblc3qlFPKvVqq7VaqpB"
    "g7WTMwNblz/7JURmLUofRatjabUcWgW9le0WULjuWQQZwg/odG5gm/oM55/fOeWmm+JXPOUpbldV6+2WtxKepBaPJRtefGFMXHoMA8UzKcQn4lyRufkU"
    "ZxXBDo7n/O8BahNm9BWLKDWH7dxIM/to9WdbfnRHraYauzEA7P5BwLm5pxMEZxGGI6TpbLduf7DEb812d9HNkiRXyVe/etMr63UzMzMTsMSjq3+lUkmr"
    "6WlVURW7NVt2NHAWxf7n0UkSmgsWrYOea/aL5MGsWNQo7mMhuZ7X/MU/1Fetyk656ab45jVr0rUjI7usVPeTufEvpuDDkAGG9etD1r7529j599Nu/Quw"
    "lf5iQBTm31OPpQXQWuOygPm5FpEuMDDwcoycNn5QYXljN3cBdmcAEEDTbB6MMUVEUtI0X3Chda9pf76WffH8POc6gdZ3a6Wk8ZGPBJs2bXK7pIJv6VH/"
    "HBdWYZNnMN8kr0asdc81+xfTRIwQFQBzP53Wf4OoAw88UL1izRqnn6Tp+W4OAgLY8vAmVxcxvPqCnzLf+nuazeuJwwfZb4XGmHxvSe8/RCNO00nIz3nQ"
    "zyRm5e7OzHdvBlCtgtauu+JsV1yoIi9pzfArXuE2bty4C99m0/n/BCpD6JBlCsfSTut55A2Rn9eHgFFxud7QpVJJxkDEt8/dFgTqlN3G6WlVFzGs/csH"
    "mZ3/GEnnQyTpT+krxsSxI3P58ulezp9QWpEJuMyCSejs/uvcvQFgfFzQOu02/Edq3e2MxzrPBnTmXOhEuHDlSler1XZp2++GnBSlLYHWKFGIy0/c6XWU"
    "xOn80I78nkQYEzQqlX3yZNq9IQiMj45mt8xM6tX1aqTOffuWg+768T/x0NbrcNkG+vs0y/o1Qoa1bsde3pJne8bkOzO1JKQq2bcDACis1bts11neMc4A"
    "Nh544O5Jo5wEOILugRI76RsR0EpQKmMfPpt+bzE5sjYtl2vZMevXhw9c8P4Fls01aLUvI+l8hyi0xLHGqHzpuTi2l3P7db8c+b6B1GqyLMS1AvVId3mf"
    "DAD76qti1xXX24fX5e9NxpVyMzfeaC8S0VKupcc94+ivMTf/DubmbyGONP2DAsoiziF0GzePFJb7pV/bD2ftbvISxfzCbl8W7AOA5z3+LqwDqE1Pm+lS"
    "yXLWBRuYm7uKdvvzKBz9xZAwtIShoz8W+gpQiKFQ+J+/iov/HEMhVvT1WVaubLP6iM7uLtMa+G/V83YsExARWTszE0ysX8/akZEN8Sfee0Un4WHC8OX0"
    "9e9PmliczcuuK6V+dUYvj5zTCJCmmmZT97AB2QcAz9u9PT4lSqn0ROcM9XqUVCo/5RPvfS/hijm2bjuDu+85gC0PawIjBFphf8WYlwJSC8WCJgwK3L/5"
    "UP7jeydJX98MzeZ/0VMxAh8APG+3EBHK4KbKZTU9NRWMj442l//n5657ePzzTe748bnMzu+POEugwf5fXW3pLhiTEOEYUjtMYfBrNJtj3e659QHA8/bk"
    "TACyi+r1iFNOMbMveMW2ADZkhYFZ0AcgTiPWQfCrA0D3uGsUEIf9BNFKUE/dndfgA4DnLSUTAMYbDcvmzdqJKHXwMwo4m2G0IJKf0iSPUVRFkwcA5xyZ"
    "24rNHtydf38/C+B5S1QGmJmR7hJihXMBSQZJBqlVZCn/56+0++cyyY8hV0Q+A/C8vbNLgAbJCyqoR+b8H2t/1eL4oNr++25d9+EzAM/bWQHgf79Reyhr"
    "89jRwgcAz/N8APA8zwcAz/N8APA8rzd+FsDzdjrpnubyeIYBHz0NIKr3YhI+AHjeEy+0QioO2b4pWHjsGnHS/V3yOoM4HwA8b29llUVUgjiLwqFEusHg"
    "sQKAAiconRC41AcAz9v7KACxFEH2xwQhIg7c46sTaXBksgynVvoA4Hl7aQhQ1mWgZ1ESAzY/JJRf3wVAO7CCk6YPAJ63F6nX667RaFCpVACzCZd9EjH7"
    "Q9ZdDPyYp0PJ9vbvMo3WP/yfgcEHAM/bs1/83XMEAA49cNnPWbbshru+99OAAcOqww+X+350j+4euPxLDjr0qcL8PL9YWOCw564W41z7h7c/COyewUAf"
    "ADxvJ7rjjjsSBQ8JwAL85IEHeKxTnB747gPbX/W/7s/6AOB5e7hqtarvu+9gMzkzA8w8Ko8f/j/y/5nuvx5GZmZgYEBYty7zAcDz9kK1/DRqu6pWA057"
    "HP/Fo/7Mafk/j69b5zMAz9uLxwP2mlOc/F4Az3sS8wHA83wA8DzPBwDP83wA8DzPBwDP83wA8DzPBwDP83wA8DzPBwDP83wA8DzPBwDP83wA8DzPBwDP"
    "83wA8DzPBwDP83wA8DzPBwDP83wA8DzPBwDP83wA8DzPBwDP83wA8DzPBwDP83wA8DzPBwDP83wA8DzPBwDP83wA8DzPBwDP83wA8DzPBwDP83wA8DzP"
    "BwDP83wA8DzvVwme0J+ulEJEoVRv/72I+qXPeyIIChQYBVb1fj2o/Dedf9wvXZ/n7fUBQCkBHCCIOJRyi09+T01PBMDhnCBi0Vp2b+N3gsIiIlgRRCRv"
    "uNJb+xdxZFaRCcDuvx7PB4Bd+q6Morzhg0Nr233slxYA8s8JsTbFud3cYIxF2RQJLKIc6DwI9HY1gFgUGpsZnLPdAClLuEeet4cEgHJZs2VLPwMDg4Rh"
    "AAyglpIyb+8KCFFk6HQOotUaACjtwssolWB6usQ6xqGdLqO/+BSesn9Isw2dVBEFi5nJjtEKOh1FHAvLl2k2P3gADz884Bu/t28EgEZDGB2dResfYO0g"
    "0EREL+kBFwGlLCIhInMEwcMAmzdv3mWZQImSTJcWe+x2C7CR1PZhM8GlCqcUzvVwLRqcA+cEaxVK3YnTs/87R/C8vbcLYMx9tNvTOHcfWbYNYzRaL+0N"
    "J+JIkgLObd1v//3v3gps3LhRyuWy7LLrWLTM3EmSfp7NDy0jbVsAXNOA7aU7AVgLc4ZWorHZ3Wh39+LPq1HzT6y3U6kn5GeuXh2ybJlhxQq30z611VIU"
    "i8KLXpQyPu5229VUq5ohAu7bbyfdyx8Dh8FDDykOPthy7712t16P53me5+1t2YcfKPM8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8"
    "z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8"
    "z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z/M8z9uL/D9arxojmkgmOAAAAABJRU5ErkJggg=="
)


# =========================================================================
#  SECTION 1 -- Core data + business logic
#  (ported from the desktop app; framework-independent)
# =========================================================================
APP_NAME = "Tenant Monitoring & Management"

DATA_DIR   = os.path.join(os.path.expanduser("~"), ".rental_manager")
DATA_FILE  = os.path.join(DATA_DIR, "data.json")
PDF_FILE   = os.path.join(DATA_DIR, "tenant_data.pdf")
EXCEL_FILE = os.path.join(DATA_DIR, "tenant_records.xlsx")
BACKUP_DIR = os.path.join(DATA_DIR, "backups")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

TENANT_FIELD_DEFAULTS = {
    "status":              "Pending",
    "payment_history":     list,
    "deposit_history":     list,
    "arrears_history":     list,
    "locked_periods":      list,
    "deposit_cycle_start": 0,
    "rent_increase_due":   0.0,
    "archived":            False,
    "vacated_date":        "",
}

MONTH_PICKER_HORIZON = 24  # how many future open months to offer in the month picker

# ── cloud mode ───────────────────────────────────────────────────────────
# This exact same file can run two ways:
#   1) PC-local (default): can be run for local testing, reads/writes the
#      local ~/.rental_manager/data.json file directly.
#   2) Cloud (CLOUD_MODE=1 + DATABASE_URL set): deployed as its own
#      always-on Render web service, storing each paired install's data as
#      a row in Postgres instead of a local file, keyed by a session id.
#      Phones/browsers pair straight to this service and read/write it
#      directly, so tenant data stays reachable with the PC fully off.
# Nothing about the business logic below (routes, status/deposit-cycle
# calculations, etc.) changes between the two modes -- only where the
# `data` dict backing it comes from is swapped out.
CLOUD_MODE = os.environ.get("CLOUD_MODE") == "1"
_cloud_pool = None


def _parse_iso_dt(s):
    """Parses an ISO-8601 timestamp string into a timezone-AWARE datetime,
    treating a naive string (no offset) as UTC. Returns None if `s` is
    falsy or unparseable.

    This exists because timestamps compared for "which edit is newer"
    come from two different clocks: the PC's own `datetime.now()` (naive,
    in whatever timezone the PC's OS is set to) and Postgres's `now()`
    (UTC). Comparing those two ISO strings directly with `>=` -- which
    /api/_sync used to do -- is only valid when both strings use the same
    UTC offset. For a PC in a timezone ahead of UTC, its naive local-time
    string sorts as lexically LATER than a genuinely more recent UTC
    string, so every push looked artificially "current" and, worse, the
    PC's own bookkeeping of "the last cloud state I've seen" got stamped
    with that inflated local time -- silently masking every real edit
    made elsewhere for the rest of that UTC offset window. Parsing both
    into actual datetimes and comparing those (Python compares
    timezone-aware datetimes correctly, honoring the offset) fixes it."""
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(str(s).replace("Z", "+00:00"))
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def _iso_ge(a, b):
    """True if timestamp string `a` is >= timestamp string `b`, comparing
    as actual datetimes (see _parse_iso_dt) rather than raw strings. A
    missing/unparseable `b` loses (anything real beats nothing); a
    missing/unparseable `a` never wins."""
    da, db = _parse_iso_dt(a), _parse_iso_dt(b)
    if da is None:
        return False
    if db is None:
        return True
    return da >= db

if CLOUD_MODE:
    import psycopg2
    import psycopg2.extras

    def _cloud_conn():
        global _cloud_pool
        db_url = os.environ.get("DATABASE_URL", "")
        if not db_url:
            raise RuntimeError("CLOUD_MODE is on but DATABASE_URL isn't set.")
        # A short-lived connection per request is plenty for a single
        # household's app and keeps this simple/robust across Render's
        # worker restarts -- no pool bookkeeping to get wrong.
        return psycopg2.connect(db_url)

    def _cloud_ensure_schema():
        with _cloud_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cloud_sessions (
                    session_id   TEXT PRIMARY KEY,
                    secret_key   TEXT NOT NULL,
                    data         JSONB NOT NULL,
                    updated_at   TIMESTAMPTZ NOT NULL DEFAULT now(),
                    updated_by   TEXT
                )
            """)
            # Per-device roster for the cloud-direct pairing path -- the
            # counterpart to the LAN model's devices.json, just keyed by
            # session_id since one Postgres row backs every household's
            # phones instead of one file per install. last_seen/first_seen
            # are epoch floats (not TIMESTAMPTZ) so they drop straight into
            # _list_devices()-style "online = now - last_seen <= timeout"
            # math and the desktop's _format_last_synced() with no
            # conversion either side.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cloud_devices (
                    session_id   TEXT NOT NULL,
                    device_id    TEXT NOT NULL,
                    label        TEXT,
                    first_seen   DOUBLE PRECISION NOT NULL,
                    last_seen    DOUBLE PRECISION NOT NULL,
                    kicked       BOOLEAN NOT NULL DEFAULT false,
                    fingerprint  TEXT,
                    PRIMARY KEY (session_id, device_id)
                )
            """)
            # fingerprint didn't exist on tables created before this change --
            # add it if missing so upgrades don't need a manual migration.
            cur.execute("""
                ALTER TABLE cloud_devices ADD COLUMN IF NOT EXISTS fingerprint TEXT
            """)
            # custom_label_locked: once a person names their own device (see
            # _cloud_set_device_label), the auto-detected user-agent label
            # must stop overwriting it on every subsequent poll.
            cur.execute("""
                ALTER TABLE cloud_devices ADD COLUMN IF NOT EXISTS custom_label_locked BOOLEAN NOT NULL DEFAULT false
            """)
            # The one-time pairing token that gates "/" -- see
            # _cloud_pairing_ok() below. A NULL token means nothing is
            # currently scannable; the desktop pushes a fresh one via
            # /api/pairing-token every time it (re)shows the QR code, and
            # it's cleared the instant it's consumed, same single-use
            # guarantee the LAN model already had.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cloud_pairing (
                    session_id  TEXT PRIMARY KEY,
                    token       TEXT
                )
            """)
            # Holds the Flask session-signing secret itself. This used to
            # only ever live in a local devices.json file next to the
            # process -- fine on the desktop app, but on a Render (or
            # similar) web service the filesystem is ephemeral: every
            # restart/redeploy/free-tier spin-down wiped it, so a brand
            # new random secret got generated each time, silently
            # invalidating every phone's pairing cookie (see
            # _cloud_pairing_ok) and dropping them back to the "Waiting
            # to connect" screen even though nothing about their pairing
            # actually changed. Storing it here instead means it survives
            # restarts exactly like cloud_sessions/cloud_pairing already
            # do. Single fixed row (id=1) -- one secret for the whole
            # service, same as before.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cloud_app_secret (
                    id          INTEGER PRIMARY KEY,
                    secret_key  TEXT NOT NULL
                )
            """)
            conn.commit()

    _cloud_ensure_schema()

    def _cloud_get_or_create_secret():
        """Race-safe get-or-create: if two workers boot at once, only one
        INSERT wins and both end up reading the same row back."""
        with _cloud_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                INSERT INTO cloud_app_secret (id, secret_key)
                VALUES (1, %s)
                ON CONFLICT (id) DO NOTHING
            """, (secrets.token_hex(32),))
            conn.commit()
            cur.execute("SELECT secret_key FROM cloud_app_secret WHERE id = 1")
            return cur.fetchone()[0]

    def _cloud_get_row(session_id):
        with _cloud_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM cloud_sessions WHERE session_id = %s", (session_id,))
            return cur.fetchone()

    def _cloud_load(session_id):
        row = _cloud_get_row(session_id)
        if not row:
            return {"units": {}, "tenants": [], "settings": {}}
        return row["data"]

    def _cloud_get_updated_at(session_id):
        """Cheap timestamp-only read (no `data` column, unlike
        _cloud_get_row) so lock_status can hand it to every device on
        every 6s ping -- see the JS pingServer() -- without paying to
        pull the full snapshot back down each time. Lets the web app
        notice a change (PC save, another phone, etc.) and apply it
        right away instead of waiting for its own slower full-refresh
        timer."""
        with _cloud_conn() as conn, conn.cursor() as cur:
            cur.execute("SELECT updated_at FROM cloud_sessions WHERE session_id = %s", (session_id,))
            row = cur.fetchone()
            return row[0].isoformat() if row and row[0] else ""

    def _cloud_save(session_id, data, updated_by=None):
        """Always stamps updated_at with the DATABASE's own now() -- never
        a client-supplied value -- so it's authoritative regardless of
        which device's clock/timezone triggered the write. Returns that
        timestamp (ISO, UTC) so callers (notably /api/_sync below) can
        hand it back to whoever pushed, instead of them relying on their
        own local clock for bookkeeping."""
        with _cloud_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                INSERT INTO cloud_sessions (session_id, secret_key, data, updated_at, updated_by)
                VALUES (%s, %s, %s, now(), %s)
                ON CONFLICT (session_id) DO UPDATE
                    SET data = EXCLUDED.data,
                        updated_at = now(),
                        updated_by = EXCLUDED.updated_by
                RETURNING updated_at
            """, (session_id, g.secret_key, json.dumps(data), updated_by))
            row = cur.fetchone()
            conn.commit()
            return row[0].isoformat() if row else None

    def _cloud_touch_device(session_id, device_id, user_agent=None, screen_hint="", fingerprint=""):
        """Cloud-mode counterpart to the LAN model's _touch_device(): same
        semantics (immediate admit, no separate approval step, label
        re-derived on every poll so it can upgrade from a generic
        'iPhone'/'Android' to a specific model, MAX_DEVICES cap per
        session) but backed by Postgres so the roster is shared by every
        phone hitting this session_id directly -- no PC involved -- and
        survives Render worker restarts.

        Returns (admitted, canonical_device_id). admitted is False only if
        the cap is full or it's kicked. canonical_device_id is normally
        just `device_id` echoed back -- except when `device_id` is brand
        new here but its `fingerprint` (stable hardware/browser signals,
        not tied to any one storage context -- see DEVICE_FINGERPRINT in
        the JS below) matches an ALREADY-known device on this roster. That
        happens whenever the same physical phone shows up through a
        storage context that doesn't share localStorage with wherever it
        was recognized before -- most commonly, iOS treats a web app
        "Added to Home Screen" as a separate storage silo from Safari
        itself, and some in-app/QR-scanner browsers do the same, even
        though it's the identical device. Rather than registering a
        second roster entry for what's really one phone, this reuses the
        existing device's row and hands its id back so the caller can
        tell the client to adopt it -- from then on that context IS that
        device, not a new one."""
        if not device_id:
            return True, device_id
        now = time.time()
        with _cloud_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT kicked, custom_label_locked FROM cloud_devices WHERE session_id=%s AND device_id=%s",
                (session_id, device_id))
            rec = cur.fetchone()
            if rec is not None:
                if rec["kicked"]:
                    return False, device_id  # not admitted -- caller also checks this directly
                if rec["custom_label_locked"]:
                    # Person named this device themselves -- never let the
                    # auto-detected user-agent label clobber it again.
                    cur.execute(
                        "UPDATE cloud_devices SET last_seen=%s, fingerprint=COALESCE(NULLIF(%s,''), fingerprint) "
                        "WHERE session_id=%s AND device_id=%s",
                        (now, fingerprint, session_id, device_id))
                    conn.commit()
                    return True, device_id
                label = _label_for_user_agent(user_agent, screen_hint)
                if label and label != "Device":
                    cur.execute(
                        "UPDATE cloud_devices SET last_seen=%s, label=%s, fingerprint=COALESCE(NULLIF(%s,''), fingerprint) "
                        "WHERE session_id=%s AND device_id=%s",
                        (now, label, fingerprint, session_id, device_id))
                else:
                    cur.execute(
                        "UPDATE cloud_devices SET last_seen=%s, fingerprint=COALESCE(NULLIF(%s,''), fingerprint) "
                        "WHERE session_id=%s AND device_id=%s",
                        (now, fingerprint, session_id, device_id))
                conn.commit()
                return True, device_id

            # Unknown device_id -- before registering it as a new phone,
            # see if an already-known, non-kicked device on this roster
            # has the same fingerprint. If so, this is the same physical
            # phone showing up from a different storage context.
            if fingerprint:
                cur.execute(
                    "SELECT device_id, custom_label_locked FROM cloud_devices "
                    "WHERE session_id=%s AND fingerprint=%s AND kicked=false "
                    "ORDER BY last_seen DESC LIMIT 1",
                    (session_id, fingerprint))
                match = cur.fetchone()
                if match:
                    canonical_id = match["device_id"]
                    if match["custom_label_locked"]:
                        cur.execute(
                            "UPDATE cloud_devices SET last_seen=%s "
                            "WHERE session_id=%s AND device_id=%s",
                            (now, session_id, canonical_id))
                        conn.commit()
                        return True, canonical_id
                    label = _label_for_user_agent(user_agent, screen_hint)
                    if label and label != "Device":
                        cur.execute(
                            "UPDATE cloud_devices SET last_seen=%s, label=%s "
                            "WHERE session_id=%s AND device_id=%s",
                            (now, label, session_id, canonical_id))
                    else:
                        cur.execute(
                            "UPDATE cloud_devices SET last_seen=%s "
                            "WHERE session_id=%s AND device_id=%s",
                            (now, session_id, canonical_id))
                    conn.commit()
                    return True, canonical_id

            cur.execute(
                "SELECT COUNT(*) AS n FROM cloud_devices WHERE session_id=%s AND kicked=false",
                (session_id,))
            if cur.fetchone()["n"] >= MAX_DEVICES:
                return False, device_id
            cur.execute("""
                INSERT INTO cloud_devices (session_id, device_id, label, first_seen, last_seen, kicked, fingerprint)
                VALUES (%s, %s, %s, %s, %s, false, %s)
                ON CONFLICT (session_id, device_id) DO NOTHING
            """, (session_id, device_id, _label_for_user_agent(user_agent, screen_hint), now, now, fingerprint))
            conn.commit()
            return True, device_id

    def _cloud_device_kicked(session_id, device_id):
        if not device_id:
            return False
        with _cloud_conn() as conn, conn.cursor() as cur:
            cur.execute(
                "SELECT kicked FROM cloud_devices WHERE session_id=%s AND device_id=%s",
                (session_id, device_id))
            row = cur.fetchone()
            return bool(row and row[0])

    def _cloud_list_devices(session_id):
        """Snapshot for the desktop app's admin list -- same shape as the
        LAN model's _list_devices() so _render_connected_devices_list()
        on the desktop side needs no changes: device_id, short_id, label,
        last_seen (epoch), online."""
        now = time.time()
        with _cloud_conn() as conn, conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT device_id, label, last_seen, custom_label_locked FROM cloud_devices
                WHERE session_id=%s AND kicked=false
                ORDER BY last_seen DESC
            """, (session_id,))
            rows = cur.fetchall()
        return [
            {
                "device_id": r["device_id"],
                "short_id": r["device_id"][:8],
                "label": r["label"] or "Unknown device",
                "custom_label_locked": bool(r["custom_label_locked"]),
                "last_seen": r["last_seen"],
                "online": (now - r["last_seen"]) <= ONLINE_TIMEOUT,
            }
            for r in rows
        ]

    def _cloud_kick_device(session_id, device_id):
        with _cloud_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                INSERT INTO cloud_devices (session_id, device_id, label, first_seen, last_seen, kicked)
                VALUES (%s, %s, 'Unknown device', %s, %s, true)
                ON CONFLICT (session_id, device_id) DO UPDATE SET kicked=true
            """, (session_id, device_id, time.time(), time.time()))
            conn.commit()

    def _cloud_set_device_label(session_id, device_id, label):
        """Lets a phone give itself a custom name (e.g. 'Mary's iPhone')
        instead of the generic auto-detected model name -- and locks it so
        _cloud_touch_device never overwrites it again on a later poll.
        Enforces that no two currently-connected (non-kicked) devices in
        this session share the same name (case-insensitively). Returns
        False (without writing anything) if that name is already taken;
        True on success."""
        now = time.time()
        with _cloud_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                SELECT 1 FROM cloud_devices
                WHERE session_id=%s AND device_id<>%s AND kicked=false
                  AND lower(label) = lower(%s)
                LIMIT 1
            """, (session_id, device_id, label))
            if cur.fetchone():
                return False
            cur.execute("""
                INSERT INTO cloud_devices (session_id, device_id, label, first_seen, last_seen, kicked, custom_label_locked)
                VALUES (%s, %s, %s, %s, %s, false, true)
                ON CONFLICT (session_id, device_id)
                DO UPDATE SET label=%s, custom_label_locked=true, last_seen=%s
            """, (session_id, device_id, label, now, now, label, now))
            conn.commit()
            return True

    def _cloud_active_device_count(session_id):
        with _cloud_conn() as conn, conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) FROM cloud_devices WHERE session_id=%s AND kicked=false",
                (session_id,))
            return cur.fetchone()[0]

    def _cloud_set_pairing_token(session_id, token):
        """Desktop calls this every time it (re)shows the QR code --
        immediately replaces whatever token was active before, so an old
        screenshot or a link copied from an earlier QR code stops working
        the moment a new one is shown, not just when it's actually used."""
        with _cloud_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                INSERT INTO cloud_pairing (session_id, token) VALUES (%s, %s)
                ON CONFLICT (session_id) DO UPDATE SET token = EXCLUDED.token
            """, (session_id, token))
            conn.commit()

    def _cloud_check_and_consume_pairing_token(session_id, token):
        """Single-use, race-safe: the UPDATE only matches (and clears) the
        token if it's still exactly what's stored, so two near-simultaneous
        requests for the same forwarded link can't both succeed -- only
        whichever one reaches Postgres first."""
        if not token:
            return False
        with _cloud_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                UPDATE cloud_pairing SET token = NULL
                WHERE session_id = %s AND token = %s
            """, (session_id, token))
            matched = cur.rowcount > 0
            conn.commit()
            return matched

    def _cloud_pairing_ok(session_id, token, device_id=""):
        """Cloud-mode counterpart of the LAN model's _pairing_ok(): '/'
        may only bootstrap the app shell for (a) a browser that already
        paired successfully before -- a persistent cookie, so reopening
        an already-installed PWA doesn't demand a fresh scan every time
        -- (b) a request presenting the current, unconsumed token, or
        (c) a request whose X-Device-Id already belongs to a known,
        non-kicked device on this session's cloud_devices roster. (c)
        is what the LAN model's _pairing_ok/_device_known already gave
        it and this was missing: the session cookie alone is what a
        browser silently drops (private/incognito tabs, iOS Safari
        clearing session-only cookies when the app is fully closed, a
        'cloud_paired_<sid>' cookie tied to a session_id that predates
        a cloud_sync.json recovery, etc.) -- and without this
        fallback, that cookie loss permanently strands an
        already-paired phone on the waiting page (its own background
        retry loop keeps hitting this same gate) until someone
        physically rescans a fresh QR code. A plain forwarded link
        (sid+key with no valid token and no recognized device) still
        gets none of the three."""
        cookie_key = "cloud_paired_" + session_id
        if session.get(cookie_key):
            return True
        if device_id:
            with _cloud_conn() as conn, conn.cursor() as cur:
                cur.execute(
                    "SELECT kicked FROM cloud_devices WHERE session_id=%s AND device_id=%s",
                    (session_id, device_id))
                row = cur.fetchone()
            if row and not row[0]:
                session[cookie_key] = True
                session.permanent = True
                return True
        if _cloud_check_and_consume_pairing_token(session_id, token):
            session[cookie_key] = True
            session.permanent = True
            return True
        return False


@app.before_request
def _cloud_gate():
    """Only active in CLOUD_MODE. Resolves which household's data this
    request is for (X-Session-Id / X-Secret-Key headers) and blocks the
    handful of routes that only make sense running on the actual PC
    (device pairing/kicking, PIN lock, shutdown, LAN QR code, etc.)."""
    if not CLOUD_MODE:
        return None

    if request.method == "OPTIONS":
        # A browser's CORS preflight for a genuine cross-origin request
        # (see cloudFetch() in the companion's JS -- reached when a page
        # is running locally with no direct-cloud pairing and falls back
        # to this service's own origin) never carries the real
        # X-Session-Id/X-Secret-Key headers, only lists them in
        # Access-Control-Request-Headers. Gating this like a normal
        # request would 401 the preflight itself, which makes the
        # browser block the real request entirely -- silently, with no
        # visible error, just a fall-through to stale local cache. Let
        # it through here; _cors_headers below decorates the response.
        return None

    path = request.path
    pc_only_prefixes = (
        "/api/device-count",
        "/api/announce-disconnect",
        "/api/unlock", "/api/lock", "/api/settings/pin",
        "/api/settings/reset", "/api/shutdown", "/connect", "/qr.png",
        "/api/cloud-config",
    )
    # /api/lock-status, /api/devices(/…/kick), and /api/pairing-token used
    # to be blanket-404'd here along with the rest of this list -- they
    # made sense only for the PC-hosted LAN companion. Now that cloud-
    # direct pairing has its own device roster (cloud_devices) and its
    # own single-use scan token (cloud_pairing), they're real, session-
    # authenticated endpoints instead: /api/lock-status is what a cloud-
    # direct phone's pingServer() polls to know the cloud service itself
    # is reachable, /api/devices is how the desktop manages that roster,
    # and /api/pairing-token is how the desktop pushes a fresh scan token
    # every time it (re)shows the QR code -- see _cloud_pairing_ok().
    if path.startswith(pc_only_prefixes):
        return Response("", status=404, mimetype="text/plain")
    if path in ("/manifest.json", "/sw.js") or path.startswith("/icon-"):
        # These carry no tenant data, so there's nothing to gate -- let
        # their own routes below answer regardless of mode.
        return None
    if path == "/":
        # Handled entirely by index() below (it checks ?sid=&key= itself
        # in CLOUD_MODE), so a phone that scanned the direct-cloud QR code
        # (see get_direct_cloud_pairing_url()) can load the app shell
        # straight from this service with the PC not involved at all.
        return None

    if not path.startswith("/api/"):
        return None

    session_id = request.headers.get("X-Session-Id", "")
    secret_key = request.headers.get("X-Secret-Key", "")
    if not session_id or not secret_key:
        return jsonify({"ok": False, "error": "session_required"}), 401

    if path == "/api/_sync" and request.method in ("PUT", "GET"):
        # A session with no row yet isn't a bad secret -- it just hasn't
        # been created by a first push yet (fresh cloud database, or a
        # phone/browser reaching this service before the PC's first
        # push has landed). PUT already treated this as "fine, this
        # push establishes it"; GET used to fall through to the generic
        # check below and get a hard bad_secret 403 instead of reaching
        # cloud_sync()'s own, already-correct {"exists": false} response
        # -- which is exactly what looked like "cloud unreachable" on
        # both the phone (blank page) and the PC (device list stuck on
        # "retrying...") even though the service was up the whole time.
        row = _cloud_get_row(session_id)
        if row and row["secret_key"] != secret_key:
            return jsonify({"ok": False, "error": "bad_secret"}), 403
        g.session_id, g.secret_key = session_id, secret_key
        return None

    row = _cloud_get_row(session_id)
    if not row or row["secret_key"] != secret_key:
        return jsonify({"ok": False, "error": "bad_secret"}), 403
    g.session_id, g.secret_key = session_id, secret_key

    # Presence tracking for the cloud device roster -- every authenticated
    # request from a phone (not the PC's own push above, which never sends
    # X-Device-Id) refreshes its last-seen time here, not just its polls
    # to /api/lock-status, so "last synced" reflects real activity.
    device_id = request.headers.get("X-Device-Id", "")
    if device_id:
        if _cloud_device_kicked(session_id, device_id):
            return jsonify({"ok": False, "error": "kicked"}), 403
        admitted, canonical_id = _cloud_touch_device(
            session_id, device_id,
            request.headers.get("User-Agent"),
            request.headers.get("X-Device-Screen", ""),
            request.headers.get("X-Device-Fingerprint", ""))
        if not admitted:
            return jsonify({"ok": False, "error": "device_limit_reached",
                             "max_devices": MAX_DEVICES}), 403
        if canonical_id != device_id:
            g.canonical_device_id = canonical_id
    return None


@app.after_request
def _cors_headers(resp):
    """CLOUD_MODE only: cloudFetch() in the companion's JS is a genuine
    cross-origin request whenever a page is running locally with no
    direct-cloud pairing and falls back to reading/writing this cloud
    service's own origin for durability. Without these headers the
    browser silently blocks it (no visible error -- api() just falls
    through to a stale local cache), which is exactly the "shows old
    data instead of live cloud data while offline" symptom this fixes.
    A wildcard origin is safe here: these routes authenticate via the
    explicit X-Session-Id / X-Secret-Key headers above, never via
    cookies, so there's no session/cookie for a hostile origin to ride
    along on."""
    if CLOUD_MODE:
        resp.headers["Access-Control-Allow-Origin"] = "*"
        resp.headers["Access-Control-Allow-Headers"] = (
            "Content-Type, X-Session-Id, X-Secret-Key, X-Device-Id, X-Device-Fingerprint")
        resp.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
        canonical_id = getattr(g, "canonical_device_id", None)
        if canonical_id:
            # Tells the client its current X-Device-Id wasn't recognized as
            # new, but matched an already-known device's fingerprint (see
            # _cloud_touch_device) -- the client adopts this id from here
            # on instead of staying registered as a separate device.
            resp.headers["X-Canonical-Device-Id"] = canonical_id
            resp.headers["Access-Control-Expose-Headers"] = "X-Canonical-Device-Id"
    return resp


# ── low level helpers ───────────────────────────────────────────────────
def parse_amount(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = re.sub(r"[^\d.\-]", "", str(value))
    if s in ("", "-", ".", "-."):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def add_months(d, n):
    n = int(n)
    total_month_index = (d.month - 1) + n
    year = d.year + total_month_index // 12
    month = total_month_index % 12 + 1
    last_day = calendar.monthrange(year, month)[1]
    day = min(d.day, last_day)
    return d.replace(year=year, month=month, day=day)


def calculate_one_month_ahead(date_string):
    try:
        d = datetime.strptime(date_string.strip(), "%Y-%m-%d").date()
        return add_months(d, 1).strftime("%Y-%m-%d")
    except (ValueError, AttributeError):
        return ""


def hash_secret(secret):
    return hashlib.sha256(secret.encode()).hexdigest()


def _parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s.strip(), "%Y-%m-%d").date()
    except (ValueError, AttributeError):
        return None


def _compute_txn_period(rec):
    to_d = _parse_date(rec.get("to_date") or "")
    pay_d = _parse_date(rec.get("date") or "")
    if to_d is not None:
        from_d = add_months(to_d, -1)
        if pay_d is not None and not (from_d <= pay_d <= to_d):
            to_d = pay_d
            from_d = add_months(to_d, -1)
    elif pay_d is not None:
        to_d = pay_d
        from_d = add_months(to_d, -1)
    else:
        return "—", "—"
    return from_d.strftime("%Y-%m-%d"), to_d.strftime("%Y-%m-%d")


def fmt_period_date(d):
    """'24 jul,26' -- day, lowercase abbreviated month, comma, 2-digit year."""
    return f"{d.day} {d.strftime('%b').lower()},{d.strftime('%y')}"


def fmt_period(from_str, to_str):
    """'24 jul,26 to 24 aug,26' from two YYYY-MM-DD strings."""
    f, t = _parse_date(from_str), _parse_date(to_str)
    if not f or not t:
        return "—"
    return f"{fmt_period_date(f)} to {fmt_period_date(t)}"


def open_months_list(t, horizon=MONTH_PICKER_HORIZON):
    """The still-payable months for tenant t, starting at the first unpaid
    month (the tenant's current due date, or their move-in date if they've
    never paid), running forward `horizon` months. Each entry's `months`
    is how many consecutive months would be paid if the picker were
    confirmed with that entry as the last one ticked."""
    anchor = _parse_date(t.get("due_date", "")) or _parse_date(t.get("entry_date", "")) or date.today()
    out = []
    d = anchor
    for i in range(horizon):
        nxt = add_months(d, 1)
        out.append({
            "months": i + 1,
            "from": d.strftime("%Y-%m-%d"),
            "to": nxt.strftime("%Y-%m-%d"),
            "label": fmt_period(d.strftime("%Y-%m-%d"), nxt.strftime("%Y-%m-%d")),
        })
        d = nxt
    return out


def cleared_months_list(t):
    """The already-paid months, from the tenant's move-in month up to (not
    including) their current due date -- shown in the month picker as
    locked/checked-off history, since a cleared month can't be selected
    again."""
    entry = _parse_date(t.get("entry_date", ""))
    due = _parse_date(t.get("due_date", ""))
    out = []
    if not entry or not due or due <= entry:
        return out
    d = entry
    while d < due:
        nxt = add_months(d, 1)
        if nxt > due:
            nxt = due
        out.append({
            "from": d.strftime("%Y-%m-%d"),
            "to": nxt.strftime("%Y-%m-%d"),
            "label": fmt_period(d.strftime("%Y-%m-%d"), nxt.strftime("%Y-%m-%d")),
        })
        d = nxt
    return out


# ── data load / save ────────────────────────────────────────────────────
def _unique_backup_path(prefix, ext):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    candidate = os.path.join(BACKUP_DIR, f"{prefix}_{ts}{ext}")
    n = 1
    while os.path.exists(candidate):
        candidate = os.path.join(BACKUP_DIR, f"{prefix}_{ts}_{n}{ext}")
        n += 1
    return candidate


def backup_current_data_file():
    if not os.path.exists(DATA_FILE):
        return None
    backup_path = _unique_backup_path("data", ".json")
    shutil.copy2(DATA_FILE, backup_path)
    return backup_path


def _raw_load():
    if CLOUD_MODE:
        return _cloud_load(g.session_id)
    if not os.path.exists(DATA_FILE):
        return {"units": {}, "tenants": [], "settings": {}}
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        try:
            corrupt_copy = _unique_backup_path("data_CORRUPT", ".json")
            shutil.copy2(DATA_FILE, corrupt_copy)
        except Exception:
            pass
        return {"units": {}, "tenants": [], "settings": {}}


def _record_content_equal(a, b):
    """True if two tenant/unit records are the same ignoring the
    `_updated_at` bookkeeping field itself (added by
    _stamp_changed_records below), so re-saving an unchanged record
    never looks like an edit. Mirrors the identically-named helper in
    the desktop app (updates.py) so both sides stamp records the same
    way."""
    a2 = {k: v for k, v in (a or {}).items() if k != "_updated_at"}
    b2 = {k: v for k, v in (b or {}).items() if k != "_updated_at"}
    return a2 == b2


def _tenant_key(t):
    return (t.get("name"), t.get("unit"), t.get("entry_date"))


def _tenant_key_str(key):
    """String form of a _tenant_key() tuple, safe to use as a JSON object
    key (tombstone dicts are stored inside data.json, which requires
    string keys). Uses a separator that can't appear in a name/unit/date
    field entered through the UI."""
    return "\x1f".join("" if p is None else str(p) for p in key)


def stamp_tenant_deleted(data, t):
    """Record a tombstone for a tenant that's about to be permanently
    removed from data["tenants"], so a later cloud merge (_merge_cloud_data)
    knows this record was deliberately deleted here rather than treating
    the OTHER side's still-existing copy of it as something new that must
    be kept -- which is exactly what silently resurrected deleted tenants
    before this existed. Call this BEFORE removing the record from the
    list, and before save_raw()/save_state()."""
    tomb = data.setdefault("deleted_tenants", {})
    tomb[_tenant_key_str(_tenant_key(t))] = datetime.now(timezone.utc).isoformat()


def stamp_unit_deleted(data, name):
    """Record a tombstone for a unit that's about to be permanently
    removed from data["units"]. See stamp_tenant_deleted() above -- same
    reasoning, just keyed by unit name (already a plain string)."""
    tomb = data.setdefault("deleted_units", {})
    tomb[name] = datetime.now(timezone.utc).isoformat()


def _stamp_changed_records(new_data, old_data):
    """Stamps `_updated_at` (ISO now) onto every tenant/unit record in
    new_data that's brand new or whose content actually changed versus
    old_data; unchanged records keep whatever `_updated_at` they
    already had. Mutates new_data in place and returns it.

    Mirrors the desktop app's identically-named helper -- this is what
    lets the PC's cloud-merge logic (_merge_cloud_data in updates.py)
    tell, record by record, whether its own copy or the phone's is
    actually newer, instead of only being able to compare whole
    snapshots (which is what used to let one side's edit to a shared
    record get silently discarded whenever the other side's snapshot
    was picked as "the base")."""
    now_iso = datetime.now(timezone.utc).isoformat()
    old_tenants = {_tenant_key(t): t for t in (old_data.get("tenants") or [])}
    for t in (new_data.get("tenants") or []):
        old_t = old_tenants.get(_tenant_key(t))
        if old_t is None or not _record_content_equal(t, old_t):
            t["_updated_at"] = now_iso
        elif "_updated_at" not in t and old_t.get("_updated_at"):
            t["_updated_at"] = old_t["_updated_at"]

    old_units = old_data.get("units") or {}
    for uk, uv in (new_data.get("units") or {}).items():
        if not isinstance(uv, dict):
            continue
        old_u = old_units.get(uk)
        if old_u is None or not _record_content_equal(uv, old_u):
            uv["_updated_at"] = now_iso
        elif "_updated_at" not in uv and isinstance(old_u, dict) and old_u.get("_updated_at"):
            uv["_updated_at"] = old_u["_updated_at"]

    # Mirrors the desktop app's settings stamping -- see the identically
    # named function in updates.py for why this exists.
    new_settings = new_data.get("settings")
    if isinstance(new_settings, dict):
        old_settings = old_data.get("settings") if isinstance(old_data.get("settings"), dict) else {}
        if not _record_content_equal(new_settings, old_settings):
            new_settings["_updated_at"] = now_iso
        elif "_updated_at" not in new_settings and old_settings.get("_updated_at"):
            new_settings["_updated_at"] = old_settings["_updated_at"]

    return new_data


def _merge_cloud_data(a, b):
    """Record-level, order-independent merge of two full snapshots (`a`
    and `b` can be passed in either order -- the result is the same
    either way). Ported 1:1 from the desktop app's identically-named
    helper so the server can do the same safe merge the PC already does
    client-side, instead of the PUT /api/_sync handler picking an
    entire snapshot as "the winner" by comparing the pushing side's own
    wall-clock timestamp against the row's stored (server-clock)
    timestamp -- that comparison is what silently dropped every PC push
    once its clock drifted even a few seconds behind the database's,
    since accepted writes are always stamped with the DB's own now().

    For each tenant/unit key:
      - present in only one side -> keep it.
      - present in both -> keep whichever copy has the newer
        `_updated_at` stamp (see _stamp_changed_records).
      - a record missing `_updated_at` entirely is treated as maximally
        stale; if neither side has a stamp, `a`'s copy wins (arbitrary
        but stable).
    `settings` is treated as a single block and resolved by its own
    `_updated_at` stamp -- whichever side's settings were actually
    edited more recently wins outright, rather than `a` always winning
    regardless of recency.

    `deleted_tenants` / `deleted_units` are tombstones: {key: deleted_at
    iso timestamp}, written by stamp_tenant_deleted()/stamp_unit_deleted()
    at the moment a record is permanently removed on either side (see
    those functions). Without this, a record present on only one side
    was always just "kept" -- indistinguishable from a genuinely NEW
    record the other side hadn't seen yet -- which is what silently
    resurrected a tenant/unit deleted on one device the next time it
    merged against the other device's still-intact copy. The fix: merge
    the tombstones themselves (newest timestamp per key wins, union of
    both sides), then drop any record whose key has a tombstone dated at
    or after that record's own `_updated_at` (a record with no stamp at
    all is treated as older than any tombstone, same "maximally stale"
    rule used elsewhere in this function) -- so a genuine edit made
    AFTER the deletion (e.g. the record was deleted, then re-added fresh
    with the same name/unit) still wins, but a stale copy that predates
    the deletion does not come back from the dead."""
    def _tenants_by_key(d):
        return {(t.get("name"), t.get("unit"), t.get("entry_date")): t
                for t in (d.get("tenants") or [])}

    a_del_t, b_del_t = (a.get("deleted_tenants") or {}), (b.get("deleted_tenants") or {})
    merged_del_t = dict(a_del_t)
    for k, ts in b_del_t.items():
        if k not in merged_del_t or _iso_ge(ts, merged_del_t[k]):
            merged_del_t[k] = ts

    a_del_u, b_del_u = (a.get("deleted_units") or {}), (b.get("deleted_units") or {})
    merged_del_u = dict(a_del_u)
    for k, ts in b_del_u.items():
        if k not in merged_del_u or _iso_ge(ts, merged_del_u[k]):
            merged_del_u[k] = ts

    def _tombstoned(record, tomb_ts):
        return tomb_ts is not None and _iso_ge(tomb_ts, record.get("_updated_at", ""))

    a_tenants, b_tenants = _tenants_by_key(a), _tenants_by_key(b)
    merged_tenants = []
    for key, ta in a_tenants.items():
        tb = b_tenants.get(key)
        winner = (ta if tb is None or _iso_ge(ta.get("_updated_at", ""), tb.get("_updated_at", "")) else tb)
        if _tombstoned(winner, merged_del_t.get(_tenant_key_str(key))):
            continue
        merged_tenants.append(winner)
    for key, tb in b_tenants.items():
        if key not in a_tenants:
            if _tombstoned(tb, merged_del_t.get(_tenant_key_str(key))):
                continue
            merged_tenants.append(tb)

    a_units, b_units = (a.get("units") or {}), (b.get("units") or {})
    merged_units = {}
    for uk, ua in a_units.items():
        ub = b_units.get(uk)
        if not isinstance(ua, dict):
            winner = ub if ub is not None else ua
        elif not isinstance(ub, dict):
            winner = ua
        else:
            winner = (ua if _iso_ge(ua.get("_updated_at", ""), ub.get("_updated_at", "")) else ub)
        if isinstance(winner, dict) and _tombstoned(winner, merged_del_u.get(uk)):
            continue
        merged_units[uk] = winner
    for uk, ub in b_units.items():
        if uk not in a_units:
            if isinstance(ub, dict) and _tombstoned(ub, merged_del_u.get(uk)):
                continue
            merged_units[uk] = ub

    a_settings, b_settings = (a.get("settings") or {}), (b.get("settings") or {})
    merged_settings = (
        a_settings if not b_settings or
        _iso_ge(a_settings.get("_updated_at", ""), b_settings.get("_updated_at", ""))
        else b_settings)

    return {
        "tenants": merged_tenants,
        "units": merged_units,
        "settings": merged_settings,
        "deleted_tenants": merged_del_t,
        "deleted_units": merged_del_u,
    }


def save_raw(data, updated_by=None):
    try:
        old_data = _raw_load()
    except Exception:
        old_data = {"units": {}, "tenants": [], "settings": {}}
    _stamp_changed_records(data, old_data)
    if CLOUD_MODE:
        return _cloud_save(g.session_id, data, updated_by=updated_by or "cloud")
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    # Same fix as the desktop app's save_data(): export_excel/export_pdf
    # rebuild the whole workbook/document from scratch, which is real
    # work -- doing it synchronously here would hold the HTTP response
    # (and so the phone's UI) hostage until both finished, on every
    # single save. The JSON write above is what everything else actually
    # depends on, so the response can go back now; exports finish a
    # moment later in the background. Deepcopy for the same reason as
    # the desktop side: `data` may still be referenced/mutated by the
    # caller after this function returns.
    export_snapshot = copy.deepcopy(data)

    def _export_worker():
        try:
            export_excel(export_snapshot)
        except Exception:
            pass
        try:
            export_pdf(export_snapshot)
        except Exception:
            pass

    threading.Thread(target=_export_worker, daemon=True).start()
    return None


def load_state():
    """Load data.json and run the same startup migrations the desktop app
    runs, returning a plain dict with 'units', 'tenants', 'settings'."""
    data = _raw_load()
    data.setdefault("units", {})
    data.setdefault("tenants", [])
    data.setdefault("settings", {})
    changed = False

    # normalize rent fields
    for t in data["tenants"]:
        raw_rent = t.get("rent", 0)
        if not isinstance(raw_rent, (int, float)):
            t["rent"] = parse_amount(raw_rent)
            changed = True
    for _name, info in data["units"].items():
        if isinstance(info, dict):
            raw_rent = info.get("rent", 0)
            if not isinstance(raw_rent, (int, float)):
                info["rent"] = parse_amount(raw_rent)
                changed = True

    # backfill schema fields
    for t in data["tenants"]:
        for field, default in TENANT_FIELD_DEFAULTS.items():
            if field not in t:
                t[field] = default() if callable(default) else default
                changed = True

    # apply due rent increases
    today = date.today()
    for unit_name, info in data["units"].items():
        if not isinstance(info, dict):
            continue
        pending = info.get("pending_rent_increase")
        if not pending:
            continue
        eff = _parse_date(str(pending.get("effective_month", "")))
        if eff is None:
            info.pop("pending_rent_increase", None)
            changed = True
            continue
        if eff <= today:
            new_rent = parse_amount(pending.get("new_rent", 0))
            info["rent"] = new_rent
            info.pop("pending_rent_increase", None)
            changed = True
            for t in data["tenants"]:
                if t.get("unit") == unit_name:
                    old_tenant_rent = parse_amount(t.get("rent", 0))
                    _bill_rent_increase_shortfall(t, old_tenant_rent, new_rent, eff)
                    t["rent"] = new_rent

    # apply deferred unit-change rent, once its effective (due) date arrives
    # -- mirrors the desktop app's apply_due_tenant_unit_changes: a tenant
    # moved to a different unit while still paid ahead keeps billing at
    # the old rate until that due date, then the new rate kicks in here.
    for t in data["tenants"]:
        pending = t.get("pending_unit_change")
        if not pending:
            continue
        eff = _parse_date(str(pending.get("effective_date", "")))
        if eff is None:
            t.pop("pending_unit_change", None)
            changed = True
            continue
        if eff <= today:
            t["rent"] = parse_amount(pending.get("rent", t.get("rent", 0)))
            t.pop("pending_unit_change", None)
            changed = True

    # auto revert expired "Confirmed" tenants back to Pending
    for t in data["tenants"]:
        if t.get("status") != "Confirmed":
            continue
        due = _parse_date(t.get("due_date", ""))
        if due and due < today:
            t["status"] = "Pending"
            t["pay_date"] = ""
            changed = True

    # deposit-in-progress tenants whose current month is already covered
    # should show as Confirmed, mirroring desktop app startup pass
    for t in data["tenants"]:
        if t.get("status") == "Confirmed":
            continue
        _, _, cleared, _ = current_deposit_cycle(t)
        if cleared:
            t["status"] = "Confirmed"
            if not t.get("pay_date"):
                t["pay_date"] = today.strftime("%Y-%m-%d")
            changed = True

    if changed:
        save_raw(data)
    return data


def save_state(data):
    save_raw(data)


# ── tenant status / deposit-cycle logic (ported 1:1 from desktop app) ──
def current_deposit_cycle(t):
    rent_target = parse_amount(t.get("rent", 0))
    full_history = t.get("deposit_history", [])
    cycle_start = t.get("deposit_cycle_start", 0)
    if not isinstance(cycle_start, int) or cycle_start < 0 or cycle_start > len(full_history):
        cycle_start = 0
    cycle_window = full_history[cycle_start:]
    active_records = [r for r in cycle_window if not r.get("_cancelled")]
    cancelled_records = [r for r in cycle_window if r.get("_cancelled")]

    paid_so_far = sum(float(r.get("amount", 0)) for r in active_records)
    current_bal = max(0.0, rent_target - paid_so_far)

    if cancelled_records:
        cancelled_sum = sum(float(r.get("amount", 0)) for r in cancelled_records)
        remaining = min(current_bal + cancelled_sum, rent_target)
        paid_so_far = max(0.0, rent_target - remaining)
        cleared = remaining <= 0
        in_progress = bool(active_records) and not cleared
    else:
        remaining = current_bal
        cleared = paid_so_far >= rent_target and rent_target > 0
        in_progress = bool(active_records) and not cleared
    return paid_so_far, remaining, cleared, in_progress


def deposit_paid_so_far(t):
    """Sum of active (non-cancelled) deposits since the last time the
    deposit cycle cleared. All deposits in this window are toward the same
    open month-window by construction (there's a single running window
    now, not separate current/next/multiple buckets), so this no longer
    needs to filter by a period label."""
    full_history = t.get("deposit_history", [])
    cycle_start = t.get("deposit_cycle_start", 0)
    if not isinstance(cycle_start, int) or cycle_start < 0 or cycle_start > len(full_history):
        cycle_start = 0
    cycle_window = full_history[cycle_start:]
    return sum(float(r.get("amount", 0)) for r in cycle_window if not r.get("_cancelled"))


def is_current_period_paid(t):
    if t.get("status") != "Confirmed":
        return False
    due = _parse_date(t.get("due_date", ""))
    if due is None:
        return False
    return due >= date.today()


def snapshot_tenant_state(t):
    return {
        "due_date": t.get("due_date", ""),
        "status": t.get("status", "Pending"),
        "pay_date": t.get("pay_date", ""),
        "locked_periods": list(t.get("locked_periods", [])),
        "deposit_cycle_start": t.get("deposit_cycle_start", 0),
        "rent_increase_due": t.get("rent_increase_due", 0),
    }


def restore_tenant_state(t, snap):
    t["due_date"] = snap.get("due_date", "")
    t["status"] = snap.get("status", "Pending")
    if snap.get("pay_date"):
        t["pay_date"] = snap["pay_date"]
    else:
        t.pop("pay_date", None)
    if snap.get("locked_periods"):
        t["locked_periods"] = list(snap["locked_periods"])
    else:
        t.pop("locked_periods", None)
    t["deposit_cycle_start"] = snap.get("deposit_cycle_start", 0)
    t["rent_increase_due"] = snap.get("rent_increase_due", 0)


def has_prior_payment_history(t):
    return bool(t.get("payment_history")) or bool(t.get("deposit_history"))


def due_date_shift_base(t, has_prior_history=None):
    if has_prior_history is None:
        has_prior_history = has_prior_payment_history(t)
    base = t.get("due_date", "") if has_prior_history else t.get("entry_date", "")
    return base or date.today().strftime("%Y-%m-%d")


def pending_reference_date_str(t):
    return t.get("due_date", "") or t.get("entry_date", "")


def status_level(t, today):
    """Returns (level, label) where level is 'paid' | 'underpaid' | 'pending'."""
    due = _parse_date(t.get("due_date", ""))
    if t.get("status") == "Confirmed":
        if due is not None and due < today:
            return "pending", "Pending"
        return "paid", "Paid"
    _, _, dep_cleared, dep_in_progress = current_deposit_cycle(t)
    if dep_cleared:
        if due is not None and due < today:
            return "pending", "Pending"
        return "paid", "Paid"
    if dep_in_progress:
        return "underpaid", "Installments"
    return "pending", "Pending"


def display_status_level(t, today):
    """Display-only refinement of status_level: splits the 'pending'
    bucket into 'pending' (never paid / not yet due) and 'overdue' (due
    date has already passed), so the Tenants list can show a distinct
    red "Overdue" chip and filter. Business logic (status_level,
    arrears, cascades, dashboard counts, etc.) is untouched -- this only
    decides how a tenant is *labelled* in the UI."""
    level, label = status_level(t, today)
    if level == "pending":
        due = _parse_date(t.get("due_date", ""))
        if due is not None and due < today:
            return "overdue", "Overdue"
    return level, label


def dashboard_recent_activity(data, today=None, limit=10):
    """Feed for the Home dashboard's Recent Activity card: a single
    reverse-chronological timeline of real actions taken in the app --
    payments and deposits received, cancelled transactions, tenants
    moved out, and rent increases scheduled -- newest first, capped at
    `limit`. Verbatim logic port of the desktop app's
    dashboard_recent_activity (icon/color are left to the client here,
    keyed off "tag"; everything else -- title, subtitle, time_label,
    tag, and the ordering/fallback rules -- matches exactly)."""
    today = today or date.today()
    tenants = data.get("tenants", [])
    events = []

    for t in tenants:
        name = t.get("name", "Unnamed Tenant")
        unit = t.get("unit") or "—"

        for hist_key, label in (("payment_history", "Full Payment"), ("deposit_history", "Deposit")):
            for r in t.get(hist_key, []):
                d = r.get("date", "")
                if not d:
                    continue
                try:
                    amt = f"UGX {int(float(r.get('amount', 0))):,}"
                except (ValueError, TypeError):
                    amt = ""
                if r.get("_cancelled"):
                    events.append({
                        "_date": d, "title": f"{label} Cancelled",
                        "subtitle": f"Unit {unit} · {name}" + (f" · {amt}" if amt else ""),
                        "time_label": d, "tag": "CANCELLED",
                    })
                else:
                    events.append({
                        "_date": d, "title": f"{label} Received",
                        "subtitle": f"Unit {unit} · {name}" + (f" · {amt}" if amt else ""),
                        "time_label": d, "tag": "PAID",
                    })

        if t.get("archived") and t.get("vacated_date"):
            events.append({
                "_date": t["vacated_date"], "title": "Tenant Moved Out",
                "subtitle": f"Unit {unit} · {name}",
                "time_label": t["vacated_date"], "tag": "ARCHIVED",
            })

    for name, info in data.get("units", {}).items():
        pending = info.get("pending_rent_increase") if isinstance(info, dict) else None
        if pending:
            eff = pending.get("effective_month") or pending.get("effective_date") or ""
            try:
                new_rent = f"UGX {int(float(pending.get('new_rent', 0))):,}"
            except (ValueError, TypeError):
                new_rent = ""
            events.append({
                "_date": eff or today.strftime("%Y-%m-%d"),
                "title": "Rent Increase Scheduled",
                "subtitle": f"Unit {name}" + (f" · {new_rent}" if new_rent else ""),
                "time_label": f"Effective {eff}" if eff else "Scheduled",
                "tag": "SCHEDULED",
            })

    events.sort(key=lambda e: e["_date"], reverse=True)

    if not events:
        overdue = []
        for t in tenants:
            if t.get("archived"):
                continue
            level, _ = display_status_level(t, today)
            if level == "overdue":
                overdue.append((t.get("due_date", ""), t))
        overdue.sort(key=lambda pair: pair[0])
        for due_str, t in overdue[:limit]:
            events.append({
                "_date": due_str, "title": "Rent Overdue",
                "subtitle": f"Unit {t.get('unit') or '—'} · {t.get('name', 'Unnamed Tenant')}",
                "time_label": f"Due {due_str}" if due_str else "Overdue",
                "tag": "OVERDUE",
            })

    return events[:limit]


def rent_increase_due(t):
    try:
        return max(0.0, float(t.get("rent_increase_due", 0) or 0))
    except (TypeError, ValueError):
        return 0.0


def _bill_rent_increase_shortfall(t, old_rent, new_rent, effective):
    if new_rent <= old_rent:
        return
    due_d = _parse_date(t.get("due_date", ""))
    if due_d is None:
        return
    months_prepaid_past_effective = (
        (due_d.year - effective.year) * 12 + (due_d.month - effective.month)
    )
    if months_prepaid_past_effective <= 0:
        return
    shortfall = (new_rent - old_rent) * months_prepaid_past_effective
    if shortfall > 0:
        t["rent_increase_due"] = float(t.get("rent_increase_due", 0) or 0) + shortfall


def apply_tenant_unit_change(record, old_unit, old_rent, due_str, new_unit, new_rent):
    """Moves a tenant to `new_unit`, deciding whether the accompanying rent
    change lands right away or is deferred. Verbatim logic port of the
    desktop app's apply_tenant_unit_change: a first-ever assignment or a
    move to a same-priced unit takes effect immediately; otherwise, if the
    tenant is still paid ahead (due date in the future), the old rent
    keeps billing until that due date and the new rate is stashed in
    record['pending_unit_change'] for load_state() to pick up later."""
    record["unit"] = new_unit
    if not old_unit or new_rent == old_rent:
        record["rent"] = new_rent
        record.pop("pending_unit_change", None)
        return
    paid_ahead = False
    due_d = _parse_date(due_str)
    if due_d is not None:
        paid_ahead = due_d > date.today()
    if paid_ahead:
        record["rent"] = old_rent
        record["pending_unit_change"] = {"rent": new_rent, "effective_date": due_str}
    else:
        record["rent"] = new_rent
        record.pop("pending_unit_change", None)


def add_old_data_records(t, records, final_state=None):
    """Backfills a tenant's pre-existing rental history in one go, for a
    tenant who was already renting before this data was tracked here.
    Verbatim logic port of the desktop app's add_old_data_records:
    history entered here looks and behaves identically to a normal
    recorded payment/deposit, minus the '_pre_state' snapshot a live
    transaction gets. `final_state` optionally sets the tenant's current
    status/due_date/pay_date/notes once every historical record is added.
    Returns {"added": n}."""
    added = 0
    today_str = date.today().strftime("%Y-%m-%d")
    for rec in records or []:
        kind = "deposit_history" if rec.get("type") == "deposit" else "payment_history"
        txn_date = (rec.get("date") or "").strip()
        if _parse_date(txn_date) is None:
            continue  # skip malformed rows rather than fail the whole batch
        if txn_date >= today_str:
            continue  # old data must predate today
        from_date = (rec.get("from_date") or "").strip()
        if from_date and _parse_date(from_date) is None:
            from_date = ""
        to_date = (rec.get("to_date") or "").strip()
        if to_date and _parse_date(to_date) is None:
            to_date = ""
        if from_date:
            # "To" always equals exactly one month after "From" -- recompute
            # rather than trust whatever the client sent, so the two can
            # never drift apart.
            to_date = add_months(_parse_date(from_date), 1).strftime("%Y-%m-%d")
        elif not from_date and to_date:
            from_date = add_months(_parse_date(to_date), -1).strftime("%Y-%m-%d")
        if from_date and from_date >= today_str:
            continue
        if to_date and to_date >= today_str:
            continue
        amount = parse_amount(rec.get("amount", 0))
        note = (rec.get("note") or "").strip()
        entry = {
            "date": txn_date, "months": max(0, int(rec.get("months", 0) or 0)),
            "amount": amount, "from_date": from_date, "to_date": to_date,
            "txn_id": f"old-data-{secrets.token_hex(4)}",
            "note": note or "Old data entered by admin.",
            "_backfilled": True,
        }
        if bool(rec.get("cancelled")):
            entry["_cancelled"] = True
            entry["cancelled_on"] = (rec.get("cancelled_on") or "").strip() or txn_date
        t.setdefault(kind, []).append(entry)
        added += 1

    # Keep each history array in chronological order so it displays and
    # exports the same as if these had been entered one at a time as they
    # actually happened, instead of clumped at the end in entry order.
    for key in ("payment_history", "deposit_history"):
        t[key] = sorted(t.get(key, []), key=lambda r: r.get("date", ""))

    if final_state:
        due_str = (final_state.get("due_date") or "").strip()
        if due_str and _parse_date(due_str) is not None:
            t["due_date"] = due_str
        status = (final_state.get("status") or "").strip()
        if status in ("Confirmed", "Pending"):
            t["status"] = status
        pay_date = (final_state.get("pay_date") or "").strip()
        if pay_date and _parse_date(pay_date) is not None:
            t["pay_date"] = pay_date
        notes = final_state.get("notes")
        if notes is not None:
            existing = (t.get("notes") or "").strip()
            addition = str(notes).strip()
            if addition:
                t["notes"] = (existing + "\n" + addition).strip() if existing else addition

    return {"added": added}


def record_arrears_payment(t, amount, method):
    pay_date = date.today().strftime("%Y-%m-%d")
    t.setdefault("arrears_history", []).append({
        "date": pay_date, "amount": amount, "method": method,
        "type": "Arrears", "txn_id": secrets.token_hex(4),
    })
    t["rent_increase_due"] = max(0.0, rent_increase_due(t) - amount)


# ── excess cascade (non-interactive web version) ────────────────────────
def apply_excess_cascade(t, excess, rent_target, base_due_str, pay_date, txn_id=None):
    """Web-simplified version of the desktop app's interactive cascade: a
    month is auto-recorded as a Full Payment whenever the remaining excess
    covers it completely, otherwise as a partial Deposit toward that month
    (matching the same rule the desktop dialog uses to decide whether
    "Full" is even offered). Returns the number of months the due date
    was advanced by."""
    if excess is None or excess <= 0 or rent_target is None or rent_target <= 0:
        return 0
    cur_due = _parse_date(base_due_str)
    if cur_due is None:
        return 0

    remaining = excess
    last_cleared = False
    full_months_shift = 0
    while remaining > 0.01:
        next_due = add_months(cur_due, 1)
        chunk = min(remaining, rent_target)
        allow_full = chunk >= rent_target - 0.01

        if allow_full:
            t.setdefault("payment_history", []).append({
                "date": pay_date, "months": 1,
                "amount": chunk,
                "from_date": cur_due.strftime("%Y-%m-%d"),
                "to_date": next_due.strftime("%Y-%m-%d"),
                "txn_id": txn_id,
            })
            cur_due = next_due
            t["due_date"] = cur_due.strftime("%Y-%m-%d")
            t["status"] = "Confirmed"
            t["pay_date"] = pay_date
            full_months_shift += 1
            last_cleared = True
            t["deposit_cycle_start"] = len(t.get("deposit_history", []))
        else:
            t.setdefault("deposit_history", []).append({
                "date": pay_date, "months": 1,
                "amount": chunk, "txn_id": txn_id,
                "from_date": cur_due.strftime("%Y-%m-%d"),
                "to_date": next_due.strftime("%Y-%m-%d"),
                "target_month": next_due.strftime("%Y-%m-%d"),
            })
            t["deposit_cycle_start"] = len(t["deposit_history"]) - 1
            last_cleared = False
        remaining -= chunk

    t["status"] = "Confirmed" if last_cleared else "Pending"
    return full_months_shift


# ── payments / deposits ─────────────────────────────────────────────────
def record_payment(t, months):
    """Mutates tenant dict t in place. `months` is the number of
    consecutive open months ticked in the month picker (1 = just the next
    due month, n = that month plus the n-1 months after it). Returns dict
    describing the result."""
    months = max(1, int(months))

    rent = parse_amount(t.get("rent", 0))
    total = rent * months

    # Net off any partial deposit/instalment already sitting on the
    # account for the current cycle -- e.g. the tenant put down half of
    # this month's rent earlier, then comes back to pay the month off in
    # full. Without this, "amount" below would always be rent*months
    # regardless of what's already been collected, double-charging for
    # whatever the deposit covered. Only the true genuine leftover (if
    # the deposit alone somehow exceeds what's being cleared right now)
    # still cascades forward onto future months, same as before.
    pre_paid, _, _, _ = current_deposit_cycle(t)
    amount_due_now = max(0.0, total - pre_paid)
    excess_after = max(0.0, pre_paid - total)
    credited_from_deposit = pre_paid - excess_after

    pre_txn_state = snapshot_tenant_state(t)
    txn_id = secrets.token_hex(4)

    old_due_str = t.get("due_date", "")
    shift_base_str = due_date_shift_base(t)
    new_due_str = old_due_str
    old_due = _parse_date(shift_base_str)
    if old_due:
        new_due_str = add_months(old_due, months).strftime("%Y-%m-%d")

    pay_date = date.today().strftime("%Y-%m-%d")
    t["pay_date"] = pay_date
    t["status"] = "Confirmed"
    if new_due_str:
        t["due_date"] = new_due_str

    record_to_date = new_due_str
    record_from_date = shift_base_str if record_to_date else ""

    t.setdefault("payment_history", []).append({
        "date": pay_date, "months": months, "amount": amount_due_now,
        "from_date": record_from_date, "to_date": record_to_date,
        "txn_id": txn_id, "_pre_state": pre_txn_state,
        "credited_from_deposit": credited_from_deposit,
    })

    if pre_paid > 0:
        t["deposit_cycle_start"] = len(t.get("deposit_history", []))
        if excess_after > 0:
            apply_excess_cascade(t, excess_after, rent, t["due_date"], pay_date, txn_id=txn_id)

    return {"amount": amount_due_now, "months": months,
            "due_date": t["due_date"], "old_due_date": old_due_str,
            "period_label": fmt_period(record_from_date, record_to_date),
            "credited_from_deposit": credited_from_deposit}


def record_deposit(t, months, instalment):
    """`months` is the number of consecutive open months ticked in the
    month picker; the instalment accumulates toward rent x months."""
    months = max(1, int(months))

    rent_target = parse_amount(t.get("rent", 0))
    dep_paid_so_far = deposit_paid_so_far(t)
    effective_target = rent_target * months
    balance_before = max(0.0, effective_target - dep_paid_so_far)
    excess = max(0.0, instalment - balance_before)
    applied_amount = instalment - excess
    new_balance = max(0.0, effective_target - dep_paid_so_far - applied_amount)
    pay_date = date.today().strftime("%Y-%m-%d")

    had_prior_history = has_prior_payment_history(t)
    pre_txn_state = snapshot_tenant_state(t)
    txn_id = secrets.token_hex(4)

    shift_base_str = due_date_shift_base(t, had_prior_history)
    target_to_str = ""
    shift_base_d = _parse_date(shift_base_str)
    if shift_base_d:
        target_to_str = add_months(shift_base_d, months).strftime("%Y-%m-%d")

    t.setdefault("deposit_history", []).append({
        "date": pay_date, "months": months, "amount": applied_amount,
        "from_date": shift_base_str, "to_date": target_to_str,
        "txn_id": txn_id, "_pre_state": pre_txn_state,
    })

    if new_balance <= 0:
        prev_start = t.get("deposit_cycle_start", 0)
        t["deposit_history"][-1]["_cycle_start_before_clear"] = prev_start
        t["deposit_cycle_start"] = len(t["deposit_history"])

        t["due_date"] = target_to_str or t.get("due_date", "")
        t["status"] = "Confirmed"
        t["pay_date"] = pay_date

    if excess > 0.01:
        apply_excess_cascade(t, excess, rent_target, t.get("due_date") or shift_base_str,
                              pay_date, txn_id=txn_id)

    return {"amount": applied_amount, "excess": excess, "new_balance": new_balance,
            "cleared": new_balance <= 0, "due_date": t.get("due_date", ""),
            "period_label": fmt_period(shift_base_str, target_to_str)}


def cancel_transaction(t, h_key, idx):
    """Cancel a payment/deposit record at t[h_key][idx], reversing every
    linked record that shares its txn_id (or falling back to legacy single
    record logic for records saved before txn_id existed)."""
    history = t.get(h_key, [])
    if idx < 0 or idx >= len(history):
        return None
    rec = history[idx]
    if rec.get("_cancelled"):
        return "already_cancelled"
    txn_id = rec.get("txn_id")

    linked = []
    if txn_id:
        for key in ("payment_history", "deposit_history"):
            for i, r in enumerate(t.get(key, [])):
                if r.get("txn_id") == txn_id and not r.get("_cancelled"):
                    linked.append((key, i, r))
    else:
        linked = [(h_key, idx, rec)]

    cancelled_on = date.today().strftime("%Y-%m-%d")
    pre_state = None
    for key, i, r in linked:
        entry = t.get(key, [])[i]
        entry["_cancelled"] = True
        entry["_type_origin"] = "deposit" if key == "deposit_history" else "payment"
        entry["cancelled_on"] = cancelled_on
        if "_pre_state" in entry and pre_state is None:
            pre_state = entry["_pre_state"]

    if pre_state is not None:
        restore_tenant_state(t, pre_state)
    else:
        # Legacy record saved before _pre_state existed -- best effort.
        if h_key == "deposit_history":
            cycle_start = t.get("deposit_cycle_start", 0)
            if isinstance(cycle_start, int) and cycle_start > idx:
                prev_start = rec.get("_cycle_start_before_clear", idx)
                t["deposit_cycle_start"] = prev_start
        t["status"] = "Pending"
        t.pop("pay_date", None)
        t.pop("locked_periods", None)

    total_amt = sum(float(r.get("amount", 0)) for _, _, r in linked)
    return {"n_records": len(linked), "total_amount": total_amt}


MONTHS = ["January", "February", "March", "April", "May", "June", "July",
          "August", "September", "October", "November", "December"]


def build_monthly_report(data, year, month):
    """Aggregates payment/deposit history for one calendar month across all
    tenants -- the same "Monthly Report" the desktop app's Reports screen
    generates, which had no equivalent here yet. Returns per-tenant summary
    rows, a flat list of individual transactions for the detail table, and
    the month's grand totals (active vs cancelled)."""
    tenants = data.get("tenants", [])
    prefix = f"{year:04d}-{month:02d}"
    month_label = f"{MONTHS[month - 1]} {year}"

    tenant_rows = []
    detail_rows = []
    grand_pay = grand_dep = grand_cancelled = 0

    for t in tenants:
        t_name = t.get("name", "—")
        t_unit = t.get("unit", "—")
        pay_active = pay_cancelled = 0
        dep_active = dep_cancelled = 0

        for rec in t.get("payment_history", []):
            if not rec.get("date", "").startswith(prefix):
                continue
            amt = int(rec.get("amount", 0))
            is_c = rec.get("_cancelled", False)
            ft, tt = _compute_txn_period(rec)
            detail_rows.append({
                "name": t_name, "unit": t_unit, "date": rec.get("date", "—"),
                "txn_type": "Full Payment" + (" (Cancelled)" if is_c else ""),
                "amount": amt, "from_d": ft, "to_d": tt,
                "is_cancelled": is_c, "cancelled_on": rec.get("cancelled_on", "—"),
            })
            if is_c:
                pay_cancelled += amt
            else:
                pay_active += amt

        for rec in t.get("deposit_history", []):
            if not rec.get("date", "").startswith(prefix):
                continue
            amt = float(rec.get("amount", 0))
            is_c = rec.get("_cancelled", False)
            ft, tt = _compute_txn_period(rec)
            detail_rows.append({
                "name": t_name, "unit": t_unit, "date": rec.get("date", "—"),
                "txn_type": "Deposit" + (" (Cancelled)" if is_c else ""),
                "amount": int(amt), "from_d": ft, "to_d": tt,
                "is_cancelled": is_c, "cancelled_on": rec.get("cancelled_on", "—"),
            })
            if is_c:
                dep_cancelled += amt
            else:
                dep_active += amt

        grand_pay += pay_active
        grand_dep += dep_active
        grand_cancelled += pay_cancelled + dep_cancelled

        if pay_active + dep_active + pay_cancelled + dep_cancelled > 0:
            tenant_rows.append({
                "name": t_name, "unit": t_unit,
                "pay_active": pay_active, "dep_active": dep_active,
                "pay_cancelled": pay_cancelled, "dep_cancelled": dep_cancelled,
            })

    detail_rows.sort(key=lambda r: (r["name"], r["date"]))
    return {
        "prefix": prefix,
        "month_label": month_label,
        "tenant_rows": tenant_rows,
        "detail_rows": detail_rows,
        "grand_pay": grand_pay,
        "grand_dep": grand_dep,
        "grand_cancelled": grand_cancelled,
        "grand_combined": grand_pay + grand_dep,
    }


def export_monthly_excel(report):
    """Writes a single "Monthly Transactions" sheet for one month, styled to
    match the desktop app's dedicated monthly export (separate from the
    all-time Transaction History workbook produced by export_excel())."""
    C_WHITE, C_GREY_H, C_CANCEL, C_TOTAL = "FFFFFF", "DDDDDD", "FFF0F0", "F0F7FF"

    def side():
        s = Side(style="thin", color="000000")
        return Border(left=s, right=s, top=s, bottom=s)

    def bold(sz=11):
        return Font(name="Calibri", bold=True, size=sz, color="000000")

    def reg(sz=10):
        return Font(name="Calibri", size=sz, color="000000")

    def strike(sz=10):
        return Font(name="Calibri", size=sz, color="000000", strike=True)

    def ctr():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def lft():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Monthly — {report['month_label']}"[:31]
    ws.sheet_view.showGridLines = True

    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = f"Monthly Transactions — {report['month_label']}"
    title.alignment = ctr()
    title.fill = PatternFill("solid", fgColor="1A73E8")
    title.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws.row_dimensions[1].height = 30

    hdrs = ["#", "TENANT", "UNIT", "DATE", "TYPE", "AMOUNT (UGX)", "FROM", "TO", "NOTE"]
    widths = [5, 24, 10, 14, 18, 20, 14, 14, 22]
    for ci, (h, cw) in enumerate(zip(hdrs, widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = bold(10)
        cell.fill = PatternFill("solid", fgColor=C_GREY_H)
        cell.alignment = ctr()
        cell.border = side()
        ws.column_dimensions[get_column_letter(ci)].width = cw
    ws.row_dimensions[2].height = 22

    row_num = 3
    for seq, dr in enumerate(report["detail_rows"], 1):
        is_c = dr["is_cancelled"]
        fgc = C_WHITE if not is_c else C_CANCEL
        note = f"Cancelled on {dr['cancelled_on']}" if is_c else ""
        row_vals = [
            (seq, ctr()), (dr["name"], lft()), (dr["unit"], ctr()),
            (dr["date"], ctr()), (dr["txn_type"], ctr()), (dr["amount"], ctr()),
            (dr["from_d"], ctr()), (dr["to_d"], ctr()), (note, lft()),
        ]
        for ci, (val, aln) in enumerate(row_vals, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.fill = PatternFill("solid", fgColor=fgc)
            cell.border = side()
            cell.alignment = aln
            cell.font = strike(10) if is_c else (bold(10) if ci in (2, 6) else reg(10))
        ws.row_dimensions[row_num].height = 20
        row_num += 1

    if row_num == 3:
        ws.cell(row=3, column=1, value="No transactions recorded this month.").font = reg(11)
        row_num += 1

    ws.merge_cells(f"A{row_num}:E{row_num}")
    tc = ws.cell(row=row_num, column=1, value="TOTALS")
    tc.font, tc.alignment, tc.border = bold(11), ctr(), side()
    tc.fill = PatternFill("solid", fgColor=C_TOTAL)
    for ci in range(2, 6):
        c2 = ws.cell(row=row_num, column=ci)
        c2.fill, c2.border = PatternFill("solid", fgColor=C_TOTAL), side()
    pay_cell = ws.cell(row=row_num, column=6,
                        value=f"Pay: {report['grand_pay']:,}  Dep: {int(report['grand_dep']):,}  "
                              f"Cancelled: {int(report['grand_cancelled']):,}")
    pay_cell.font, pay_cell.alignment, pay_cell.border = bold(10), ctr(), side()
    pay_cell.fill = PatternFill("solid", fgColor=C_TOTAL)
    for ci in range(7, 10):
        c3 = ws.cell(row=row_num, column=ci, value="")
        c3.fill, c3.border = PatternFill("solid", fgColor=C_TOTAL), side()
    ws.row_dimensions[row_num].height = 22

    monthly_path = os.path.join(DATA_DIR, f"monthly_{report['prefix'].replace('-', '_')}.xlsx")
    wb.save(monthly_path)
    return monthly_path


# ── exports (ported near-verbatim from desktop app) ─────────────────────
def export_excel(data):
    wb = openpyxl.Workbook()
    C_WHITE = "FFFFFF"
    C_BLACK = "000000"

    def thin_border():
        s = Side(style="thin", color="000000")
        return Border(left=s, right=s, top=s, bottom=s)

    def bold(sz=11):
        return Font(name="Calibri", bold=True, size=sz, color=C_BLACK)

    def reg(sz=10):
        return Font(name="Calibri", size=sz, color=C_BLACK)

    def cancelled_font(sz=10):
        return Font(name="Calibri", size=sz, color=C_BLACK, strike=True)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    tenants = data.get("tenants", [])
    ws = wb.active
    ws.title = "Transaction History"
    ws.sheet_view.showGridLines = True

    hdrs = ["#", "TENANT NAME", "UNIT", "DATE OF PAYMENT", "TYPE OF PAYMENT",
            "AMOUNT (UGX)", "FROM", "TO"]
    col_widths = [5, 24, 10, 18, 20, 22, 16, 16]
    for ci, (h, cw) in enumerate(zip(hdrs, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = bold(10)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = cw
    ws.row_dimensions[1].height = 24

    row_num = 2
    seq = 0
    for t in tenants:
        all_txns = []
        for rec in t.get("payment_history", []):
            all_txns.append({"_type": "payment", **rec})
        for rec in t.get("deposit_history", []):
            all_txns.append({"_type": "deposit", **rec})
        all_txns.sort(key=lambda r: r.get("date", ""), reverse=True)

        for rec in all_txns:
            seq += 1
            is_cancelled = rec.get("_cancelled", False)
            if rec["_type"] == "deposit":
                type_str = "Deposit" + (" (Cancelled)" if is_cancelled else "")
            else:
                type_str = "Full Payment" + (" (Cancelled)" if is_cancelled else "")
            try:
                amt_str = f"{int(rec.get('amount', 0)):,}"
            except Exception:
                amt_str = "—"
            ft, tt = _compute_txn_period(rec)
            row_vals = [
                (seq, center()), (t.get("name", "—"), left()), (t.get("unit", "—"), center()),
                (rec.get("date", "—"), center()), (type_str, center()),
                (amt_str, center()), (ft, center()), (tt, center()),
            ]
            for ci, (val, aln) in enumerate(row_vals, 1):
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.fill = PatternFill("solid", fgColor=C_WHITE)
                cell.border = thin_border()
                cell.alignment = aln
                if is_cancelled:
                    cell.font = cancelled_font(10)
                else:
                    cell.font = bold(10) if ci in (2, 6) else reg(10)
            ws.row_dimensions[row_num].height = 20
            row_num += 1

    if row_num == 2:
        ws.cell(row=2, column=1, value="No transactions recorded yet.").font = reg(11)

    wb.save(EXCEL_FILE)
    return EXCEL_FILE


def export_pdf(data):
    doc = SimpleDocTemplate(
        PDF_FILE, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm, topMargin=20 * mm, bottomMargin=20 * mm,
    )
    W = A4[0] - 36 * mm
    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=18,
                         textColor=colors.HexColor("#0E4F4F"), spaceAfter=4)
    H2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13,
                         textColor=colors.HexColor("#1FAD9F"), spaceAfter=2, spaceBefore=10)
    MUTED = ParagraphStyle("MUTED", parent=styles["Normal"], fontSize=9,
                            textColor=colors.HexColor("#6E8482"))
    BOLD = ParagraphStyle("BOLD", parent=styles["Normal"], fontSize=10, fontName="Helvetica-Bold")

    story = []
    today_str = date.today().strftime("%d %B %Y")
    story.append(Spacer(1, 10 * mm))
    story.append(Paragraph(APP_NAME, H1))
    story.append(Paragraph(f"Data Export — {today_str}", MUTED))
    story.append(HRFlowable(width=W, thickness=1, color=colors.HexColor("#1FAD9F"), spaceAfter=8))

    tenants = data.get("tenants", [])
    units = data.get("units", {})
    total_collected = int(
        sum(int(rec.get("amount", 0)) for t in tenants for rec in t.get("payment_history", [])
            if not rec.get("_cancelled", False))
        + sum(float(rec.get("amount", 0)) for t in tenants for rec in t.get("deposit_history", [])
              if not rec.get("_cancelled", False))
    )
    today = date.today()

    def _days_status(t):
        ref_str = t.get("due_date", "") or t.get("entry_date", "")
        due = _parse_date(ref_str)
        if due is None:
            return "—"
        rem = (due - today).days
        if t.get("status") == "Confirmed" and rem >= 0:
            return f"Paid  |  {rem}d until next due"
        return f"{'Overdue by ' + str(abs(rem)) + 'd' if rem < 0 else str(rem) + 'd remaining'}"

    summary_data = [
        ["Total Tenants", "Units", "Total Income", "Export Date"],
        [str(len(tenants)), str(len(units)), f"UGX {total_collected:,}", today_str],
    ]
    ts = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E4F4F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#EAF7F4")),
        ("FONTSIZE", (0, 1), (-1, 1), 11),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#1FAD9F")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B0D4D0")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ])
    story.append(Table(summary_data, colWidths=[W / 4] * 4, style=ts))
    story.append(Spacer(1, 8 * mm))

    if units:
        story.append(Paragraph("Units", H2))
        house_rows = [["Unit", "Monthly Rent (UGX)", "Location", "Current Tenant"]]
        for h, val in sorted(units.items()):
            rent = parse_amount(val["rent"] if isinstance(val, dict) else val)
            location = val.get("location", "—") if isinstance(val, dict) else "—"
            tenant_name = next((t.get("name", "Unnamed") for t in tenants if t.get("unit") == h), "Vacant")
            house_rows.append([h, f"{int(rent):,}", location or "—", tenant_name])
        ht = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1FAD9F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0FAF8")]),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#B0D4D0")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0E8E4")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])
        story.append(Table(house_rows, colWidths=[W * 0.15, W * 0.22, W * 0.28, W * 0.35], style=ht))
        story.append(Spacer(1, 8 * mm))

    story.append(Paragraph("Tenant Records", H2))
    if not tenants:
        story.append(Paragraph("No tenants on record.", MUTED))
    else:
        for t in tenants:
            history = t.get("payment_history", [])
            dep_history = t.get("deposit_history", [])
            paid_total = int(
                sum(int(r.get("amount", 0)) for r in history if not r.get("_cancelled", False))
                + sum(float(r.get("amount", 0)) for r in dep_history if not r.get("_cancelled", False))
            )
            story.append(Spacer(1, 3 * mm))
            story.append(Paragraph(f"<b>{t.get('name', 'Unnamed Tenant')}</b>", BOLD))
            story.append(Paragraph(
                f"Unit: {t.get('unit', '—')}  |  Rent: UGX {int(parse_amount(t.get('rent', 0))):,}/mo  |  "
                f"{_days_status(t)}", MUTED))
            info_rows = [
                ["Phone", t.get("phone", "—"), "Email", t.get("email", "—")],
                ["Occupation", t.get("occupation", "—"), "", ""],
                ["Entry Date", t.get("entry_date", "—"), "Due Date", t.get("due_date", "—")],
                ["Emergency", t.get("emergency_contact", "—"), "Emg. Phone", t.get("emergency_phone", "—")],
            ]
            it = TableStyle([
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6E8482")),
                ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#6E8482")),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor("#F6FBFA"), colors.white]),
                ("BOX", (0, 0), (-1, -1), 0.3, colors.HexColor("#D0E8E4")),
                ("INNERGRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#E0F0EC")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ])
            story.append(Spacer(1, 1 * mm))
            story.append(Table(info_rows, colWidths=[W * 0.16, W * 0.34, W * 0.16, W * 0.34], style=it))
            if history:
                story.append(Spacer(1, 2 * mm))
                story.append(Paragraph(f"Payment History  |  Total Income: UGX {paid_total:,}", BOLD))
                pay_rows = [["#", "Date Paid", "Period", "Months", "Amount (UGX)"]]
                for n, rec in enumerate(reversed(history), 1):
                    pay_rows.append([str(n), rec.get("date", "—"),
                                      fmt_period(rec.get("from_date", ""), rec.get("to_date", "")),
                                      str(rec.get("months", 1)), f"{int(rec.get('amount', 0)):,}"])
                pt = TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAF7F4")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("ALIGN", (4, 0), (4, -1), "RIGHT"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6FBFA")]),
                    ("BOX", (0, 0), (-1, -1), 0.3, colors.HexColor("#B0D4D0")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#D8EEEA")),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ])
                story.append(Table(pay_rows, colWidths=[W * 0.06, W * 0.18, W * 0.32, W * 0.12, W * 0.32], style=pt))
            else:
                story.append(Paragraph("No payments recorded yet.", MUTED))
            story.append(HRFlowable(width=W, thickness=0.5, color=colors.HexColor("#C8E0DC"),
                                     spaceAfter=4, spaceBefore=4))
    doc.build(story)
    return PDF_FILE


# ── watchlist / alerts ───────────────────────────────────────────────────
def get_watchlist_tenants(tenants, max_days=10):
    today = date.today()
    items = []
    for t in tenants:
        if status_level(t, today)[0] == "paid":
            continue
        due_str = pending_reference_date_str(t)
        due = _parse_date(due_str)
        if due is None:
            continue
        days_left = (due - today).days
        if days_left <= max_days:
            items.append((t, days_left))
    items.sort(key=lambda pair: pair[1])
    return items

# =========================================================================
#  SECTION 2 -- Flask app + REST API
# =========================================================================
# (app itself is created near the top of the file, before the CLOUD_MODE
# gate's @app.before_request decorator needs it)

# ── device roster + secret key persistence ─────────────────────────────
# Everything about "who is paired" used to live only in this process's
# memory, so restarting the companion (which used to happen on every
# desktop-app "Connect" click) wiped every paired phone's session AND
# forgot every device it had ever seen -- forcing a rescan for a phone
# that had already scanned once before. This file makes both survive a
# restart: the signing secret (so old session cookies keep verifying)
# and the device roster (so a known phone is still known).
DEVICES_FILE = os.path.join(DATA_DIR, "devices.json")
_devices_io_lock = threading.Lock()


def _load_devices_file():
    if not os.path.exists(DEVICES_FILE):
        return {"secret_key": None, "devices": {}}
    try:
        with open(DEVICES_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
        if not isinstance(raw, dict):
            raise ValueError("bad devices.json")
        raw.setdefault("secret_key", None)
        raw.setdefault("devices", {})
        return raw
    except Exception:
        return {"secret_key": None, "devices": {}}


def _save_devices_file_locked():
    """Caller must hold _devices_io_lock."""
    tmp = DEVICES_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(_devices_state, f, indent=2)
    os.replace(tmp, DEVICES_FILE)


_devices_state = _load_devices_file()

# A hardcoded fallback secret would be visible to anyone reading this
# file's source (it ships as plain .py, not compiled/obfuscated) --
# which would let them forge a validly-signed session cookie without
# ever entering the PIN or scanning a QR code. So it's still generated
# randomly -- just once, then reused on every later restart so existing
# session cookies and paired devices don't all become invalid just
# because the process restarted. RENTAL_APP_SECRET still overrides it
# for anyone who wants to manage that themselves.
#
# WHERE it's persisted differs by mode:
#   - CLOUD_MODE: in Postgres (cloud_app_secret table), NOT the local
#     devices.json file -- a Render (or similar) web service's local disk
#     is wiped on every restart/redeploy/free-tier spin-down, which used
#     to silently rotate this secret each time, invalidating every
#     phone's pairing cookie and sending them back to the "Waiting to
#     connect" screen for no visible reason.
#   - PC-local mode: devices.json next to the process, as before -- the
#     desktop app's own disk is not ephemeral, so this was never broken
#     there.
if os.environ.get("RENTAL_APP_SECRET"):
    app.secret_key = os.environ["RENTAL_APP_SECRET"]
elif CLOUD_MODE:
    app.secret_key = _cloud_get_or_create_secret()
else:
    if not _devices_state.get("secret_key"):
        _devices_state["secret_key"] = secrets.token_hex(32)
        with _devices_io_lock:
            _save_devices_file_locked()
    app.secret_key = _devices_state["secret_key"]

app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["PERMANENT_SESSION_LIFETIME"] = 60 * 60 * 24 * 180  # 180 days

# Set by the desktop app (via /api/announce-disconnect) right before it
# tears this server down deliberately, e.g. Settings → Disconnect. Lets
# already-connected phones tell that apart from a plain network/PC
# outage: they poll /api/lock-status and see this flag while the server
# is still reachable, instead of only ever finding out via a failed
# request (which is indistinguishable from "PC is off").
_disconnect_state = {"announced": False}

# ── QR-only pairing gate ────────────────────────────────────────────────
# A QR code is nothing more than an encoded URL -- there's no way for a
# server to tell "this request came from a camera scan" apart from "this
# request came from someone pasting/forwarding that same URL", since
# they're literally the same HTTP request. What IS achievable, and is
# what this implements: make the link single-use and short-lived, so a
# forwarded/screenshotted/copied link stops working (serves nothing, not
# even an error that reveals this is a rental-management app) as soon as
# either (a) the original scan succeeds, or (b) the desktop app shows a
# fresh code (e.g. a phone connects, or the rotation timer below ticks).
# The desktop app pushes the currently-valid token here every time it
# displays a QR code (see /api/pairing-token below); the "/" route only
# ever serves the app shell to a request presenting that exact,
# not-yet-consumed token, OR to a request whose X-Device-Id already
# belongs to a known, non-kicked device -- so a phone that scanned once
# never needs to scan again, only a brand-new or disconnected device
# needs a live token.
_pairing_state = {"token": None, "consumed": False}
_pairing_lock = threading.Lock()


def _pairing_ok(device_id=""):
    """True if this request may load the app shell -- see the module
    comment above. Consumes the active token on a successful match so it
    can never work a second time for anyone else it gets forwarded to."""
    if session.get("paired"):
        return True
    if device_id and _device_known(device_id):
        # Recognized via X-Device-Id (e.g. the waiting page's background
        # retry, or an /api/ call) rather than the one-time token. Set the
        # session cookie here too, so the NEXT plain top-level navigation
        # (which can't carry the X-Device-Id header at all) is recognized
        # via session.get("paired") above instead of falling through to
        # the waiting page again.
        session["paired"] = True
        session.permanent = True
        return True
    token = request.args.get("pt", "")
    with _pairing_lock:
        active = _pairing_state["token"]
        matched = bool(active) and bool(token) and token == active and not _pairing_state["consumed"]
        if matched:
            _pairing_state["consumed"] = True
    if matched:
        session["paired"] = True
        session.permanent = True
        if device_id:
            _readmit_device(device_id)
    return matched


def _readmit_device(device_id):
    """Clears a device's 'kicked' flag if -- and only if -- it just
    matched a fresh, valid, single-use pairing token (see the caller
    above): that can only happen because the desktop app explicitly
    showed a new QR code and this device scanned it, i.e. an admin
    re-invited it. Nothing else (a lingering session cookie, a replayed
    /api/ call, the device's own retry loop) can undo a kick this way --
    without this, a disconnected device stayed blocked forever even
    after scanning a brand-new code, since nothing else ever cleared the
    flag. Still respects MAX_DEVICES: if the cap is already full, it
    stays kicked until a slot is freed."""
    import time
    with _devices_lock:
        devices = _devices_dict()
        rec = devices.get(device_id)
        if rec is None or not rec.get("kicked"):
            return
        if sum(1 for r in devices.values() if not r.get("kicked")) >= MAX_DEVICES:
            return
        rec["kicked"] = False
        rec["last_seen"] = time.time()
        _save_devices_locked()

# Tracks which phones/browsers have ever paired with this app, persisted
# to disk (DEVICES_FILE) so the roster survives the companion process
# being stopped and restarted -- e.g. every time the desktop app is
# closed and reopened. Each client generates a random id on first load
# (see DEVICE_ID in the JS below) and sends it as X-Device-Id on every
# /api/ request. Scanning the QR code (or opening an already-known
# device's saved link) grants full access immediately -- there is no
# separate "approve" step; the token gate above is what stands in for
# that. A device stays listed, occupying one of the MAX_DEVICES slots,
# until it is explicitly disconnected from the desktop app -- going
# offline (no internet, phone off, app closed) does NOT drop it from the
# roster, only its displayed status changes.
MAX_DEVICES = 3           # Settings -> Connect Phone allows this many phones
                          # connected at once. A 4th scan sees a "device
                          # limit reached" screen; disconnect one from the
                          # desktop app's device list to free a slot.
ONLINE_TIMEOUT = 20       # seconds since last contact before a device's
                          # displayed status flips from "online" to
                          # "offline" -- display only, never removes it.
_devices_lock = threading.Lock()


def _devices_dict():
    """The live device roster, backed by _devices_state (loaded from /
    persisted to DEVICES_FILE). Caller must hold _devices_lock for any
    read that needs to stay consistent with a following write."""
    return _devices_state["devices"]


def _save_devices_locked():
    """Persist the current roster to disk. Caller must hold _devices_lock."""
    with _devices_io_lock:
        _save_devices_file_locked()


def _device_known(device_id):
    """Non-mutating: has this device ever paired and not since been
    disconnected? Used by _pairing_ok so a previously-scanned phone can
    reload the app shell without a fresh token."""
    if not device_id:
        return False
    with _devices_lock:
        rec = _devices_dict().get(device_id)
        return bool(rec and not rec.get("kicked"))


def _device_was_kicked(device_id):
    """Non-mutating: has this device been explicitly disconnected from
    the desktop app's device list?"""
    if not device_id:
        return False
    with _devices_lock:
        rec = _devices_dict().get(device_id)
        return bool(rec and rec.get("kicked"))


# Best-effort map of iPhone physical screen resolution ("WIDTHxHEIGHT")
# to model name(s). Apple's mobile Safari deliberately doesn't expose the
# exact hardware model in its User-Agent (unlike Android), so this is the
# closest available signal -- and it's inherently approximate, since
# several generations share an identical screen. Where more than one
# model shares a resolution, all of them are listed.
_IPHONE_SCREEN_MODELS = {
    "1170x2532": "iPhone 12/12 Pro/13/13 Pro/14",
    "1284x2778": "iPhone 12 Pro Max/13 Pro Max/14 Plus",
    "1179x2556": "iPhone 14 Pro/15/15 Pro",
    "1290x2796": "iPhone 14 Pro Max/15 Plus/15 Pro Max",
    "1206x2622": "iPhone 16/16 Pro",
    "1320x2868": "iPhone 16 Plus/16 Pro Max",
    "1080x2340": "iPhone 12 mini/13 mini",
    "828x1792": "iPhone 11/XR",
    "1125x2436": "iPhone X/XS/11 Pro",
    "1242x2688": "iPhone XS Max/11 Pro Max",
    "750x1334": "iPhone 6/6s/7/8/SE (2nd/3rd gen)",
    "1080x1920": "iPhone 6/7/8 Plus",
    "640x1136": "iPhone 5/5s/5c/SE (1st gen)",
}


def _guess_ios_model(is_ipad, screen_hint):
    """`screen_hint` is the "WIDTHxHEIGHT@DPR" string sent by the client
    (see DEVICE_SCREEN_HINT in the JS below). Converts logical CSS pixels
    back to physical pixels (x DPR) and looks that up. Falls back to the
    generic platform name if the resolution isn't recognized or wasn't
    sent."""
    if is_ipad:
        return "iPad"  # too many iPad models share resolutions to guess further
    try:
        dims, dpr = screen_hint.split("@")
        w, h = (int(n) for n in dims.split("x"))
        dpr = float(dpr)
        w, h = round(w * dpr), round(h * dpr)
        key = f"{min(w,h)}x{max(w,h)}"
        return _IPHONE_SCREEN_MODELS.get(key, "iPhone")
    except Exception:
        return "iPhone"


def _label_for_user_agent(ua, screen_hint=""):
    """Turns a raw User-Agent (plus an optional screen-resolution hint)
    into a short, phone-model-only label for the admin device list --
    e.g. 'iPhone 12', 'Tecno KI5k', 'itel A662L'. Browser name is
    intentionally never included (admin request: only the device/phone
    type should be shown, not which browser it's using)."""
    ua = ua or ""
    if "iPhone" in ua or "iPad" in ua:
        return _guess_ios_model("iPad" in ua, screen_hint)
    if "Android" in ua:
        # Standard Android UA shape: "...; Android <ver>; <MODEL>) ...",
        # sometimes followed by "Build/...". The model segment is where
        # manufacturers put their real marketing/model name (e.g. "Tecno
        # KI5k", "itel A662L", "SM-A125F", "Redmi Note 11") -- this is
        # the one platform where the exact device name really is
        # available, so use it verbatim instead of just "Android".
        m = re.search(r"Android\s+[\d.]+\s*;\s*([^;)]+?)(?:\s+Build/[^;)]+)?\)", ua)
        if m:
            model = m.group(1).strip()
            if model and model.lower() not in ("k", ""):
                return model
        return "Android"
    if "Windows" in ua:
        return "Windows PC"
    if "Macintosh" in ua:
        return "Mac"
    return "Device"


def _touch_device(device_id, user_agent=None, screen_hint="", fingerprint=""):
    """Refresh an already-known device's last-seen time, or register a
    brand-new one -- immediately admitted, no separate approval step --
    if there's room under MAX_DEVICES. Returns (admitted, canonical_id):
    admitted is False only if it was turned away because the cap is
    already full or it was explicitly disconnected from the desktop
    app's device list. Existing devices (online or not) are never
    evicted to make room for a newcomer -- a slot only frees up when the
    desktop app disconnects one.

    Same fingerprint-based reconciliation as the cloud counterpart (see
    _cloud_touch_device): an unrecognized device_id whose fingerprint
    matches an already-known device reuses that device's slot instead of
    registering a new one, and canonical_id tells the caller which id
    that was."""
    if not device_id:
        return True, device_id
    import time
    with _devices_lock:
        devices = _devices_dict()
        rec = devices.get(device_id)
        if rec is not None and rec.get("kicked"):
            return False, device_id
        now = time.time()
        if rec is not None:
            rec["last_seen"] = now
            # Re-derive the label on every poll rather than only at first
            # registration: the very first request from a device can
            # arrive before its screen-hint header is populated in some
            # edge cases, so this lets the label upgrade from a generic
            # "iPhone"/"Android" to a specific model as soon as a usable
            # hint shows up. Skipped entirely once the person has given
            # this device a custom name (see _set_local_device_label) --
            # that name should never be silently overwritten again.
            if not rec.get("custom_label_locked"):
                fresh_label = _label_for_user_agent(user_agent, screen_hint)
                if fresh_label and fresh_label not in ("Device",):
                    rec["label"] = fresh_label
            if fingerprint:
                rec["fingerprint"] = fingerprint
            _save_devices_locked()
            return True, device_id
        if fingerprint:
            match_id = next(
                (d for d, r in devices.items()
                 if not r.get("kicked") and r.get("fingerprint") == fingerprint),
                None)
            if match_id:
                match_rec = devices[match_id]
                match_rec["last_seen"] = now
                if not match_rec.get("custom_label_locked"):
                    fresh_label = _label_for_user_agent(user_agent, screen_hint)
                    if fresh_label and fresh_label not in ("Device",):
                        match_rec["label"] = fresh_label
                _save_devices_locked()
                return True, match_id
        if sum(1 for r in devices.values() if not r.get("kicked")) >= MAX_DEVICES:
            return False, device_id
        devices[device_id] = {
            "first_seen": now,
            "last_seen": now,
            "label": _label_for_user_agent(user_agent, screen_hint),
            "kicked": False,
            "fingerprint": fingerprint,
        }
        _save_devices_locked()
        return True, device_id


def _device_admitted(device_id):
    """Non-mutating check: is this device allowed to use every other
    /api/ call? True for any known, non-kicked device -- pairing (scan
    or already-known link) is the only gate; there is no further
    approval step."""
    if not device_id:
        return True
    with _devices_lock:
        rec = _devices_dict().get(device_id)
        return bool(rec and not rec.get("kicked"))


def _active_device_count():
    """How many devices are currently on the roster (online or not) --
    what the Connect Phone card's 'Connected -- N/3' status and the
    MAX_DEVICES cap are based on. A device only stops counting once it
    is explicitly disconnected."""
    with _devices_lock:
        return sum(1 for rec in _devices_dict().values() if not rec.get("kicked"))


def _list_devices():
    """Snapshot of every currently-tracked (non-kicked) device for the
    desktop app's admin list: id (full, used to target a kick), a short
    id for display, a human label, whether it's currently reachable
    (online), and when it last synced (epoch seconds) -- devices stay
    listed indefinitely once offline, they are never dropped by a
    timeout, only by an explicit Disconnect."""
    import time
    with _devices_lock:
        now = time.time()
        return [
            {
                "device_id": d,
                "short_id": d[:8],
                "label": rec.get("label") or "Unknown device",
                "custom_label_locked": bool(rec.get("custom_label_locked")),
                "last_seen": rec["last_seen"],
                "online": (now - rec["last_seen"]) <= ONLINE_TIMEOUT,
            }
            for d, rec in sorted(_devices_dict().items(),
                                  key=lambda kv: kv[1]["last_seen"], reverse=True)
            if not rec.get("kicked")
        ]


def _kick_device(device_id):
    """Forcibly disconnects one specific phone/browser (desktop app's
    Connect Phone -> device list -> Disconnect), freeing its slot.
    Marked kicked (rather than deleted outright) so it's permanently
    blocked from silently re-admitting itself on its next poll -- the
    next request from that device_id gets turned away instead. It clears
    only if the phone/browser generates a fresh device id (e.g. clearing
    site data) or is manually removed from devices.json."""
    with _devices_lock:
        rec = _devices_dict().get(device_id)
        if rec is not None:
            rec["kicked"] = True
        else:
            _devices_dict()[device_id] = {
                "first_seen": 0, "last_seen": 0,
                "label": "Unknown device", "kicked": True,
            }
        _save_devices_locked()


def _set_local_device_label(device_id, label):
    """Lets a phone give itself a custom name (e.g. 'Mary's iPhone')
    instead of the generic auto-detected model name -- and locks it so
    _touch_device never overwrites it again on a later poll. Enforces
    that no two currently-connected (non-kicked) devices share the same
    name (case-insensitively), so the admin's device list -- and the
    phones themselves -- can always tell devices apart. Returns False
    (without writing anything) if that name is already taken by another
    device; True on success."""
    import time
    with _devices_lock:
        devices = _devices_dict()
        lname = label.strip().lower()
        if any(other_id != device_id and not other.get("kicked")
               and (other.get("label") or "").strip().lower() == lname
               for other_id, other in devices.items()):
            return False
        rec = devices.get(device_id)
        now = time.time()
        if rec is None:
            rec = {"first_seen": now, "last_seen": now, "kicked": False, "fingerprint": ""}
            devices[device_id] = rec
        rec["label"] = label
        rec["custom_label_locked"] = True
        rec["last_seen"] = now
        _save_devices_locked()
        return True


# ── auth helpers ─────────────────────────────────────────────────────────
def _pin_required(data):
    return bool(data.get("settings", {}).get("pin_hash"))


def _authed():
    return session.get("unlocked") is True


@app.before_request
def _guard():
    if CLOUD_MODE:
        return None
    if request.path.startswith("/api/"):
        if request.path == "/api/lock-status":
            return None
        if (request.path == "/api/devices" or request.path.startswith("/api/devices/")
                or request.path in ("/api/shutdown", "/api/announce-disconnect",
                                     "/api/device-count", "/api/pairing-token")) and (
                request.remote_addr in ("127.0.0.1", "::1")
                or (request.remote_addr or "").startswith("::ffff:127.")):
            # Only the desktop app (calling from the same PC) may stop the
            # server, announce it's about to, list/kick devices, or push a
            # fresh pairing token, this way; a phone reaching over the LAN
            # still needs to unlock first like any other /api/ call.
            return None
        device_id = request.headers.get("X-Device-Id", "")
        if device_id and _device_was_kicked(device_id):
            return jsonify({"error": "kicked"}), 403
        if device_id and not _device_admitted(device_id):
            return jsonify({"error": "device_limit_reached",
                             "max_devices": MAX_DEVICES}), 403
        if request.path == "/api/unlock":
            return None
        data = load_state()
        if _pin_required(data) and not _authed():
            return jsonify({"error": "locked"}), 401
    return None


# ── idempotency guard for mutating requests ───────────────────────────────
# A POST/PUT/DELETE can reach this server twice for reasons that have
# nothing to do with someone tapping a button twice: a request that times
# out or gets a transient 502/503/504 is automatically retried over a
# second transport by the frontend's api() (see cloudFetch fallback), and
# a change queued for offline sync is replayed once the device reconnects.
# In both cases the FIRST attempt may already have been applied here even
# though the client never saw a clean response for it -- without this,
# the retry runs the same mutation again (a payment recorded twice, a
# tenant deleted twice, etc). The frontend tags every mutating call with
# a client-generated X-Idempotency-Key that stays the same across retries
# of one logical action. The first request carrying a given key runs
# normally and its response is cached below; any later request with that
# same key gets the exact cached response replayed back instead of
# re-running the handler.
_idempotency_lock = threading.Lock()
_idempotency_cache = {}   # key -> (stored_at, status_code, body_bytes, mimetype)
_IDEMPOTENCY_TTL = 600     # seconds a cached response stays replayable
_IDEMPOTENCY_MAX = 1000    # hard cap so this can never grow unbounded


def _idempotency_prune_locked():
    if len(_idempotency_cache) <= _IDEMPOTENCY_MAX:
        return
    stale = sorted(_idempotency_cache.items(), key=lambda kv: kv[1][0])
    for k, _ in stale[: len(_idempotency_cache) - _IDEMPOTENCY_MAX]:
        _idempotency_cache.pop(k, None)


@app.before_request
def _idempotency_replay():
    if request.method not in ("POST", "PUT", "DELETE"):
        return None
    key = request.headers.get("X-Idempotency-Key", "")
    if not key:
        return None
    now = time.time()
    with _idempotency_lock:
        hit = _idempotency_cache.get(key)
        if hit and (now - hit[0]) <= _IDEMPOTENCY_TTL:
            _, status_code, body, mimetype = hit
            return Response(body, status=status_code, mimetype=mimetype)
    return None


@app.after_request
def _idempotency_store(resp):
    if request.method in ("POST", "PUT", "DELETE"):
        key = request.headers.get("X-Idempotency-Key", "")
        if key:
            try:
                body = resp.get_data()
                with _idempotency_lock:
                    _idempotency_cache[key] = (time.time(), resp.status_code, body, resp.mimetype)
                    _idempotency_prune_locked()
            except Exception:
                pass
    return resp


# ── page ─────────────────────────────────────────────────────────────────
WAITING_HTML = """<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1, viewport-fit=cover">
<title>Tenant Management</title>
<meta name="theme-color" content="#14181A">
<meta name="color-scheme" content="dark">
<link rel="icon" href="icon-192.png?v=2">
<link rel="apple-touch-icon" href="icon-256.png?v=2">
<style>
  html,body{margin:0;padding:0;height:100%;background:radial-gradient(120% 100% at 50% 0%,#1C2124 0%,#14181A 55%,#0F1416 100%);color:#fff;font-family:'Segoe UI',system-ui,-apple-system,sans-serif;}
  .wrap{min-height:100vh;min-height:100dvh;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:32px;}
  .mark{width:56px;height:56px;border-radius:16px;margin-bottom:18px;background:#fff;overflow:hidden;
        display:flex;align-items:center;justify-content:center;box-shadow:0 12px 28px -10px rgba(0,0,0,.55);}
  .mark img{width:100%;height:100%;display:block;object-fit:cover;}
  h1{font-family:'Segoe UI',system-ui,-apple-system,sans-serif;font-weight:600;font-size:21px;margin:0 0 8px;letter-spacing:.2px;}
  p{font-size:13.5px;line-height:1.55;opacity:.68;max-width:280px;margin:0;}
  .spin{margin-top:26px;width:20px;height:20px;border-radius:50%;border:2.5px solid rgba(255,255,255,.18);
        border-top-color:#22D3A6;animation:sp 0.85s linear infinite;}
  @media (prefers-reduced-motion: reduce){.spin{animation-duration:2.4s;}}
  @keyframes sp{to{transform:rotate(360deg);}}
</style></head>
<body><div class="wrap">
  <div class="mark"><img src="icon-128.png?v=2" alt=""></div>
  <h1>Tenant Management</h1>
  <p>Waiting to connect. Scan the QR code shown in Connect Phone on the PC to open this device's tenant data.</p>
  <div class="spin"></div>
</div>
<script>
  // Not paired yet (or a forwarded/expired link) -- reveals nothing about
  // tenant data. Quietly retries in the background so this page moves on
  // by itself the moment the desktop app admits this device, without the
  // person having to notice and reload manually.
  setInterval(function () {
    fetch(location.pathname + location.search, {headers:{'X-Device-Id': (function(){
      try { return localStorage.getItem('rm_device_id_v1') || ''; } catch(e) { return ''; }
    })()}}).then(function (r) { if (r.ok) location.reload(); }).catch(function () {});
  }, 4000);
</script>
</body></html>"""


# ── page ─────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    if CLOUD_MODE:
        # sid+key identify WHICH household's data this is and are what
        # every subsequent /api/ call authenticates with going forward --
        # they can't be single-use themselves, or the phone couldn't keep
        # syncing after this first load. Getting INTO the app the first
        # time is gated separately, by pt: the one-time token the desktop
        # pushes every time it (re)shows the QR code (see
        # _cloud_pairing_ok() above). A link with valid sid+key but no
        # valid pt -- forwarded, screenshotted, or just reopened by
        # someone who was never shown the QR -- gets the same branded
        # waiting page a stale LAN link gets, never the app shell.
        session_id = request.args.get("sid", "")
        secret_key = request.args.get("key", "")
        pt = request.args.get("pt", "")
        row = _cloud_get_row(session_id) if session_id else None
        if not session_id or not secret_key or not row or row["secret_key"] != secret_key:
            return Response("", status=404, mimetype="text/plain")
        if not _cloud_pairing_ok(session_id, pt, request.headers.get("X-Device-Id", "")):
            return Response(WAITING_HTML, status=404, mimetype="text/html")
        bootstrap = (
            "<script>window.__CLOUD_DIRECT__=" + json.dumps({
                "sessionId": session_id, "secretKey": secret_key,
            }) + ";</script>\n"
        )
        html = INDEX_HTML.replace("<head>", "<head>\n" + bootstrap, 1)
        # Carry sid/key onto the manifest request too. manifest.json's
        # start_url is resolved against the MANIFEST's own URL, not this
        # page's -- so without this, an installed home-screen icon always
        # launches a bare "/" with no sid/key at all. index() then 404s
        # with an EMPTY body (see the sid/key check above), which is
        # exactly the blank white page on launch. Passing them through
        # here lets manifest() below bake them into start_url itself.
        qs = f"sid={quote(session_id)}&key={quote(secret_key)}"
        html = html.replace(
            'href="manifest.json"', f'href="manifest.json?{qs}"', 1)
        return Response(html, mimetype="text/html")
    if not _pairing_ok(request.headers.get("X-Device-Id", "")):
        # No token, a stale/wrong one, or one already used once -- this
        # used to be a totally bare, unlabelled 404 so a forwarded/copied
        # link couldn't even tell this was a rental-management app. That
        # also meant the browser tab (and "Add to Home Screen") had no
        # <title> to show and fell back to displaying the raw URL
        # instead. This branded waiting page keeps the same "no tenant
        # data leaks to an unpaired visitor" guarantee, but shows the
        # Tenant Management name/icon instead of the bare link, and
        # quietly retries in the background until it's admitted.
        return Response(WAITING_HTML, status=404, mimetype="text/html")
    return Response(INDEX_HTML, mimetype="text/html")


# ── PWA: manifest, service worker, icon ───────────────────────────────────
# These three routes are what let a phone or desktop browser "install" this
# page (Chrome/Edge: address-bar install icon or ⋮ → Install app; Safari:
# Share → Add to Home Screen) so it opens full-screen with its own icon,
# just like a native app — no separate hosting or build step required,
# since it's all served straight from this same process.
@app.route("/icon-<int:size>.png")
def app_icon(size):
    # Served at native (512) resolution regardless of the requested size --
    # browsers/OSes downscale a too-large PWA icon just fine, and that's
    # simpler than maintaining separate pre-baked assets per size. What
    # actually caused the "shows a plain letter instead of the icon" bug
    # was that the old icon was ~92% transparent; this one is a fully
    # opaque, edge-to-edge background so it satisfies both plain ("any")
    # and "maskable" icon requirements.
    icon_bytes = base64.b64decode(APP_ICON_512_B64)
    return Response(icon_bytes, mimetype="image/png")


@app.route("/icon-splash.png")
def app_icon_splash():
    # Transparent variant used only by the boot splash's breathing icon
    # (see .splash-house / #splashScreen in INDEX_HTML) so it blends with
    # the splash background instead of always showing a white box,
    # matching light/dark theme. NOT used for the manifest/home-screen
    # icon -- that one deliberately stays opaque, see app_icon() above.
    icon_bytes = base64.b64decode(APP_ICON_SPLASH_B64)
    return Response(icon_bytes, mimetype="image/png")


@app.route("/manifest.json")
def manifest():
    # In CLOUD_MODE, index() now forwards this page's sid/key onto the
    # manifest request (see its "?sid=&key=" href rewrite above) so they
    # can be baked into start_url here. Without this, start_url resolves
    # to a bare "/" -- valid per spec, since these fields resolve against
    # the MANIFEST's own URL rather than the page's -- and a home-screen
    # launch of that bare "/" gets index()'s empty-body 404 (no session
    # means no household to show), i.e. the reported blank white page.
    sid = request.args.get("sid", "")
    key = request.args.get("key", "")
    start_url = f"./?sid={quote(sid)}&key={quote(key)}" if (CLOUD_MODE and sid and key) else "."
    return jsonify({
        "name": "Tenant Management",
        "short_name": "Tenant Management",
        "description": "Manage tenants, payments, and units on the go.",
        # Relative, not "/": per the manifest spec these all resolve
        # against the manifest's OWN url, not the page's.
        "start_url": start_url,
        "scope": ".",
        "display": "standalone",
        "orientation": "portrait",
        "background_color": "#14181A",
        "theme_color": "#14181A",
        "icons": [
            # Separate "any" and "maskable" entries (rather than one combined
            # "any maskable" purpose) per the manifest spec's own guidance --
            # some platforms pick the wrong one when both are declared on
            # the same entry. The icon itself is a fully opaque square with
            # the house glyph inside the safe zone, so it works either way.
            {"src": "icon-192.png?v=2", "sizes": "192x192", "type": "image/png", "purpose": "any"},
            {"src": "icon-192.png?v=2", "sizes": "192x192", "type": "image/png", "purpose": "maskable"},
            {"src": "icon-256.png?v=2", "sizes": "256x256", "type": "image/png", "purpose": "any"},
            {"src": "icon-256.png?v=2", "sizes": "256x256", "type": "image/png", "purpose": "maskable"},
            {"src": "icon-512.png?v=2", "sizes": "512x512", "type": "image/png", "purpose": "any"},
            {"src": "icon-512.png?v=2", "sizes": "512x512", "type": "image/png", "purpose": "maskable"},
        ],
    })


@app.route("/sw.js")
def service_worker():
    # Cache-first for the app shell (so it still opens with no signal),
    # but always goes to the network for /api/* so tenant data is never
    # served stale. Scope is "/" via the header below so it can control
    # the whole app, not just its own folder.
    js = """
// Bumped so every existing install picks up this fix on its next
// activate() instead of continuing to serve a stale precached "./" that
// (pre-fix) never had sid/key on it in the first place.
const CACHE = 'rental-app-shell-v5';
// sid/key travel here via this script's OWN url (see the
// navigator.serviceWorker.register('sw.js?sid=...&key=...') call) --
// self.location.search is that query string, readable from inside the
// worker regardless of which page registered it.
const PARAMS = self.location.search;
const ROOT_URL = './' + PARAMS;
const MANIFEST_URL = './manifest.json' + PARAMS;
// Relative, not root-absolute: resolved against this script's own URL.
const SHELL_URLS = [ROOT_URL, MANIFEST_URL, './icon-192.png?v=2'];

self.addEventListener('install', (evt) => {
  self.skipWaiting();
  evt.waitUntil(caches.open(CACHE).then((c) => c.addAll(SHELL_URLS)));
});

self.addEventListener('activate', (evt) => {
  evt.waitUntil(
    caches.keys().then((keys) =>
      Promise.all(keys.filter((k) => k !== CACHE).map((k) => caches.delete(k)))
    ).then(() => self.clients.claim())
  );
});

self.addEventListener('fetch', (evt) => {
  const url = new URL(evt.request.url);
  if (url.pathname.includes('/api/')) return;

  // Network-first for the app shell: whenever the phone can reach the
  // server, it always gets whatever HTML/JS is currently running there,
  // so edits to this app show up the next time it's opened. Falls back
  // to cache whenever the network request fails outright (no Wi-Fi/data,
  // or the browser was closed and reopened somewhere without a
  // connection) so the app keeps opening instead of showing nothing --
  // and if even THIS exact request was never cached before, falls back
  // to ROOT_URL (the paired shell, WITH sid/key) rather than the bare
  // worker scope, which used to have no session on it at all and is
  // what produced the blank white page after installing to the home
  // screen.
  if (evt.request.method === 'GET') {
    evt.respondWith(
      fetch(evt.request).then((resp) => {
        if (resp && resp.ok) {
          caches.open(CACHE).then((c) => c.put(evt.request, resp.clone()));
          return resp;
        }
        // Not a real page -- our own not-yet-paired waiting page, or a
        // genuine server error. Prefer the last cached shell if there is
        // one, so the app still opens.
        return caches.match(evt.request).then((cached) => cached || caches.match(ROOT_URL) || resp);
      }).catch(() => caches.match(evt.request).then((cached) => cached || caches.match(ROOT_URL)))
    );
  }
});
"""
    return Response(js, mimetype="application/javascript",
                     headers={"Service-Worker-Allowed": "/"})


# ── cloud durability / sync ─────────────────────────────────────────────
CLOUD_SYNC_FILE = os.path.join(DATA_DIR, "cloud_sync.json")


def _load_cloud_sync_config():
    """Local-only file the desktop app owns (creates/updates), telling the
    phone where the cloud fallback service lives and how to authenticate
    to it. Read-only from here -- if it doesn't exist yet, cloud fallback
    just isn't configured for this install."""
    if not os.path.exists(CLOUD_SYNC_FILE):
        return None
    try:
        with open(CLOUD_SYNC_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        if cfg.get("cloud_base_url") and cfg.get("session_id") and cfg.get("secret_key"):
            return cfg
    except Exception:
        pass
    return None


@app.route("/api/cloud-config")
def cloud_config():
    """PC-mode only: tells the phone (while it can still reach the PC)
    where the cloud fallback is and how to authenticate to it, so the
    phone can keep working -- reading AND writing -- once the PC goes
    offline. The phone's own api() wrapper caches this response like any
    other GET, so it's still available from the phone's local cache even
    after the PC is gone."""
    cfg = _load_cloud_sync_config()
    if not cfg:
        return jsonify({"configured": False})
    return jsonify({
        "configured": True,
        "cloud_base_url": cfg["cloud_base_url"].rstrip("/"),
        "session_id": cfg["session_id"],
        "secret_key": cfg["secret_key"],
    })


@app.route("/api/_sync", methods=["GET", "PUT"])
def cloud_sync():
    """Cloud-mode only (see _cloud_gate). GET returns the stored snapshot
    for reconciliation; PUT accepts a full snapshot push -- from either
    the PC (after a local save) or, in principle, anywhere else -- and
    merges it record-by-record against whatever's already stored (see
    _merge_cloud_data) rather than treating either side's whole snapshot
    as "the winner"."""
    if not CLOUD_MODE:
        return Response("", status=404, mimetype="text/plain")

    if request.method == "GET":
        row = _cloud_get_row(g.session_id)
        if not row:
            return jsonify({"exists": False})
        return jsonify({
            "exists": True,
            "data": row["data"],
            "updated_at": row["updated_at"].isoformat(),
            "updated_by": row["updated_by"],
        })

    body = request.get_json(silent=True) or {}
    incoming_data = body.get("data")
    updated_by = body.get("updated_by", "unknown")
    if not isinstance(incoming_data, dict):
        return jsonify({"ok": False, "error": "bad_request"}), 400

    # Merged record-by-record against the currently stored snapshot,
    # rather than being accepted or rejected wholesale based on comparing
    # the PUSHING side's own wall-clock timestamp (body.get("updated_at"))
    # against the row's stored timestamp. That comparison used to be the
    # gate here, but every ACCEPTED write is stamped with the database's
    # own now() (see _cloud_save) -- an authoritative, ever-advancing
    # clock -- while the pushing side (e.g. the PC, syncing every ~20s)
    # sends its own local clock's reading. Even a few seconds of
    # client/server clock drift meant that, after the first successful
    # push, every later push from that same side looked "not new enough"
    # and was silently discarded forever -- while writes landing directly
    # on this server (from the phone/web app) always kept advancing the
    # row's timestamp, so pulls from that side (reconcile) kept working.
    # That's what made syncing look one-directional: phone → PC worked,
    # PC → phone silently stopped. Merging here removes the clock
    # comparison from the equation entirely; per-record `_updated_at`
    # stamps (see _stamp_changed_records) already make the merge safe
    # regardless of which side's clock is ahead.
    row = _cloud_get_row(g.session_id)
    merged = _merge_cloud_data(incoming_data, row["data"]) if row else incoming_data
    server_updated_at = save_raw(merged, updated_by=updated_by)
    return jsonify({"ok": True, "stored": True, "updated_at": server_updated_at, "data": merged})



@app.route("/api/lock-status")
def lock_status():
    if CLOUD_MODE:
        # No PC in the loop here at all -- this only ever confirms the
        # cloud service itself is reachable (which is what a cloud-direct
        # phone's isOnline badge actually means: "works with the PC off"),
        # plus this device's own kicked/roster status. _cloud_gate already
        # ran the touch/kicked check above before this handler is reached.
        device_id = request.headers.get("X-Device-Id", "")
        return jsonify({
            "pin_set": False,
            "unlocked": True,
            "app_name": APP_NAME,
            "disconnecting": False,
            "kicked": _cloud_device_kicked(g.session_id, device_id),
            "device_count": _cloud_active_device_count(g.session_id),
            "max_devices": MAX_DEVICES,
            "device_limit_reached": False,
            "data_updated_at": _cloud_get_updated_at(g.session_id),
        })
    data = load_state()
    device_id = request.headers.get("X-Device-Id", "")
    kicked = _device_was_kicked(device_id)
    admitted = True
    canonical_device_id = None
    if not _disconnect_state["announced"] and not kicked:
        admitted, canonical_device_id = _touch_device(
            device_id,
            request.headers.get("User-Agent"),
            request.headers.get("X-Device-Screen", ""),
            request.headers.get("X-Device-Fingerprint", ""),
        )
    try:
        data_updated_at = str(os.path.getmtime(DATA_FILE))
    except OSError:
        data_updated_at = ""
    resp = {
        "pin_set": _pin_required(data),
        "unlocked": _authed(),
        "app_name": APP_NAME,
        "disconnecting": _disconnect_state["announced"],
        "kicked": kicked,
        "device_count": _active_device_count(),
        "max_devices": MAX_DEVICES,
        "device_limit_reached": (not admitted) and not kicked,
        "data_updated_at": data_updated_at,
    }
    if canonical_device_id and canonical_device_id != device_id:
        resp["canonical_device_id"] = canonical_device_id
    return jsonify(resp)


@app.route("/api/device-count")
def device_count():
    """Polled locally (127.0.0.1) by the desktop app's Connect Phone panel
    to tell Pending (server up, 0 phones actively polling) apart from
    Connected (>=1 phone actively polling), and to show how many out of
    the MAX_DEVICES cap are currently in use."""
    return jsonify({"device_count": _active_device_count(), "max_devices": MAX_DEVICES})


@app.route("/api/devices")
def devices_list():
    """The desktop app's Connect Phone card admin list -- which phones/
    browsers are currently connected, with enough of an id shown
    (short_id) and a friendly label to tell them apart, plus a device_id
    to target with /api/devices/<id>/kick. In CLOUD_MODE this is
    session-authenticated (X-Session-Id/X-Secret-Key) rather than
    loopback-only, since the desktop app reaches it over the same public
    URL every phone does -- there's no "same machine" signal to check
    instead, the household's secret_key IS the admin credential here,
    same as it already is for every other cloud endpoint."""
    if CLOUD_MODE:
        return jsonify({"devices": _cloud_list_devices(g.session_id)})
    return jsonify({"devices": _list_devices()})


@app.route("/api/devices/<device_id>/kick", methods=["POST"])
def devices_kick(device_id):
    """Forcibly disconnects one specific phone/browser -- the admin
    'Disconnect' action next to a device in the desktop app's Connect
    Phone card, as opposed to the single Disconnect button that tears
    the whole companion down for every device. See devices_list() above
    re: CLOUD_MODE auth."""
    if CLOUD_MODE:
        _cloud_kick_device(g.session_id, device_id)
    else:
        _kick_device(device_id)
    return jsonify({"ok": True})


@app.route("/api/devices/label", methods=["POST"])
def set_device_label():
    """Lets the calling device (identified by its own X-Device-Id header)
    give itself a custom name -- e.g. 'Mary's iPhone' instead of the
    auto-detected 'iPhone 12' -- so it's recognizable in the Connected
    Devices list. Persisted server-side (and locked against being
    overwritten by auto-detection again), so it only needs to be typed
    once, ever, from any device that's already paired -- not re-entered
    each time the web app is opened. Must be unique among currently
    connected devices -- this is what the client-side compulsory naming
    prompt relies on to actually differentiate phones from each other,
    not just cosmetically label them."""
    device_id = request.headers.get("X-Device-Id", "")
    body = request.get_json(force=True) or {}
    label = (body.get("label") or "").strip()[:40]
    if not device_id:
        return jsonify({"ok": False, "error": "Missing device id."}), 400
    if not label:
        return jsonify({"ok": False, "error": "Please enter a name."}), 400
    if CLOUD_MODE:
        ok = _cloud_set_device_label(g.session_id, device_id, label)
    else:
        ok = _set_local_device_label(device_id, label)
    if not ok:
        return jsonify({"ok": False, "error": "That name is already used by another connected phone. Please choose a different, unique name."}), 409
    return jsonify({"ok": True, "label": label})


@app.route("/api/pairing-token", methods=["POST"])
def set_pairing_token():
    """The desktop app calls this every time it displays a QR code,
    handing over the one token that will be accepted by '/' -- see
    _pairing_ok() / _cloud_pairing_ok() above. Immediately invalidates
    whatever token was active before, so an old screenshot or a link
    copied from an earlier QR code stops working the moment a new one is
    shown, not just when it's actually used."""
    body = request.get_json(force=True) or {}
    token = (body.get("token") or "").strip()
    if CLOUD_MODE:
        _cloud_set_pairing_token(g.session_id, token or None)
        return jsonify({"ok": True})
    with _pairing_lock:
        _pairing_state["token"] = token or None
        _pairing_state["consumed"] = False
    return jsonify({"ok": True})


@app.route("/api/announce-disconnect", methods=["POST"])
def announce_disconnect():
    """Called by the desktop app right before it stops this server via
    Settings → Disconnect. Marks the flag picked up by lock_status()
    above so connected phones can show a proper 'Disconnected' state
    instead of just going offline."""
    _disconnect_state["announced"] = True
    return jsonify({"ok": True})


@app.route("/api/unlock", methods=["POST"])
def unlock():
    body = request.get_json(force=True) or {}
    pin = (body.get("pin") or "").strip()
    data = load_state()
    pin_hash = data.get("settings", {}).get("pin_hash", "")
    if not pin_hash:
        session["unlocked"] = True
        return jsonify({"ok": True})
    if hash_secret(pin) == pin_hash:
        session["unlocked"] = True
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Incorrect PIN."}), 400


@app.route("/api/lock", methods=["POST"])
def lock():
    session["unlocked"] = False
    return jsonify({"ok": True})


@app.route("/api/settings/pin", methods=["POST"])
def set_pin():
    body = request.get_json(force=True) or {}
    new_pin = (body.get("new_pin") or "").strip()
    current_pin = (body.get("current_pin") or "").strip()
    data = load_state()
    settings = data.setdefault("settings", {})
    existing_hash = settings.get("pin_hash", "")

    if existing_hash and hash_secret(current_pin) != existing_hash:
        return jsonify({"ok": False, "error": "Current PIN is incorrect."}), 400
    if not new_pin or not new_pin.isdigit() or len(new_pin) < 4:
        return jsonify({"ok": False, "error": "PIN must be at least 4 digits."}), 400

    settings["pin_hash"] = hash_secret(new_pin)
    settings["lock_mode"] = "pin"
    save_state(data)
    session["unlocked"] = True
    return jsonify({"ok": True})


@app.route("/api/settings/pin", methods=["DELETE"])
def remove_pin():
    body = request.get_json(force=True) or {}
    current_pin = (body.get("current_pin") or "").strip()
    data = load_state()
    settings = data.setdefault("settings", {})
    existing_hash = settings.get("pin_hash", "")
    if existing_hash and hash_secret(current_pin) != existing_hash:
        return jsonify({"ok": False, "error": "Current PIN is incorrect."}), 400
    settings.pop("pin_hash", None)
    settings.pop("lock_mode", None)
    save_state(data)
    return jsonify({"ok": True})


# ── dashboard ────────────────────────────────────────────────────────────
@app.route("/api/dashboard")
def dashboard():
    data = load_state()
    tenants = _with_index(data["tenants"])
    active_tenants = [t for t in tenants if not t.get("archived")]
    units = data["units"]
    today = date.today()
    month_name = today.strftime("%B")
    this_month = today.strftime("%Y-%m")

    total_collected = int(
        sum(int(r.get("amount", 0)) for t in tenants for r in t.get("payment_history", [])
            if not r.get("_cancelled", False))
        + sum(float(r.get("amount", 0)) for t in tenants for r in t.get("deposit_history", [])
              if not r.get("_cancelled", False))
    )
    counts = {"paid": 0, "underpaid": 0, "pending": 0}
    for t in active_tenants:
        level, _ = status_level(t, today)
        counts[level] += 1

    # Archived (moved-out) tenants keep their history for records/exports
    # but don't count toward "occupied", same as the desktop app.
    occupied_units = {t.get("unit") for t in active_tenants if t.get("unit")}
    vacant = [u for u in units if u not in occupied_units]
    occupied = len(units) - len(vacant)

    # This-calendar-month income breakdown — mirrors the desktop app's
    # Dashboard "TOTAL INCOME" card exactly (full payments + deposits,
    # minus anything cancelled, all restricted to the current month).
    full_payment_total = int(
        sum(int(r.get("amount", 0)) for t in tenants for r in t.get("payment_history", [])
            if r.get("date", "").startswith(this_month) and not r.get("_cancelled", False)))
    deposit_total = int(
        sum(float(r.get("amount", 0)) for t in tenants for r in t.get("deposit_history", [])
            if r.get("date", "").startswith(this_month) and not r.get("_cancelled", False)))
    cancelled_total = int(
        sum(int(r.get("amount", 0)) for t in tenants for r in t.get("payment_history", [])
            if r.get("date", "").startswith(this_month) and r.get("_cancelled", False))
        + sum(float(r.get("amount", 0)) for t in tenants for r in t.get("deposit_history", [])
              if r.get("date", "").startswith(this_month) and r.get("_cancelled", False)))
    month_income = full_payment_total + deposit_total

    watchlist = get_watchlist_tenants(tenants, max_days=10)
    watchlist_out = [_tenant_summary(t, today) | {"days_left": d} for t, d in watchlist[:8]]

    # First few tenant names, for the Home dashboard's avatar-stack card
    # (mirrors the desktop app's DashboardPage, which shows the first 3
    # active tenants as overlapping avatars plus a "+N" overflow bubble).
    tenant_names = [t.get("name", "Unnamed") for t in active_tenants[:3]]

    # ── Revenue history (trailing 12 months) + a simple next-month
    # forecast, for the Home dashboard's Revenue Forecast card. Mirrors
    # the desktop app's revenue_history()/revenue_forecast(): predicted
    # income is the average of the last up-to-3 months on record, growth
    # compares the two most recent months.
    def _month_income_amount(tenants_list, yy, mm):
        prefix = f"{yy:04d}-{mm:02d}"
        total = sum(
            int(r.get("amount", 0)) for t in tenants_list for r in t.get("payment_history", [])
            if r.get("date", "").startswith(prefix) and not r.get("_cancelled", False)
        )
        total += sum(
            float(r.get("amount", 0)) for t in tenants_list for r in t.get("deposit_history", [])
            if r.get("date", "").startswith(prefix) and not r.get("_cancelled", False)
        )
        return int(total)

    yy, mm = today.year, today.month
    ym_list = []
    for _ in range(12):
        ym_list.append((yy, mm))
        mm -= 1
        if mm == 0:
            mm, yy = 12, yy - 1
    ym_list.reverse()
    revenue_history = [
        {"label": date(yy_, mm_, 1).strftime("%b"), "amount": _month_income_amount(tenants, yy_, mm_)}
        for (yy_, mm_) in ym_list
    ]
    recent_amounts = [h["amount"] for h in revenue_history[-3:]]
    revenue_predicted = int(sum(recent_amounts) / len(recent_amounts)) if recent_amounts else 0
    if len(revenue_history) >= 2 and revenue_history[-2]["amount"] > 0:
        revenue_growth = ((revenue_history[-1]["amount"] - revenue_history[-2]["amount"])
                           / revenue_history[-2]["amount"] * 100)
    else:
        revenue_growth = 0.0

    recent_activity = dashboard_recent_activity(data, today, limit=10)

    return jsonify({
        "app_name": APP_NAME,
        "total_tenants": len(tenants),
        "total_units": len(units),
        "vacant_units": len(vacant),
        "occupied_units": occupied,
        "total_collected": total_collected,
        "counts": counts,
        "month_name": month_name,
        "month_income": month_income,
        "full_payment_total": full_payment_total,
        "deposit_total": deposit_total,
        "cancelled_total": cancelled_total,
        "watchlist": watchlist_out,
        "tenant_names": tenant_names,
        "revenue_history": revenue_history,
        "revenue_predicted": revenue_predicted,
        "revenue_growth": revenue_growth,
        "recent_activity": recent_activity,
    })


def _tenant_summary(t, today):
    level, label = display_status_level(t, today)
    due = _parse_date(t.get("due_date", "")) or _parse_date(t.get("entry_date", ""))
    days_left = (due - today).days if due else None
    dep_paid, dep_remaining, dep_cleared, dep_in_progress = current_deposit_cycle(t)
    return {
        "index": t.get("_idx"),
        "name": t.get("name", "Unnamed"),
        "unit": t.get("unit", "—"),
        "phone": t.get("phone", ""),
        "rent": parse_amount(t.get("rent", 0)),
        "status": t.get("status", "Pending"),
        "level": level,
        "label": label,
        "due_date": t.get("due_date", ""),
        "entry_date": t.get("entry_date", ""),
        "days_left": days_left,
        "deposit_paid": dep_paid,
        "deposit_remaining": dep_remaining,
        "rent_increase_due": rent_increase_due(t),
        "archived": bool(t.get("archived")),
        "vacated_date": t.get("vacated_date", ""),
    }


def _with_index(tenants):
    for i, t in enumerate(tenants):
        t["_idx"] = i
    return tenants


# ── tenants ──────────────────────────────────────────────────────────────
@app.route("/api/tenants")
def list_tenants():
    data = load_state()
    tenants = _with_index(data["tenants"])
    today = date.today()
    q = (request.args.get("q") or "").strip().lower()
    flt = (request.args.get("filter") or "all").lower()

    out = []
    for t in tenants:
        # Archived (moved-out) tenants only ever show under the
        # "Archived" filter, never mixed into the active-tenant filters
        # -- mirrors the desktop app's TenantsPage filtering.
        is_archived = bool(t.get("archived"))
        if flt == "archived":
            if not is_archived:
                continue
        elif is_archived:
            continue
        summary = _tenant_summary(t, today)
        if q and q not in summary["name"].lower() and q not in summary["unit"].lower():
            continue
        if flt not in ("all", "archived") and summary["level"] != flt:
            continue
        out.append(summary)
    out.sort(key=lambda s: (s["days_left"] if s["days_left"] is not None else 99999))
    return jsonify({"tenants": out})


@app.route("/api/tenants/<int:idx>")
def get_tenant(idx):
    data = load_state()
    tenants = data["tenants"]
    if idx < 0 or idx >= len(tenants):
        return jsonify({"error": "not found"}), 404
    t = tenants[idx]
    today = date.today()
    detail = _tenant_summary(t, today) | {
        "index": idx,
        "email": t.get("email", ""),
        "occupation": t.get("occupation", ""),
        "emergency_contact": t.get("emergency_contact", ""),
        "emergency_phone": t.get("emergency_phone", ""),
        "notes": t.get("notes", ""),
        "payment_history": list(reversed(t.get("payment_history", []))),
        "deposit_history": list(reversed(t.get("deposit_history", []))),
        "arrears_history": list(reversed(t.get("arrears_history", []))),
    }
    return jsonify({"tenant": detail})


@app.route("/api/tenants/<int:idx>/months")
def get_tenant_months(idx):
    """Feeds the month-picker dropdown: the already-paid months (from the
    tenant's move-in month up to their current due date -- shown locked,
    can't be re-selected) and the still-open months going forward (shown
    as tickable checkboxes, must be ticked in order starting from the
    first one)."""
    data = load_state()
    tenants = data["tenants"]
    if idx < 0 or idx >= len(tenants):
        return jsonify({"error": "not found"}), 404
    t = tenants[idx]
    try:
        horizon = min(60, max(1, int(request.args.get("horizon", MONTH_PICKER_HORIZON))))
    except (ValueError, TypeError):
        horizon = MONTH_PICKER_HORIZON
    pre_paid, _, _, _ = current_deposit_cycle(t)
    return jsonify({
        "cleared": cleared_months_list(t),
        "open": open_months_list(t, horizon),
        "deposit_paid": pre_paid,
        # Straight from the source record, not the (never-populated)
        # client-side state.tenants cache -- see loadMonthPicker() in the
        # front end, which used to read state.tenants[idx].rent and always
        # got undefined/0 since nothing ever wrote to that array.
        "rent": parse_amount(t.get("rent", 0)),
    })


@app.route("/api/history")
def history():
    """Mirrors the desktop app's Records tab: one card per tenant (Entry
    Date + a merged, date-descending ledger of their payment_history and
    deposit_history — cancelled records flagged in place, exactly as on
    desktop), filtered by the same name/unit/phone search."""
    data = load_state()
    tenants = _with_index(data["tenants"])
    q = (request.args.get("q") or "").strip().lower()

    out = []
    for t in tenants:
        name = t.get("name", "")
        unit = t.get("unit", "")
        phone = t.get("phone", "")
        if q and q not in name.lower() and q not in unit.lower() and q not in phone.lower():
            continue
        txns = []
        for rec in t.get("payment_history", []):
            txns.append({**rec, "kind": "payment"})
        for rec in t.get("deposit_history", []):
            txns.append({**rec, "kind": "deposit"})
        txns.sort(key=lambda r: r.get("date", ""), reverse=True)
        out_txns = []
        for rec in txns:
            frm, to = _compute_txn_period(rec)
            out_txns.append({
                "date": rec.get("date", "—"),
                "kind": rec.get("kind"),
                "cancelled": bool(rec.get("_cancelled", False)),
                "amount": float(rec.get("amount", 0) or 0),
                "from": frm, "to": to,
            })
        out.append({
            "index": t.get("_idx"), "name": name, "unit": unit,
            "entry_date": t.get("entry_date", "") or "—",
            "transactions": out_txns,
        })
    out.sort(key=lambda r: r["name"].lower())
    return jsonify({"tenants": out})


@app.route("/api/units/vacant")
def vacant_units():
    data = load_state()
    occupied = {t.get("unit") for t in data["tenants"] if t.get("unit") and not t.get("archived")}
    exclude = request.args.get("exclude")
    if exclude:
        occupied.discard(exclude)
    out = []
    for name, info in data["units"].items():
        if name in occupied:
            continue
        rent = parse_amount(info.get("rent", 0) if isinstance(info, dict) else info)
        out.append({"name": name, "rent": rent})
    out.sort(key=lambda u: u["name"])
    return jsonify({"units": out})


@app.route("/api/tenants", methods=["POST"])
def add_tenant():
    body = request.get_json(force=True) or {}
    name = (body.get("name") or "").strip().upper()
    phone = (body.get("phone") or "").strip()
    unit = (body.get("unit") or "").strip()
    rent_str = str(body.get("rent") or "").strip()
    entry_str = (body.get("entry_date") or "").strip()
    notes = (body.get("notes") or "").strip()
    replace = bool(body.get("replace"))

    if not name:
        return jsonify({"ok": False, "error": "Tenant name is required."}), 400
    if not unit:
        return jsonify({"ok": False, "error": "Please select a unit."}), 400
    if not phone:
        return jsonify({"ok": False, "error": "Phone number is required."}), 400
    if rent_str and not re.search(r"\d", rent_str):
        return jsonify({"ok": False, "error": "Enter a valid rent amount."}), 400
    try:
        datetime.strptime(entry_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({"ok": False, "error": "Move-in date must be in YYYY-MM-DD format."}), 400

    rent = parse_amount(rent_str) if rent_str else 0.0
    record = {
        "name": name, "phone": phone,
        "email": (body.get("email") or "").strip(),
        "occupation": (body.get("occupation") or "").strip(),
        "emergency_contact": (body.get("emergency_contact") or "").strip(),
        "emergency_phone": (body.get("emergency_phone") or "").strip(),
        "unit": unit, "rent": rent,
        "entry_date": entry_str, "due_date": "",
        "status": "Pending", "pay_date": "", "notes": notes,
        "payment_history": [], "deposit_history": [], "arrears_history": [],
        "locked_periods": [], "deposit_cycle_start": 0, "rent_increase_due": 0.0,
    }

    data = load_state()
    for i, t in enumerate(data["tenants"]):
        if t.get("unit") == unit and not t.get("archived"):
            if not replace:
                return jsonify({"ok": False, "error": "unit_taken",
                                 "message": f"Unit {unit} is already assigned to "
                                            f"{t.get('name', 'Unnamed Tenant')}."}), 409
            # Archive the outgoing tenant (their history is kept, just
            # marked moved-out) instead of overwriting their record —
            # mirrors the desktop app's open_add_dialog.
            t["archived"] = True
            t["vacated_date"] = date.today().strftime("%Y-%m-%d")
            break

    data["tenants"].append(record)
    save_state(data)
    return jsonify({"ok": True, "index": len(data["tenants"]) - 1})


@app.route("/api/tenants/<int:idx>", methods=["PUT"])
def edit_tenant(idx):
    data = load_state()
    tenants = data["tenants"]
    if idx < 0 or idx >= len(tenants):
        return jsonify({"error": "not found"}), 404
    t = tenants[idx]
    body = request.get_json(force=True) or {}

    editable = ["name", "phone", "email", "occupation", "emergency_contact",
                "emergency_phone", "entry_date", "due_date", "notes"]
    for key in editable:
        if key in body:
            val = str(body[key]).strip()
            t[key] = val.upper() if key == "name" else val

    # Unit reassignment — mirrors the desktop app's edit-tenant flow: a
    # move to a different unit is allowed as long as that unit isn't
    # already occupied by another active tenant, and the rent change
    # is handled by apply_tenant_unit_change (immediate, or deferred to
    # the tenant's due date if they're still paid ahead).
    unit_changed = False
    if "unit" in body:
        new_unit = str(body["unit"]).strip()
        if new_unit and new_unit != t.get("unit", ""):
            for other_i, other in enumerate(tenants):
                if other_i != idx and other.get("unit") == new_unit and not other.get("archived"):
                    return jsonify({"ok": False, "error": "unit_taken",
                                     "message": f"Unit {new_unit} is already assigned to "
                                                f"{other.get('name', 'Unnamed Tenant')}."}), 409
            old_unit = t.get("unit", "")
            old_rent = parse_amount(t.get("rent", 0))
            due_str = t.get("due_date", "")
            if "rent" in body and str(body["rent"]).strip():
                new_rent = parse_amount(body["rent"])
            else:
                unit_info = data["units"].get(new_unit)
                new_rent = parse_amount(unit_info.get("rent", old_rent)) if isinstance(unit_info, dict) else old_rent
            apply_tenant_unit_change(t, old_unit, old_rent, due_str, new_unit, new_rent)
            unit_changed = True

    if "rent" in body and not unit_changed:
        rent_str = str(body["rent"]).strip()
        if rent_str and not re.search(r"\d", rent_str):
            return jsonify({"ok": False, "error": "Enter a valid rent amount."}), 400
        t["rent"] = parse_amount(rent_str) if rent_str else 0.0

    save_state(data)
    return jsonify({"ok": True})


@app.route("/api/tenants/<int:idx>", methods=["DELETE"])
def delete_tenant(idx):
    data = load_state()
    tenants = data["tenants"]
    if idx < 0 or idx >= len(tenants):
        return jsonify({"error": "not found"}), 404
    stamp_tenant_deleted(data, tenants[idx])
    del tenants[idx]
    save_state(data)
    return jsonify({"ok": True})


@app.route("/api/tenants/<int:idx>/move-out", methods=["POST"])
def move_out_tenant(idx):
    """Marks a tenant as moved out: their unit becomes free and their
    history is kept, just archived — a reversible-in-spirit alternative
    to permanent Delete. Mirrors the desktop app's TenantsPage.move_out."""
    data = load_state()
    tenants = data["tenants"]
    if idx < 0 or idx >= len(tenants):
        return jsonify({"error": "not found"}), 404
    t = tenants[idx]
    if t.get("archived"):
        return jsonify({"ok": False, "error": "This tenant is already marked as moved out."}), 400
    t["archived"] = True
    t["vacated_date"] = date.today().strftime("%Y-%m-%d")
    save_state(data)
    return jsonify({"ok": True, "vacated_date": t["vacated_date"]})


@app.route("/api/tenants/<int:idx>/old-data", methods=["POST"])
def old_data_tenant(idx):
    """Backfills a tenant's pre-existing rental history (payments/
    deposits from before this app was tracking them) in one batch, and
    optionally sets their current status/due date/pay date/notes once
    every historical record is in. Mirrors the desktop app's Add Old
    Data dialog / add_old_data_records."""
    data = load_state()
    tenants = data["tenants"]
    if idx < 0 or idx >= len(tenants):
        return jsonify({"error": "not found"}), 404
    t = tenants[idx]
    body = request.get_json(force=True) or {}
    records = body.get("records") or []
    if not isinstance(records, list) or not records:
        return jsonify({"ok": False, "error": "Add at least one record."}), 400
    final_state = body.get("final_state") if isinstance(body.get("final_state"), dict) else None
    result = add_old_data_records(t, records, final_state=final_state)
    save_state(data)
    return jsonify({"ok": True, "added": result["added"]})


# ── payments / deposits ───────────────────────────────────────────────────
@app.route("/api/tenants/<int:idx>/payment", methods=["POST"])
def do_record_payment(idx):
    data = load_state()
    tenants = data["tenants"]
    if idx < 0 or idx >= len(tenants):
        return jsonify({"error": "not found"}), 404
    t = tenants[idx]

    # Once a tenant has an instalment/deposit plan in progress, full
    # "Pay Rent" is locked until that balance reaches zero -- enforced
    # here (not just hidden in the UI) so it holds for every caller,
    # including direct API calls, not just the phone-app button.
    _, dep_remaining, dep_cleared, dep_in_progress = current_deposit_cycle(t)
    if dep_in_progress and not dep_cleared:
        return jsonify({"ok": False,
            "error": f"An instalment plan is in progress (UGX {int(dep_remaining):,} "
                     f"still remaining) — record instalments until the balance reaches "
                     f"zero before paying rent in full."}), 400

    body = request.get_json(force=True) or {}
    try:
        months = int(body.get("months", 1))
        if months < 1:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "Select at least one month."}), 400

    result = record_payment(t, months)
    save_state(data)
    return jsonify({"ok": True, "result": result})


@app.route("/api/tenants/<int:idx>/deposit", methods=["POST"])
def do_record_deposit(idx):
    data = load_state()
    tenants = data["tenants"]
    if idx < 0 or idx >= len(tenants):
        return jsonify({"error": "not found"}), 404
    t = tenants[idx]
    body = request.get_json(force=True) or {}
    try:
        months = int(body.get("months", 1))
        if months < 1:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "Select at least one month."}), 400
    try:
        instalment = float(str(body.get("amount", "")).strip().replace(",", ""))
        if instalment <= 0:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "Enter a valid deposit amount greater than zero."}), 400

    result = record_deposit(t, months, instalment)
    save_state(data)
    return jsonify({"ok": True, "result": result})


@app.route("/api/tenants/<int:idx>/cancel", methods=["POST"])
def do_cancel(idx):
    data = load_state()
    tenants = data["tenants"]
    if idx < 0 or idx >= len(tenants):
        return jsonify({"error": "not found"}), 404
    t = tenants[idx]
    body = request.get_json(force=True) or {}
    h_key = body.get("history_key")   # "payment_history" or "deposit_history"
    rec_idx = body.get("record_index")
    if h_key not in ("payment_history", "deposit_history") or rec_idx is None:
        return jsonify({"ok": False, "error": "Invalid record reference."}), 400

    result = cancel_transaction(t, h_key, int(rec_idx))
    if result is None:
        return jsonify({"ok": False, "error": "Record not found."}), 404
    if result == "already_cancelled":
        return jsonify({"ok": False, "error": "That transaction has already been reversed."}), 400
    save_state(data)
    return jsonify({"ok": True, "result": result})


@app.route("/api/tenants/<int:idx>/arrears", methods=["POST"])
def do_clear_arrears(idx):
    data = load_state()
    tenants = data["tenants"]
    if idx < 0 or idx >= len(tenants):
        return jsonify({"error": "not found"}), 404
    t = tenants[idx]
    body = request.get_json(force=True) or {}
    method = body.get("method", "Full")
    due = rent_increase_due(t)
    if due <= 0:
        return jsonify({"ok": False, "error": "No arrears outstanding."}), 400
    try:
        amt = float(str(body.get("amount", "")).strip())
        if amt <= 0:
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "Enter a valid amount."}), 400
    if amt > due + 0.01:
        return jsonify({"ok": False, "error": f"Amount can't exceed the outstanding UGX {int(due):,}."}), 400
    if method == "Full" and amt < due - 0.01:
        return jsonify({"ok": False, "error": "Full clearance must cover the entire outstanding amount."}), 400

    record_arrears_payment(t, amt, method)
    save_state(data)
    return jsonify({"ok": True})


# ── units ────────────────────────────────────────────────────────────────
@app.route("/api/units")
def list_units():
    data = load_state()
    occupants = {t.get("unit"): t.get("name") for t in data["tenants"] if not t.get("archived")}
    out = []
    for name, info in data["units"].items():
        rent = parse_amount(info.get("rent", 0) if isinstance(info, dict) else info)
        location = info.get("location", "") if isinstance(info, dict) else ""
        pending = info.get("pending_rent_increase") if isinstance(info, dict) else None
        out.append({
            "name": name, "rent": rent, "location": location,
            "occupant": occupants.get(name),
            "pending_rent_increase": pending,
        })
    out.sort(key=lambda u: u["name"])
    return jsonify({"units": out})


@app.route("/api/units", methods=["POST"])
def add_unit():
    body = request.get_json(force=True) or {}
    name = (body.get("name") or "").strip()
    rent_str = str(body.get("rent") or "").strip()
    location = (body.get("location") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "Enter a unit ID."}), 400
    data = load_state()
    if name in data["units"]:
        return jsonify({"ok": False, "error": f"Unit '{name}' already exists."}), 409
    if rent_str and not re.search(r"\d", rent_str):
        return jsonify({"ok": False, "error": "Enter a valid rent amount."}), 400
    rent = parse_amount(rent_str) if rent_str else 0.0
    data["units"][name] = {"rent": rent, "location": location}
    save_state(data)
    return jsonify({"ok": True})


@app.route("/api/units/<name>", methods=["PUT"])
def edit_unit(name):
    body = request.get_json(force=True) or {}
    data = load_state()
    if name not in data["units"]:
        return jsonify({"ok": False, "error": "Unit not found."}), 404
    rent_str = str(body.get("rent") or "").strip()
    location = (body.get("location") or "").strip()
    if rent_str and not re.search(r"\d", rent_str):
        return jsonify({"ok": False, "error": "Enter a valid rent amount."}), 400
    rent = parse_amount(rent_str) if rent_str else 0.0
    existing = data["units"].get(name, {})
    pending = existing.get("pending_rent_increase") if isinstance(existing, dict) else None
    updated = {"rent": rent, "location": location}
    if pending:
        updated["pending_rent_increase"] = pending
    data["units"][name] = updated
    save_state(data)
    return jsonify({"ok": True})


@app.route("/api/units/<name>", methods=["DELETE"])
def delete_unit(name):
    data = load_state()
    if name not in data["units"]:
        return jsonify({"ok": False, "error": "Unit not found."}), 404
    stamp_unit_deleted(data, name)
    del data["units"][name]
    save_state(data)
    return jsonify({"ok": True})


@app.route("/api/units/<name>/increase-rent", methods=["POST"])
def increase_rent(name):
    body = request.get_json(force=True) or {}
    data = load_state()
    info = data["units"].setdefault(name, {"rent": 0, "location": ""})
    if not isinstance(info, dict):
        info = {"rent": parse_amount(info), "location": ""}
        data["units"][name] = info
    cur_rent = parse_amount(info.get("rent", 0))

    new_rent_str = str(body.get("new_rent") or "").strip()
    if not re.search(r"\d", new_rent_str):
        return jsonify({"ok": False, "error": "Enter a valid rent amount."}), 400
    new_rent = parse_amount(new_rent_str)
    if new_rent <= cur_rent:
        return jsonify({"ok": False,
                         "error": f"New rent must be greater than the current rent of UGX {int(cur_rent):,}."}), 400

    eff_str = (body.get("effective_month") or "").strip()
    try:
        datetime.strptime(eff_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({"ok": False, "error": "Invalid effective month."}), 400

    info["pending_rent_increase"] = {"new_rent": new_rent, "effective_month": eff_str}
    save_state(data)   # load_state (called next time) applies it immediately if due
    data = load_state()  # re-run migrations now so an immediate increase takes effect right away
    return jsonify({"ok": True})


# ── alerts ───────────────────────────────────────────────────────────────
@app.route("/api/alerts")
def alerts():
    data = load_state()
    tenants = _with_index(data["tenants"])
    today = date.today()
    items = get_watchlist_tenants(tenants, max_days=36500)
    out = [_tenant_summary(t, today) | {"days_left": d} for t, d in items]
    return jsonify({"alerts": out})


# ── exports ──────────────────────────────────────────────────────────────
@app.route("/api/monthly-report")
def monthly_report():
    try:
        year = int(request.args.get("year", date.today().year))
        month = int(request.args.get("month", date.today().month))
        if not (1 <= month <= 12):
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid year/month."}), 400
    data = load_state()
    report = build_monthly_report(data, year, month)
    return jsonify(report)


@app.route("/api/export/monthly-excel")
def do_export_monthly_excel():
    try:
        year = int(request.args.get("year", date.today().year))
        month = int(request.args.get("month", date.today().month))
        if not (1 <= month <= 12):
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid year/month."}), 400
    data = load_state()
    report = build_monthly_report(data, year, month)
    path = export_monthly_excel(report)
    return send_file(path, as_attachment=True, download_name=os.path.basename(path))


@app.route("/api/export/excel")
def do_export_excel():
    data = load_state()
    path = export_excel(data)
    return send_file(path, as_attachment=True, download_name="tenant_records.xlsx")


@app.route("/api/export/pdf")
def do_export_pdf():
    data = load_state()
    path = export_pdf(data)
    return send_file(path, as_attachment=True, download_name="tenant_data.pdf")


# ── settings / data management ────────────────────────────────────────────
@app.route("/api/shutdown", methods=["POST"])
def shutdown_server():
    """Gracefully stops this process. Runs independently of the desktop
    GUI (see module docstring), so the desktop app calls this over HTTP
    to stop it — via 'Stop Phone Access' — rather than needing to hold a
    direct handle to this process, which it may not have if the web
    companion was started in an earlier desktop-app session and is still
    running after that session's window was closed."""
    def _die():
        import time as _t
        _t.sleep(0.3)  # give the HTTP response below a moment to flush
        os._exit(0)
    threading.Thread(target=_die, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/settings/reset", methods=["POST"])
def reset_data():
    """Wipes all tenant/unit data. Requires the admin PIN that was
    originally set on the PC (settings.pin_hash -- the same one used to
    lock/unlock this app) to be re-entered, so a device that merely has
    an open, unlocked session can't nuke everything on its own; a backup
    of the current data is always taken first regardless."""
    data = load_state()
    pin_hash = data.get("settings", {}).get("pin_hash", "")
    if not pin_hash:
        return jsonify({"ok": False,
                         "error": "No admin PIN is set on the PC yet. Set one from "
                                  "the desktop app's Settings before resetting data "
                                  "from here."}), 400
    body = request.get_json(force=True) or {}
    entered_pin = (body.get("admin_pin") or "").strip()
    if not entered_pin or hash_secret(entered_pin) != pin_hash:
        return jsonify({"ok": False, "error": "Incorrect admin PIN."}), 400
    backup_current_data_file()
    save_state({"units": {}, "tenants": [], "settings": {}})
    session["unlocked"] = True
    return jsonify({"ok": True})

# =========================================================================
#  SECTION 4 -- Mobile frontend (single-page app, embedded as one string)
# =========================================================================
INDEX_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<!-- Sets data-theme before anything paints, so the picked theme (saved
     choice, or the OS preference the very first time) applies immediately
     instead of flashing the wrong palette for a frame. -->
<script>
(function(){
  try {
    var saved = localStorage.getItem('rm_theme');
    var mode = (saved === 'light' || saved === 'dark') ? saved
      : (window.matchMedia && window.matchMedia('(prefers-color-scheme: dark)').matches ? 'dark' : 'light');
    document.documentElement.setAttribute('data-theme', mode);
  } catch (e) {
    document.documentElement.setAttribute('data-theme', 'light');
  }
})();
</script>
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1, viewport-fit=cover">
<meta name="format-detection" content="telephone=no">
<title>Tenant Management</title>
<!-- PWA: lets this page be saved/installed to a phone home screen or
     desktop browser (Chrome/Edge install icon, Safari "Add to Home
     Screen") so it launches full-screen with its own icon, for easy
     repeat access without re-typing the LAN address each time. -->
<link rel="manifest" href="manifest.json">
<meta name="theme-color" content="#14181A">
<meta name="color-scheme" content="light dark">
<link rel="icon" href="icon-192.png?v=2">
<link rel="apple-touch-icon" href="icon-256.png?v=2">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="Tenant Management">
<style>
  :root{
    /* ── Brand tokens that don't change between light/dark ── */
    --teal-deep:#0C7F69;
    --teal:#0F9E82;
    --teal-bright:#22D3A6;
    --brass:#F59E0B;
    --danger:#EF4444;
    --warn:#F59E0B;
    --good:#22C55E;

    --radius-sm:12px;
    --radius:18px;
    --radius-lg:26px;

    --font-display:'Segoe UI',system-ui,-apple-system,sans-serif;
    --font-body:'Segoe UI',system-ui,-apple-system,sans-serif;
    --font-mono:'Consolas',ui-monospace,SFMono-Regular,Menlo,monospace;
  }

  /* ── Light mode: bright white, chosen explicitly in Settings or the
     default the first time the app is opened. A step up in brightness
     from the old warm-paper light palette, and with slightly darkened
     muted/line tones so text still passes contrast on pure white. ── */
  :root, html[data-theme="light"]{
    --teal-soft:#E7F3F0;
    --teal-soft2:#F1F7F5;
    --brass-soft:#F6EDDC;
    --ink:#1E2624;
    --muted:#6B756F;
    --line:#E7E2D8;
    --card:#FFFFFF;
    --card-2:#F3F5F2;
    --bg:#FFFFFF;
    --danger-soft:#FBE7EA;
    --warn-soft:#FBF0DA;
    --amber-soft:#FBF1DC;
    --amber-line:#E7CB93;
    --good-soft:#E3F3EC;
    --shadow-sm:0 1px 2px rgba(16,45,40,.06);
    --shadow-md:0 14px 34px -18px rgba(16,45,40,.4);
    --shadow-lift:0 22px 44px -20px rgba(16,45,40,.45);

    /* dashboard tile accents — one tonal family instead of a primary
       color wheel, differentiated by icon/label, not hue. Solid values
       on purpose: the JS layer reuses these as plain text colors too
       (not just tile backgrounds), where a gradient wouldn't render. */
    --accent-tenants:#1A73E8;
    --accent-units:#F9A825;
    --accent-income:#2E7D32;
    --accent-alerts:#E65100;
    --link-blue:#3B82F6;
    color-scheme:light;
  }

  /* ── Dark mode: the previous default palette, now an explicit choice
     in Settings rather than only following the OS. Dashboard accent
     colors are brightened here specifically -- the light-mode accents
     above are dark, low-luminance tones meant for a white card, and
     were nearly invisible against the dark card background (this was
     the "some text colors are not visible" issue). ── */
  html[data-theme="dark"]{
    --teal-soft:#173934;
    --teal-soft2:#123531;
    --brass-soft:#2E2517;
    --ink:#ECEDEC;
    --muted:#8B9592;
    --line:#2A3033;
    --card:#1C2124;
    --card-2:#1B2225;
    --bg:#14181A;
    --danger-soft:#3A1A20;
    --warn-soft:#332612;
    --amber-soft:#33280f;
    --amber-line:#6b5424;
    --good-soft:#153327;
    --shadow-sm:0 1px 2px rgba(0,0,0,.3);
    --shadow-md:0 14px 34px -16px rgba(0,0,0,.55);
    --shadow-lift:0 22px 44px -18px rgba(0,0,0,.6);

    --accent-tenants:var(--teal-bright);
    --accent-units:#FBBF24;
    --accent-income:#34D399;
    --accent-alerts:#F87171;
    --link-blue:#60A5FA;
    color-scheme:dark;
  }
  *{box-sizing:border-box;-webkit-tap-highlight-color:transparent;}
  html,body{margin:0;padding:0;}
  body{
    background:var(--bg);
    color:var(--ink);
    font-family:var(--font-body);
    -webkit-font-smoothing:antialiased;
    text-rendering:optimizeLegibility;
    padding-bottom:104px;
    min-height:100vh;
    min-height:100dvh;
    transition:background .2s ease,color .2s ease;
  }
  h1,h2,h3,.disp{font-family:var(--font-display);font-weight:600;letter-spacing:.1px;}
  .app{max-width:520px;margin:0 auto;position:relative;}

  /* ── header: frosted glass, safe-area aware ───────────────────── */
  header.top{
    position:sticky;top:0;z-index:20;
    padding:calc(env(safe-area-inset-top) + 14px) 18px 14px;
    background:linear-gradient(165deg,rgba(12,127,105,.96),rgba(27,42,46,.98));
    -webkit-backdrop-filter:saturate(160%) blur(14px);
    backdrop-filter:saturate(160%) blur(14px);
    color:#fff;
    display:flex;align-items:center;justify-content:space-between;
    border-bottom:1px solid rgba(255,255,255,.08);
  }
  header.top .brand{display:flex;align-items:center;gap:11px;min-width:0;}
  header.top .brand .mark{
    width:36px;height:36px;border-radius:11px;flex-shrink:0;
    background:#fff;overflow:hidden;
    box-shadow:0 6px 16px -6px rgba(0,0,0,.5);
    display:flex;align-items:center;justify-content:center;
  }
  header.top .brand .mark img{width:100%;height:100%;display:block;object-fit:cover;}
  header.top h1{font-size:16px;margin:0;font-weight:600;letter-spacing:.1px;}
  header.top .sub{font-size:11px;opacity:.7;margin-top:2px;font-family:var(--font-body);}
  .icon-btn{
    width:38px;height:38px;border-radius:12px;background:rgba(255,255,255,.12);
    border:none;color:#fff;display:flex;align-items:center;justify-content:center;
    font-size:16px;cursor:pointer;transition:transform .12s ease,background .15s ease;
  }
  .icon-btn:active{transform:scale(.9);background:rgba(255,255,255,.22);}
  .icon-btn:disabled{opacity:.5;cursor:default;}

  /* ── Home bar: on the Dashboard tab only, the header switches to the
     same plain "page title + today's date, hairline rule below" bar
     used atop the desktop app's Dashboard page, instead of the frosted
     teal brand bar shown on every other screen. ───────────────────── */
  header.top.dash-style{
    background:var(--card);
    -webkit-backdrop-filter:none;backdrop-filter:none;
    color:var(--ink);
    border-bottom:1px solid var(--line);
    box-shadow:none;
  }
  header.top.dash-style .brand .mark{display:none;}
  header.top.dash-style h1{font-size:21px;font-weight:800;letter-spacing:0;color:var(--ink);}
  header.top.dash-style .sub{font-size:13px;opacity:1;color:var(--muted);margin-top:3px;font-weight:400;}
  header.top.dash-style .icon-btn{background:var(--card-2);color:var(--ink);}
  header.top.dash-style .icon-btn:active{background:var(--line);}
  main{padding:16px 16px 8px;}

  .card{
    background:var(--card);border-radius:var(--radius);
    box-shadow:var(--shadow-sm);
    padding:16px;margin-bottom:14px;border:1px solid var(--line);
    transition:background .2s ease,border-color .2s ease;
  }
  .grid2{display:grid;grid-template-columns:1fr 1fr;gap:12px;}
  .stat{padding:14px;border-radius:var(--radius-sm);background:var(--card-2);}
  .stat .num{font-family:var(--font-mono);font-size:21px;font-weight:600;color:var(--teal-deep);font-variant-numeric:tabular-nums;}
  .stat .lbl{font-size:11.5px;color:var(--muted);margin-top:3px;}

  /* ── Dashboard "ledger tile" cards — the signature element: a
     colored ticket-stub header perforated from a plain white body,
     figures set in mono for a receipt-like feel ───────────────── */
  /* ── Dashboard KPI cards — ported from the desktop app's Dashboard
     tab: plain white cards, a small tinted icon badge + caps label up
     top, a big bold value, and a compact body slot (avatar stack /
     occupancy bar / percentage breakdown / status pills) unique to
     each card. The whole card is a tap target, same as the desktop
     app (no separate footer link). ───────────────────────────────── */
  .kpi-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:2px;}
  .kpi-card{
    background:var(--card);border:1px solid var(--line);border-radius:16px;
    padding:18px 16px;cursor:pointer;min-width:0;
    display:flex;flex-direction:column;gap:10px;
    transition:transform .12s ease,background .15s ease;
  }
  .kpi-card:active{transform:scale(.97);background:var(--card-2);}
  .kpi-top{display:flex;align-items:flex-start;justify-content:space-between;gap:10px;}
  .kpi-label{
    font-size:11px;font-weight:700;letter-spacing:.4px;text-transform:uppercase;
    color:var(--muted);line-height:1.3;padding-right:2px;
  }
  .kpi-icon{
    width:38px;height:38px;border-radius:10px;flex-shrink:0;
    display:flex;align-items:center;justify-content:center;font-size:15px;
  }
  .kpi-value{font-size:26px;font-weight:800;color:var(--ink);line-height:1.05;word-break:break-word;font-variant-numeric:tabular-nums;}
  .kpi-avatars{display:flex;align-items:center;}
  .kpi-avatar{
    width:28px;height:28px;border-radius:50%;border:2px solid var(--card);
    display:flex;align-items:center;justify-content:center;color:#fff;
    font-size:10px;font-weight:700;margin-left:-9px;flex-shrink:0;
  }
  .kpi-avatar:first-child{margin-left:0;}
  .kpi-progress{height:8px;border-radius:999px;background:var(--line);overflow:hidden;}
  .kpi-progress > div{height:100%;border-radius:999px;}
  .kpi-occ-row{display:flex;justify-content:space-between;font-size:11px;color:var(--muted);}
  .kpi-breakdown{display:flex;flex-direction:column;gap:6px;}
  .kpi-breakdown-row{display:flex;justify-content:space-between;font-size:11px;}
  .kpi-breakdown-row .l{color:var(--ink);}
  .kpi-breakdown-row .v{font-weight:700;color:var(--ink);}
  .kpi-breakdown-row.danger .l,.kpi-breakdown-row.danger .v{color:var(--danger);}
  .kpi-pills{display:flex;gap:8px;flex-wrap:wrap;}
  .kpi-pill{padding:4px 12px;border-radius:11px;font-size:10px;font-weight:700;}

  /* ── Revenue Forecast card — ported from the desktop app's
     RevenueForecastCard: title + 6/12-month toggle, bar+line chart,
     predicted-income / growth-rate footer stats. ─────────────────── */
  .rf-card{
    background:var(--card);border:1px solid var(--line);border-radius:16px;
    padding:18px 16px;margin-top:14px;
  }
  .rf-header{display:flex;align-items:center;justify-content:space-between;gap:8px;flex-wrap:wrap;margin-bottom:8px;}
  .rf-title{font-size:16px;font-weight:700;color:var(--ink);}
  .rf-toggle{display:flex;gap:4px;}
  .rf-toggle-btn{
    border:none;cursor:pointer;padding:5px 10px;border-radius:7px;
    font-size:11px;font-weight:700;font-family:var(--font-body);
    background:var(--card-2);color:var(--muted);
  }
  .rf-toggle-btn.active{background:var(--teal);color:#fff;}
  .rf-empty{padding:30px 0;text-align:center;color:var(--muted);font-size:13px;}
  .rf-footer{display:flex;gap:28px;margin-top:12px;flex-wrap:wrap;}
  .rf-stat .dot-row{display:flex;align-items:center;gap:6px;}
  .rf-stat .dot{width:8px;height:8px;border-radius:50%;flex-shrink:0;}
  .rf-stat .lbl{font-size:10px;font-weight:700;letter-spacing:.5px;text-transform:uppercase;color:var(--muted);}
  .rf-stat .val{font-size:18px;font-weight:800;color:var(--ink);margin-top:3px;}

  /* ── Recent Activity card — ported from the desktop app's
     RecentActivityCard: a reverse-chronological feed of real actions
     (payments, cancellations, move-outs, rent increases), each row
     icon+title+subtitle+tag / time. ──── */
  .ra-card{
    background:var(--card);border:1px solid var(--line);border-radius:16px;
    padding:16px 16px 4px;margin-top:14px;
  }
  .ra-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:2px;}
  .ra-title{font-size:16px;font-weight:700;color:var(--ink);}
  .ra-viewall{background:none;border:none;cursor:pointer;color:var(--link-blue,#3B82F6);font-size:12px;font-weight:700;font-family:var(--font-body);}
  .ra-row{display:flex;gap:10px;padding:9px 0;border-bottom:1px solid var(--line);}
  .ra-row:last-child{border-bottom:none;}
  .ra-icon{width:32px;height:32px;border-radius:9px;flex-shrink:0;display:flex;align-items:center;justify-content:center;font-size:14px;}
  .ra-text{flex:1;min-width:0;}
  .ra-text .t{font-size:12px;font-weight:700;color:var(--ink);}
  .ra-text .s{font-size:11px;color:var(--muted);margin-top:1px;}
  .ra-tag{display:inline-block;margin-top:4px;padding:1px 7px;border-radius:8px;font-size:9px;font-weight:700;}
  .ra-time{font-size:10px;color:var(--muted);white-space:nowrap;flex-shrink:0;padding-top:1px;}
  .ra-empty{text-align:center;color:var(--muted);font-size:13px;padding:14px 0;}
  .ra-footer-btn{
    display:block;width:100%;text-align:left;background:none;cursor:pointer;
    border:none;border-top:1px solid var(--line);color:var(--muted);
    font-size:12px;font-weight:600;font-family:var(--font-body);padding:9px 0;
  }
  .ra-footer-btn:active{color:var(--ink);}

  .btn{
    border:none;border-radius:var(--radius-sm);font-weight:600;font-size:14px;
    padding:12px 16px;cursor:pointer;display:inline-flex;align-items:center;gap:8px;justify-content:center;
    transition:transform .1s ease,filter .15s ease,background .15s ease;
  }
  .btn:active{transform:scale(.97);}
  .btn-primary{background:linear-gradient(155deg,var(--teal-bright),var(--teal));color:#fff;box-shadow:0 8px 18px -8px rgba(23,143,130,.55);}
  .btn-primary:active{filter:brightness(.92);}
  .btn-ghost{background:var(--teal-soft);color:var(--teal-deep);}
  .btn-danger{background:var(--danger-soft);color:var(--danger);}
  .btn-full{width:100%;}
  .btn:disabled{opacity:.5;}
  input,select,textarea{
    width:100%;border:1.5px solid var(--line);border-radius:var(--radius-sm);
    padding:11px 12px;font-size:15px;font-family:inherit;background:var(--card);color:var(--ink);
    transition:border-color .15s ease;
  }
  input:focus,select:focus,textarea:focus{outline:none;border-color:var(--teal);box-shadow:0 0 0 3px rgba(23,143,130,.15);}
  label.field{display:block;font-size:12.5px;color:var(--muted);font-weight:600;margin:10px 0 5px;}
  .row{display:flex;gap:10px;}
  .row > *{flex:1;}
  .chip{
    display:inline-flex;align-items:center;gap:5px;font-size:12px;font-weight:600;
    padding:5px 10px;border-radius:999px;
  }
  .chip-paid{background:var(--good-soft);color:var(--good);}
  .chip-underpaid{background:rgba(59,130,246,.14);color:var(--link-blue);}
  .chip-pending{background:var(--warn-soft);color:var(--warn);}
  .chip-overdue{background:var(--danger-soft);color:var(--danger);}
  .tenant-row{
    display:flex;align-items:center;gap:12px;padding:12px 4px;border-bottom:1px solid var(--line);cursor:pointer;
    transition:background .12s ease;
  }
  .tenant-row:active{background:var(--card-2);}
  .tenant-row:last-child{border-bottom:none;}
  .avatar{
    width:42px;height:42px;border-radius:12px;background:var(--teal-soft);color:var(--teal-deep);
    display:flex;align-items:center;justify-content:center;font-weight:700;font-size:14px;flex-shrink:0;
    font-family:var(--font-display);
  }
  .tenant-row .meta{flex:1;min-width:0;}
  .tenant-row .name{font-weight:600;font-size:14.5px;}
  .tenant-row .sub{font-size:12px;color:var(--muted);margin-top:1px;}
  .searchbar{
    display:flex;align-items:center;gap:8px;background:var(--card);border:1.5px solid var(--line);
    border-radius:var(--radius-sm);padding:10px 12px;margin-bottom:12px;box-shadow:var(--shadow-sm);
  }
  .menu-row{
    display:flex;align-items:center;gap:14px;padding:15px 16px;cursor:pointer;
    border-bottom:1px solid var(--line);transition:background .12s ease;
  }
  .menu-row:last-child{border-bottom:none;}
  .menu-row:active{background:var(--card-2);}
  .menu-icon{
    font-size:17px;flex-shrink:0;width:36px;height:36px;border-radius:11px;
    background:var(--teal-soft);display:flex;align-items:center;justify-content:center;
  }
  .menu-text{flex:1;min-width:0;}
  .menu-title{font-weight:600;font-size:14.5px;}
  .menu-sub{font-size:12px;color:var(--muted);margin-top:2px;}
  .menu-chevron{font-size:18px;color:var(--muted);}
  .hist-empty{
    background:var(--card-2);border-radius:var(--radius-sm);padding:12px 14px;
    font-size:12.5px;color:var(--muted);
  }
  .hist-table{border:1px solid var(--line);border-radius:var(--radius-sm);overflow:hidden;}
  .hist-hdr{
    display:grid;grid-template-columns:1.1fr 1fr .9fr 1.4fr;gap:4px;
    background:var(--card-2);padding:8px 10px;font-size:10px;font-weight:700;
    letter-spacing:.3px;text-transform:uppercase;color:var(--muted);
  }
  .hist-row{
    display:grid;grid-template-columns:1.1fr 1fr .9fr 1.4fr;gap:4px;
    padding:9px 10px;font-size:12px;font-family:var(--font-mono);border-top:1px solid var(--line);align-items:center;
  }
  .hist-row.hist-cancelled{background:var(--danger-soft);color:var(--danger);}
  .hist-period{color:var(--muted);font-size:11px;font-family:var(--font-body);}
  .hist-row.hist-cancelled .hist-period{color:var(--danger);opacity:.8;}
  .month-picker-control{
    width:100%;border:1.5px solid var(--line);border-radius:var(--radius-sm);padding:11px 12px;
    font-size:14px;background:var(--card);color:var(--ink);display:flex;align-items:center;
    justify-content:space-between;cursor:pointer;
  }
  .month-picker-control .mp-caret{color:var(--muted);font-size:12px;}
  .select-native-wrap{position:relative;flex:1;}
  .select-native-wrap select{
    width:100%;appearance:none;-webkit-appearance:none;-moz-appearance:none;
    border:1.5px solid var(--line);border-radius:var(--radius-sm);padding:11px 30px 11px 12px;
    font-size:14px;background:var(--card);color:var(--ink);
  }
  .select-native-wrap::after{
    content:'▾';position:absolute;right:12px;top:50%;transform:translateY(-50%);
    color:var(--muted);font-size:12px;pointer-events:none;
  }
  .month-picker-panel{
    border:1.5px solid var(--line);border-radius:var(--radius-sm);margin-top:6px;overflow:hidden;background:var(--card);
  }
  .month-picker-list{max-height:230px;overflow-y:auto;}
  .month-row{
    display:flex;align-items:center;gap:10px;padding:10px 12px;font-size:13.5px;
    border-bottom:1px solid var(--line);cursor:pointer;
  }
  .month-row:last-child{border-bottom:none;}
  .month-row input[type=checkbox]{width:17px;height:17px;flex-shrink:0;pointer-events:none;accent-color:var(--teal);}
  .month-row .month-label{flex:1;}
  .month-row.cleared{background:var(--card-2);color:var(--muted);cursor:default;}
  .month-row.cleared .month-tag{
    font-size:10px;font-weight:700;color:var(--teal-deep);background:var(--teal-soft);
    padding:2px 7px;border-radius:999px;text-transform:uppercase;letter-spacing:.3px;
  }
  .month-row.open:active{background:var(--card-2);}
  .month-row.open.ticked{background:var(--card-2);}
  .month-picker-actions{
    display:flex;gap:8px;padding:10px 12px;border-top:1px solid var(--line);background:var(--card-2);
  }
  .month-picker-actions .btn{padding:9px 14px;font-size:13px;}
  .mp-summary-row{display:flex;gap:8px;padding:10px 12px 0;}
  .mp-summary-box{flex:1;background:var(--card-2);border:1.5px solid var(--line);border-radius:10px;padding:8px 10px;}
  .mp-summary-box .mp-summary-label{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.3px;font-weight:700;}
  .mp-summary-box .mp-summary-value{font-size:14.5px;font-weight:600;color:var(--ink);margin-top:2px;font-family:var(--font-mono);}
  .searchbar input{border:none;padding:0;font-size:14px;background:transparent;}
  .filters{display:flex;gap:8px;overflow-x:auto;padding-bottom:2px;margin-bottom:14px;-ms-overflow-style:none;scrollbar-width:none;}
  .filters::-webkit-scrollbar{display:none;}
  .filter-pill{
    flex-shrink:0;padding:7px 14px;border-radius:999px;background:var(--card);border:1.5px solid var(--line);
    font-size:12.5px;font-weight:600;color:var(--muted);cursor:pointer;transition:background .15s ease,color .15s ease,border-color .15s ease;
  }
  .filter-pill.active{background:var(--teal-deep);border-color:var(--teal-deep);color:#fff;}

  /* ── bottom nav: solid black bar, white labels ────────────────── */
  nav.tabbar{
    position:fixed;left:12px;right:12px;bottom:calc(10px + env(safe-area-inset-bottom));
    background:#0A0A0A;
    border:1px solid #000;border-radius:22px;
    display:flex;z-index:30;box-shadow:var(--shadow-lift);
    max-width:496px;margin:0 auto;
  }
  nav.tabbar .tab{
    flex:1;padding:10px 4px;display:flex;flex-direction:column;align-items:center;gap:3px;
    color:#FFFFFF;font-size:10.5px;font-weight:600;cursor:pointer;
    transform:scale(.9);opacity:.72;
    transition:transform .18s cubic-bezier(.34,1.56,.64,1),opacity .18s ease,color .18s ease;
  }
  nav.tabbar .tab.active{
    color:var(--teal-bright);transform:scale(1.08);opacity:1;
  }
  nav.tabbar .tab .ic{font-size:18px;transition:font-size .18s ease;}
  nav.tabbar .tab.active .ic{font-size:21px;}
  .empty{text-align:center;padding:36px 12px;color:var(--muted);}
  .empty .big{font-size:32px;margin-bottom:8px;}
  .section-title{font-size:12.5px;font-weight:700;color:var(--teal-deep);text-transform:uppercase;letter-spacing:.5px;margin:4px 0 10px;font-family:var(--font-body);}
  .modal-backdrop{
    position:fixed;inset:0;background:rgba(8,20,18,.55);z-index:100;
    -webkit-backdrop-filter:blur(2px);backdrop-filter:blur(2px);
    display:flex;align-items:flex-end;justify-content:center;
  }
  .modal{
    background:var(--card);border-radius:26px 26px 0 0;width:100%;max-width:520px;max-height:88vh;overflow-y:auto;
    padding:10px 18px calc(20px + env(safe-area-inset-bottom));animation:slideup .22s cubic-bezier(.2,.8,.3,1);
    box-shadow:0 -20px 50px -20px rgba(0,0,0,.35);
  }
  .modal-handle{width:36px;height:4px;border-radius:999px;background:var(--line);margin:0 auto 14px;}
  @keyframes slideup{from{transform:translateY(28px);opacity:0;}to{transform:translateY(0);opacity:1;}}
  .modal h2{font-size:17px;margin:0 0 4px;font-family:var(--font-display);}
  .modal .desc{font-size:13px;color:var(--muted);margin-bottom:14px;line-height:1.4;}
  .modal .close-x{position:absolute;top:14px;right:14px;}
  .err{color:var(--danger);font-size:12.5px;margin-top:8px;min-height:1px;}
  .toast{
    position:fixed;bottom:112px;left:50%;transform:translateX(-50%);
    background:var(--teal-deep);color:#fff;padding:11px 18px;border-radius:var(--radius-sm);font-size:13.5px;
    z-index:200;max-width:88%;text-align:center;box-shadow:0 14px 30px -10px rgba(0,0,0,.4);
    animation:toastin .18s ease-out;
  }
  @keyframes toastin{from{opacity:0;transform:translate(-50%,8px);}to{opacity:1;transform:translate(-50%,0);}}
  /* ── boot splash: shown the instant the page loads, hidden as soon as
     init() knows whether to show the lock screen, a blocking screen, or
     the app itself. Uses the theme background so it never flashes a
     mismatched color before/after the lock screen or app appear. ── */
  .splash-screen{
    position:fixed;inset:0;z-index:3000;background:var(--bg);color:var(--teal-deep);
    display:flex;flex-direction:column;align-items:center;justify-content:center;gap:14px;
  }
  html[data-theme="dark"] .splash-screen{color:var(--teal-bright);}
  /* Static fallback for prefers-reduced-motion (see hideSplash()) -- a
     plain fade instead of the lunge/wipe below. */
  .splash-screen.hidden{opacity:0;pointer-events:none;transition:opacity .25s ease;}
  .splash-house{
    width:64px;height:64px;overflow:hidden;
    display:flex;align-items:center;justify-content:center;
    animation:splash-grow 1.1s ease-in-out infinite;transform-origin:center;
  }
  .splash-house img{width:100%;height:100%;display:block;object-fit:cover;}
  @keyframes splash-grow{
    0%,100%{transform:scale(.72);opacity:.55;}
    50%{transform:scale(1.18);opacity:1;}
  }
  .splash-label{font-family:var(--font-display);font-weight:600;font-size:15px;color:var(--muted);letter-spacing:.2px;transition:opacity .2s ease;}
  /* Exit sequence, triggered once the dashboard's real data is actually
     ready (see hideSplash()): the breathing icon stops pulsing and takes
     one last lunge toward the viewer, scaling past the edges of the
     screen as it fades ("magnifies completely out") -- then, instead of
     just cutting to the app underneath, a circular wipe opens from that
     same center point so the real data is revealed through an expanding
     porthole rather than appearing abruptly. Both are pure CSS transforms
     (no layout thrash), so they stay smooth even on low-end phones. */
  .splash-screen.exiting .splash-house{
    animation: splash-lunge .55s cubic-bezier(.5,0,.85,0) forwards;
  }
  .splash-screen.exiting .splash-label{ opacity:0; }
  @keyframes splash-lunge{
    0%{transform:scale(1);opacity:1;}
    70%{transform:scale(7);opacity:1;}
    100%{transform:scale(13);opacity:0;}
  }
  .splash-screen.wiping{
    animation: splash-wipe .5s cubic-bezier(.4,0,.2,1) forwards;
  }
  @keyframes splash-wipe{
    from{clip-path:circle(150% at 50% 50%);}
    to{clip-path:circle(0% at 50% 50%);}
  }
  @media (prefers-reduced-motion: reduce){
    .splash-screen.exiting .splash-house{animation:none;opacity:0;transition:opacity .25s ease;}
    .splash-screen.wiping{animation:none;}
  }
  .lockscreen{
    position:fixed;inset:0;background:radial-gradient(120% 100% at 50% 0%,#1B2225 0%,#14181A 55%,#0F1416 100%);
    z-index:1000;display:flex;flex-direction:column;align-items:center;justify-content:center;color:#fff;padding:24px;
  }
  .pin-dots{display:flex;gap:12px;margin:22px 0;}
  .pin-dots .d{width:12px;height:12px;border-radius:50%;border:2px solid rgba(255,255,255,.35);transition:background .15s ease,border-color .15s ease,transform .15s ease;}
  .pin-dots .d.filled{background:var(--teal-bright);border-color:var(--teal-bright);transform:scale(1.1);}
  .keypad{display:grid;grid-template-columns:repeat(3,64px);gap:14px;margin-top:10px;}
  .keypad button{
    width:64px;height:64px;border-radius:50%;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.1);color:#fff;
    font-size:19px;font-weight:600;cursor:pointer;transition:transform .1s ease,background .15s ease;font-family:var(--font-body);
  }
  .keypad button:active{background:rgba(255,255,255,.22);transform:scale(.94);}
  .disc-screen{
    position:fixed;inset:0;background:radial-gradient(120% 100% at 50% 0%,#1B2225 0%,#14181A 55%,#0F1416 100%);
    z-index:1100;display:flex;flex-direction:column;align-items:center;justify-content:center;
    color:#fff;padding:32px;text-align:center;
  }
  .disc-screen .icon{font-size:46px;margin-bottom:16px;}
  .disc-screen h1{margin:0 0 10px;font-size:20px;font-family:var(--font-display);}
  .disc-screen p{margin:0;font-size:13.5px;opacity:.75;max-width:320px;line-height:1.55;}
  .disc-screen .spin{
    margin-top:24px;width:20px;height:20px;border-radius:50%;
    border:2.5px solid rgba(255,255,255,.2);border-top-color:var(--teal-bright);
    animation:disc-spin 0.85s linear infinite;
  }
  @keyframes disc-spin{to{transform:rotate(360deg);}}
  @media (prefers-reduced-motion: reduce){
    .spin,.disc-screen .spin,.ptr-indicator.spinning{animation-duration:2.4s;}
    .modal{animation:none;}
    .toast{animation:none;}
  }
  .txn-item{display:flex;justify-content:space-between;align-items:center;padding:10px 0;border-bottom:1px solid var(--line);font-family:var(--font-mono);font-size:13px;}
  .txn-item:last-child{border-bottom:none;}
  .txn-item .amt{font-weight:600;font-size:14px;}
  .txn-item .amt.cancelled{text-decoration:line-through;color:var(--muted);font-weight:500;}
  .badge-dot{
    position:absolute;top:-4px;right:-4px;background:var(--danger);color:#fff;font-size:10px;font-weight:700;
    min-width:16px;height:16px;border-radius:8px;display:flex;align-items:center;justify-content:center;padding:0 3px;
  }
  a.linklike{color:var(--teal-deep);font-weight:600;text-decoration:none;font-size:13px;}
  .progress-bar{height:8px;border-radius:6px;background:var(--teal-soft);overflow:hidden;margin-top:6px;}
  .progress-bar > div{height:100%;background:linear-gradient(90deg,var(--teal),var(--teal-bright));}
  hr.sep{border:none;border-top:1px solid var(--line);margin:14px 0;}
  .sync-badge{
    font-size:11px;font-weight:700;padding:5px 10px;border-radius:999px;
    background:rgba(255,255,255,.14);color:#fff;white-space:nowrap;margin-right:8px;
  }
  .sync-badge.offline{background:#C1791F;color:#2A1B00;}
  .sync-badge.syncing{background:#2A9BD9;color:#fff;}
  .sync-badge.online{background:#22946A;color:#fff;}
  .sync-badge.disconnected{background:#C24343;color:#fff;}
  .ptr-indicator{
    text-align:center;font-size:12px;color:var(--muted);font-weight:700;
    height:0;overflow:hidden;transition:height .15s ease;
  }
  .ptr-indicator.visible{height:34px;line-height:34px;}
  .ptr-indicator.ready{color:var(--teal-deep);}
  .ptr-indicator.spinning{animation:ptr-pulse 1s ease-in-out infinite;}
  @keyframes ptr-pulse{0%,100%{opacity:.5;}50%{opacity:1;}}
  .pending-note{
    background:var(--amber-soft);color:var(--warn);border:1px solid var(--amber-line);border-radius:var(--radius-sm);
    padding:10px 12px;font-size:12.5px;font-weight:600;margin-bottom:12px;
  }
</style>
</head>
<body>

<div id="splashScreen" class="splash-screen">
  <div class="splash-house">
    <img src="icon-splash.png?v=2" alt="">
  </div>
  <div class="splash-label">Tenant Management</div>
</div>

<div id="lockscreen" class="lockscreen" style="display:none;">
  <div style="font-size:15px;opacity:.85;">Tenant Management</div>
  <h1 style="margin:4px 0 0;font-size:20px;" id="lockTitle">Enter PIN</h1>
  <div class="pin-dots" id="pinDots"></div>
  <div class="err" id="lockErr" style="color:#FFB4C0;min-height:16px;"></div>
  <div class="keypad" id="keypad"></div>
</div>

<div id="discscreen" class="disc-screen" style="display:none;">
  <div class="icon" id="discIcon">🔌</div>
  <h1 id="discTitle">Disconnected</h1>
  <p id="discMsg">Sharing was turned off on the PC, so no tenant data is available on this phone right now. This closes automatically once it's reconnected.</p>
  <div class="spin"></div>
</div>

<div id="devlimitscreen" class="disc-screen" style="display:none;">
  <div class="icon">📵</div>
  <h1>Device Limit Reached</h1>
  <p id="devlimitMsg">Up to 3 phones can be connected at once, and that limit is currently full. This connects automatically as soon as one of the other devices disconnects — no need to scan again.</p>
  <div class="spin"></div>
</div>

<div class="app" id="app" style="display:none;">
  <header class="top">
    <div class="brand">
      <div class="mark"><img src="icon-128.png?v=2" alt=""></div>
      <div>
        <h1 id="headerTitle">Tenant Management</h1>
        <div class="sub" id="headerSub">—</div>
      </div>
    </div>
    <div id="syncBadge" class="sync-badge" style="display:none;"></div>
    <button class="icon-btn" onclick="lockNow()" title="Lock">🔒</button>
  </header>
  <div id="ptrIndicator" class="ptr-indicator">↓ Pull to refresh</div>
  <main id="main"></main>
</div>

<nav class="tabbar" id="tabbar" style="display:none;">
  <div class="tab" data-tab="dashboard"><span class="ic"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" style="display:block;vertical-align:middle;flex-shrink:0"><polyline points="3,3 11,3 11,11 3,11 3,3"/><polyline points="13,3 21,3 21,11 13,11 13,3"/><polyline points="3,13 11,13 11,21 3,21 3,13"/><polyline points="13,13 21,13 21,21 13,21 13,13"/></svg></span>Home</div>
  <div class="tab" data-tab="tenants"><span class="ic"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" style="display:block;vertical-align:middle;flex-shrink:0"><polyline points="9,3.5 6.7,4.4 5.7,6.6 6.7,8.8 9,9.7 11.3,8.8 12.3,6.6 11.3,4.4 9,3.5"/><polyline points="2.5,19.5 3.2,15.5 6.3,13.3 11.7,13.3 14.8,15.5 15.5,19.5"/><polyline points="15,4.2 17,5.3 17.5,7.2 16.7,9 15,9.7"/><polyline points="16.5,13.6 19.3,15.3 20,19.5"/></svg></span>Tenants</div>
  <div class="tab" data-tab="units"><span class="ic"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" style="display:block;vertical-align:middle;flex-shrink:0"><polyline points="4,21 4,3.5 14,3.5 14,21"/><polyline points="14,9 20,9 20,21"/><polyline points="7,7 7,7.3"/><polyline points="10.5,7 10.5,7.3"/><polyline points="7,11 7,11.3"/><polyline points="10.5,11 10.5,11.3"/><polyline points="7,15 7,15.3"/><polyline points="10.5,15 10.5,15.3"/><polyline points="7,21 7,17.5 10.5,17.5 10.5,21"/></svg></span>Units</div>
  <div class="tab" data-tab="alerts" style="position:relative;"><span class="ic"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" style="display:block;vertical-align:middle;flex-shrink:0"><polyline points="12,2.7 12,4.2"/><polyline points="5,10.5 5,15 3,18.5 21,18.5 19,15 19,10.5 16.9,5.9 12,4.9 7.1,5.9 5,10.5"/><polyline points="9.3,18.5 9.3,20 10.6,21.3 13.4,21.3 14.7,20 14.7,18.5"/></svg></span>Alerts<span class="badge-dot" id="alertDot" style="display:none;"></span></div>
  <div class="tab" data-tab="more"><span class="ic"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" style="display:block;vertical-align:middle;flex-shrink:0"><polyline points="12,8.3 9.9,9.1 8.7,11 8.9,13.2 10.5,14.7 12.7,15 14.6,13.8 15.3,11.7 14.5,9.6 12,8.3"/><polyline points="12,2.8 12,5.3"/><polyline points="12,18.7 12,21.2"/><polyline points="3.8,12 6.3,12"/><polyline points="17.7,12 20.2,12"/><polyline points="6.2,6.2 7.9,7.9"/><polyline points="16.1,16.1 17.8,17.8"/><polyline points="6.2,17.8 7.9,16.1"/><polyline points="16.1,7.9 17.8,6.2"/></svg></span>More</div>
</nav>

<div id="modalRoot"></div>
<div id="toastRoot"></div>

<script>
const $ = (sel, el=document) => el.querySelector(sel);
const $$ = (sel, el=document) => [...el.querySelectorAll(sel)];
const fmt = n => 'UGX ' + Math.round(n||0).toLocaleString();
const todayStr = () => new Date().toISOString().slice(0,10);

// ── theme (Light/Dark, set in Settings) ────────────────────────────────
// The bootstrap <script> in <head> already set data-theme before this
// script ever runs, so getTheme() just reads it back.
function getTheme() {
  return document.documentElement.getAttribute('data-theme') || 'light';
}
function setTheme(mode) {
  if (mode !== 'light' && mode !== 'dark') return;
  document.documentElement.setAttribute('data-theme', mode);
  try { localStorage.setItem('rm_theme', mode); } catch (e) {}
  if (state.tab === 'settings') renderSettings();
}

function parseISO(s) {
  if (!s) return null;
  const m = /^(\d{4})-(\d{2})-(\d{2})/.exec(s);
  if (!m) return null;
  return new Date(Date.UTC(+m[1], +m[2]-1, +m[3]));
}
function addMonthsJS(d, n) {
  const day = d.getUTCDate();
  const nd = new Date(Date.UTC(d.getUTCFullYear(), d.getUTCMonth()+n, 1));
  const lastDay = new Date(Date.UTC(nd.getUTCFullYear(), nd.getUTCMonth()+1, 0)).getUTCDate();
  nd.setUTCDate(Math.min(day, lastDay));
  return nd;
}
function diffDays(a, b) { return Math.round((a - b) / 86400000); }

// Mirrors the desktop app's "Lease Summary" card: a progress bar from
// last-due-date to next-due-date while paid, or a plain "Days Overdue"
// figure once pending — same math as ModernRentalApp's tenant detail.
function leaseProgressBlock(t) {
  const today = parseISO(todayStr());
  if (t.level === 'pending') {
    const refD = parseISO(t.due_date || t.entry_date);
    const overdueTxt = refD ? `${diffDays(today, refD)} day(s)` : '—';
    return `<div class="card">
      <div class="section-title" style="margin-top:0;">Days Overdue</div>
      <div style="font-size:24px;font-weight:700;color:var(--danger);">${overdueTxt}</div>
    </div>`;
  }
  const dueD = parseISO(t.due_date);
  if (!dueD) return '';
  const fromD    = addMonthsJS(dueD, -1);
  const remDays  = diffDays(dueD, today);
  const totalDays = diffDays(dueD, fromD) || 1;
  const elapsed   = diffDays(today, fromD);
  const pct       = Math.max(0, Math.min(1, elapsed / totalDays));
  const pctPx     = Math.round(pct * 100);
  const daysNum   = Math.abs(remDays);
  const circSize  = daysNum < 100 ? '22px' : '15px';
  return `<div class="card">
    <div class="section-title" style="margin-top:0;">Lease Progress</div>
    <div style="display:flex;align-items:center;gap:16px;">
      <div style="flex:1;min-width:0;">
        <div style="font-size:13px;font-weight:700;margin-bottom:6px;">~${pctPx}%</div>
        <div class="progress-bar"><div style="width:${pctPx}%"></div></div>
        <div style="display:flex;justify-content:space-between;font-size:11px;color:var(--muted);margin-top:6px;">
          <span>↑ Start: ${fromD.toISOString().slice(0,10)}</span>
          <span>↑ End: ${dueD.toISOString().slice(0,10)}</span>
        </div>
      </div>
      <div style="text-align:center;flex-shrink:0;">
        <div style="width:64px;height:64px;border-radius:50%;border:3px solid var(--line);
                    display:flex;align-items:center;justify-content:center;font-weight:700;
                    font-size:${circSize};color:${remDays<0?'var(--danger)':'var(--ink)'};">${daysNum}</div>
        <div style="font-size:10px;font-weight:700;color:var(--muted);margin-top:4px;">${remDays<0?'OVERDUE':'DAYS LEFT'}</div>
      </div>
    </div>
  </div>`;
}

let state = { tab:'dashboard', tenants:[], selectedIdx:null, filter:'all', q:'', revenueRange:6 };

// ── Offline queue + cache ───────────────────────────────────────────
// While this phone/browser can't reach the PC (PC is off, or off Wi-Fi),
// reads are served from the last-known-good copy of each GET response,
// and any Add/Edit/Payment/etc. is saved to a local queue instead of
// failing outright. As soon as the PC is reachable again, the queue is
// replayed against the real backend automatically, in the order it was
// made, and the screen refreshes with the authoritative server data.
const CACHE_KEY = 'rm_offline_cache_v1';
const QUEUE_KEY = 'rm_offline_queue_v1';
const DEVICE_ID_KEY = 'rm_device_id_v1';
function getDeviceId() {
  let id = localStorage.getItem(DEVICE_ID_KEY);
  if (!id) {
    id = (crypto.randomUUID ? crypto.randomUUID() : (Date.now() + '-' + Math.random().toString(16).slice(2)));
    try { localStorage.setItem(DEVICE_ID_KEY, id); } catch(e) {}
  }
  return id;
}
let DEVICE_ID = getDeviceId();
function adoptCanonicalDeviceId(canonical) {
  // The server recognizes phones primarily by this localStorage-held id,
  // but that storage isn't always the same across contexts on the SAME
  // physical phone: iOS treats a web app "Added to Home Screen" as a
  // separate storage silo from Safari itself, and some in-app/QR-scanner
  // browsers do too. When the server's fingerprint match (see
  // DEVICE_FINGERPRINT below) says this context is really an
  // already-known device, it hands back that device's real id -- adopt
  // it here so this context IS that device from now on, not a new one.
  if (!canonical || canonical === DEVICE_ID) return;
  DEVICE_ID = canonical;
  try { localStorage.setItem(DEVICE_ID_KEY, canonical); } catch(e) {}
}
// Safari deliberately omits the specific iPhone/iPad model from its
// User-Agent string (unlike Android, which usually names the exact
// model), so there's no reliable way to ask the browser "which iPhone is
// this". Physical screen resolution is the best available hint -- not
// perfect (several generations share the same screen size) but enough
// for the server to make a reasonable guess instead of just "iPhone".
const DEVICE_SCREEN_HINT = `${screen.width}x${screen.height}@${window.devicePixelRatio || 1}`;
// A small, stable hash of signals that stay the same for this physical
// device/browser combination regardless of which storage silo a given
// page load happens to be in (Safari tab vs. home-screen install vs. an
// in-app browser) -- used server-side to recognize "this is the same
// phone as an already-known device" even when localStorage itself
// isn't shared between those contexts. Not meant to be unique across
// ALL devices (two identical phone models could coincide) -- just
// stable for THIS one across contexts, which is what actually matters
// here: the fallback only ever kicks in for an otherwise-unrecognized
// device_id, so a coincidental match just means one fewer duplicate
// entry on the household's device list, never a security boundary.
function _fnv1aHash(str) {
  let h = 0x811c9dc5;
  for (let i = 0; i < str.length; i++) {
    h ^= str.charCodeAt(i);
    h = Math.imul(h, 0x01000193);
  }
  return (h >>> 0).toString(16);
}
const DEVICE_FINGERPRINT = _fnv1aHash([
  navigator.userAgent,
  screen.width + 'x' + screen.height,
  navigator.platform || '',
  navigator.hardwareConcurrency || '',
  navigator.maxTouchPoints || '',
  navigator.language || '',
].join('|'));
let isOnline = true;
let syncing = false;
let tempIdCounter = 0;

function loadCache() { try { return JSON.parse(localStorage.getItem(CACHE_KEY)) || {}; } catch(e) { return {}; } }
function saveCache(c) { try { localStorage.setItem(CACHE_KEY, JSON.stringify(c)); } catch(e) {} }
function cacheGet(path) { return loadCache()[path]; }
function cacheSet(path, data) { const c = loadCache(); c[path] = data; saveCache(c); }
function cacheDeletePrefix(prefix) {
  const c = loadCache();
  Object.keys(c).forEach(k => { if (k.indexOf(prefix) === 0) delete c[k]; });
  saveCache(c);
}
function loadQueue() { try { return JSON.parse(localStorage.getItem(QUEUE_KEY)) || []; } catch(e) { return []; } }
function saveQueue(q) { try { localStorage.setItem(QUEUE_KEY, JSON.stringify(q)); } catch(e) {} updateSyncBadge(); }

function updateSyncBadge() {
  const el = $('#syncBadge');
  if (!el) return;
  const pending = loadQueue().length;
  el.style.display = 'inline-block';
  if (syncing) {
    el.className = 'sync-badge syncing';
    el.textContent = '⟳ Syncing…';
  } else if (explicitDisconnect) {
    // The PC deliberately turned sharing off (Settings → Disconnect) --
    // distinct from a plain network drop, per the three states this
    // badge is meant to represent (Online / Offline / Disconnected).
    el.className = 'sync-badge disconnected';
    el.textContent = '⛔ Disconnected';
  } else if (!isOnline) {
    el.className = 'sync-badge offline';
    el.textContent = pending ? `📴 Offline · ${pending} pending` : '📴 Offline';
  } else {
    el.className = 'sync-badge online';
    el.textContent = pending ? `🟢 Online · ${pending} pending` : '🟢 Online';
  }
}

let booted = false;
let explicitDisconnect = false;
let wasKicked = false;
let wasPending = false;

// Four distinct "not showing you data right now" states:
//
// 1. Explicit disconnect — the desktop app tapped Settings → Disconnect
//    and told this server so (see /api/announce-disconnect) *before*
//    actually shutting it down, so this phone catches it on its next
//    poll while still connected. Blocking screen, no tenant data shown,
//    matching what the PC is telling every phone: sharing is off.
// 2. Kicked — the desktop app's Connect Phone device list disconnected
//    just THIS device (see /api/devices/<id>/kick) while everyone else
//    stays connected. Also blocking, but won't clear itself the way
//    explicit-disconnect does: this device id is permanently blocked
//    server-side, so it needs a fresh scan to come back.
// 3. Pending approval — reaching this page at all (QR scan or a
//    forwarded link — the server genuinely cannot tell them apart) is
//    NOT enough on its own. Every new device id lands here first and
//    stays blocked, showing no tenant data whatsoever, until someone
//    with the PC in front of them explicitly approves it from Settings
//    → Connect Phone. This is what makes sharing the link pointless:
//    the recipient just sees "Waiting for approval" forever unless the
//    PC owner deliberately lets that specific device in.
// 4. Network/PC loss — the server just stopped answering, with no such
//    warning (PC powered off, lost Wi-Fi/internet, etc). Non-blocking:
//    keep showing the last-known tenant data from cache, keep the app
//    fully usable, and queue any changes to sync automatically once
//    reachable again. See the offline queue + cache section above.
function setOnline(v) {
  const wasOffline = !isOnline;
  isOnline = v;
  if (v) {
    explicitDisconnect = false;
    wasKicked = false;
    wasPending = false;
    hideDisconnected();
    if (wasOffline) flushQueue();
  }
  updateSyncBadge();
}

// A single failed or successful request is not, on its own, reliable
// proof the PC has actually gone offline or come back -- a slow response
// or one dropped packet happens on a perfectly fine connection too. These
// two only flip the visible badge once the SAME direction has been seen
// STATUS_CONFIRM_THRESHOLD times in a row (across every request source --
// the api() helper, the background ping loop, PIN unlock, etc.), so the
// status only changes once it's actually confirmed either way.
const STATUS_CONFIRM_THRESHOLD = 2;
let _onlineConfirmCount = 0;
let _offlineConfirmCount = 0;
function reportReachable() {
  _offlineConfirmCount = 0;
  if (isOnline) { _onlineConfirmCount = 0; return; }
  _onlineConfirmCount += 1;
  if (_onlineConfirmCount >= STATUS_CONFIRM_THRESHOLD) {
    _onlineConfirmCount = 0;
    setOnline(true);
  }
}
function reportUnreachable() {
  _onlineConfirmCount = 0;
  if (!isOnline) { _offlineConfirmCount = 0; return; }
  _offlineConfirmCount += 1;
  if (_offlineConfirmCount >= STATUS_CONFIRM_THRESHOLD) {
    _offlineConfirmCount = 0;
    setOnline(false);
  }
}

function handleExplicitDisconnect() {
  isOnline = false;
  explicitDisconnect = true;
  wasKicked = false;
  wasPending = false;
  updateSyncBadge();
  $('#discIcon').textContent = '🔌';
  $('#discTitle').textContent = 'Disconnected';
  $('#discMsg').textContent = "Sharing was turned off on the PC, so no tenant data is available on this phone right now. This closes automatically once it's reconnected.";
  showDisconnected();
}

function handleKicked() {
  isOnline = false;
  wasKicked = true;
  explicitDisconnect = false;
  wasPending = false;
  updateSyncBadge();
  $('#discIcon').textContent = '🚫';
  $('#discTitle').textContent = 'Disconnected by Admin';
  $('#discMsg').textContent = "This device was disconnected from the PC's Connect Phone list. It won't reconnect automatically — scan the QR code again on the PC if access should be restored.";
  showDisconnected();
}

function handlePendingApproval() {
  isOnline = false;
  wasPending = true;
  explicitDisconnect = false;
  wasKicked = false;
  updateSyncBadge();
  $('#discIcon').textContent = '🔐';
  $('#discTitle').textContent = 'Waiting for Approval';
  $('#discMsg').textContent = "Opening this link isn't enough on its own — someone needs to approve this device from Settings → Connect Phone on the PC before any tenant data is shown here. This closes automatically once approved.";
  showDisconnected();
}

function showDisconnected() {
  if (explicitDisconnect || wasKicked || wasPending) { $('#discscreen').style.display = 'flex'; }
}
function hideDisconnected() {
  $('#discscreen').style.display = 'none';
}

function showDeviceLimit(maxDevices) {
  $('#lockscreen').style.display = 'none';
  $('#app').style.display = 'none';
  $('#tabbar').style.display = 'none';
  $('#discscreen').style.display = 'none';
  const n = maxDevices || 3;
  $('#devlimitMsg').textContent =
    `Up to ${n} phones can be connected at once, and that limit is currently full. This connects automatically as soon as one of the other devices disconnects — no need to scan again.`;
  $('#devlimitscreen').style.display = 'flex';
}
function hideDeviceLimit() {
  $('#devlimitscreen').style.display = 'none';
}

// Single retry loop behind every blocking state above (kicked, explicit
// disconnect, pending approval, device limit) — re-checks lock-status
// every 5s and either re-shows whichever blocking screen still applies,
// or clears everything and proceeds to the lock screen / app once
// whatever was blocking has genuinely resolved.
let blockedRetryTimer = null;
function enterBlockedState(ls) {
  hideSplash();
  if (ls.kicked) { handleKicked(); }
  else if (ls.disconnecting) { handleExplicitDisconnect(); }
  else if (ls.pending_approval) { handlePendingApproval(); }
  else if (ls.device_limit_reached) { showDeviceLimit(ls.max_devices); }
  else { return; }
  scheduleBlockedRetry();
}
function scheduleBlockedRetry() {
  if (blockedRetryTimer) return; // already waiting on one
  blockedRetryTimer = setTimeout(async () => {
    blockedRetryTimer = null;
    try {
      const res = await fetchTimeout('/api/lock-status', {headers: {'X-Device-Id': DEVICE_ID}}, 3500);
      const ls = await res.json();
      if (ls && (ls.kicked || ls.disconnecting || ls.pending_approval || ls.device_limit_reached)) {
        enterBlockedState(ls);
        return;
      }
      hideDeviceLimit();
      hideDisconnected();
      cacheSet('/api/lock-status', ls);
      setOnline(true);
      if (ls.pin_set && !ls.unlocked) { showLock(); } else { hideLock(); boot(); }
      updateSyncBadge();
    } catch (e) {
      scheduleBlockedRetry();
    }
  }, 5000);
}

// Fetch with a short timeout — an unreachable LAN IP can otherwise hang
// for a very long time before the browser gives up on its own.
async function fetchTimeout(path, opts={}, ms=4000) {
  const ctrl = new AbortController();
  const timer = setTimeout(() => ctrl.abort(), ms);
  // Every call site already sends X-Device-Id; adding the screen hint
  // and fingerprint here once means every request tags along the
  // model-guessing hint and the cross-context device-recognition signal
  // without having to touch every fetchTimeout(...) call individually.
  const headers = { ...(opts.headers || {}), 'X-Device-Screen': DEVICE_SCREEN_HINT, 'X-Device-Fingerprint': DEVICE_FINGERPRINT };
  if (CLOUD_DIRECT && path.charAt(0) === '/' && path.indexOf('/api/') === 0) {
    headers['X-Session-Id'] = CLOUD_DIRECT.sessionId;
    headers['X-Secret-Key'] = CLOUD_DIRECT.secretKey;
  }
  try {
    const res = await fetch(path, {...opts, headers, signal: ctrl.signal});
    adoptCanonicalDeviceId(res.headers.get('X-Canonical-Device-Id'));
    return res;
  } finally {
    clearTimeout(timer);
  }
}

// Skipped whenever a refresh could disrupt someone actively typing/
// editing, or when there's nothing to refresh anyway. Shared by both the
// slow 90s baseline timer below and the fast change-detection in
// pingServer(), so "don't interrupt what the person is doing" is defined
// in exactly one place.
function refreshIfSafe() {
  if (document.hidden) return;
  const lockEl = $('#lockscreen');
  if (lockEl && lockEl.style.display !== 'none') return;
  const modalEl = $('#modalRoot');
  if (modalEl && modalEl.innerHTML.trim() !== '') return;
  if (state.tab === 'add-tenant') return;
  render();
}

// Last data_updated_at we've seen from the server (via /api/lock-status).
// null until the first successful ping, so that first ping only primes
// this value instead of firing a refresh (there's nothing "new" about
// data the app hasn't loaded yet).
let lastKnownDataUpdatedAt = null;

// Budget for the online/offline check itself. A literal ~20ms round trip
// isn't achievable over a real network (even a fast one) -- anything
// that tight would report "Offline" almost constantly even when the PC
// is fine, since plain latency alone usually exceeds it. 2.5s is short
// enough that a genuinely offline PC is detected quickly, while still
// giving a normal request room to actually complete.
const STATUS_CHECK_TIMEOUT_MS = 2500;
// Once a check finishes (success or failure), that Online/Offline
// status is held/displayed for this long before the next check fires --
// the continuous loop below is: check -> show result for 10s -> check
// again, forever.
const STATUS_DISPLAY_MS = 10000;

async function pingServer() {
  try {
    const res = await fetchTimeout('/api/lock-status',
      {cache:'no-store', headers: {'X-Device-Id': DEVICE_ID}}, STATUS_CHECK_TIMEOUT_MS);
    if (res.ok) {
      const data = await res.json().catch(() => ({}));
      adoptCanonicalDeviceId(data && data.canonical_device_id);
      if (data && (data.kicked || data.disconnecting || data.pending_approval || data.device_limit_reached)) {
        enterBlockedState(data);
        return false;
      }
      reportReachable();
      // /api/lock-status is cheap and already polled on this loop for
      // the online badge -- data_updated_at rides along on that same
      // request instead of needing its own poll. The moment it moves
      // past what we last saw (a save on the PC, another phone, etc.),
      // apply the change right away rather than waiting for the slow
      // 90s baseline refresh below.
      const stamp = data && data.data_updated_at;
      if (stamp) {
        if (lastKnownDataUpdatedAt !== null && stamp !== lastKnownDataUpdatedAt) {
          refreshIfSafe();
        }
        lastKnownDataUpdatedAt = stamp;
      }
      return true;
    }
  } catch(e) {}
  // Unreachable, and never told us it was deliberate — treat as a
  // network/PC drop, not a disconnect: keep working from cache.
  //
  // A single miss on its own isn't enough to call it "offline" --
  // reportUnreachable() only actually flips the badge once the same
  // outcome has been confirmed STATUS_CONFIRM_THRESHOLD times in a row
  // (shared with every other request source, not just this loop).
  reportUnreachable();
  return false;
}

// Continuous check/display loop: run a check, then wait STATUS_DISPLAY_MS
// before the next one -- and repeat forever. Uses a self-scheduling
// setTimeout (rather than setInterval) so a slow check can never overlap
// with the next one; the wait always starts only after the previous
// check has actually finished.
(function statusCheckLoop() {
  pingServer().finally(() => setTimeout(statusCheckLoop, STATUS_DISPLAY_MS));
})();

// Baseline safety-net refresh -- catches anything data_updated_at-based
// detection in pingServer() might miss (e.g. this device was hidden/
// backgrounded when the change stamp ticked over) without needing its
// own tight polling interval, since real-time pickup is already handled
// above. 90s, as originally designed.
setInterval(function () {
  refreshIfSafe();
}, 90000);
window.addEventListener('online', pingServer);
window.addEventListener('offline', () => setOnline(false));

// Applies a locally-made change to the cached GET data immediately, so
// the UI reflects it right away even though the server hasn't seen it
// yet. Kept deliberately simple: it only touches the fields it's sure
// about, and otherwise just flags the record "pending sync" rather than
// guessing at recalculated totals (rent cascades, balances, etc. are
// only ever trusted once the real server has computed them).
function applyOptimisticPatch(path, method, bodyObj) {
  if (method === 'POST' && path === '/api/tenants') {
    tempIdCounter -= 1;
    const tempIdx = tempIdCounter;
    const tenant = {
      index: tempIdx, name: bodyObj.name || '(unnamed)', unit: bodyObj.unit || '',
      level: 'pending', label: 'Pending sync', days_left: null, _pending: true,
    };
    const c = loadCache();
    Object.keys(c).forEach(k => {
      if ((k === '/api/tenants' || k.indexOf('/api/tenants?') === 0) && c[k] && Array.isArray(c[k].tenants)) {
        c[k].tenants = [tenant, ...c[k].tenants];
      }
    });
    saveCache(c);
    cacheSet('/api/tenants/' + tempIdx, { tenant: {
      ...tenant, phone: bodyObj.phone||'', email: bodyObj.email||'',
      occupation: bodyObj.occupation||'', emergency_contact: bodyObj.emergency_contact||'',
      emergency_phone: bodyObj.emergency_phone||'', rent: parseFloat(bodyObj.rent)||0,
      entry_date: bodyObj.entry_date||'', notes: bodyObj.notes||'',
      deposit_paid: 0, deposit_remaining: 0,
    }});
    return tempIdx;
  }
  if (method === 'POST' && path === '/api/units') {
    const list = cacheGet('/api/units');
    if (list && Array.isArray(list.units)) {
      list.units.push({ name: bodyObj.name || '(unnamed)', rent: parseFloat(bodyObj.rent)||0, vacant: true, _pending: true });
      cacheSet('/api/units', list);
    }
    return null;
  }
  // Edits / payments / deposits / cancellations / arrears clearing on an
  // *existing* tenant or unit: don't try to recompute the financial
  // result locally — just flag it as pending so the person can see the
  // change was captured and is waiting to sync, without risking showing
  // a wrong number in the meantime.
  const tenantMatch = path.match(/^\/api\/tenants\/(-?\d+)/);
  if (tenantMatch) {
    const idx = parseInt(tenantMatch[1], 10);
    const single = cacheGet('/api/tenants/' + idx);
    if (single && single.tenant) { single.tenant._pendingSync = true; cacheSet('/api/tenants/' + idx, single); }
    const c = loadCache();
    Object.keys(c).forEach(k => {
      if ((k === '/api/tenants' || k.indexOf('/api/tenants?') === 0) && c[k] && Array.isArray(c[k].tenants)) {
        const row = c[k].tenants.find(t => t.index === idx);
        if (row) row._pendingSync = true;
      }
    });
    saveCache(c);
  }
  const unitMatch = path.match(/^\/api\/units\/([^/]+)/);
  if (unitMatch) {
    const list = cacheGet('/api/units');
    if (list && Array.isArray(list.units)) {
      const row = list.units.find(u => u.name === decodeURIComponent(unitMatch[1]));
      if (row) { row._pendingSync = true; cacheSet('/api/units', list); }
    }
  }
  return null;
}

async function flushQueue() {
  if (syncing) return;
  const q = loadQueue();
  if (!q.length) { updateSyncBadge(); return; }
  syncing = true; updateSyncBadge();
  let progressed = false;
  while (loadQueue().length) {
    const current = loadQueue();
    const op = current[0];
    try {
      const res = await fetchTimeout(op.path, {
        method: op.method,
        headers: {'Content-Type':'application/json', 'X-Idempotency-Key': op.idemKey || _newIdemKey()},
        body: op.body,
      }, 6000);
      const data = await res.json().catch(() => ({}));
      if (!res.ok && res.status === 401) { syncing = false; updateSyncBadge(); showLock(); return; }
      // Whether the server accepted or rejected it (e.g. validation
      // error), it's been *seen* — drop it from the queue either way so
      // one bad entry can't block everything behind it forever.
      current.shift();
      saveQueue(current);
      progressed = true;
    } catch (err) {
      // Still unreachable — stop here, keep this and everything after it
      // queued in order, and try again on the next successful ping.
      syncing = false; setOnline(false); updateSyncBadge();
      return;
    }
  }
  // Fully flushed: the local cache may now be stale/wrong in ways that
  // are hard to patch precisely (server-computed totals, cascades,
  // real indexes replacing temp ones) — simplest safe move is to drop
  // the cached GET data so the next screen render pulls fresh truth.
  if (progressed) cacheDeletePrefix('/api/');
  syncing = false; updateSyncBadge();
  toast('All changes synced ✓');
  render();
}

// ── PWA: service worker + "Add to Home Screen" / install prompt ───────
// Registering the worker is what makes Chrome/Edge treat this as an
// installable app at all; the beforeinstallprompt listener captures the
// browser's native install prompt so it can be replayed from the
// Settings button instead of only appearing as a small address-bar icon.
if ('serviceWorker' in navigator) {
  window.addEventListener('load', () => {
    // Relative registration/scope keeps this correct regardless of what
    // path the app happens to be served under. sid/key are forwarded
    // onto the SW script's own URL (readable inside it as
    // self.location.search) so the worker can precache -- and fall back
    // to -- the actual paired root instead of a bare "/", which is what
    // caused the blank white page after installing to the home screen.
    const cd = window.__CLOUD_DIRECT__;
    const swUrl = cd
      ? 'sw.js?sid=' + encodeURIComponent(cd.sessionId) + '&key=' + encodeURIComponent(cd.secretKey)
      : 'sw.js';
    navigator.serviceWorker.register(swUrl).catch(()=>{});
  });
}
let deferredInstallPrompt = null;
const isStandalone = () =>
  window.matchMedia('(display-mode: standalone)').matches || window.navigator.standalone === true;
window.addEventListener('beforeinstallprompt', (e) => {
  e.preventDefault();
  deferredInstallPrompt = e;
  updateInstallUI();
});
window.addEventListener('appinstalled', () => { deferredInstallPrompt = null; updateInstallUI(); });
function updateInstallUI() {
  const btn = $('#installBtn'), hint = $('#installHint');
  if (!btn || !hint) return; // Settings isn't the active page right now
  if (isStandalone()) {
    btn.style.display = 'none';
    hint.textContent = "Already installed \u2014 you're using the installed app right now.";
  } else if (deferredInstallPrompt) {
    btn.style.display = 'block';
    hint.textContent = '';
  } else {
    btn.style.display = 'none';
    const ua = navigator.userAgent || '';
    hint.textContent = /iPhone|iPad|iPod/.test(ua)
      ? 'On iPhone/iPad: tap the Share icon, then "Add to Home Screen".'
      : "Use your browser's menu (\u22ee or \u2026) and choose \\\"Install app\\\" or \\\"Add to Home Screen\\\".";
  }
}
async function triggerInstall() {
  if (!deferredInstallPrompt) return;
  deferredInstallPrompt.prompt();
  await deferredInstallPrompt.userChoice;
  deferredInstallPrompt = null;
  updateInstallUI();
}

// ── offline fallback shapes ──────────────────────────────────────────
// Every render*() function destructures a specific set of fields out of
// whatever api() gives it (d.tenant, d.counts.pending, d.alerts.length,
// etc). If a GET fails while offline AND there's no cached copy of that
// exact path yet (a tab that was never opened while online, or a
// tenant-detail id that was never viewed), returning one generic
// {tenants:[], units:[]} object for every endpoint left most callers
// reading a field that simply wasn't there (d.counts undefined, d.alerts
// undefined, d.tenant undefined...) which threw mid-render. That throw
// was swallowed by render()'s try/catch, so the screen was just left on
// whatever it last showed (often the "Loading…" placeholder) forever --
// this is the "offline, then reload, then nothing" bug. Shaping the
// fallback per endpoint means every render function gets the empty
// structure it actually expects instead of crashing.
function offlineDefaultFor(path) {
  const base = { error: 'offline', offline: true, no_cache: true };
  if (path.indexOf('/api/dashboard') === 0) {
    return { ...base, total_tenants: 0, total_units: 0, occupied_units: 0, vacant_units: 0,
      month_name: '', month_income: 0, full_payment_total: 0, deposit_total: 0,
      cancelled_total: 0, counts: { pending: 0, paid: 0, underpaid: 0 }, watchlist: [] };
  }
  if (/^\/api\/tenants\/[^/]+$/.test(path)) {
    return { ...base, tenant: null };
  }
  if (path.indexOf('/api/tenants') === 0) {
    return { ...base, tenants: [] };
  }
  if (path.indexOf('/api/units') === 0) {
    return { ...base, units: [] };
  }
  if (path.indexOf('/api/alerts') === 0) {
    return { ...base, alerts: [] };
  }
  if (path.indexOf('/api/history') === 0) {
    return { ...base, tenants: [] };
  }
  return { ...base, tenants: [], units: [] };
}

// ── cloud fallback ──────────────────────────────────────────────────
// When this page was loaded straight from the cloud service via the
// direct-cloud QR code (?sid=&key=), CLOUD_DIRECT is embedded before
// anything else in <head> -- see index()'s CLOUD_MODE branch in app.py
// -- and cloudCfg below is configured immediately from it, so every
// read/write always targets the cloud service, with the PC out of the
// picture entirely. If this page was instead loaded directly against a
// PC running app.py locally (e.g. for local testing) with no cloud
// pairing, /api/cloud-config tells us where a configured cloud service
// lives so reads/writes can still reach it once learned. We cache that
// response like any other GET (see cacheSet below) so it's still known
// even once the PC that served it goes away.
const CLOUD_DIRECT = window.__CLOUD_DIRECT__ || null;

let cloudCfg = CLOUD_DIRECT ? {
  configured: true,
  cloud_base_url: window.location.origin,
  session_id: CLOUD_DIRECT.sessionId,
  secret_key: CLOUD_DIRECT.secretKey,
} : null;

async function loadCloudConfig() {
  if (CLOUD_DIRECT) return; // already configured above; nothing to fetch
  const cached = cacheGet('/api/cloud-config');
  if (cached && cached.configured) cloudCfg = cached;
  try {
    const res = await fetchTimeout('/api/cloud-config', {headers:{'X-Device-Id':DEVICE_ID}}, 4000);
    if (res.ok) {
      const data = await res.json();
      cacheSet('/api/cloud-config', data);
      if (data.configured) cloudCfg = data;
    }
  } catch (e) { /* PC unreachable right now -- keep whatever was cached */ }
}

async function cloudFetch(path, opts = {}) {
  if (!cloudCfg || !cloudCfg.configured) throw new Error('cloud_not_configured');
  const headers = {
    'Content-Type': 'application/json',
    'X-Session-Id': cloudCfg.session_id,
    'X-Secret-Key': cloudCfg.secret_key,
    'X-Device-Id': DEVICE_ID,
    'X-Device-Screen': DEVICE_SCREEN_HINT,
    ...(opts.headers || {}),
  };
  const res = await fetchTimeout(cloudCfg.cloud_base_url + path, { ...opts, headers }, 6000);
  const data = await res.json().catch(() => ({}));
  if (!res.ok) { data.ok = false; data.error = data.error || 'cloud_request_failed'; }
  return data;
}

// ── api helper ───────────────────────────────────────────────────────
async function api(path, opts={}) {
  const method = (opts.method || 'GET').toUpperCase();

  if (method === 'GET') {
    if (/^\/api\/tenants\/-\d+/.test(path)) {
      // Temp/offline-created record — never exists on the server, so
      // don't waste time trying to reach it.
      const cached = cacheGet(path);
      if (cached !== undefined) return cached;
      return { error: 'not_found' };
    }

    // ── Cloud-first ──────────────────────────────────────────────────
    // Reads are served from the cloud database whenever it's reachable,
    // full stop. When this page was loaded via the direct-cloud QR code,
    // that's the only path that's ever configured, so it's also the
    // only place reads ever come from. Every successful read here
    // refreshes the local cache too, so the app still has something to
    // show even with no connectivity at all.
    //
    // loadCloudConfig() otherwise only ever ran once, at boot() -- if
    // that single attempt missed (briefly unreachable at that exact
    // moment), cloudCfg stayed unconfigured for the whole session.
    // Retrying it here, every time it's missing, means a later
    // reconnect actually gets picked up instead of needing a reload.
    if (!cloudCfg || !cloudCfg.configured) {
      await loadCloudConfig();
    }
    if (cloudCfg && cloudCfg.configured) {
      try {
        const cloudData = await cloudFetch(path);
        if (cloudData && cloudData.ok !== false) {
          cacheSet(path, cloudData);
          reportReachable();
          return cloudData;
        }
        // Cloud answered but rejected the request (e.g. bad session) --
        // that's still a reachable cloud, not an offline device, so
        // don't flip the badge to offline for this.
      } catch (e) {
        // Cloud genuinely unreachable -- this IS what "offline" means
        // for a cloud-direct phone (there's no other server to fall
        // back to), so the badge needs to reflect it here, not only
        // from the separate pingServer() polling loop.
        reportUnreachable();
      }
    }

    // ── Local/PC fallback ────────────────────────────────────────────
    // Only reached when the cloud isn't configured yet or isn't
    // reachable right now -- e.g. this page is running directly against
    // a local app.py instance with no cloud pairing set up. Still
    // responsible for the isOnline flag -- that badge specifically
    // tracks "is the server this page came from reachable", which is
    // exactly what pingServer() also polls independently every 6s.
    try {
      const res = await fetchTimeout(path, {headers:{'Content-Type':'application/json','X-Device-Id':DEVICE_ID}, ...opts}, 4000);
      if (res.status === 401) { showLock(); throw new Error('locked'); }
      if (res.status === 403) {
        const data403 = await res.json().catch(()=>({}));
        if (data403 && data403.error === 'kicked') {
          handleKicked(); throw new Error('kicked');
        }
        if (data403 && data403.error === 'pending_approval') {
          handlePendingApproval(); throw new Error('pending_approval');
        }
        if (data403 && data403.error === 'device_limit_reached') {
          showDeviceLimit(data403.max_devices); throw new Error('device_limit');
        }
      }
      if (res.status === 502 || res.status === 503 || res.status === 504) {
        // A genuine gateway/server error, not a real answer from the
        // app. Treat exactly like a network failure below: fall back to
        // cache, don't let this overwrite the real cached data or get
        // reported as "Online".
        throw new Error('gateway_unreachable');
      }
      const data = await res.json().catch(()=>({}));
      if (!res.ok && data.ok !== false) { data.error = data.error || 'Request failed.'; }
      cacheSet(path, data);
      reportReachable();
      return data;
    } catch (err) {
      if (err && (err.message === 'locked' || err.message === 'device_limit' || err.message === 'kicked' || err.message === 'pending_approval')) throw err;
      reportUnreachable();
      const cached = cacheGet(path);
      if (cached !== undefined) return cached;
      return offlineDefaultFor(path);
    }
  }

  // Mutating request (POST/PUT/DELETE) -- one idempotency key covers
  // every attempt at THIS action (first try, cloud-transport retry
  // below, and a later offline-queue replay if it comes to that), so
  // the server can tell "this already happened, here's what I did last
  // time" apart from "this is genuinely a new request".
  const idemKey = (opts.headers && opts.headers['X-Idempotency-Key']) || _newIdemKey();
  try {
    const res = await fetchTimeout(path, {headers:{'Content-Type':'application/json','X-Device-Id':DEVICE_ID,'X-Idempotency-Key':idemKey}, ...opts}, 6000);
    if (res.status === 401) { showLock(); throw new Error('locked'); }
    if (res.status === 403) {
      const data403 = await res.json().catch(()=>({}));
      if (data403 && data403.error === 'kicked') {
        handleKicked(); throw new Error('kicked');
      }
      if (data403 && data403.error === 'pending_approval') {
        handlePendingApproval(); throw new Error('pending_approval');
      }
      if (data403 && data403.error === 'device_limit_reached') {
        showDeviceLimit(data403.max_devices); throw new Error('device_limit');
      }
    }
    if (res.status === 502 || res.status === 503 || res.status === 504) {
      // A genuine gateway/server error, not a real answer -- same
      // situation as a network failure, so fall into the catch block
      // below and queue this change to sync once reconnected.
      throw new Error('gateway_unreachable');
    }
    const data = await res.json().catch(()=>({}));
    if (!res.ok && data.ok !== false) { data.error = data.error || 'Request failed.'; }
    reportReachable();
    return data;
  } catch (err) {
    if (err && (err.message === 'locked' || err.message === 'device_limit' || err.message === 'kicked' || err.message === 'pending_approval')) throw err;
    reportUnreachable();
    if (!cloudCfg || !cloudCfg.configured) {
      await loadCloudConfig();
    }
    if (cloudCfg && cloudCfg.configured) {
      try {
        const cloudData = await cloudFetch(path, { method, body: opts.body, headers: {'X-Idempotency-Key': idemKey} });
        if (cloudData && cloudData.ok !== false) {
          toast('Saved directly to the cloud — PC will sync when it reconnects.');
          return cloudData;
        }
      } catch (e2) { /* cloud unreachable too -- fall through to local queue */ }
    }
    let bodyObj = {};
    try { bodyObj = opts.body ? JSON.parse(opts.body) : {}; } catch(e) {}
    const tempIdx = applyOptimisticPatch(path, method, bodyObj);
    const q = loadQueue();
    q.push({ path, method, body: opts.body || null, ts: Date.now(), idemKey });
    saveQueue(q);
    toast('Offline — change saved, will sync once reconnected.');
    return { ok: true, offline: true, queued: true, index: tempIdx };
  }
}
function toast(msg) {
  const el = document.createElement('div');
  el.className = 'toast';
  el.textContent = msg;
  $('#toastRoot').appendChild(el);
  setTimeout(()=>el.remove(), 2600);
}

// ── lock screen ──────────────────────────────────────────────────────
let pinBuffer = '';
function buildKeypad() {
  const kp = $('#keypad'); kp.innerHTML = '';
  const keys = ['1','2','3','4','5','6','7','8','9','','0','⌫'];
  keys.forEach(k => {
    const b = document.createElement('button');
    b.textContent = k;
    if (!k) { b.style.visibility='hidden'; }
    else b.onclick = () => onKey(k);
    kp.appendChild(b);
  });
}
function renderDots() {
  const dots = $('#pinDots'); dots.innerHTML = '';
  for (let i=0;i<Math.max(4,pinBuffer.length);i++) {
    const d = document.createElement('div');
    d.className = 'd' + (i < pinBuffer.length ? ' filled':'');
    dots.appendChild(d);
  }
}
async function onKey(k) {
  $('#lockErr').textContent = '';
  if (k === '⌫') { pinBuffer = pinBuffer.slice(0,-1); renderDots(); return; }
  if (pinBuffer.length >= 8) return;
  pinBuffer += k;
  renderDots();
  if (pinBuffer.length >= 4) {
    try {
      const res = await fetchTimeout('/api/unlock', {method:'POST', headers:{'Content-Type':'application/json','X-Device-Id':DEVICE_ID}, body: JSON.stringify({pin: pinBuffer})}, 4000);
      const d = await res.json();
      if (d.ok) { pinBuffer=''; setOnline(true); hideLock(); boot(); }
      else if (d.error === 'kicked') {
        handleKicked(); pinBuffer=''; renderDots();
      }
      else if (d.error === 'pending_approval') {
        handlePendingApproval(); pinBuffer=''; renderDots();
      }
      else if (d.error === 'device_limit_reached') {
        showDeviceLimit(d.max_devices); pinBuffer=''; renderDots();
      }
      else if (pinBuffer.length >= 8 || d.error) {
        $('#lockErr').textContent = d.error || 'Incorrect PIN.';
        pinBuffer=''; renderDots();
      }
    } catch (err) {
      setOnline(false);
      $('#lockErr').textContent = "Can't reach the PC to verify your PIN right now — try again once it's back online.";
      pinBuffer=''; renderDots();
    }
  }
}
function showLock() {
  hideSplash();
  $('#lockscreen').style.display = 'flex';
  $('#app').style.display = 'none';
  $('#tabbar').style.display = 'none';
  pinBuffer = ''; renderDots();
  buildKeypad();
}
function hideLock() {
  $('#lockscreen').style.display = 'none';
  $('#app').style.display = 'block';
  $('#tabbar').style.display = 'flex';
}
function lockNow() {
  fetchTimeout('/api/lock', {method:'POST'}, 3000).catch(()=>{}).then(()=>showLock());
}

// ── modal helper ─────────────────────────────────────────────────────
function openModal(html, opts) {
  const dismissible = !opts || opts.dismissible !== false;
  const root = $('#modalRoot');
  root.innerHTML = `<div class="modal-backdrop" ${dismissible ? `onclick="if(event.target===this) closeModal()"` : ''}>
    <div class="modal" style="position:relative;">
      <div class="modal-handle"></div>
      ${dismissible ? `<button class="icon-btn close-x" style="background:var(--teal-soft);color:var(--teal-deep);" onclick="closeModal()">✕</button>` : ''}
      ${html}
    </div>
  </div>`;
}
function closeModal() { $('#modalRoot').innerHTML = ''; }

// ── double-submit guard ──────────────────────────────────────────────
// Fixes the "tapped Save/Confirm twice because it felt slow" bug: a
// second tap before the first request finished used to re-run the same
// handler and duplicate the change (double payment, double edit,
// double cancellation, etc). This intercepts every click on any
// <button> that still has its inline onclick -- covers .btn (Save,
// Confirm, Record Payment...) as well as icon-buttons like the ↺
// cancel-transaction control and the lock button.
//
// Two independent locks, not just one:
//  1. btn.disabled -- the normal, visible "this button is busy" state.
//  2. _inFlightActions -- keyed by the action's own onclick code (e.g.
//     "submitPayment(7)"), not by the DOM node. This is the one that
//     actually matters for a genuine double-tap: if anything in between
//     the two taps causes this button to be re-rendered (a fresh
//     re-render swaps in a brand-new, non-disabled button element),
//     lock #1 alone would miss it because it lives on the old node.
//     Locking on the action text itself closes that gap regardless of
//     which physical button element the second tap lands on.
// A short MIN_LOCK_MS floor keeps the lock held for at least that long
// even if the action resolves instantly, so two taps landing in the
// same handful of milliseconds can't both slip through before the lock
// is set.
const _inFlightActions = new Set();
const MIN_LOCK_MS = 400;
function _sleep(ms) { return new Promise(r => setTimeout(r, ms)); }
function _newIdemKey() {
  return (window.crypto && crypto.randomUUID) ? crypto.randomUUID()
    : (Date.now().toString(36) + '-' + Math.random().toString(36).slice(2));
}

// ── Double-tap guard ─────────────────────────────────────────────────
// Wrap any action handler's body in _beginAction(key)/_endAction(key) so a
// second tap on the same action button while the first tap is still being
// processed is ignored instead of firing the action again.
const _actionKeys = new Set();
function _beginAction(key) {
  if (_actionKeys.has(key)) return false;
  _actionKeys.add(key);
  return true;
}
function _endAction(key) { _actionKeys.delete(key); }
document.addEventListener('click', function (e) {
  const btn = e.target.closest('button[onclick]');
  if (!btn) return;
  const code = btn.getAttribute('onclick');
  if (!code) return;
  if (btn.disabled || _inFlightActions.has(code)) {
    // Already running (or just ran) -- swallow this tap entirely so it
    // has no effect at all, rather than letting it fall through to any
    // other handler.
    e.preventDefault();
    e.stopImmediatePropagation();
    return;
  }
  e.preventDefault();
  e.stopImmediatePropagation();
  btn.disabled = true;
  _inFlightActions.add(code);
  Promise.all([
    Promise.resolve().then(() => new Function(code).call(btn)).catch(err => console.error(err)),
    _sleep(MIN_LOCK_MS),
  ]).finally(() => {
    _inFlightActions.delete(code);
    if (document.body.contains(btn)) btn.disabled = false;
  });
}, true);

// ── tabs ─────────────────────────────────────────────────────────────
$$('.tab').forEach(t => t.addEventListener('click', () => switchTab(t.dataset.tab)));
function switchTab(tab) {
  state.tab = tab;
  const moreGroup = ['more', 'history', 'settings'];
  $$('.tab').forEach(t => {
    const dt = t.dataset.tab;
    const active = (dt === tab) || (dt === 'more' && moreGroup.includes(tab));
    t.classList.toggle('active', active);
  });
  return render();
}

async function refreshAlertBadge() {
  try {
    const d = await api('/api/dashboard');
    const n = d.watchlist ? d.watchlist.length : 0;
    const dot = $('#alertDot');
    if (n>0) { dot.style.display='flex'; dot.textContent = n>9?'9+':n; }
    else dot.style.display='none';
  } catch(e) {}
}

// ── header bar (mirrors the desktop app's per-page title bar: plain
// title left, contextual text right, hairline rule below -- switched on
// only for the Dashboard/Home tab, since every other tab keeps the
// frosted brand bar) ───────────────────────────────────────────────────
const WEEKDAY_NAMES = ['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'];
function todayFullLabel() {
  const d = new Date();
  return `${WEEKDAY_NAMES[d.getDay()]}, ${String(d.getDate()).padStart(2,'0')} ${MONTH_NAMES[d.getMonth()]} ${d.getFullYear()}`;
}
function applyHeaderStyle() {
  const isDash = state.tab === 'dashboard';
  $('header.top').classList.toggle('dash-style', isDash);
  $('#headerTitle').textContent = isDash ? 'Dashboard' : 'Tenant Management';
}

// ── render router ────────────────────────────────────────────────────
async function render() {
  applyHeaderStyle();
  const main = $('#main');
  // Don't blank the screen immediately -- on a fast LAN the fetch below
  // usually resolves in well under 100ms, so clearing main first just
  // produces a visible flash/flicker on every tab-bar tap for no benefit.
  // Only show the Loading placeholder if this render is actually slow.
  let loadingShown = false;
  const loadingTimer = setTimeout(() => {
    loadingShown = true;
    main.innerHTML = '<div class="empty">Loading…</div>';
  }, 220);
  try {
    if (state.tab === 'dashboard') await renderDashboard();
    else if (state.tab === 'tenants') await renderTenants();
    else if (state.tab === 'units') await renderUnits();
    else if (state.tab === 'alerts') await renderAlerts();
    else if (state.tab === 'more') renderMoreMenu();
    else if (state.tab === 'history') await renderHistory();
    else if (state.tab === 'settings') await renderSettings();
    else if (state.tab === 'tenant-detail') await renderTenantDetail(state.selectedIdx);
    else if (state.tab === 'add-tenant') renderAddTenant();
  } catch(e) {
    // Previously this just logged and returned, leaving whatever was on
    // screen (often the "Loading…" placeholder from above) stuck there
    // forever with no way out. Now it always leaves the person with a
    // visible, actionable screen instead of a silent dead end.
    console.error(e);
    main.innerHTML = `<div class="empty">
      <div class="big">${isOnline ? '⚠️' : '📴'}</div>
      ${isOnline ? "Something went wrong loading this." : "You're offline and this hasn't loaded before, so there's nothing cached to show yet."}
      <div style="margin-top:14px;"><button class="btn btn-primary" onclick="render()">Try Again</button></div>
    </div>`;
  }
  clearTimeout(loadingTimer);
  refreshAlertBadge();
}

// ── DASHBOARD (Home) ───────────────────────────────────────────────────
// Ported from the desktop app's DashboardPage/DashboardCard: four white
// KPI cards in a 2×2 grid, same accent colors (fixed, not theme-tinted —
// the desktop app's CARD_BLUE/AMBER/GREEN/ORANGE don't change between
// light/dark either), same labels, same per-card content (avatar stack,
// occupancy bar, income % breakdown, status pills), whole card tappable.
const KPI_BLUE   = '#1A73E8';   // Total Tenants
const KPI_AMBER  = '#F9A825';   // Total Units
const KPI_GREEN  = '#2E7D32';   // Total Income
const KPI_ORANGE = '#E65100';   // Payment Alerts
const AVATAR_COLORS = ['#0F9E82','#3B82F6','#F97362','#8E7CF0','#F5A623','#34C795','#EF5A78','#0E8AA0'];

// ── Icon set — same stroke-only Feather/Lucide-traced glyphs as the
// desktop app's _outline_icon() (identical point coordinates, just
// drawn as SVG <polyline>s here instead of QPainter paths), so every
// icon that has a desktop-app counterpart looks the same on both.
// Glyphs the desktop app itself renders as emoji (💰, and the Recent
// Activity feed's ✅/❌/📦/📈/📅) are deliberately left as emoji here
// too, to match — not everything the desktop app shows is an outline
// icon there either.
const ICON_POLYLINES = {
  grid: [[[3,3],[11,3],[11,11],[3,11],[3,3]],[[13,3],[21,3],[21,11],[13,11],[13,3]],
         [[3,13],[11,13],[11,21],[3,21],[3,13]],[[13,13],[21,13],[21,21],[13,21],[13,13]]],
  people: [[[9,3.5],[6.7,4.4],[5.7,6.6],[6.7,8.8],[9,9.7],[11.3,8.8],[12.3,6.6],[11.3,4.4],[9,3.5]],
           [[2.5,19.5],[3.2,15.5],[6.3,13.3],[11.7,13.3],[14.8,15.5],[15.5,19.5]],
           [[15,4.2],[17,5.3],[17.5,7.2],[16.7,9],[15,9.7]],
           [[16.5,13.6],[19.3,15.3],[20,19.5]]],
  building: [[[4,21],[4,3.5],[14,3.5],[14,21]],[[14,9],[20,9],[20,21]],
             [[7,7],[7,7.3]],[[10.5,7],[10.5,7.3]],[[7,11],[7,11.3]],[[10.5,11],[10.5,11.3]],
             [[7,15],[7,15.3]],[[10.5,15],[10.5,15.3]],[[7,21],[7,17.5],[10.5,17.5],[10.5,21]]],
  bell: [[[12,2.7],[12,4.2]],
         [[5,10.5],[5,15],[3,18.5],[21,18.5],[19,15],[19,10.5],[16.9,5.9],[12,4.9],[7.1,5.9],[5,10.5]],
         [[9.3,18.5],[9.3,20],[10.6,21.3],[13.4,21.3],[14.7,20],[14.7,18.5]]],
  gear: [[[12,8.3],[9.9,9.1],[8.7,11],[8.9,13.2],[10.5,14.7],[12.7,15],[14.6,13.8],[15.3,11.7],[14.5,9.6],[12,8.3]],
         [[12,2.8],[12,5.3]],[[12,18.7],[12,21.2]],[[3.8,12],[6.3,12]],[[17.7,12],[20.2,12]],
         [[6.2,6.2],[7.9,7.9]],[[16.1,16.1],[17.8,17.8]],[[6.2,17.8],[7.9,16.1]],[[16.1,7.9],[17.8,6.2]]],
  edit: [[[3,21],[4,19],[19,4],[21,6],[6,21],[3,21]],[[7.5,17.5],[11,21]]],
  trash: [[[3,6],[21,6]],[[8,6],[8,4],[10,2],[14,2],[16,4],[16,6]],
          [[19,6],[19,20],[17,22],[7,22],[5,20],[5,6]],[[10,11],[10,17]],[[14,11],[14,17]]],
  clock: [[[12,21],[7,19.8],[3.5,16],[2.5,11],[4.5,6.3],[9,3.2],[14,3.5],[18.5,6.7],[20.5,11.5],[19,16.5],[15.5,19.8],[12,21]],
          [[12,7.5],[12,12.5],[16,15]]],
  calendar: [[[3,5.5],[21,5.5],[21,20.5],[3,20.5],[3,5.5]],[[3,9.5],[21,9.5]],[[7,3],[7,7.5]],[[17,3],[17,7.5]]],
  alert: [[[12,2.5],[22.5,20.5],[1.5,20.5],[12,2.5]],[[12,9],[12,14]],[[12,17],[12,17.2]]],
  phone: [[[5.5,3.5],[9,3.5],[10.5,8],[8,9.5],[9.5,12.5],[11.5,14.5],[14.5,16],[16,13.5],[20.5,15],[20.5,18.5],[19,20.5],[16,20.5],[8.5,17.5],[3.5,8.5],[5.5,3.5]]],
  mail: [[[3,5],[21,5],[21,19],[3,19],[3,5]],[[3,5.5],[12,13],[21,5.5]]],
  user: [[[12,3.5],[9.2,4.6],[8,7.3],[9.2,10],[12,11.1],[14.8,10],[16,7.3],[14.8,4.6],[12,3.5]],
         [[4,21],[4.8,16.5],[8.5,14],[15.5,14],[19.2,16.5],[20,21]]],
};
function svgIcon(kind, size, color, strokeWidth) {
  const polys = ICON_POLYLINES[kind];
  if (!polys) return '';
  size = size || 18; color = color || 'currentColor'; strokeWidth = strokeWidth || 1.8;
  const parts = polys.map(p => `<polyline points="${p.map(xy => xy.join(',')).join(' ')}"/>`).join('');
  return `<svg width="${size}" height="${size}" viewBox="0 0 24 24" fill="none" stroke="${color}" `
       + `stroke-width="${strokeWidth}" stroke-linecap="round" stroke-linejoin="round" `
       + `style="display:block;vertical-align:middle;flex-shrink:0">${parts}</svg>`;
}

function kpiAlpha(hex, a) {
  const r = parseInt(hex.slice(1,3),16), g = parseInt(hex.slice(3,5),16), b = parseInt(hex.slice(5,7),16);
  return `rgba(${r},${g},${b},${a})`;
}
// Mirrors the desktop app's initials_of(): first letter of the first
// word + first letter of the last word, or the 2nd letter of the same
// word if there's only one.
function initialsOf(name) {
  const parts = (name || '').trim().split(/\s+/).filter(Boolean);
  if (!parts.length) return '?';
  const first = parts[0][0];
  const last = parts.length > 1 ? parts[parts.length - 1][0] : (parts[0][1] || '');
  return (first + last).toUpperCase();
}
// Mirrors the desktop app's avatar_color(): deterministic sum-of-codepoints
// hash into AVATAR_COLORS, so a given name always lands on the same color.
function avatarColorFor(name) {
  let h = 0;
  for (const ch of (name || '')) h += ch.codePointAt(0);
  return AVATAR_COLORS[h % AVATAR_COLORS.length];
}

function kpiCard({accent, iconAlpha, icon, label, value, valueColor, tab, bodyHtml}) {
  return `<div class="kpi-card" onclick="switchTab('${tab}')">
    <div class="kpi-top">
      <div class="kpi-label">${label}</div>
      <div class="kpi-icon" style="background:${kpiAlpha(accent, iconAlpha)};color:${accent}">${icon}</div>
    </div>
    <div class="kpi-value"${valueColor ? ` style="color:${valueColor}"` : ''}>${value}</div>
    ${bodyHtml || ''}
  </div>`;
}

// ── Revenue Forecast card — ported from the desktop app's
// RevenueForecastCard/MiniRevenueChart: a hand-drawn bar+smooth-line
// combo chart (here an inline SVG instead of QPainter), a 6/12-month
// toggle, and predicted-income / growth-rate footer stats.
let _dashData = null;

function buildRevenueChartSvg(history) {
  if (!history.length || !history.some(d => d.amount > 0)) {
    return `<div class="rf-empty">No revenue recorded yet</div>`;
  }
  const w = 320, h = 150, padL = 6, padR = 6, padT = 10, padB = 22;
  const chartW = w - padL - padR, chartH = h - padT - padB;
  const n = history.length;
  const maxAmt = Math.max(...history.map(d => d.amount)) * 1.15 || 1;
  const slotW = chartW / n;
  const barW = Math.min(26, slotW * 0.42);
  const yFor = amt => padT + chartH - (amt / maxAmt) * chartH;
  const centers = history.map((_, i) => padL + slotW * i + slotW / 2);

  const bars = history.map((d, i) => {
    const by = yFor(d.amount);
    const bh = Math.max(2, (padT + chartH) - by);
    return `<rect x="${(centers[i]-barW/2).toFixed(1)}" y="${by.toFixed(1)}" width="${barW.toFixed(1)}" height="${bh.toFixed(1)}" rx="6" style="fill:${kpiAlpha(KPI_BLUE,0.16)}"/>`;
  }).join('');

  const points = history.map((d, i) => [centers[i], yFor(d.amount)]);
  let path = `M ${points[0][0].toFixed(1)} ${points[0][1].toFixed(1)}`;
  for (let i = 1; i < points.length; i++) {
    const [x0, y0] = points[i-1], [x1, y1] = points[i];
    const midX = (x0 + x1) / 2;
    path += ` C ${midX.toFixed(1)} ${y0.toFixed(1)}, ${midX.toFixed(1)} ${y1.toFixed(1)}, ${x1.toFixed(1)} ${y1.toFixed(1)}`;
  }
  const dots = points.map(([x,y]) => `<circle cx="${x.toFixed(1)}" cy="${y.toFixed(1)}" r="3" style="fill:${KPI_BLUE}"/>`).join('');
  const labels = history.map((d, i) => `<text x="${centers[i].toFixed(1)}" y="${h-6}" text-anchor="middle" font-size="8" style="fill:var(--muted)">${escapeHtml(d.label)}</text>`).join('');

  return `<svg viewBox="0 0 ${w} ${h}" xmlns="http://www.w3.org/2000/svg" style="width:100%;height:auto;display:block;">
    ${bars}
    <path d="${path}" fill="none" style="stroke:${KPI_BLUE}" stroke-width="2.4" stroke-linecap="round" stroke-linejoin="round"/>
    ${dots}${labels}
  </svg>`;
}

function rfCardHtml(d) {
  const months = state.revenueRange;
  const history = (d.revenue_history || []).slice(-months);
  const growth = d.revenue_growth || 0;
  const arrow = growth >= 0 ? '▲' : '▼';
  const growthColor = growth >= 0 ? 'var(--good)' : 'var(--danger)';
  return `<div class="rf-card">
    <div class="rf-header">
      <div class="rf-title">Revenue Forecast</div>
      <div class="rf-toggle">
        <button class="rf-toggle-btn ${months===6?'active':''}" onclick="setRevenueRange(6)">6 Months</button>
        <button class="rf-toggle-btn ${months===12?'active':''}" onclick="setRevenueRange(12)">12 Months</button>
      </div>
    </div>
    ${buildRevenueChartSvg(history)}
    <div class="rf-footer">
      <div class="rf-stat">
        <div class="dot-row"><span class="dot" style="background:${KPI_BLUE}"></span><span class="lbl">Predicted Income</span></div>
        <div class="val">${d.revenue_predicted ? fmt(d.revenue_predicted) : 'UGX 0'}</div>
      </div>
      <div class="rf-stat">
        <div class="dot-row"><span class="dot" style="background:var(--muted)"></span><span class="lbl">Growth Rate</span></div>
        <div class="val" style="color:${growthColor}">${arrow} ${Math.abs(growth).toFixed(1)}%</div>
      </div>
    </div>
  </div>`;
}
function setRevenueRange(months) {
  state.revenueRange = months;
  const el = $('#rfCardWrap');
  if (el && _dashData) el.innerHTML = rfCardHtml(_dashData);
}

// ── Recent Activity card — ported from the desktop app's
// RecentActivityCard/dashboard_recent_activity: a single reverse-
// chronological timeline of real actions (payments/deposits received,
// cancellations, tenants moved out, rent increases scheduled), newest
// first; overdue tenants only ever appear as a fallback, and only when
// there's no real activity at all yet. Tag values and their colors/
// icons mirror the desktop app's event dicts exactly (PAID/CANCELLED/
// ARCHIVED/SCHEDULED/OVERDUE).
const RA_TAG_STYLE = {
  PAID:      { tint: '#22C55E', color: 'var(--good)',   icon: '✅' },
  CANCELLED: { tint: '#EF4444', color: 'var(--danger)', icon: '❌' },
  ARCHIVED:  { tint: '#9AA5B1', color: 'var(--muted)',  icon: '📦' },
  SCHEDULED: { tint: '#F59E0B', color: 'var(--warn)',   icon: '📈' },
  OVERDUE:   { tint: '#EF4444', color: 'var(--danger)', icon: '📅' },
};
function raCardHtml(d) {
  const items = d.recent_activity || [];
  const rows = items.length ? items.map(it => {
    const style = RA_TAG_STYLE[it.tag] || RA_TAG_STYLE.PAID;
    return `<div class="ra-row">
      <div class="ra-icon" style="background:${kpiAlpha(style.tint,0.10)}">${style.icon}</div>
      <div class="ra-text">
        <div class="t">${escapeHtml(it.title)}</div>
        <div class="s">${escapeHtml(it.subtitle)}</div>
        <span class="ra-tag" style="background:${kpiAlpha(style.tint,0.13)};color:${style.color}">${escapeHtml(it.tag)}</span>
      </div>
      <div class="ra-time">${escapeHtml(it.time_label)}</div>
    </div>`;
  }).join('') : `<div class="ra-empty">No recent activity yet.</div>`;
  return `<div class="ra-card">
    <div class="ra-header">
      <div class="ra-title">Recent Activity</div>
      <button class="ra-viewall" onclick="switchTab('alerts')">View All</button>
    </div>
    ${rows}
    <button class="ra-footer-btn" onclick="switchTab('alerts')">See Full Notification Log ›</button>
  </div>`;
}

async function renderDashboard() {
  // Matches the desktop app's Dashboard header, which always shows
  // today's date on the right regardless of load state.
  $('#headerSub').textContent = todayFullLabel();
  const d = await api('/api/dashboard');
  if (d.no_cache) {
    // Offline and the dashboard was never loaded on this device before --
    // showing "0 tenants · 0 units" here would look like real data instead
    // of "we don't know yet", so say that plainly instead.
    $('#main').innerHTML = `<div class="empty"><div class="big">📴</div>You're offline, and this device hasn't loaded any data yet. Connect to the same Wi-Fi as the PC (or check its internet), then try again.</div>`;
    return;
  }
  _dashData = d;

  const shownNames = (d.tenant_names || []).slice(0, 3);
  const overflow = Math.max(0, d.total_tenants - shownNames.length);
  const avatarsBody = `<div class="kpi-avatars">${
    shownNames.map(n => `<div class="kpi-avatar" style="background:${avatarColorFor(n)}">${initialsOf(n)}</div>`).join('')
  }${overflow > 0 ? `<div class="kpi-avatar" style="background:${KPI_BLUE}">+${overflow}</div>` : ''}</div>`;

  const occFrac = d.total_units ? (d.occupied_units / d.total_units) : 0;
  const unitsBody = `
    <div class="kpi-progress"><div style="width:${Math.round(occFrac*100)}%;background:${KPI_BLUE}"></div></div>
    <div class="kpi-occ-row"><span>${d.occupied_units} Occupied</span><span>${d.vacant_units} Vacant</span></div>`;

  const gross = d.full_payment_total + d.deposit_total + d.cancelled_total;
  const pct = amt => gross ? Math.round(amt / gross * 100) : 0;
  const incomeBody = `<div class="kpi-breakdown">
    <div class="kpi-breakdown-row"><span class="l">Full</span><span class="v">${pct(d.full_payment_total)}%</span></div>
    <div class="kpi-breakdown-row"><span class="l">Deposits</span><span class="v">${pct(d.deposit_total)}%</span></div>
    <div class="kpi-breakdown-row danger"><span class="l">Cancelled</span><span class="v">${pct(d.cancelled_total)}%</span></div>
  </div>`;

  const alertsBody = `<div class="kpi-pills">
    <div class="kpi-pill" style="background:${kpiAlpha('#F59E0B',0.13)};color:var(--warn)">PENDING: ${d.counts.pending}</div>
    <div class="kpi-pill" style="background:${kpiAlpha('#22C55E',0.13)};color:var(--good)">PAID: ${d.counts.paid}</div>
  </div>`;

  $('#main').innerHTML = `
    <div class="kpi-grid">
      ${kpiCard({
        accent: KPI_BLUE, iconAlpha: 0.10, icon: svgIcon('people'), label:'Total Tenants', value: d.total_tenants,
        tab:'tenants', bodyHtml: avatarsBody
      })}
      ${kpiCard({
        accent: KPI_AMBER, iconAlpha: 0.13, icon: svgIcon('building'), label:'Total Units', value: d.total_units,
        tab:'units', bodyHtml: unitsBody
      })}
      ${kpiCard({
        accent: KPI_GREEN, iconAlpha: 0.13, icon:'💰', label:`Total Income — ${d.month_name}`,
        value: fmt(d.month_income), tab:'history', bodyHtml: incomeBody
      })}
      ${kpiCard({
        accent: KPI_ORANGE, iconAlpha: 0.13, icon: svgIcon('bell'), label:'Payment Alerts', value: d.counts.pending,
        valueColor:'var(--danger)', tab:'alerts', bodyHtml: alertsBody
      })}
    </div>
    <div id="rfCardWrap">${rfCardHtml(d)}</div>
    ${raCardHtml(d)}
  `;
}

function chipClassFor(level) {
  if (level === 'paid') return 'chip-paid';
  if (level === 'underpaid') return 'chip-underpaid';
  if (level === 'overdue') return 'chip-overdue';
  return 'chip-pending';
}
function tenantRowHtml(t) {
  const chipClass = chipClassFor(t.level);
  const daysTxt = t.days_left===null||t.days_left===undefined ? '' :
    (t.days_left<0 ? `Overdue ${Math.abs(t.days_left)}d` : `${t.days_left}d left`);
  const initials = (t.name||'?').split(' ').filter(Boolean).slice(0,2).map(w=>w[0]).join('').toUpperCase();
  const chip = (t._pending || t._pendingSync)
    ? `<span class="chip chip-underpaid">${svgIcon('clock',12)} Pending sync</span>`
    : (t.archived
        ? `<span class="chip" style="background:var(--card-2);color:var(--muted);">Moved out</span>`
        : `<span class="chip ${chipClass}">${t.label}</span>`);
  return `<div class="tenant-row" onclick="openTenant(${t.index})">
    <div class="avatar">${initials||'?'}</div>
    <div class="meta">
      <div class="name">${escapeHtml(t.name)}</div>
      <div class="sub">${escapeHtml(t.unit)} · ${t.archived ? ('Moved out ' + (t.vacated_date||'')) : daysTxt}</div>
    </div>
    ${chip}
  </div>`;
}
function escapeHtml(s){ const d=document.createElement('div'); d.textContent=s??''; return d.innerHTML; }

const MONTH_ABBR = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
function fmtPeriodDate(d) {
  return `${d.getDate()} ${MONTH_ABBR[d.getMonth()]}, ${String(d.getFullYear()).slice(-2)}`;
}
function abbrevPeriod(fromIso, toIso) {
  // Must always agree with the record's actual from/to fields (as shown
  // in the desktop app's Records export), so this renders the exact
  // dates rather than collapsing them to a month/year label -- rounding
  // to whole months previously hid the real (often mid-month) cycle
  // boundaries and could visibly disagree with the underlying dates.
  const f = _parseIsoDate(fromIso), t = _parseIsoDate(toIso);
  if (!f || !t) return `${fromIso || '—'} to ${toIso || '—'}`;
  return `${fmtPeriodDate(f)} to ${fmtPeriodDate(t)}`;
}
function _parseIsoDate(s) {
  if (!s || !/^\d{4}-\d{2}-\d{2}$/.test(s)) return null;
  const [y, m, d] = s.split('-').map(Number);
  return new Date(y, m - 1, d);
}

function openTenant(idx) { state.selectedIdx = idx; state.tab = 'tenant-detail'; render(); }

// ── TENANTS LIST ─────────────────────────────────────────────────────
async function renderTenants() {
  const params = new URLSearchParams({q: state.q, filter: state.filter});
  const d = await api('/api/tenants?' + params.toString());
  $('#headerSub').textContent = `${d.tenants.length} shown`;
  const filters = [['all','All'],['paid','Paid'],['underpaid','Instalments'],['pending','Pending'],['overdue','Overdue'],['archived','Archived']];
  const emptyMsg = d.no_cache
    ? `<div class="empty"><div class="big">📴</div>You're offline and this hasn't loaded before, so there's nothing cached to show yet.</div>`
    : `<div class="empty"><div class="big">${svgIcon('user',40)}</div>No tenants match.</div>`;
  const rows = d.tenants.map(tenantRowHtml).join('') || emptyMsg;
  $('#main').innerHTML = `
    <div class="searchbar"><span>🔎</span><input id="searchInput" placeholder="Search name or unit" value="${escapeHtml(state.q)}"></div>
    <div class="filters">${filters.map(([k,l])=>`<div class="filter-pill ${state.filter===k?'active':''}" data-f="${k}">${l}</div>`).join('')}</div>
    <button class="btn btn-primary btn-full" style="margin-bottom:14px;" onclick="state.tab='add-tenant'; render();">＋ Add Tenant</button>
    <div class="card" style="padding:4px 12px;">${rows}</div>
  `;
  $('#searchInput').addEventListener('input', debounce(e => { state.q = e.target.value; renderTenants(); }, 300));
  $$('.filter-pill').forEach(p => p.addEventListener('click', () => { state.filter = p.dataset.f; renderTenants(); }));
}
function debounce(fn, ms) { let t; return (...a)=>{ clearTimeout(t); t=setTimeout(()=>fn(...a), ms); }; }

// ── ADD TENANT ───────────────────────────────────────────────────────
async function renderAddTenant() {
  $('#headerSub').textContent = 'New tenant';
  const u = await api('/api/units/vacant');
  const opts = u.units.map(x => `<option value="${escapeHtml(x.name)}" data-rent="${x.rent}">${escapeHtml(x.name)} — ${fmt(x.rent)}/mo</option>`).join('');
  $('#main').innerHTML = `
    <button class="btn btn-ghost" style="margin-bottom:12px;" onclick="state.tab='tenants'; render();">← Back</button>
    <div class="card">
      <h2 style="margin-top:0;">Add Tenant</h2>
      <label class="field">Full Name *</label><input id="f_name">
      <label class="field">Unit *</label><select id="f_unit"><option value="">Select a vacant unit…</option>${opts}</select>
      <label class="field">Phone *</label><input id="f_phone">
      <div class="row"><div><label class="field">Email</label><input id="f_email"></div>
      <div><label class="field">Occupation</label><input id="f_occupation"></div></div>
      <div class="row"><div><label class="field">Emergency Contact</label><input id="f_emergency_contact"></div>
      <div><label class="field">Emergency Phone</label><input id="f_emergency_phone"></div></div>
      <label class="field">Monthly Rent (UGX)</label><input id="f_rent" inputmode="numeric">
      <label class="field">Move-in Date *</label><input id="f_entry_date" type="date" value="${todayStr()}">
      <label class="field">Notes</label><textarea id="f_notes" rows="2"></textarea>
      <div class="err" id="addErr"></div>
      <button class="btn btn-primary btn-full" style="margin-top:12px;" onclick="submitAddTenant()">Save Tenant</button>
    </div>
  `;
  $('#f_unit').addEventListener('change', e => {
    const opt = e.target.selectedOptions[0];
    if (opt && opt.dataset.rent) $('#f_rent').value = opt.dataset.rent;
  });
}
async function submitAddTenant(replace=false) {
  if (!replace && !_beginAction('submitAddTenant')) return;
  try {
    const body = {
      name: $('#f_name').value, unit: $('#f_unit').value, phone: $('#f_phone').value,
      email: $('#f_email').value, occupation: $('#f_occupation').value,
      emergency_contact: $('#f_emergency_contact').value, emergency_phone: $('#f_emergency_phone').value,
      rent: $('#f_rent').value, entry_date: $('#f_entry_date').value, notes: $('#f_notes').value, replace,
    };
    const d = await api('/api/tenants', {method:'POST', body: JSON.stringify(body)});
    if (d.ok) { toast('Tenant saved.'); state.tab='tenants'; render(); return; }
    if (d.error === 'unit_taken') {
      if (confirm(d.message + ' The outgoing tenant will be archived (history kept) and this new tenant added. Continue?')) return await submitAddTenant(true);
      return;
    }
    $('#addErr').textContent = d.error || 'Could not save tenant.';
  } finally {
    if (!replace) _endAction('submitAddTenant');
  }
}

// ── TENANT DETAIL ────────────────────────────────────────────────────
async function renderTenantDetail(idx) {
  const d = await api('/api/tenants/' + idx);
  const t = d.tenant;
  if (!t) {
    // Offline, and this specific tenant was never viewed/cached before —
    // nothing safe to render, so say so plainly instead of crashing on
    // t.unit below.
    $('#headerSub').textContent = 'Tenant';
    $('#main').innerHTML = `
      <button class="btn btn-ghost" style="margin-bottom:12px;" onclick="state.tab='tenants'; render();">← Back</button>
      <div class="empty"><div class="big">📴</div>You're offline and this tenant hasn't loaded before, so there's nothing cached to show yet.</div>`;
    return;
  }
  $('#headerSub').textContent = t.unit;

  if (t._pending) {
    // Created while offline — has no real record on the server yet, so
    // there's nothing to safely edit/pay against until it syncs.
    $('#main').innerHTML = `
      <button class="btn btn-ghost" style="margin-bottom:12px;" onclick="state.tab='tenants'; render();">← Back</button>
      <div class="pending-note">${svgIcon('clock',13)} This tenant was added while offline and hasn't synced to the PC yet. It'll sync automatically once reconnected — editing and payments open up after that.</div>
      <div class="card">
        <h2 style="margin-top:0;">${escapeHtml(t.name)}</h2>
        <div style="font-size:13.5px;color:var(--muted);">${escapeHtml(t.unit)} · ${fmt(t.rent)}/mo</div>
        <hr class="sep">
        <div style="font-size:13px;"><b>Phone:</b> ${escapeHtml(t.phone||'—')}</div>
        <div style="font-size:13px;margin-top:4px;"><b>Move-in:</b> ${escapeHtml(t.entry_date||'—')}</div>
      </div>`;
    return;
  }
  const pendingBanner = t._pendingSync
    ? `<div class="pending-note">${svgIcon('clock',13)} A change to this tenant is queued and will sync as soon as this PC is reachable again. Figures below are from the last sync.</div>`
    : '';
  const chipClass = chipClassFor(t.level);

  const depPct = t.rent>0 ? Math.min(100, Math.round((t.deposit_paid/t.rent)*100)) : 0;
  const depBlock = (t.level==='underpaid') ? `
    <div class="card">
      <div class="section-title" style="margin-top:0;">Instalment Progress</div>
      <div style="display:flex;justify-content:space-between;font-size:13px;"><span>${fmt(t.deposit_paid)} paid</span><span>${fmt(t.deposit_remaining)} left</span></div>
      <div class="progress-bar"><div style="width:${depPct}%"></div></div>
    </div>` : '';

  const arrearsBlock = t.rent_increase_due > 0 ? `
    <div class="card" style="border-color:var(--amber-line);background:var(--amber-soft);">
      <div style="display:flex;justify-content:space-between;align-items:center;">
        <div><b>Rent Increase Arrears</b><div class="sub" style="color:var(--muted);font-size:12px;">${fmt(t.rent_increase_due)} owed from a mid-cycle increase</div></div>
        <button class="btn btn-ghost" onclick="openClearArrears(${idx}, ${t.rent_increase_due})">Clear</button>
      </div>
    </div>` : '';

  const payHist = t.payment_history.map((r,i)=>txnRow(t, r, 'payment_history', origIdx(t.payment_history,i))).join('');
  const depHist = t.deposit_history.map((r,i)=>txnRow(t, r, 'deposit_history', origIdx(t.deposit_history,i))).join('');

  $('#main').innerHTML = `
    <button class="btn btn-ghost" style="margin-bottom:12px;" onclick="state.tab='tenants'; render();">← Back</button>
    ${pendingBanner}
    <div class="card">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;">
        <div>
          <h2 style="margin:0 0 2px;">${escapeHtml(t.name)}</h2>
          <div class="sub" style="color:var(--muted);font-size:13px;">${escapeHtml(t.unit)} · ${fmt(t.rent)}/mo</div>
        </div>
        <div style="display:flex;flex-direction:column;align-items:flex-end;gap:6px;">
          <span class="chip ${chipClass}">${t.label}</span>
          ${t.archived ? `<span class="chip" style="background:var(--card-2);color:var(--muted);">Moved out ${escapeHtml(t.vacated_date||'')}</span>` : ''}
        </div>
      </div>
      <hr class="sep">
      <div style="font-size:13px;line-height:1.9;">
        <div>${svgIcon('phone',13)} ${escapeHtml(t.phone||'—')}</div>
        <div>${svgIcon('mail',13)} ${escapeHtml(t.email||'—')}</div>
        <div>💼 ${escapeHtml(t.occupation||'—')}</div>
        <div>🏁 Move-in: ${escapeHtml(t.entry_date||'—')}</div>
        <div>${svgIcon('calendar',13)} Due date: ${escapeHtml(t.due_date||'Not yet set')}</div>
      </div>
      <div class="row" style="margin-top:10px;">
        <button class="btn btn-ghost" onclick="openEditTenant(${idx})">${svgIcon('edit',13)} Edit Tenant</button>
        ${!t.archived ? `<button class="btn btn-ghost" onclick="moveOutTenant(${idx})">🏚 Move Out</button>` : ''}
      </div>
      <div class="row" style="margin-top:10px;">
        <button class="btn btn-danger btn-full" onclick="deleteTenant(${idx})">${svgIcon('trash',13)} Delete Tenant</button>
      </div>
      <div style="text-align:right;margin-top:10px;">
        <button onclick="openAddOldData(${idx})" style="background:none;border:none;color:var(--muted);font-size:12px;font-weight:600;font-family:var(--font-body);cursor:pointer;padding:0;">+ Add Old Data</button>
      </div>
    </div>

    ${leaseProgressBlock(t)}
    ${depBlock}
    ${arrearsBlock}

    <div class="card">
      <div class="row">
        ${t.level === 'underpaid'
          ? `<button class="btn btn-primary" disabled style="opacity:.5;cursor:not-allowed;">🔒 Pay Rent</button>`
          : `<button class="btn btn-primary" onclick="openPaymentModal(${idx})">💳 Pay Rent</button>`}
      </div>
      <div class="row" style="margin-top:10px;">
        <button class="btn btn-ghost" onclick="openDepositModal(${idx})">＋ Record Instalment</button>
      </div>
      ${t.level === 'underpaid'
        ? `<div class="sub" style="color:var(--muted);font-size:12px;margin-top:8px;">🔒 Pay Rent is locked while an instalment plan is in progress. Keep recording instalments until the balance reaches zero.</div>`
        : ''}
    </div>

    <div class="section-title">Payment History</div>
    <div class="card" style="padding:4px 14px;">${payHist || '<div class="empty" style="padding:16px;">No full payments yet.</div>'}</div>

    <div class="section-title">Instalment / Deposit History</div>
    <div class="card" style="padding:4px 14px;">${depHist || '<div class="empty" style="padding:16px;">No instalments yet.</div>'}</div>
  `;
}
// history arrays are returned reversed (most-recent-first); recover original index for cancel calls
function origIdx(arr, displayI) { return arr.length - 1 - displayI; }

function txnRow(t, r, key, origI) {
  const cancelled = r._cancelled;
  const label = key==='payment_history' ? 'Full Payment' : 'Deposit';
  const periodTxt = (r.from_date && r.to_date) ? abbrevPeriod(r.from_date, r.to_date) : '';
  const monthsTxt = r.months ? `${r.months} month${r.months==1?'':'s'}` : '';
  const subTxt = [escapeHtml(r.date||''), monthsTxt, periodTxt].filter(Boolean).join(' · ');
  return `<div class="txn-item">
    <div>
      <div style="font-size:13px;">${label}${cancelled?' <span style="color:var(--danger);font-weight:600;">(Cancelled)</span>':''}</div>
      <div class="sub" style="font-size:11.5px;color:var(--muted);">${subTxt}</div>
    </div>
    <div style="display:flex;align-items:center;gap:10px;">
      <span class="amt ${cancelled?'cancelled':''}">${fmt(r.amount)}</span>
      ${!cancelled ? `<button class="icon-btn" style="width:30px;height:30px;background:var(--teal-soft);color:var(--teal-deep);font-size:13px;" onclick="cancelTxn(${t.index}, '${key}', ${origI})">↺</button>`:''}
    </div>
  </div>`;
}

async function cancelTxn(idx, key, recIdx) {
  const _k = `cancelTxn:${idx}:${key}:${recIdx}`;
  if (!_beginAction(_k)) return;
  try {
    if (!confirm("Cancel this record and reverse its effect on the tenant's account?")) return;
    const d = await api(`/api/tenants/${idx}/cancel`, {method:'POST', body: JSON.stringify({history_key:key, record_index:recIdx})});
    if (d.ok) { toast(`Reversed ${fmt(d.result.total_amount)}.`); state.selectedIdx=idx; state.tab='tenant-detail'; render(); }
    else toast(d.error || 'Could not cancel.');
  } finally {
    _endAction(_k);
  }
}

function openEditTenant(idx) {
  api('/api/tenants/'+idx).then(async d => {
    const t = d.tenant;
    // Vacant units + this tenant's own current unit (so "keep the same
    // unit" is always a valid option) -- mirrors the desktop app's edit
    // dialog, which lets a tenant be reassigned to a different unit.
    const u = await api('/api/units/vacant?exclude=' + encodeURIComponent(t.unit));
    const unitOpts = u.units.map(x =>
      `<option value="${escapeHtml(x.name)}" data-rent="${x.rent}" ${x.name===t.unit?'selected':''}>${escapeHtml(x.name)} — ${fmt(x.rent)}/mo</option>`
    ).join('');
    const archivedNote = t.archived
      ? `<div class="pending-note">🏚 This tenant moved out on ${escapeHtml(t.vacated_date||'—')}. They're archived, so they won't show in active tenant lists — editing still works.</div>`
      : '';
    openModal(`
      <h2>Edit Tenant</h2>
      ${archivedNote}
      <label class="field">Full Name</label><input id="e_name" value="${escapeHtml(t.name)}">
      <label class="field">Unit</label><select id="e_unit">${unitOpts}</select>
      <div class="row"><div><label class="field">Phone</label><input id="e_phone" value="${escapeHtml(t.phone)}"></div>
      <div><label class="field">Email</label><input id="e_email" value="${escapeHtml(t.email)}"></div></div>
      <label class="field">Occupation</label><input id="e_occupation" value="${escapeHtml(t.occupation)}">
      <div class="row"><div><label class="field">Emergency Contact</label><input id="e_emergency_contact" value="${escapeHtml(t.emergency_contact)}"></div>
      <div><label class="field">Emergency Phone</label><input id="e_emergency_phone" value="${escapeHtml(t.emergency_phone)}"></div></div>
      <div class="row"><div><label class="field">Move-in Date</label><input id="e_entry_date" type="date" value="${t.entry_date}"></div>
      <div><label class="field">Due Date</label><input id="e_due_date" type="date" value="${t.due_date}"></div></div>
      <label class="field">Rent (UGX)</label><input id="e_rent" value="${t.rent}">
      <div class="sub" style="font-size:11px;color:var(--muted);margin-top:2px;">Changing units auto-fills that unit's rent. If the tenant is still paid ahead, the old rent keeps billing until their due date, then the new rate kicks in.</div>
      <label class="field">Notes</label><textarea id="e_notes" rows="2">${escapeHtml(t.notes)}</textarea>
      <div class="err" id="editErr"></div>
      <button class="btn btn-primary btn-full" style="margin-top:12px;" onclick="submitEditTenant(${idx})">Save Changes</button>
    `);
    $('#e_unit').addEventListener('change', e => {
      const opt = e.target.selectedOptions[0];
      if (opt && opt.dataset.rent) $('#e_rent').value = opt.dataset.rent;
    });
  });
}
async function submitEditTenant(idx) {
  const _k = `submitEditTenant:${idx}`;
  if (!_beginAction(_k)) return;
  try {
    const body = {
      name: $('#e_name').value, unit: $('#e_unit').value, phone: $('#e_phone').value, email: $('#e_email').value,
      occupation: $('#e_occupation').value, emergency_contact: $('#e_emergency_contact').value,
      emergency_phone: $('#e_emergency_phone').value, entry_date: $('#e_entry_date').value,
      due_date: $('#e_due_date').value, rent: $('#e_rent').value, notes: $('#e_notes').value,
    };
    const d = await api('/api/tenants/'+idx, {method:'PUT', body: JSON.stringify(body)});
    if (d.ok) { closeModal(); toast('Tenant updated.'); state.selectedIdx=idx; state.tab='tenant-detail'; render(); }
    else $('#editErr').textContent = d.message || d.error || 'Could not save.';
  } finally {
    _endAction(_k);
  }
}
async function moveOutTenant(idx) {
  const _k = `moveOutTenant:${idx}`;
  if (!_beginAction(_k)) return;
  try {
    if (!confirm('Mark this tenant as moved out? Their unit becomes free and their history is kept, just archived.')) return;
    const d = await api('/api/tenants/'+idx+'/move-out', {method:'POST'});
    if (d.ok) { toast('Marked as moved out.'); state.selectedIdx=idx; state.tab='tenant-detail'; render(); }
    else toast(d.error || 'Could not update.');
  } finally {
    _endAction(_k);
  }
}

// ── ADD OLD DATA (backfill pre-existing history) ────────────────────
// Ported from the desktop app's Add Old Data dialog: batch in one or
// more historical payment/deposit records for a tenant who was already
// renting before this data was tracked here, then optionally set their
// current status/due date once every row is in.
let _oldDataRows = [];
function openAddOldData(idx) {
  _oldDataRows = [];
  openModal(`
    <h2>Add Old Data</h2>
    <div class="sub" style="font-size:12px;color:var(--muted);margin-top:-8px;margin-bottom:10px;">Backfill this tenant's pre-existing rental history — payments or instalments from before this app was tracking them. Every date must be in the past.</div>
    <div class="row"><div><label class="field">Type</label>
      <select id="od_type"><option value="payment">Full Payment</option><option value="deposit">Instalment / Deposit</option></select></div>
      <div><label class="field">Amount (UGX)</label><input id="od_amount" inputmode="numeric"></div>
    </div>
    <div class="row"><div><label class="field">Date Paid</label><input id="od_date" type="date"></div>
      <div><label class="field">Covers From</label><input id="od_from" type="date"></div>
    </div>
    <div class="row"><div><label class="field">Months Covered</label><input id="od_months" inputmode="numeric" value="1"></div>
      <div><label class="field">Note</label><input id="od_note" placeholder="Optional"></div>
    </div>
    <button class="btn btn-ghost btn-full" style="margin-top:6px;" onclick="addOldDataRow()">＋ Add This Record</button>
    <div id="od_rows" style="margin-top:12px;"></div>
    <hr class="sep">
    <div class="sub" style="font-size:11px;color:var(--muted);margin-bottom:6px;">Optional — set once every record above is in:</div>
    <div class="row"><div><label class="field">Current Status</label>
      <select id="od_status"><option value="">Leave as-is</option><option value="Confirmed">Confirmed</option><option value="Pending">Pending</option></select></div>
      <div><label class="field">Current Due Date</label><input id="od_due_date" type="date"></div>
    </div>
    <div class="err" id="oldDataErr"></div>
    <button class="btn btn-primary btn-full" style="margin-top:12px;" onclick="submitOldData(${idx})">Save All Records</button>
  `);
  renderOldDataRows();
}
function addOldDataRow() {
  const type = $('#od_type').value;
  const amount = $('#od_amount').value;
  const date = $('#od_date').value;
  const from = $('#od_from').value;
  const months = $('#od_months').value;
  const note = $('#od_note').value;
  if (!date) { toast('Enter the date paid.'); return; }
  if (!amount) { toast('Enter an amount.'); return; }
  _oldDataRows.push({type, amount, date, from_date: from, months, note});
  $('#od_amount').value = ''; $('#od_date').value = ''; $('#od_from').value = ''; $('#od_note').value = '';
  renderOldDataRows();
}
function removeOldDataRow(i) { _oldDataRows.splice(i,1); renderOldDataRows(); }
function renderOldDataRows() {
  const el = $('#od_rows');
  if (!el) return;
  if (!_oldDataRows.length) { el.innerHTML = `<div class="sub" style="font-size:12px;color:var(--muted);">No records added yet.</div>`; return; }
  el.innerHTML = _oldDataRows.map((r,i) => `
    <div class="txn-item">
      <div>
        <div style="font-size:13px;">${r.type==='deposit'?'Instalment':'Full Payment'}</div>
        <div class="sub" style="font-size:11.5px;color:var(--muted);">${escapeHtml(r.date)}${r.months?` · ${r.months} mo`:''}</div>
      </div>
      <div style="display:flex;align-items:center;gap:10px;">
        <span class="amt">${fmt(r.amount)}</span>
        <button class="icon-btn" style="width:30px;height:30px;background:var(--card-2);color:var(--danger);font-size:13px;" onclick="removeOldDataRow(${i})">✕</button>
      </div>
    </div>`).join('');
}
async function submitOldData(idx) {
  const _k = `submitOldData:${idx}`;
  if (!_beginAction(_k)) return;
  try {
    if (!_oldDataRows.length) { $('#oldDataErr').textContent = 'Add at least one record first.'; return; }
    const final_state = {};
    if ($('#od_status').value) final_state.status = $('#od_status').value;
    if ($('#od_due_date').value) final_state.due_date = $('#od_due_date').value;
    const d = await api(`/api/tenants/${idx}/old-data`, {method:'POST', body: JSON.stringify({records: _oldDataRows, final_state})});
    if (d.ok) { closeModal(); toast(`Added ${d.added} old record(s).`); state.selectedIdx=idx; state.tab='tenant-detail'; render(); }
    else $('#oldDataErr').textContent = d.error || 'Could not save.';
  } finally {
    _endAction(_k);
  }
}
async function deleteTenant(idx) {
  const _k = `deleteTenant:${idx}`;
  if (!_beginAction(_k)) return;
  try {
    if (!confirm('Permanently delete this tenant? This cannot be undone.')) return;
    const d = await api('/api/tenants/'+idx, {method:'DELETE'});
    if (d.ok) { closeModal(); toast('Tenant deleted.'); state.tab='tenants'; render(); }
  } finally {
    _endAction(_k);
  }
}

// ── Month picker (shared by Pay Rent + Record Instalment) ──────────────
// Months are always ticked contiguously starting from the tenant's first
// still-open month: ticking a row ticks it and every open row before it;
// unticking a row unticks it and every open row after it. Already-paid
// (cleared) months are shown for context but can't be selected again.
let _mp = { open: [], cleared: [], selected: 0, pending: 0, rent: 0, prepaid: 0 };

async function loadMonthPicker(idx) {
  const d = await api(`/api/tenants/${idx}/months`);
  // Rent now comes straight from this same response (see get_tenant_months
  // server-side) instead of state.tenants[idx].rent -- that array was never
  // actually populated anywhere in this file, so rent was always 0 here,
  // which made the Amount field always compute to 0 regardless of how many
  // months were ticked.
  const rent = d.rent || 0;
  _mp = { open: d.open || [], cleared: d.cleared || [], selected: 0, pending: 0, rent, prepaid: d.deposit_paid || 0 };
}

function monthPickerHtml(locked) {
  if (locked) {
    // Used when an instalment balance is already outstanding on the
    // current month: no month-selection dropdown at all -- the current
    // month is the only thing payable until its balance clears, so there's
    // nothing to choose. Fields still populate live via updateMonthPickerFields.
    return `
      <label class="field">Month</label>
      <div class="month-picker-control" style="cursor:default;">
        <span id="mp_summary" style="color:var(--ink);"></span>
      </div>
      <div class="mp-summary-row">
        <div class="mp-summary-box">
          <div class="mp-summary-label">Months</div>
          <div class="mp-summary-value" id="mp_months_val">1</div>
        </div>
        <div class="mp-summary-box">
          <div class="mp-summary-label">Amount</div>
          <div class="mp-summary-value" id="mp_amount_val">${fmt(0)}</div>
        </div>
      </div>
      <div class="sub" style="color:var(--muted);font-size:12px;margin-top:6px;">Clear this month's balance before you can select another period.</div>
      <div class="sub" id="mp_credit_note" style="color:var(--muted);font-size:12px;margin-top:2px;margin-bottom:8px;display:none;"></div>`;
  }
  return `
    <label class="field">Month(s)</label>
    <div class="month-picker-control" id="mp_control" onclick="toggleMonthPickerPanel()">
      <span id="mp_summary" style="color:var(--muted);">Select month(s)</span>
      <span class="mp-caret">▾</span>
    </div>
    <div class="month-picker-panel" id="mp_panel" style="display:none;">
      <div class="month-picker-list" id="mp_list"></div>
      <div class="mp-summary-row">
        <div class="mp-summary-box">
          <div class="mp-summary-label">Months</div>
          <div class="mp-summary-value" id="mp_months_val">0</div>
        </div>
        <div class="mp-summary-box">
          <div class="mp-summary-label">Amount</div>
          <div class="mp-summary-value" id="mp_amount_val">${fmt(0)}</div>
        </div>
      </div>
      <div class="sub" id="mp_credit_note" style="color:var(--muted);font-size:12px;margin-top:-6px;margin-bottom:8px;display:none;"></div>
      <div class="month-picker-actions">
        <button class="btn btn-ghost" type="button" style="flex:1;" onclick="cancelMonthPicker()">Cancel</button>
        <button class="btn btn-primary" type="button" style="flex:1;" onclick="confirmMonthPicker()">OK</button>
      </div>
    </div>`;
}

function renderMonthPickerList() {
  const list = $('#mp_list');
  // Locked mode has no #mp_list (no dropdown to render) -- still fall
  // through to updateMonthPickerFields() so Months/Amount/Deposit/Balance
  // populate immediately instead of only after list markup exists.
  if (list) {
    const clearedHtml = _mp.cleared.map(m => `
      <div class="month-row cleared">
        <input type="checkbox" checked disabled>
        <span class="month-label">${escapeHtml(m.label)}</span>
        <span class="month-tag">Paid</span>
      </div>`).join('');
    const openHtml = _mp.open.map((m, i) => `
      <div class="month-row open ${i < _mp.pending ? 'ticked' : ''}" onclick="toggleMonthRow(${i})">
        <input type="checkbox" ${i < _mp.pending ? 'checked' : ''}>
        <span class="month-label">${escapeHtml(m.label)}</span>
      </div>`).join('');
    list.innerHTML = clearedHtml + openHtml;
  }
  updateMonthPickerFields();
}

function mpNetAmount() {
  // The actual amount still to be cleared for the ticked months -- not a
  // flat rent*months -- if a deposit/instalment has already been paid
  // toward the current open month, that credit is subtracted here so the
  // figure matches what's really left to collect.
  const gross = (_mp.rent || 0) * _mp.pending;
  const credit = Math.min(_mp.prepaid || 0, gross);
  return Math.max(0, gross - credit);
}

function updateMonthPickerFields() {
  // Live "Months" / "Amount" summary inside the picker panel itself, so
  // the total is visible BEFORE confirming with OK -- updates on every
  // tick/untick, not just after confirming the selection.
  const monthsVal = $('#mp_months_val');
  const amountVal = $('#mp_amount_val');
  const creditNote = $('#mp_credit_note');
  const net = mpNetAmount();
  const credit = Math.min(_mp.prepaid || 0, (_mp.rent || 0) * _mp.pending);
  if (monthsVal) monthsVal.textContent = String(_mp.pending);
  if (amountVal) amountVal.textContent = fmt(net);
  // The Deposit modal has its own editable "Deposit" field (Pay Rent
  // doesn't -- it always charges the full computed amount). It was
  // starting out blank/zero instead of pre-filled with the computed
  // amount = rent * months. Keep it in sync with the month picker here,
  // but only while the user hasn't typed their own custom deposit --
  // once they touch the field (see the 'input' listener in
  // openDepositModal), this stops overwriting what they entered, since
  // a deposit is often a partial amount rather than the full total.
  const depositEl = $('#d_deposit');
  if (depositEl && !depositEl.dataset.userEdited) {
    depositEl.value = net > 0 ? String(net) : '';
  }
  updateDepositBalance();
  if (creditNote) {
    if (credit > 0) {
      creditNote.style.display = 'block';
      creditNote.textContent = `${fmt(credit)} already paid toward this is credited here.`;
    } else {
      creditNote.style.display = 'none';
    }
  }
}

function updateDepositBalance() {
  // "Balance" = Amount due for the ticked months minus whatever the admin
  // types into "Deposit" -- recomputed on every keystroke and every month
  // tick/untick so it always reflects amount - deposit.
  const depositEl = $('#d_deposit');
  const balanceEl = $('#d_balance');
  if (!depositEl || !balanceEl) return;
  const net = mpNetAmount();
  const deposit = parseFloat(depositEl.value) || 0;
  const balance = Math.max(0, net - deposit);
  balanceEl.textContent = fmt(balance);
}

function toggleMonthRow(i) {
  // Ticking row i ticks it and every open month before it (keeps the
  // selection contiguous from the first open month); unticking it drops
  // it and every open month after it, for the same reason.
  _mp.pending = (i < _mp.pending) ? i : i + 1;
  renderMonthPickerList();
}

function toggleMonthPickerPanel() {
  const panel = $('#mp_panel');
  if (panel.style.display !== 'none') { panel.style.display = 'none'; return; }
  _mp.pending = _mp.selected;
  renderMonthPickerList();
  panel.style.display = 'block';
}

function cancelMonthPicker() {
  _mp.pending = _mp.selected;
  $('#mp_panel').style.display = 'none';
}

function confirmMonthPicker() {
  _mp.selected = _mp.pending;
  $('#mp_panel').style.display = 'none';
  updateMonthPickerSummary();
}

function updateMonthPickerSummary() {
  const el = $('#mp_summary');
  if (!el) return;
  if (_mp.selected <= 0) {
    el.textContent = 'Select month(s)';
    el.style.color = 'var(--muted)';
  } else {
    const first = _mp.open[0], last = _mp.open[_mp.selected - 1];
    el.textContent = abbrevPeriod(first.from, last.to);
    el.style.color = 'var(--ink)';
  }
  const btn = document.getElementById('txnSubmitBtn');
  if (btn) btn.disabled = _mp.selected <= 0;
}

async function openPaymentModal(idx) {
  await loadMonthPicker(idx);
  openModal(`
    <h2>Record Payment</h2>
    <div class="desc">Select the month(s) being paid in full — the due date moves forward by however many months you pick.</div>
    ${monthPickerHtml()}
    <div class="err" id="payErr"></div>
    <button class="btn btn-primary btn-full" style="margin-top:12px;" id="txnSubmitBtn" disabled onclick="submitPayment(${idx})">✓ Record Payment</button>
  `);
  renderMonthPickerList();
}
async function submitPayment(idx) {
  const _k = `submitPayment:${idx}`;
  if (!_beginAction(_k)) return;
  try {
    if (_mp.selected <= 0) { $('#payErr').textContent = 'Select at least one month.'; return; }
    const d = await api(`/api/tenants/${idx}/payment`, {method:'POST', body: JSON.stringify({months: _mp.selected})});
    if (d.ok) { closeModal(); toast(`${fmt(d.result.amount)} recorded. New due date: ${d.result.due_date}`); state.selectedIdx=idx; state.tab='tenant-detail'; render(); }
    else $('#payErr').textContent = d.error || 'Could not record payment.';
  } finally {
    _endAction(_k);
  }
}

async function openDepositModal(idx) {
  await loadMonthPicker(idx);
  // If there's already a partial balance on the current month (an
  // instalment plan mid-way through), lock the picker to that one month --
  // no dropdown, no picking further months -- until it clears to zero.
  const locked = (_mp.prepaid || 0) > 0;
  if (locked) { _mp.pending = 1; _mp.selected = 1; }
  openModal(`
    <h2>Record Instalment / Deposit</h2>
    <div class="desc">${locked ? 'Continue clearing the balance for the current month below.' : "Partial payments accumulate toward the selected month(s)' rent."}</div>
    ${monthPickerHtml(locked)}
    <label class="field">Deposit (UGX)</label><input id="d_deposit" inputmode="numeric">
    <label class="field">Balance (UGX)</label>
    <div class="month-picker-control" style="cursor:default;">
      <span id="d_balance" style="font-family:var(--font-mono);color:var(--ink);">${fmt(0)}</span>
    </div>
    <div class="err" id="depErr"></div>
    <button class="btn btn-primary btn-full" style="margin-top:12px;" id="txnSubmitBtn" ${locked ? '' : 'disabled'} onclick="submitDeposit(${idx})">＋ Record Instalment</button>
  `);
  renderMonthPickerList();
  updateMonthPickerSummary();
  const depositEl = $('#d_deposit');
  if (depositEl) {
    depositEl.addEventListener('input', () => { depositEl.dataset.userEdited = '1'; updateDepositBalance(); });
  }
}
async function submitDeposit(idx) {
  const _k = `submitDeposit:${idx}`;
  if (!_beginAction(_k)) return;
  try {
    if (_mp.selected <= 0) { $('#depErr').textContent = 'Select at least one month.'; return; }
    const amount = $('#d_deposit').value;
    const d = await api(`/api/tenants/${idx}/deposit`, {method:'POST', body: JSON.stringify({months: _mp.selected, amount})});
    if (d.ok) {
      closeModal();
      toast(d.result.cleared ? `Cleared! ${fmt(d.result.amount)} recorded.` : `${fmt(d.result.amount)} recorded, ${fmt(d.result.new_balance)} left.`);
      state.selectedIdx=idx; state.tab='tenant-detail'; render();
    } else $('#depErr').textContent = d.error || 'Could not record deposit.';
  } finally {
    _endAction(_k);
  }
}

function openClearArrears(idx, due) {
  openModal(`
    <h2>Clear Rent Increase Arrears</h2>
    <div class="desc">Outstanding: <b>${fmt(due)}</b>. This is separate from rent and does not affect the due date.</div>
    <label class="field">Method</label>
    <select id="a_method"><option value="Full">Pay in Full</option><option value="Deposit">Partial (Deposit)</option></select>
    <label class="field">Amount (UGX)</label><input id="a_amount" value="${Math.round(due)}">
    <div class="err" id="arrErr"></div>
    <button class="btn btn-primary btn-full" style="margin-top:8px;" onclick="submitArrears(${idx})">Record Payment</button>
  `);
  $('#a_method').addEventListener('change', e => { if (e.target.value==='Full') $('#a_amount').value = Math.round(due); });
}
async function submitArrears(idx) {
  const _k = `submitArrears:${idx}`;
  if (!_beginAction(_k)) return;
  try {
    const method = $('#a_method').value, amount = $('#a_amount').value;
    const d = await api(`/api/tenants/${idx}/arrears`, {method:'POST', body: JSON.stringify({method, amount})});
    if (d.ok) { closeModal(); toast('Arrears payment recorded.'); state.selectedIdx=idx; state.tab='tenant-detail'; render(); }
    else $('#arrErr').textContent = d.error || 'Could not record.';
  } finally {
    _endAction(_k);
  }
}

// ── UNITS ────────────────────────────────────────────────────────────
function unitRowHtml(u) {
  return `
    <div class="tenant-row" style="cursor:default;">
      <div class="avatar">${svgIcon('building',18,'currentColor')}</div>
      <div class="meta">
        <div class="name">${escapeHtml(u.name)} ${u.pending_rent_increase?`<span style="font-size:11px;color:var(--warn);font-weight:700;">↑ scheduled</span>`:''}</div>
        <div class="sub">${fmt(u.rent)}/mo · ${u.occupant ? 'Occupied: '+escapeHtml(u.occupant) : 'Vacant'}${u.location?' · '+escapeHtml(u.location):''}</div>
      </div>
      <div style="display:flex;gap:6px;">
        <button class="icon-btn" style="background:var(--teal-soft);color:var(--teal-deep);font-size:13px;width:32px;height:32px;" onclick="openEditUnit('${encodeURIComponent(u.name)}')">${svgIcon('edit',14)}</button>
        <button class="icon-btn" style="background:var(--teal-soft);color:var(--teal-deep);font-size:13px;width:32px;height:32px;" onclick="openIncreaseRent('${encodeURIComponent(u.name)}', ${u.rent})">↑</button>
      </div>
    </div>`;
}
async function renderUnits() {
  const d = await api('/api/units');
  $('#headerSub').textContent = `${d.units.length} units`;
  const unitsEmptyMsg = d.no_cache
    ? `<div class="empty"><div class="big">📴</div>You're offline and this hasn't loaded before, so there's nothing cached to show yet.</div>`
    : `<div class="empty"><div class="big">${svgIcon('building',40)}</div>No units yet.</div>`;
  if (!d.units.length) {
    $('#main').innerHTML = `
      <button class="btn btn-primary btn-full" style="margin-bottom:14px;" onclick="openAddUnit()">＋ Add Unit</button>
      <div class="card" style="padding:4px 12px;">${unitsEmptyMsg}</div>
    `;
    return;
  }
  // Split into Occupied / Vacant sections rather than one flat list, so
  // it's immediately clear at a glance how many units are free.
  const occupied = d.units.filter(u => u.occupant);
  const vacant = d.units.filter(u => !u.occupant);
  const section = (title, color, list) => list.length ? `
    <div class="section-title" style="display:flex;align-items:center;gap:8px;margin:${title==='Occupied'?'0':'18px'} 0 6px;">
      <span>${title}</span>
      <span style="background:${color};color:#fff;border-radius:999px;padding:1px 9px;font-size:12px;font-weight:700;">${list.length}</span>
    </div>
    <div class="card" style="padding:4px 12px;">${list.map(unitRowHtml).join('')}</div>` : '';
  $('#main').innerHTML = `
    <button class="btn btn-primary btn-full" style="margin-bottom:14px;" onclick="openAddUnit()">＋ Add Unit</button>
    ${section('Occupied', 'var(--teal-deep)', occupied)}
    ${section('Vacant', 'var(--muted)', vacant)}
  `;
}
function openAddUnit() {
  openModal(`
    <h2>Add Unit</h2>
    <label class="field">Unit ID</label><input id="u_name" placeholder="e.g. A1">
    <label class="field">Monthly Rent (UGX)</label><input id="u_rent" inputmode="numeric">
    <label class="field">Location</label><input id="u_location">
    <div class="err" id="unitErr"></div>
    <button class="btn btn-primary btn-full" style="margin-top:8px;" onclick="submitAddUnit()">Save Unit</button>
  `);
}
async function submitAddUnit() {
  if (!_beginAction('submitAddUnit')) return;
  try {
    const body = {name: $('#u_name').value, rent: $('#u_rent').value, location: $('#u_location').value};
    const d = await api('/api/units', {method:'POST', body: JSON.stringify(body)});
    if (d.ok) { closeModal(); toast('Unit added.'); renderUnits(); } else $('#unitErr').textContent = d.error;
  } finally {
    _endAction('submitAddUnit');
  }
}
async function openEditUnit(nameEnc) {
  const name = decodeURIComponent(nameEnc);
  const d = await api('/api/units');
  const u = d.units.find(x => x.name===name);
  openModal(`
    <h2>Edit Unit — ${escapeHtml(name)}</h2>
    <label class="field">Monthly Rent (UGX)</label><input id="eu_rent" value="${u.rent}">
    <label class="field">Location</label><input id="eu_location" value="${escapeHtml(u.location||'')}">
    <div class="err" id="euErr"></div>
    <button class="btn btn-primary btn-full" style="margin-top:8px;" onclick="submitEditUnit('${nameEnc}')">Save Changes</button>
    <button class="btn btn-danger btn-full" style="margin-top:10px;" onclick="deleteUnit('${nameEnc}')">Delete Unit</button>
  `);
}
async function submitEditUnit(nameEnc) {
  const _k = `submitEditUnit:${nameEnc}`;
  if (!_beginAction(_k)) return;
  try {
    const name = decodeURIComponent(nameEnc);
    const body = {rent: $('#eu_rent').value, location: $('#eu_location').value};
    const d = await api('/api/units/'+encodeURIComponent(name), {method:'PUT', body: JSON.stringify(body)});
    if (d.ok) { closeModal(); toast('Unit updated.'); renderUnits(); } else $('#euErr').textContent = d.error;
  } finally {
    _endAction(_k);
  }
}
async function deleteUnit(nameEnc) {
  const _k = `deleteUnit:${nameEnc}`;
  if (!_beginAction(_k)) return;
  try {
    const name = decodeURIComponent(nameEnc);
    if (!confirm(`Permanently remove unit '${name}'?`)) return;
    const d = await api('/api/units/'+encodeURIComponent(name), {method:'DELETE'});
    if (d.ok) { closeModal(); toast('Unit removed.'); renderUnits(); }
  } finally {
    _endAction(_k);
  }
}
function openIncreaseRent(nameEnc, currentRent) {
  const name = decodeURIComponent(nameEnc);
  const months = [];
  const base = new Date(); base.setDate(1);
  for (let i=0;i<24;i++) {
    const m = new Date(base.getFullYear(), base.getMonth()+i, 1);
    const val = m.toISOString().slice(0,10);
    const lbl = m.toLocaleString('default',{month:'short', year:'numeric'});
    months.push(`<option value="${val}" ${i===1?'selected':''}>${lbl}</option>`);
  }
  openModal(`
    <h2>Increase Rent — ${escapeHtml(name)}</h2>
    <div class="desc">Current rent: ${fmt(currentRent)}/month. The occupying tenant's billed rent updates automatically that month.</div>
    <label class="field">New Monthly Rent (UGX)</label><input id="ir_rent">
    <label class="field">Effective From</label><select id="ir_month">${months.join('')}</select>
    <div class="err" id="irErr"></div>
    <button class="btn btn-primary btn-full" style="margin-top:8px;" onclick="submitIncreaseRent('${nameEnc}')">Confirm Increase</button>
  `);
}
async function submitIncreaseRent(nameEnc) {
  const _k = `submitIncreaseRent:${nameEnc}`;
  if (!_beginAction(_k)) return;
  try {
    const name = decodeURIComponent(nameEnc);
    const body = {new_rent: $('#ir_rent').value, effective_month: $('#ir_month').value};
    const d = await api('/api/units/'+encodeURIComponent(name)+'/increase-rent', {method:'POST', body: JSON.stringify(body)});
    if (d.ok) { closeModal(); toast('Rent increase scheduled.'); renderUnits(); } else $('#irErr').textContent = d.error;
  } finally {
    _endAction(_k);
  }
}

// ── ALERTS ───────────────────────────────────────────────────────────
async function renderAlerts() {
  const d = await api('/api/alerts');
  $('#headerSub').textContent = `${d.alerts.length} tenant(s) to watch`;
  const alertsEmptyMsg = d.no_cache
    ? `<div class="empty"><div class="big">📴</div>You're offline and this hasn't loaded before, so there's nothing cached to show yet.</div>`
    : `<div class="empty"><div class="big">✅</div>No alerts. Everyone's paid up.</div>`;
  const rows = d.alerts.map(tenantRowHtml).join('') || alertsEmptyMsg;
  $('#main').innerHTML = `<div class="section-title">Overdue &amp; Upcoming</div><div class="card" style="padding:4px 12px;">${rows}</div>`;
}

// ── SETTINGS ─────────────────────────────────────────────────────────
function renderMoreMenu() {
  $('#headerSub').textContent = 'More';
  $('#main').innerHTML = `
    <div class="card" style="padding:0;overflow:hidden;">
      <div class="menu-row" onclick="switchTab('history')">
        <span class="menu-icon">📜</span>
        <div class="menu-text">
          <div class="menu-title">Transaction History</div>
          <div class="menu-sub">Every tenant's payment &amp; deposit ledger</div>
        </div>
        <span class="menu-chevron">›</span>
      </div>
      <div class="menu-row" onclick="switchTab('settings')">
        <span class="menu-icon">${svgIcon('gear',18)}</span>
        <div class="menu-text">
          <div class="menu-title">Settings</div>
          <div class="menu-sub">App, security, reports &amp; data</div>
        </div>
        <span class="menu-chevron">›</span>
      </div>
    </div>
  `;
}

async function renderHistory() {
  const q = state.historyQ || '';
  $('#headerSub').textContent = 'Transaction History';
  const [d, dash] = await Promise.all([
    api('/api/history' + (q ? ('?q=' + encodeURIComponent(q)) : '')),
    api('/api/dashboard'),
  ]);
  const tenants = d.tenants || [];

  const incomeCard = `<div class="card" style="background:var(--teal-soft2);border-color:var(--teal-soft);">
    <div class="section-title" style="margin-top:0;">Monthly Income — ${escapeHtml(dash.month_name || '')}</div>
    <div style="font-family:var(--font-mono);font-size:24px;font-weight:600;color:var(--teal-deep);">${fmt(dash.month_income || 0)}</div>
    <div style="display:flex;gap:16px;margin-top:8px;font-size:12px;">
      <div><b style="color:var(--good);">${fmt(dash.full_payment_total || 0)}</b><div style="color:var(--muted);">Full</div></div>
      <div><b style="color:var(--teal-deep);">${fmt(dash.deposit_total || 0)}</b><div style="color:var(--muted);">Deposits</div></div>
      <div><b style="color:var(--danger);">${fmt(dash.cancelled_total || 0)}</b><div style="color:var(--muted);">Cancelled</div></div>
    </div>
  </div>`;

  const monthlyCard = `<div class="section-title">Monthly Transactions</div>
  <div class="card">
    <div class="sub" style="color:var(--muted);font-size:12.5px;margin-bottom:10px;">
      Generate a summary of payments and deposits for one specific month.
    </div>
    <div style="display:flex;gap:8px;">
      <div class="select-native-wrap"><select id="mrMonth"></select></div>
      <div class="select-native-wrap"><select id="mrYear"></select></div>
    </div>
    <button class="btn btn-primary btn-full" style="margin-top:10px;" onclick="generateMonthlyReport()">Generate</button>
    <div id="mrResults" style="margin-top:12px;"></div>
  </div>`;

  const rowsHtml = (t) => {
    if (!t.transactions.length) {
      return `<div class="hist-empty">No transactions recorded yet.</div>`;
    }
    return `<div class="hist-table">
      <div class="hist-hdr">
        <div>Date</div><div>Type</div><div>Amount</div><div>Period</div>
      </div>
      ${t.transactions.map(rec => {
        const cancelled = rec.cancelled;
        const typeStr = (rec.kind === 'deposit' ? 'Deposit' : 'Full') + (cancelled ? '  ✕' : '');
        const rowCls = cancelled ? 'hist-row hist-cancelled' : 'hist-row';
        return `<div class="${rowCls}">
          <div>${escapeHtml(rec.date)}</div>
          <div>${typeStr}</div>
          <div>${fmt(rec.amount)}</div>
          <div class="hist-period">${escapeHtml(abbrevPeriod(rec.from, rec.to))}</div>
        </div>`;
      }).join('')}
    </div>`;
  };

  const cardsHtml = tenants.length
    ? tenants.map(t => `
      <div class="card">
        <div style="display:flex;justify-content:space-between;align-items:flex-start;">
          <div>
            <div style="font-weight:700;font-size:15px;">${escapeHtml(t.name)}</div>
            <div class="sub" style="color:var(--muted);font-size:12.5px;">${svgIcon('building',12)} ${escapeHtml(t.unit)}</div>
          </div>
          <button class="btn btn-ghost" style="padding:6px 10px;font-size:12px;" onclick="openTenant(${t.index})">Profile</button>
        </div>
        <div style="margin-top:10px;font-size:11px;color:var(--muted);text-transform:uppercase;font-weight:700;">Entry Date</div>
        <div style="font-size:14px;font-weight:700;margin-top:2px;">${escapeHtml(t.entry_date)}</div>
        <div style="margin-top:12px;font-size:11px;color:var(--muted);text-transform:uppercase;font-weight:700;">Transaction History</div>
        <div style="margin-top:8px;">${rowsHtml(t)}</div>
      </div>`).join('')
    : `<div class="empty">No tenants match your search.</div>`;

  $('#main').innerHTML = `
    <button class="btn btn-ghost" style="margin-bottom:12px;" onclick="switchTab('more');">← Back</button>
    <div class="searchbar"><span>🔎</span><input id="historySearch" placeholder="Search name, unit or phone…" value="${escapeHtml(q)}"></div>
    ${incomeCard}
    ${monthlyCard}
    ${cardsHtml}
  `;
  $('#historySearch').addEventListener('input', debounce(e => { state.historyQ = e.target.value; renderHistory(); }, 300));
  populateMonthlyPickers();
}

async function renderSettings() {
  $('#headerSub').textContent = 'Settings';
  const theme = getTheme();
  $('#main').innerHTML = `
    <button class="btn btn-ghost" style="margin-bottom:12px;" onclick="switchTab('more');">← Back</button>
    <div class="section-title">Appearance</div>
    <div class="card">
      <div class="sub" style="color:var(--muted);font-size:12.5px;margin-bottom:10px;">
        Choose how Tenant Management looks on this device.
      </div>
      <div style="display:flex;gap:8px;">
        <div class="filter-pill ${theme==='light'?'active':''}" style="flex:1;text-align:center;" onclick="setTheme('light')">☀️ Light</div>
        <div class="filter-pill ${theme==='dark'?'active':''}" style="flex:1;text-align:center;" onclick="setTheme('dark')">🌙 Dark</div>
      </div>
    </div>
    <div class="section-title">Get the App</div>
    <div class="card" id="installCard">
      <div class="sub" style="color:var(--muted);font-size:12.5px;margin-bottom:10px;">
        Save this to your phone's home screen or browser for one-tap access — no need to retype the address each time.
      </div>
      <button class="btn btn-primary btn-full" id="installBtn" onclick="triggerInstall()" style="display:none;">📲 Install App</button>
      <div class="sub" id="installHint" style="color:var(--muted);font-size:11.5px;margin-top:8px;"></div>
    </div>
    <div class="section-title">Reports</div>
    <div class="card">
      <a class="linklike" href="api/export/excel">⬇ Download Excel (.xlsx)</a><br><br>
      <a class="linklike" href="api/export/pdf">⬇ Download PDF Report</a>
    </div>
    <div class="section-title">Danger Zone</div>
    <div class="card">
      <div class="sub" style="color:var(--muted);font-size:12.5px;margin-bottom:10px;">
        Permanently erases every tenant, unit, and transaction shared with the desktop app. A backup of the current data is kept automatically before wiping, and the PC's admin PIN is required.
      </div>
      <button class="btn btn-danger btn-full" onclick="openResetData()">${svgIcon('trash',13)} Reset All Data</button>
    </div>
    <div class="section-title">About</div>
    <div class="card sub" style="color:var(--muted);font-size:12.5px;">
      Tenant Monitoring &amp; Management — Web Edition.<br>Shares data with the desktop app on this PC.<br><br>
      Device naming, connected-device management, and the app lock/PIN are all controlled from the desktop app now, not from here.
    </div>
  `;
  updateInstallUI();
}

// ── DEVICE NAMING (prompted once, right after a QR pairing scan) ──────
// PIN/lock and the connected-devices roster are admin-only now and are
// managed from the desktop app's Settings -> Connect Phone panel instead
// of from here -- a phone can still name ITSELF (device_id is already
// tied to it, so this is just a courtesy so the admin can tell devices
// apart in that PC-side panel), but nothing else about security or the
// device list is editable from the web app.
async function maybePromptDeviceName() {
  try {
    const d = await api('/api/devices');
    const devices = (d && d.devices) || [];
    const mine = devices.find(dev => dev.device_id === DEVICE_ID);
    // custom_label_locked only becomes true once a person has actually
    // typed and saved a name -- unlike `label`, which is never empty
    // (the server always auto-fills it with a detected model name like
    // "iPhone" or the generic "Device"), so checking `label` here used
    // to make this prompt effectively never fire. Runs on every boot,
    // not just right after a fresh QR-code pairing, so a phone that
    // connected before this requirement existed -- or that dismissed/
    // missed it somehow -- is still asked the next time it opens the
    // app, until it actually has a unique name on file.
    if (mine && mine.custom_label_locked) return;
  } catch (e) { return; }
  openModal(`
    <h2>Name This Device</h2>
    <div class="desc">Every connected phone needs its own unique name so the admin -- and this phone itself -- can always tell it apart from any other connected phone.</div>
    <label class="field">Device Name</label>
    <input id="dev_my_label" placeholder="e.g. Mary's iPhone">
    <div class="err" id="devLabelErr"></div>
    <button class="btn btn-primary btn-full" style="margin-top:8px;" onclick="submitDeviceLabel()">Save Name</button>
  `, {dismissible: false});
}
async function submitDeviceLabel() {
  if (!_beginAction('submitDeviceLabel')) return;
  try {
    const label = $('#dev_my_label').value.trim();
    if (!label) { $('#devLabelErr').textContent = 'Please enter a name.'; return; }
    const d = await api('/api/devices/label', {method:'POST', body: JSON.stringify({label})});
    if (d.ok) { closeModal(); toast('Device name saved.'); }
    else $('#devLabelErr').textContent = d.error || 'Could not save name -- that name may already be taken by another connected phone.';
  } finally {
    _endAction('submitDeviceLabel');
  }
}

// ── DANGER ZONE: reset data ──────────────────────────────────────────
function openResetData() {
  openModal(`
    <h2>Reset All Data</h2>
    <div class="desc">This permanently deletes every tenant, unit, and transaction on both this app and the desktop app. A backup of the current data file is saved automatically first. This cannot be undone from here.</div>
    <label class="field">Admin PIN (set on the PC)</label><input id="reset_admin_pin" type="password" inputmode="numeric">
    <div class="err" id="resetErr"></div>
    <button class="btn btn-danger btn-full" style="margin-top:8px;" onclick="submitResetData()">Yes, Reset Everything</button>
  `);
}
async function submitResetData() {
  if (!_beginAction('submitResetData')) return;
  try {
    const adminPin = $('#reset_admin_pin').value.trim();
    if (!adminPin) { $('#resetErr').textContent = 'Enter the admin PIN set on the PC.'; return; }
    if (!confirm('Are you absolutely sure? All tenants and units will be erased.')) return;
    const d = await api('/api/settings/reset', {method:'POST', body: JSON.stringify({admin_pin: adminPin})});
    if (d.ok) { closeModal(); toast('All data reset.'); cacheDeletePrefix('/api/'); switchTab('dashboard'); }
    else $('#resetErr').textContent = d.error || 'Could not reset data.';
  } finally {
    _endAction('submitResetData');
  }
}

const MONTH_NAMES = ['January','February','March','April','May','June','July',
                      'August','September','October','November','December'];

function populateMonthlyPickers() {
  const now = new Date();
  const monthSel = $('#mrMonth'), yearSel = $('#mrYear');
  if (!monthSel || !yearSel) return;
  monthSel.innerHTML = MONTH_NAMES.map((m, i) =>
    `<option value="${i+1}" ${i === now.getMonth() ? 'selected' : ''}>${m}</option>`).join('');
  const startYear = now.getFullYear() - 3;
  let yearsHtml = '';
  for (let y = startYear; y <= now.getFullYear() + 1; y++) {
    yearsHtml += `<option value="${y}" ${y === now.getFullYear() ? 'selected' : ''}>${y}</option>`;
  }
  yearSel.innerHTML = yearsHtml;
}

async function generateMonthlyReport() {
  if (!_beginAction('generateMonthlyReport')) return;
  const month = $('#mrMonth').value, year = $('#mrYear').value;
  const box = $('#mrResults');
  box.innerHTML = `<div class="sub" style="color:var(--muted);">Generating…</div>`;
  try {
    const res = await fetchTimeout(`/api/monthly-report?year=${year}&month=${month}`, {}, 6000);
    const report = await res.json();
    if (report.error) { box.innerHTML = `<div class="sub" style="color:var(--danger);">${escapeHtml(report.error)}</div>`; return; }
    const rowsHtml = report.tenant_rows.length
      ? report.tenant_rows.map(r => `
        <div class="hist-row">
          <div>${escapeHtml(r.name)} <span style="color:var(--muted);">(${escapeHtml(r.unit)})</span></div>
          <div style="text-align:right;font-weight:700;">${fmt(r.pay_active + r.dep_active)}</div>
        </div>`).join('')
      : `<div class="hist-empty">No transactions in ${escapeHtml(report.month_label)}.</div>`;
    box.innerHTML = `
      <div style="font-family:var(--font-mono);font-size:20px;font-weight:600;color:var(--teal-deep);margin-top:6px;">
        ${fmt(report.grand_combined)}
      </div>
      <div style="display:flex;gap:16px;margin:6px 0 12px;font-size:12px;">
        <div><b style="color:var(--good);">${fmt(report.grand_pay)}</b><div style="color:var(--muted);">Full</div></div>
        <div><b style="color:var(--teal-deep);">${fmt(report.grand_dep)}</b><div style="color:var(--muted);">Deposits</div></div>
        <div><b style="color:var(--danger);">${fmt(report.grand_cancelled)}</b><div style="color:var(--muted);">Cancelled</div></div>
      </div>
      ${rowsHtml}
      <a class="linklike" style="display:block;margin-top:12px;" href="api/export/monthly-excel?year=${year}&month=${month}">⬇ Download Monthly Excel (.xlsx)</a>
    `;
  } catch (e) {
    box.innerHTML = `<div class="sub" style="color:var(--danger);">Couldn't generate the report — check the connection and try again.</div>`;
  } finally {
    _endAction('generateMonthlyReport');
  }
}

// ── pull-to-refresh ──────────────────────────────────────────────────
// Only engages when the page is scrolled all the way to the top (so it
// never fights with normal scrolling further down), and re-runs whatever
// tab is currently showing by re-dispatching switchTab, which re-fetches
// from the server (bypassing nothing — /api/* is never cached by the
// service worker, see sw.js above) and re-renders with fresh data.
(function setupPullToRefresh() {
  const main = $('#main');
  const indicator = $('#ptrIndicator');
  if (!main || !indicator) return;
  const THRESHOLD = 70;
  let startY = null, pulling = false, refreshing = false;

  main.addEventListener('touchstart', (e) => {
    if (refreshing || main.scrollTop > 0) { startY = null; return; }
    startY = e.touches[0].clientY;
    pulling = true;
  }, { passive: true });

  main.addEventListener('touchmove', (e) => {
    if (!pulling || startY === null || refreshing) return;
    const dy = e.touches[0].clientY - startY;
    if (dy <= 0) return;
    indicator.classList.add('visible');
    if (dy > THRESHOLD) {
      indicator.classList.add('ready');
      indicator.textContent = '↑ Release to refresh';
    } else {
      indicator.classList.remove('ready');
      indicator.textContent = '↓ Pull to refresh';
    }
  }, { passive: true });

  main.addEventListener('touchend', async (e) => {
    if (!pulling || startY === null || refreshing) { pulling = false; return; }
    const ready = indicator.classList.contains('ready');
    pulling = false;
    startY = null;
    if (!ready) {
      indicator.classList.remove('visible', 'ready');
      return;
    }
    refreshing = true;
    indicator.textContent = '⟳ Refreshing…';
    indicator.classList.add('spinning');
    await pingServer();
    switchTab(state.tab);
    setTimeout(() => {
      indicator.classList.remove('visible', 'ready', 'spinning');
      refreshing = false;
    }, 400);
  });
})();

// ── boot ─────────────────────────────────────────────────────────────
function hideSplash() {
  const s = $('#splashScreen');
  if (!s || s.dataset.hidden) return;
  s.dataset.hidden = '1';
  if (window.matchMedia('(prefers-reduced-motion: reduce)').matches) {
    s.classList.add('hidden');
    setTimeout(() => { if (s) s.style.display = 'none'; }, 300);
    return;
  }
  // Stage 1: icon lunges toward the viewer and fades (.exiting, 550ms).
  // Stage 2, started partway through stage 1 so the two read as one
  // continuous motion rather than two separate steps: a circular wipe
  // opens from the same center point, revealing the now-ready dashboard
  // underneath (.wiping, 500ms). See the matching @keyframes above.
  s.classList.add('exiting');
  setTimeout(() => { if (s) s.classList.add('wiping'); }, 300);
  setTimeout(() => { if (s) s.style.display = 'none'; }, 850);
}
// Safety net: never leave the splash stuck on screen if something above
// goes wrong (e.g. an unexpected error before init() reaches a branch
// that hides it).
setTimeout(hideSplash, 8000);

async function boot() {
  booted = true;
  loadCloudConfig();
  // Wait for the dashboard's own data to actually be ready before the
  // splash disappears, rather than hiding it as soon as the lock-status
  // check alone comes back.
  await switchTab('dashboard');
  hideSplash();
  maybePromptDeviceName();
}
async function init() {
  let ls;
  try {
    const res = await fetchTimeout('/api/lock-status', {headers: {'X-Device-Id': DEVICE_ID}}, 3500);
    ls = await res.json();
    adoptCanonicalDeviceId(ls && ls.canonical_device_id);
    if (ls && (ls.kicked || ls.disconnecting || ls.pending_approval || ls.device_limit_reached)) {
      enterBlockedState(ls);
      return;
    }
    cacheSet('/api/lock-status', ls);
    setOnline(true);
  } catch (err) {
    setOnline(false);
    ls = cacheGet('/api/lock-status') || { pin_set: false, unlocked: true };
  }
  if (ls.pin_set && !ls.unlocked) { showLock(); }
  else { hideLock(); boot(); }
  updateSyncBadge();
}
init();
</script>
</body>
</html>
"""


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    no_browser = os.environ.get("RM_NO_BROWSER") == "1"
    print(f"\n  {APP_NAME} -- web edition (single file)")
    print(f"  Data file: {DATA_FILE}")
    print(f"  Local:  http://127.0.0.1:{port}")
    if CLOUD_MODE:
        print("  Running in CLOUD_MODE -- reachable at this service's public URL; "
              "phones pair to it directly via the QR code the desktop app shows "
              "under Settings -> Connect Phone.\n")
    else:
        print("  Not running in CLOUD_MODE (no DATABASE_URL) -- this is a local "
              "test run only. Deploy with CLOUD_MODE=1 and a DATABASE_URL for "
              "phones to be able to reach it.\n")

    if not no_browser:
        # Standalone run (person double-clicked/ran this file themselves) —
        # open the app on this PC as a convenience.
        def _open_local_page():
            try:
                webbrowser.open(f"http://127.0.0.1:{port}")
            except Exception:
                pass
        threading.Timer(1.0, _open_local_page).start()
    # else: launched from the desktop app's Settings → Connect Phone, which
    # already shows its own QR code in-window — nothing should open on the PC.

    app.run(host="0.0.0.0", port=port, debug=False)
